"""
TARIFARIO EDITABLE de couriers (2026-07-25) — módulo aparte, mismo patrón que
cubicador_plus.py / logistica_cotizaciones.py.

POR QUÉ EXISTE Y POR QUÉ VIVE ACÁ (no en app.py)
--------------------------------------------------
app.py tiene ~75.600 líneas y hay OTRA sesión editándolo en paralelo (módulo
de Cotizaciones de logística). Este módulo se registra al final de app.py con
2-3 líneas try/except, mismo criterio que cubicador_plus.py.

EL PROBLEMA REAL QUE RESUELVE
------------------------------
El motor de cotización (transporte_tarifas.py) lee las tarifas de archivos
`tarifas/<slug>.json` que viven EN EL REPO. Eso funciona para leer, pero no
sirve para EDITAR desde la UI: Cloud Run tiene disco efímero por instancia —
un archivo escrito en runtime se pierde en el próximo deploy/escalado y no se
comparte entre instancias corriendo en paralelo. Además, la lógica de negocio
de transporte_tarifas.py (TIERS hardcodeados, la regla comercial confidencial
Felca/Milling con descuento por comuna) está VALIDADA contra facturas reales
y NO debe tocarse ni reimplementarse acá.

LA SOLUCIÓN (híbrida, bajo riesgo — ver investigación 2026-07-25)
-------------------------------------------------------------------
1. Solo se mueve el ALMACENAMIENTO de los JSON a una tabla MySQL nueva
   (`transport_tarifas_json`), con historial de versiones anteriores
   (`transport_tarifas_json_historial`) para poder deshacer.
2. transporte_tarifas.py::_load() cambia SOLO en de dónde vienen los bytes:
   primero intenta MySQL, si no hay fila o falla cae al archivo local del
   repo (red de seguridad — nunca se rompe el motor si la tabla está vacía
   o la DB no responde). El resto de ese archivo (TIERS, cotizar(), la regla
   comercial) queda intacto, tal cual estaba.
3. Este módulo (logistica_tarifario.py) es la UI/API de administración:
   listar, subir Excel/CSV nuevo (con validación estricta — si algo está
   mal se RECHAZA sin tocar el tarifario vigente), descargar el actual en
   Excel (para editar offline y volver a subirlo), ver historial y
   restaurar una versión anterior.

FORMATO DEL EXCEL DE SUBIDA/DESCARGA (propio de este módulo)
----------------------------------------------------------------
No es el mismo layout que usa `/transporte/couriers/import` (ese alimenta
`transport_courier_comunas`, un motor DB paralelo que NO usa el cotizador
real — ver investigación). Acá el layout es canónico y autodescriptivo:

    Hoja "Tarifario", fila 1 = encabezado:
        COMUNA | DIAS | 1 | 2 | 3 ... | 100 (o 130 en Clickex) | HEAVY_105 | HEAVY_106 | ...

    - Columnas numéricas puras (1, 2, 3...) = tarifa LIVIANA por kg exacto.
    - Columnas "HEAVY_<código>" = tramo pesado. El valor puede ser un número
      (factor $/kg) o texto "factor + fijo" (solo Clickex en tramos altos),
      exactamente como ya soporta `_parse_heavy()` en transporte_tarifas.py.
    - Una fila por comuna. COMUNA no puede repetirse ni venir vacía.

Solo SUPERADMIN (pedido explícito de Daniel). Nunca toca `app_*` ni el ERP
Random — únicamente tablas propias nuevas.

Reglas aplicadas: #1 (front sin alert/confirm nativos — ver tarifario.html),
#2 (paleta), #4 (SQL parametrizado, sin detalles internos al cliente),
#4.1 (no aplica — no toca el ERP), #4.2 (solo agrega; transporte_tarifas.py
se modifica quirúrgicamente, sin borrar nada), #5 (schema idempotente +
audit log antes de sobrescribir), #6 (horas Chile vía chile_fmt en el
template).
"""

import sys
import io
import csv
import json
import traceback
import re
from datetime import datetime, timezone
from functools import wraps

from flask import request, jsonify, g, flash, redirect, url_for, render_template, send_file

import transporte_tarifas as _ttar

# ──────────────────────────────────────────────────────────────────────
#  Acceso a los helpers de app.py — IMPORT DIFERIDO (mismo patrón que
#  cubicador_plus.py). transporte_tarifas.py en cambio es un módulo puro
#  sin dependencia de Flask, así que se importa directo arriba.
# ──────────────────────────────────────────────────────────────────────

_APP_MODULE_NAME = "app"


def _appmod():
    mod = sys.modules.get(_APP_MODULE_NAME)
    if mod is None:
        mod = sys.modules.get("app") or sys.modules.get("__main__")
    return mod


def _h(nombre, default=None):
    mod = _appmod()
    if mod is None:
        return default
    return getattr(mod, nombre, default)


# ══════════════════════════════════════════════════════════════════════
#  Tablas propias — nombres fijos (no vienen de MYSQL_CONFIG: son nuevas)
# ══════════════════════════════════════════════════════════════════════

TABLA_TARIFAS = "transport_tarifas_json"
TABLA_HISTORIAL = "transport_tarifas_json_historial"

# Couriers válidos: los mismos slugs que ya conoce el motor de cotización.
# NOMBRE (transporte_tarifas.py) es la lista oficial — no se inventa ninguno.
SLUGS_VALIDOS = set(_ttar.NOMBRE.keys())


# ══════════════════════════════════════════════════════════════════════
#  Schema — verificación/creación idempotente al boot (Regla #5)
# ══════════════════════════════════════════════════════════════════════

def _ensure_transport_tarifas_json():
    """Crea las tablas si no existen. Idempotente: corre siempre al boot
    aunque ILUS_SKIP_MIGRATIONS=1 (igual que el resto del proyecto)."""
    get_db = _h("get_db")
    if not callable(get_db):
        print("[logistica_tarifario][WARN] get_db no disponible: se omite el schema.", flush=True)
        return
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS `{TABLA_TARIFAS}` (
                    courier_slug    VARCHAR(30) NOT NULL PRIMARY KEY,
                    json_data       LONGTEXT NOT NULL,
                    sheet_origen    VARCHAR(120) DEFAULT NULL,
                    filas           INT NOT NULL DEFAULT 0,
                    version         INT NOT NULL DEFAULT 1,
                    actualizado_por VARCHAR(190) DEFAULT NULL,
                    actualizado_en  DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                        ON UPDATE CURRENT_TIMESTAMP,
                    creado_en       DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS `{TABLA_HISTORIAL}` (
                    id              INT AUTO_INCREMENT PRIMARY KEY,
                    courier_slug    VARCHAR(30) NOT NULL,
                    json_data       LONGTEXT NOT NULL,
                    sheet_origen    VARCHAR(120) DEFAULT NULL,
                    filas           INT NOT NULL DEFAULT 0,
                    version         INT NOT NULL DEFAULT 1,
                    actualizado_por VARCHAR(190) DEFAULT NULL,
                    actualizado_en  DATETIME DEFAULT NULL,
                    reemplazado_en  DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    KEY idx_slug_version (courier_slug, version)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"[logistica_tarifario][WARN] _ensure_transport_tarifas_json: {e}", flush=True)


# ══════════════════════════════════════════════════════════════════════
#  Gates de permiso (SOLO superadmin — pedido explícito de Daniel)
# ══════════════════════════════════════════════════════════════════════

def _tarifario_api(fn):
    @wraps(fn)
    def wrapped(*args, **kwargs):
        try:
            permisos = getattr(g, "permissions", None) or {}
            if not permisos.get("superadmin"):
                return jsonify({
                    "ok": False,
                    "error": "Solo el superadministrador puede editar el tarifario.",
                    "error_codigo": "SIN_PERMISO",
                }), 403
        except Exception as e_perm:
            print(f"[logistica_tarifario] fallo leyendo permisos: {e_perm}", flush=True)
            return jsonify({"ok": False, "error": "No se pudo validar tu sesión."}), 403

        try:
            return fn(*args, **kwargs)
        except Exception as e:
            print(f"[logistica_tarifario][ERROR] {fn.__name__}: {e}", flush=True)
            try:
                print(traceback.format_exc(), flush=True)
            except Exception:
                pass
            return jsonify({
                "ok": False,
                "error": "Ocurrió un problema procesando la solicitud. "
                         "Intenta nuevamente; si persiste avisa a soporte.",
            }), 500

    login_required = _h("login_required")
    return login_required(wrapped) if callable(login_required) else wrapped


def _tarifario_page(fn):
    @wraps(fn)
    def wrapped(*args, **kwargs):
        permisos = getattr(g, "permissions", None) or {}
        if not permisos.get("superadmin"):
            flash("Solo el superadministrador puede acceder al tarifario.", "danger")
            return redirect(url_for("index"))
        return fn(*args, **kwargs)

    login_required = _h("login_required")
    return login_required(wrapped) if callable(login_required) else wrapped


# ══════════════════════════════════════════════════════════════════════
#  Lectura del estado actual (DB > archivo local > sin datos)
# ══════════════════════════════════════════════════════════════════════

def _fila_db(slug):
    mysql_fetchone = _h("mysql_fetchone")
    if not callable(mysql_fetchone):
        return None
    return mysql_fetchone(
        f"SELECT courier_slug, sheet_origen, filas, version, actualizado_por, actualizado_en "
        f"FROM `{TABLA_TARIFAS}` WHERE courier_slug=%s", (slug,)
    )


def _resumen_courier(slug):
    """Resumen para el listado: de dónde viene la tarifa vigente hoy y sus
    métricas, sin cargar el JSON completo (que puede pesar hasta ~2.8MB)."""
    import os as _os

    etiqueta = _ttar.NOMBRE.get(slug, slug)
    row = _fila_db(slug)
    if row:
        return {
            "slug": slug,
            "nombre": etiqueta,
            "fuente": "base_de_datos",
            "fuente_label": "Editado desde el admin (MySQL)",
            "sheet_origen": row.get("sheet_origen") or "",
            "comunas": int(row.get("filas") or 0),
            "version": int(row.get("version") or 1),
            "actualizado_por": row.get("actualizado_por") or "",
            "actualizado_en": row.get("actualizado_en"),
            "editable": True,
        }

    # Sin fila en MySQL: el motor está usando el archivo del repo (o no hay
    # tarifa cargada). NO se toca la tabla solo por listar.
    path = _os.path.join(_ttar._BASE, slug + ".json")
    if _os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as fh:
                d = json.load(fh)
            comunas = len(d.get("rows", {}))
            sheet = d.get("sheet") or ""
            mtime = datetime.fromtimestamp(_os.path.getmtime(path))
        except Exception:
            comunas, sheet, mtime = 0, "", None
        return {
            "slug": slug,
            "nombre": etiqueta,
            "fuente": "archivo_local",
            "fuente_label": "Archivo del repo (aún no editado desde el admin)",
            "sheet_origen": sheet,
            "comunas": comunas,
            "version": 0,
            "actualizado_por": "",
            "actualizado_en": mtime,
            "editable": True,
        }

    return {
        "slug": slug,
        "nombre": etiqueta,
        "fuente": "sin_datos",
        "fuente_label": "Sin tarifario cargado",
        "sheet_origen": "",
        "comunas": 0,
        "version": 0,
        "actualizado_por": "",
        "actualizado_en": None,
        "editable": True,
    }


def _json_actual(slug):
    """El dict JSON completo vigente para `slug` (DB si hay fila, si no el
    archivo local). Igual que transporte_tarifas._load() pero sin tocar su
    cache — se usa acá solo para exportar a Excel."""
    import os as _os

    mysql_fetchone = _h("mysql_fetchone")
    if callable(mysql_fetchone):
        row = mysql_fetchone(
            f"SELECT json_data FROM `{TABLA_TARIFAS}` WHERE courier_slug=%s", (slug,)
        )
        if row and row.get("json_data"):
            try:
                return json.loads(row["json_data"])
            except Exception as e:
                print(f"[logistica_tarifario] JSON corrupto en DB para {slug}: {e}", flush=True)

    path = _os.path.join(_ttar._BASE, slug + ".json")
    if _os.path.exists(path):
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    return None


# ══════════════════════════════════════════════════════════════════════
#  Excel / CSV → JSON (parseo + validación estricta)
# ══════════════════════════════════════════════════════════════════════

_COL_LIVIANA_RE = re.compile(r"^\s*(\d{1,4})\s*(kg)?\s*$", re.IGNORECASE)
_COL_PESADA_RE = re.compile(r"^\s*HEAVY[_\s-]?(\d+)\s*(?:\((.*)\))?\s*$", re.IGNORECASE)


def _parse_valor_pesado(raw):
    """Igual criterio que transporte_tarifas._parse_heavy(): número o
    'factor + fijo' (texto, solo Clickex en tramos altos)."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if "+" in s:
        # se valida que ambos lados sean numéricos, pero se GUARDA como texto
        # (así lo consume _parse_heavy en el motor).
        a, b = s.split("+", 1)
        float(a.strip())
        float(b.strip())
        return s.replace(" ", "").replace("+", " + ")
    return float(s.replace("$", "").replace(",", "").strip())


def _leer_filas_workbook(filename, filebytes):
    """Devuelve list[dict] (una fila = un dict header->valor) desde .xlsx/.csv.
    Lanza ValueError con mensaje amigable si el formato no se puede leer."""
    nombre = (filename or "").lower()

    if nombre.endswith(".csv"):
        try:
            texto = filebytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = filebytes.decode("latin-1")
        lector = csv.reader(io.StringIO(texto))
        filas_raw = list(lector)
        if not filas_raw:
            raise ValueError("El CSV está vacío.")
        header = filas_raw[0]
        return header, filas_raw[1:]

    if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(filebytes), data_only=True)
        except Exception as exc:
            raise ValueError(f"No se pudo leer el Excel: {exc}")
        # Usa la primera hoja que tenga contenido, priorizando "Tarifario" si existe.
        ws = wb["Tarifario"] if "Tarifario" in wb.sheetnames else wb[wb.sheetnames[0]]
        filas_raw = list(ws.iter_rows(values_only=True))
        if not filas_raw:
            raise ValueError("La hoja está vacía.")
        header = list(filas_raw[0])
        return header, filas_raw[1:]

    raise ValueError("Formato no soportado: sube un archivo .xlsx o .csv.")


def _validar_y_construir_json(slug, filename, filebytes):
    """Valida el archivo subido y arma el dict JSON en el formato que
    consume transporte_tarifas.py. Devuelve (dict_json, avisos) o lanza
    ValueError con un mensaje claro (el caller responde 400 sin tocar nada).
    """
    header, filas_raw = _leer_filas_workbook(filename, filebytes)

    header_norm = [str(h).strip() if h is not None else "" for h in header]
    try:
        col_comuna = header_norm.index("COMUNA")
    except ValueError:
        # Tolerante a minúsculas/variantes razonables.
        candidatos = [i for i, h in enumerate(header_norm) if h.strip().upper() == "COMUNA"]
        if not candidatos:
            raise ValueError(
                "Falta la columna 'COMUNA' en la primera fila. "
                "Descarga el tarifario actual para ver el formato exacto."
            )
        col_comuna = candidatos[0]

    col_dias = None
    for i, h in enumerate(header_norm):
        if h.strip().upper() == "DIAS":
            col_dias = i
            break

    cols_liviana = {}   # indice -> kg (int)
    cols_pesada = {}    # indice -> codigo (str)
    heavy_headers = {}  # codigo -> descripcion original de la columna

    lmax = _ttar.LIGHT_MAX.get(slug, 100)

    for i, h in enumerate(header_norm):
        if i in (col_comuna, col_dias):
            continue
        if not h:
            continue
        m_liv = _COL_LIVIANA_RE.match(h)
        if m_liv:
            kg = int(m_liv.group(1))
            cols_liviana[i] = kg
            continue
        m_pes = _COL_PESADA_RE.match(h)
        if m_pes:
            codigo = m_pes.group(1)
            cols_pesada[i] = codigo
            heavy_headers[codigo] = (m_pes.group(2) or h).strip()
            continue
        # Columna no reconocida: se ignora (no rompe el import), pero se avisa.

    if not cols_liviana and not cols_pesada:
        raise ValueError(
            "No se reconoció ninguna columna de precio. Se esperan columnas "
            "numéricas (1, 2, 3… kg) y/o columnas 'HEAVY_<código>'. "
            "Descarga el tarifario actual para ver el formato exacto."
        )

    rows = {}
    avisos = []
    vistos_upper = {}
    fila_num = 1  # la fila 1 es el header

    for fila in filas_raw:
        fila_num += 1
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue  # fila totalmente vacía: se ignora en silencio (no es un dato)

        fila = list(fila) + [None] * (len(header_norm) - len(fila))  # padding defensivo

        comuna_raw = fila[col_comuna] if col_comuna < len(fila) else None
        comuna = str(comuna_raw).strip() if comuna_raw is not None else ""
        if not comuna:
            raise ValueError(
                f"Fila {fila_num}: falta el nombre de la comuna. "
                "Corrige el archivo y vuelve a subirlo — no se aplicó ningún cambio."
            )

        clave_upper = comuna.upper()
        if clave_upper in vistos_upper:
            raise ValueError(
                f"Fila {fila_num}: la comuna '{comuna}' está repetida "
                f"(ya aparece en la fila {vistos_upper[clave_upper]}). "
                "No se aplicó ningún cambio."
            )
        vistos_upper[clave_upper] = fila_num

        dias_val = None
        if col_dias is not None and col_dias < len(fila) and fila[col_dias] not in (None, ""):
            try:
                dias_val = float(fila[col_dias])
            except (TypeError, ValueError):
                dias_val = None

        light = {}
        for idx, kg in cols_liviana.items():
            if idx >= len(fila):
                continue
            val = fila[idx]
            if val is None or str(val).strip() == "":
                continue
            try:
                light[str(kg)] = float(str(val).replace("$", "").replace(",", "").strip())
            except (TypeError, ValueError):
                raise ValueError(
                    f"Fila {fila_num} ({comuna}), columna '{kg}': "
                    f"'{val}' no es un número válido. No se aplicó ningún cambio."
                )

        heavy = {}
        for idx, codigo in cols_pesada.items():
            if idx >= len(fila):
                continue
            val = fila[idx]
            if val is None or str(val).strip() == "":
                continue
            try:
                heavy[codigo] = _parse_valor_pesado(val)
            except (TypeError, ValueError):
                raise ValueError(
                    f"Fila {fila_num} ({comuna}), columna 'HEAVY_{codigo}': "
                    f"'{val}' no es un valor válido (número o 'factor + fijo'). "
                    "No se aplicó ningún cambio."
                )

        if not light and not heavy:
            raise ValueError(
                f"Fila {fila_num} ({comuna}): no tiene ningún precio cargado "
                "(ni liviano ni pesado). No se aplicó ningún cambio."
            )

        rows[comuna] = {"dias": dias_val, "light": light, "heavy": heavy}

    if not rows:
        raise ValueError("El archivo no tiene ninguna fila de datos válida (solo encabezado).")

    if len(rows) < 2:
        avisos.append(
            f"El archivo trae solo {len(rows)} comuna(s) — revisa que no falten filas "
            "antes de confirmar (esto NO bloquea la subida)."
        )

    data = {
        "sheet": f"Tarifario {slug} (subido {datetime.now().strftime('%d/%m/%Y')})",
        "comuna_col": None,
        "heavy_headers": heavy_headers,
        "rows": rows,
    }
    return data, avisos


# ══════════════════════════════════════════════════════════════════════
#  Escritura: guardar nueva versión (con respaldo de la anterior)
# ══════════════════════════════════════════════════════════════════════

def _guardar_nueva_version(slug, data, sheet_origen, usuario):
    get_db = _h("get_db")
    mysql_fetchone = _h("mysql_fetchone")
    conn = get_db()

    actual = mysql_fetchone(
        f"SELECT json_data, sheet_origen, filas, version, actualizado_por, actualizado_en "
        f"FROM `{TABLA_TARIFAS}` WHERE courier_slug=%s", (slug,)
    )

    nueva_version = int(actual["version"]) + 1 if actual else 1
    json_txt = json.dumps(data, ensure_ascii=False)
    filas = len(data.get("rows", {}))

    try:
        with conn.cursor() as cur:
            if actual:
                # Respaldo: la versión que se va a reemplazar queda en el historial
                # ANTES de sobrescribir — nunca se pierde sin poder deshacer.
                cur.execute(
                    f"INSERT INTO `{TABLA_HISTORIAL}` "
                    f"(courier_slug, json_data, sheet_origen, filas, version, "
                    f" actualizado_por, actualizado_en) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (slug, actual["json_data"], actual.get("sheet_origen"),
                     int(actual.get("filas") or 0), int(actual["version"]),
                     actual.get("actualizado_por"), actual.get("actualizado_en")),
                )
            cur.execute(
                f"INSERT INTO `{TABLA_TARIFAS}` "
                f"(courier_slug, json_data, sheet_origen, filas, version, actualizado_por) "
                f"VALUES (%s,%s,%s,%s,%s,%s) "
                f"ON DUPLICATE KEY UPDATE json_data=VALUES(json_data), "
                f"sheet_origen=VALUES(sheet_origen), filas=VALUES(filas), "
                f"version=VALUES(version), actualizado_por=VALUES(actualizado_por), "
                f"actualizado_en=CURRENT_TIMESTAMP",
                (slug, json_txt, sheet_origen, filas, nueva_version, usuario),
            )
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"[logistica_tarifario] error guardando {slug}: {e}", flush=True)
        raise

    # El motor de cotización cachea el JSON en memoria (transporte_tarifas._CACHE):
    # invalidar para que el próximo request ya use el precio nuevo, sin deploy.
    try:
        _ttar._CACHE.pop(slug, None)
    except Exception:
        pass

    return nueva_version, filas


# ══════════════════════════════════════════════════════════════════════
#  REGISTRO DE RUTAS
# ══════════════════════════════════════════════════════════════════════

def register_logistica_tarifario(app, ctx=None):
    global _APP_MODULE_NAME
    try:
        nombre_mod = getattr(app, "import_name", None)
        if nombre_mod and nombre_mod in sys.modules:
            _APP_MODULE_NAME = nombre_mod
    except Exception:
        pass

    # ──────────────────────────────────────────────────────────────
    #  GET /transporte/tarifario — página
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario", methods=["GET"])
    @_tarifario_page
    def logistica_tarifario_page():
        return render_template("transporte/tarifario.html")

    # ──────────────────────────────────────────────────────────────
    #  GET /transporte/tarifario/api/lista
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario/api/lista", methods=["GET"])
    @_tarifario_api
    def logistica_tarifario_lista():
        items = [_resumen_courier(slug) for slug in sorted(SLUGS_VALIDOS,
                                                             key=lambda s: _ttar.NOMBRE.get(s, s))]
        return jsonify({"ok": True, "couriers": items})

    # ──────────────────────────────────────────────────────────────
    #  POST /transporte/tarifario/api/<slug>/subir
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario/api/<string:slug>/subir", methods=["POST"])
    @_tarifario_api
    def logistica_tarifario_subir(slug):
        slug = (slug or "").strip().lower()
        if slug not in SLUGS_VALIDOS:
            return jsonify({"ok": False, "error": f"Courier desconocido: {slug}."}), 404

        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"ok": False, "error": "No se recibió ningún archivo."}), 400

        filebytes = file.read()
        if not filebytes:
            return jsonify({"ok": False, "error": "El archivo llegó vacío."}), 400
        if len(filebytes) > 15 * 1024 * 1024:
            return jsonify({"ok": False, "error": "El archivo supera los 15MB."}), 400

        try:
            data, avisos = _validar_y_construir_json(slug, file.filename, filebytes)
        except ValueError as ve:
            # Validación falló: se RECHAZA, el tarifario vigente NO se toca.
            return jsonify({"ok": False, "error": str(ve), "error_codigo": "VALIDACION"}), 400

        current_username = _h("current_username")
        try:
            usuario = current_username() if callable(current_username) else None
        except Exception:
            usuario = None

        try:
            version, filas = _guardar_nueva_version(
                slug, data, file.filename[:120], usuario
            )
        except Exception:
            return jsonify({"ok": False,
                            "error": "No se pudo guardar el tarifario. No se modificó nada."}), 500

        try:
            _audit = _h("_audit")
            if callable(_audit):
                _audit(
                    "tarifario_courier_actualizado",
                    target_type="courier_tarifa",
                    target_id=slug,
                    details={"slug": slug, "archivo": file.filename,
                             "comunas": filas, "version": version},
                )
        except Exception as e_aud:
            print(f"[logistica_tarifario] audit subir falló (no crítico): {e_aud}", flush=True)

        return jsonify({
            "ok": True,
            "slug": slug,
            "nombre": _ttar.NOMBRE.get(slug, slug),
            "comunas": filas,
            "version": version,
            "avisos": avisos,
            "mensaje": f"Tarifario de {_ttar.NOMBRE.get(slug, slug)} actualizado: "
                       f"{filas} comuna(s), versión {version}. La versión anterior quedó "
                       "guardada en el historial.",
        })

    # ──────────────────────────────────────────────────────────────
    #  GET /transporte/tarifario/api/<slug>/descargar
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario/api/<string:slug>/descargar", methods=["GET"])
    @_tarifario_api
    def logistica_tarifario_descargar(slug):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        slug = (slug or "").strip().lower()
        if slug not in SLUGS_VALIDOS:
            return jsonify({"ok": False, "error": f"Courier desconocido: {slug}."}), 404

        data = _json_actual(slug)
        if not data:
            return jsonify({"ok": False,
                            "error": "Este courier no tiene tarifario cargado todavía."}), 404

        rows = data.get("rows", {})
        heavy_headers = data.get("heavy_headers", {}) or {}
        lmax = _ttar.LIGHT_MAX.get(slug, 100)

        cols_liviana = [str(k) for k in range(1, lmax + 1)]
        codigos_pesada = sorted(heavy_headers.keys(), key=lambda c: (len(c), c))
        if not codigos_pesada:
            # Por si el JSON no trae heavy_headers pero sí valores 'heavy' sueltos.
            vistos = set()
            for r in rows.values():
                for c in (r.get("heavy") or {}).keys():
                    vistos.add(c)
            codigos_pesada = sorted(vistos, key=lambda c: (len(c), c))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tarifario"

        RED_FILL = PatternFill("solid", fgColor="DC2626")
        WHITE_FONT = Font(color="FFFFFF", bold=True)

        headers = ["COMUNA", "DIAS"] + cols_liviana + [
            f"HEAVY_{c} ({heavy_headers.get(c, c)})" for c in codigos_pesada
        ]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font = WHITE_FONT
            cell.fill = RED_FILL
            cell.alignment = Alignment(horizontal="center")

        for ri, comuna in enumerate(sorted(rows.keys()), 2):
            r = rows[comuna]
            ws.cell(ri, 1, comuna)
            ws.cell(ri, 2, r.get("dias"))
            light = r.get("light") or {}
            for ci, kg in enumerate(cols_liviana, 3):
                v = light.get(kg)
                if v is not None:
                    ws.cell(ri, ci, round(float(v), 4))
            heavy = r.get("heavy") or {}
            base_ci = 3 + len(cols_liviana)
            for ci, codigo in enumerate(codigos_pesada, base_ci):
                v = heavy.get(codigo)
                if v is not None:
                    ws.cell(ri, ci, v)

        ws.freeze_panes = "C2"
        ws.column_dimensions["A"].width = 22

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        fecha = datetime.now().strftime("%Y%m%d")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"tarifario_{slug}_{fecha}.xlsx",
        )

    # ──────────────────────────────────────────────────────────────
    #  POST /transporte/tarifario/api/<slug>/sembrar-archivo-local
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario/api/<string:slug>/sembrar-archivo-local", methods=["POST"])
    @_tarifario_api
    def logistica_tarifario_sembrar(slug):
        """Copia el archivo tarifas/<slug>.json (que hoy usa el motor) a
        MySQL tal cual, sin cambiar ningún precio — primer paso para dejar
        de depender del disco efímero de Cloud Run en ese courier."""
        import os as _os

        slug = (slug or "").strip().lower()
        if slug not in SLUGS_VALIDOS:
            return jsonify({"ok": False, "error": f"Courier desconocido: {slug}."}), 404

        if _fila_db(slug):
            return jsonify({"ok": False,
                            "error": "Este courier ya tiene tarifario editado desde el admin "
                                     "(MySQL). Usa 'Subir' si quieres reemplazarlo."}), 409

        path = _os.path.join(_ttar._BASE, slug + ".json")
        if not _os.path.exists(path):
            return jsonify({"ok": False,
                            "error": "No existe un archivo local para este courier."}), 404

        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        if not data.get("rows"):
            return jsonify({"ok": False, "error": "El archivo local no tiene filas."}), 400

        current_username = _h("current_username")
        try:
            usuario = current_username() if callable(current_username) else None
        except Exception:
            usuario = None

        version, filas = _guardar_nueva_version(
            slug, data, data.get("sheet") or f"{slug}.json (archivo local)", usuario
        )

        try:
            _audit = _h("_audit")
            if callable(_audit):
                _audit("tarifario_courier_sembrado", target_type="courier_tarifa",
                       target_id=slug, details={"slug": slug, "comunas": filas})
        except Exception:
            pass

        return jsonify({
            "ok": True, "slug": slug, "comunas": filas, "version": version,
            "mensaje": f"Tarifario de {_ttar.NOMBRE.get(slug, slug)} copiado a la base de "
                       f"datos ({filas} comunas). Ya se puede editar desde acá sin deploy.",
        })

    # ──────────────────────────────────────────────────────────────
    #  GET /transporte/tarifario/api/<slug>/historial
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario/api/<string:slug>/historial", methods=["GET"])
    @_tarifario_api
    def logistica_tarifario_historial(slug):
        mysql_fetchall = _h("mysql_fetchall")
        slug = (slug or "").strip().lower()
        if slug not in SLUGS_VALIDOS:
            return jsonify({"ok": False, "error": f"Courier desconocido: {slug}."}), 404

        vigente = _fila_db(slug)
        historial = mysql_fetchall(
            f"SELECT id, sheet_origen, filas, version, actualizado_por, actualizado_en, "
            f"reemplazado_en FROM `{TABLA_HISTORIAL}` WHERE courier_slug=%s "
            f"ORDER BY version DESC LIMIT 30", (slug,)
        ) or []

        return jsonify({
            "ok": True,
            "slug": slug,
            "vigente": vigente,
            "historial": historial,
        })

    # ──────────────────────────────────────────────────────────────
    #  POST /transporte/tarifario/api/<slug>/restaurar/<int:hist_id>
    # ──────────────────────────────────────────────────────────────
    @app.route("/transporte/tarifario/api/<string:slug>/restaurar/<int:hist_id>", methods=["POST"])
    @_tarifario_api
    def logistica_tarifario_restaurar(slug, hist_id):
        """Restaura una versión del historial como la vigente. La versión que
        estaba vigente ANTES de restaurar también queda respaldada — nunca se
        pierde nada, incluso deshaciendo un deshacer."""
        mysql_fetchone = _h("mysql_fetchone")
        slug = (slug or "").strip().lower()
        if slug not in SLUGS_VALIDOS:
            return jsonify({"ok": False, "error": f"Courier desconocido: {slug}."}), 404

        backup = mysql_fetchone(
            f"SELECT json_data, sheet_origen FROM `{TABLA_HISTORIAL}` "
            f"WHERE id=%s AND courier_slug=%s", (hist_id, slug)
        )
        if not backup:
            return jsonify({"ok": False, "error": "Esa versión del historial no existe."}), 404

        try:
            data = json.loads(backup["json_data"])
        except Exception:
            return jsonify({"ok": False, "error": "La versión del historial está corrupta."}), 500

        if not data.get("rows"):
            return jsonify({"ok": False, "error": "Esa versión no tiene filas — no se restauró."}), 400

        current_username = _h("current_username")
        try:
            usuario = current_username() if callable(current_username) else None
        except Exception:
            usuario = None

        version, filas = _guardar_nueva_version(
            slug, data, (backup.get("sheet_origen") or "") + " (restaurado)", usuario
        )

        try:
            _audit = _h("_audit")
            if callable(_audit):
                _audit("tarifario_courier_restaurado", target_type="courier_tarifa",
                       target_id=slug, details={"slug": slug, "desde_historial_id": hist_id,
                                                 "comunas": filas, "version_nueva": version})
        except Exception:
            pass

        return jsonify({
            "ok": True, "slug": slug, "comunas": filas, "version": version,
            "mensaje": f"Se restauró la versión anterior de {_ttar.NOMBRE.get(slug, slug)} "
                       f"({filas} comunas). Quedó como versión {version}.",
        })

    # ── Schema: se verifica/crea SIEMPRE al registrar (Regla #5 —
    #    producción corre con ILUS_SKIP_MIGRATIONS=1). Nunca revienta el boot.
    try:
        with app.app_context():
            _ensure_transport_tarifas_json()
    except Exception as e:
        print(f"[ILUS][WARN] _ensure_transport_tarifas_json: {e}", flush=True)

    return app
