import os
import re
import secrets
import threading
import time
from datetime import datetime, timedelta

from flask import flash, has_request_context, jsonify, redirect, render_template, request, send_from_directory, url_for
from werkzeug.utils import secure_filename


def _public_base_url():
    """Base URL pública absoluta, SEGURA para usar dentro de threads.

    `url_for(..., _external=True)` necesita un request activo (o SERVER_NAME
    configurado); en threads daemon lanza "Unable to build URLs outside an
    active request without SERVER_NAME" y el email sale sin link (verificado
    en logs de Cloud Run 2026-06). Patrón:
      - Con request activo → usamos request.url_root (respeta dominio/proxy).
      - Sin request (thread/cron) → env ILUS_PUBLIC_BASE_URL con default a la
        URL pública de Cloud Run.
    """
    try:
        if has_request_context():
            return (request.url_root or "").rstrip("/")
    except Exception:
        pass
    return os.environ.get(
        "ILUS_PUBLIC_BASE_URL",
        "https://ilus-app-469212710544.southamerica-west1.run.app",
    ).rstrip("/")


PICKUP_STATUS = {
    "solicitud_recibida": "Solicitud recibida",
    "en_revision": "En revision",
    "informacion_incompleta": "Informacion incompleta",
    "propuesta_enviada": "Propuesta enviada",
    "esperando_cliente": "Esperando respuesta",
    "agenda_confirmada": "Agenda confirmada",
    "reagendada": "Reagendada",
    "rechazada": "Rechazada",
    "en_preparacion": "En preparacion",
    "retirada": "Retirada",
    "fallida": "Fallida",
    "cerrada": "Cerrada",
}

PICKUP_STATUS_COLORS = {
    "solicitud_recibida": "primary",
    "en_revision": "info",
    "informacion_incompleta": "warning",
    "propuesta_enviada": "warning",
    "esperando_cliente": "secondary",
    "agenda_confirmada": "success",
    "reagendada": "warning",
    "rechazada": "danger",
    "en_preparacion": "dark",
    "retirada": "success",
    "fallida": "danger",
    "cerrada": "secondary",
}

# ════════════════════════════════════════════════════════════════════
#  MODELO CANÓNICO DE TRACKING — UNA SOLA FUENTE DE VERDAD
#  Daniel 2026-06-15: el correo, el seguimiento público del cliente y el
#  stepper interno deben mostrar LOS MISMOS hitos. Estos 5 hitos (+ el
#  terminal negativo "Cancelada") son el modelo único. Tanto el email
#  (_ret_stepper en app.py, que importa PICKUP_JOURNEY) como el seguimiento
#  (templates/retiros/public_tracking.html, que recibe journey/journey_idx)
#  y el stepper interno se alinean a esto. NO duplicar etiquetas: editar acá.
# ════════════════════════════════════════════════════════════════════
PICKUP_JOURNEY = [
    {"label": "Solicitada",  "sub": "Recibimos tu solicitud",    "emoji": "📩", "icon": "envelope-paper"},
    {"label": "Propuesta",   "sub": "Esperando tu confirmación",  "emoji": "📅", "icon": "calendar-event"},
    {"label": "Confirmada",  "sub": "Cita acordada",              "emoji": "✅", "icon": "calendar-check"},
    {"label": "Preparación", "sub": "Bodega lista",               "emoji": "📦", "icon": "box-seam"},
    {"label": "Retirada",    "sub": "Completado",                 "emoji": "🎉", "icon": "check2-circle"},
]
# Estado canónico → índice del hito (0-4). -1 = terminal NEGATIVO (Cancelada).
PICKUP_JOURNEY_IDX = {
    "solicitud_recibida": 0, "en_revision": 0, "informacion_incompleta": 0,
    "propuesta_enviada": 1, "esperando_cliente": 1, "reagendada": 1,
    "agenda_confirmada": 2,
    "en_preparacion": 3,
    "retirada": 4, "cerrada": 4,
    "rechazada": -1, "fallida": -1,
}

def pickup_journey_idx(status):
    """Índice del hito canónico (0-4) para un estado; -1 = cancelado/fallido."""
    return PICKUP_JOURNEY_IDX.get(status or "", 0)

# ── PIPELINE_GROUPS (2026-05-26 Daniel) ──────────────────────────────
# Agrupación VISUAL del kanban del monitor de retiros: los 12 estados
# de PICKUP_STATUS se consolidan en 6 columnas para reducir saturación.
# Las cards mantienen su badge real (status_badge) dentro de cada columna
# para no perder granularidad operacional.
# La última columna SIEMPRE es "Retirada" (terminal exitosa) como acordó
# Daniel — refuerza el foco del operador en el outcome positivo.
#
# IMPORTANTE: NO cambia los estados de BD ni los endpoints de transición.
# Solo cambia cómo el template agrupa visualmente las columnas.
PIPELINE_GROUPS = [
    {
        "key": "por_revisar",
        "label": "Por revisar",
        "statuses": ["solicitud_recibida", "en_revision", "informacion_incompleta"],
        "accent": "#6b7280",   # gris neutro
        "icon": "bi-inbox",
    },
    {
        "key": "propuesta",
        "label": "Propuesta",
        "statuses": ["propuesta_enviada", "esperando_cliente"],
        "accent": "#3b82f6",   # azul info
        "icon": "bi-envelope-paper",
    },
    {
        "key": "agendada",
        "label": "Agendada",
        "statuses": ["agenda_confirmada", "reagendada"],
        "accent": "#8b5cf6",   # violeta
        "icon": "bi-calendar-check",
    },
    {
        "key": "preparacion",
        "label": "En preparacion",
        "statuses": ["en_preparacion"],
        "accent": "#f59e0b",   # ambar
        "icon": "bi-box-seam",
    },
    {
        "key": "canceladas",
        "label": "Canceladas",
        "statuses": ["rechazada", "fallida", "cerrada"],
        "accent": "#9ca3af",   # gris frio
        "icon": "bi-x-octagon",
    },
    {
        "key": "retirada",
        "label": "Retirada",
        "statuses": ["retirada"],
        "accent": "#16a34a",   # verde exito (terminal final)
        "icon": "bi-check2-circle",
    },
]

PICKUP_RELATIONS = [
    ("dueno", "Dueno / titular"),
    ("comprador", "Comprador"),
    ("familiar", "Familiar autorizado"),
    ("chofer", "Chofer"),
    ("transporte", "Transporte externo"),
    ("autorizado", "Persona autorizada"),
    ("otro", "Otro"),
]


# ── VALIDADORES CHILENOS ──────────────────────────────────────────────

def _clean_rut(rut: str) -> str:
    """Normaliza RUT: solo números y K, en mayúsculas."""
    return re.sub(r"[^0-9kK]", "", str(rut or "")).upper()


def _calc_dv(num: str) -> str:
    """Calcula el dígito verificador del RUT (módulo 11)."""
    suma, mul = 0, 2
    for ch in reversed(num):
        suma += int(ch) * mul
        mul = 2 if mul == 7 else mul + 1
    r = 11 - (suma % 11)
    return "0" if r == 11 else "K" if r == 10 else str(r)


def is_valid_rut(rut: str) -> bool:
    """Valida RUT chileno usando módulo 11. Acepta cualquier formato."""
    c = _clean_rut(rut)
    if not 8 <= len(c) <= 9:
        return False
    num, dv = c[:-1], c[-1]
    return num.isdigit() and _calc_dv(num) == dv


def format_rut(rut: str) -> str:
    """Formatea RUT como '12.345.678-9'."""
    c = _clean_rut(rut)
    if len(c) < 2:
        return c
    num, dv = c[:-1], c[-1]
    rev = "".join(reversed(num))
    parts = [rev[i:i+3] for i in range(0, len(rev), 3)]
    return ".".join("".join(reversed(p)) for p in reversed(parts)) + "-" + dv


def is_valid_cl_phone(phone: str) -> bool:
    """Valida teléfono chileno: +56 9 XXXX XXXX (móvil)."""
    c = re.sub(r"[^\d+]", "", str(phone or "")).lstrip("+")
    return bool(re.match(r"^(56)?9\d{8}$", c))


def format_cl_phone(phone: str) -> str:
    """Devuelve teléfono normalizado a formato '+56 9 XXXX XXXX' (legible).
    Si no se puede formatear, devuelve el input original."""
    c = re.sub(r"[^\d+]", "", str(phone or "")).lstrip("+")
    if c.startswith("56"):
        c = c[2:]
    if c.startswith("9") and len(c) == 9:
        # Formato legible: +56 9 1234 5678
        return f"+56 {c[0]} {c[1:5]} {c[5:9]}"
    return phone


# Regex robusto de email (alineado con app.py _EMAIL_RE).
# - permite letras/numeros/._%+-
# - dominio con letras/numeros/.-
# - TLD de 2+ caracteres (alfabético)
# Rechaza: doble @, espacios, sin TLD válido.
_PICKUP_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


def is_valid_email(email: str) -> bool:
    """Valida email con regex robusto. Limita longitud a 200 chars."""
    if not email:
        return False
    e = str(email).strip().lower()
    if not e or len(e) > 200:
        return False
    # Rechaza espacios internos, doble @ y otros casos edge
    if " " in e or e.count("@") != 1:
        return False
    return bool(_PICKUP_EMAIL_RE.match(e))


def register_pickup_routes(app, ctx):
    mysql_fetchone = ctx["mysql_fetchone"]
    mysql_fetchall = ctx["mysql_fetchall"]
    mysql_execute = ctx["mysql_execute"]
    get_db = ctx["get_db"]
    # get_mysql: conexión directa para transacciones explícitas con
    # SELECT FOR UPDATE (usado en confirmación de retiros para evitar
    # double-booking). Ver pickup_public_tracking acción 'confirm'.
    get_mysql = ctx["get_mysql"]
    require_permission = ctx["require_permission"]
    EMAIL_RE = ctx["EMAIL_RE"]
    BASE_DIR = ctx["BASE_DIR"]
    g = ctx["g"]
    _ilus_email_html = ctx["_ilus_email_html"]
    _send_ilus_email = ctx["_send_ilus_email"]
    _get_wa_cfg = ctx["_get_wa_cfg"]
    _send_whatsapp = ctx["_send_whatsapp"]
    # Brand config (Daniel 2026-05-23): para notificación al operador en
    # caso de que el cliente rechace la propuesta. Opcional — si no viene,
    # caemos al default fijo "soportetec@sphs.cl".
    _get_brand_cfg = ctx.get("_get_brand_cfg") or (
        lambda: {"support_email": "soportetec@sphs.cl"}
    )
    # Rate limiter de app.py (persistido en BD, cross-worker). Si no está
    # disponible (entorno de tests), creamos un no-op para no romper.
    _rate_limited = ctx.get("rate_limited") or (
        lambda *a, **kw: (lambda fn: fn)
    )
    # Daniel dio de baja Twilio (mayo 2026). El canal WhatsApp sólo dispara
    # si la env var COMM_CANALES_ACTIVOS lo incluye. Email vía Resend va
    # siempre — eso es lo importante para el flujo de retiros.
    _canal_activo = ctx.get("_canal_activo") or (lambda canal: (canal or "").lower() == "email")

    REQ = ctx["PICKUP_REQUESTS_TABLE"]
    PKG = ctx["PICKUP_PACKAGES_TABLE"]
    PROP = ctx["PICKUP_PROPOSALS_TABLE"]
    LOG = ctx["PICKUP_LOGS_TABLE"]
    ATT = ctx["PICKUP_ATTACHMENTS_TABLE"]
    SIG = ctx["PICKUP_SIGNATURES_TABLE"]
    SET = ctx["PICKUP_SETTINGS_TABLE"]
    TPL = ctx["PICKUP_TEMPLATES_TABLE"]

    upload_dir = os.path.join(BASE_DIR, "static", "uploads", "retiros")
    os.makedirs(upload_dir, exist_ok=True)

    def ensure_marketing_columns():
        for col in ("hero_image_1", "hero_image_2", "hero_image_3"):
            try:
                mysql_execute(f"ALTER TABLE `{SET}` ADD COLUMN `{col}` VARCHAR(260) NULL")
            except Exception:
                pass
        # 2026-06-09: emails internos (CSV) que reciben alertas del equipo
        # de retiros (alta pública, respuesta del cliente, sin saldo).
        # Idempotente: si la columna ya existe, el ALTER falla y seguimos.
        try:
            mysql_execute(
                f"ALTER TABLE `{SET}` ADD COLUMN notify_emails TEXT NULL "
                f"COMMENT 'Emails internos (CSV) para alertas del equipo de retiros'"
            )
        except Exception:
            pass

    # 2026-06-10 — FIX RAÍZ: estas migraciones de boot corren en IMPORT,
    # donde NO hay contexto Flask. get_db() usa `g` → RuntimeError que el
    # try/except se tragaba EN SILENCIO desde siempre en Cloud Run (la
    # columna notify_emails nunca se creó → 500 en guardar settings).
    # app.app_context() habilita `g` fuera de un request.
    with app.app_context():
        ensure_marketing_columns()

    def ensure_reminder_columns():
        """Migración: timestamp del envío del recordatorio 24h."""
        try:
            mysql_execute(
                f"ALTER TABLE `{REQ}` ADD COLUMN reminder_24h_sent DATETIME NULL "
                f"COMMENT 'Timestamp del recordatorio 24h enviado al cliente'"
            )
        except Exception:
            pass

    with app.app_context():
        ensure_reminder_columns()

    # ════════════════════════════════════════════════════════════════════
    #  MIGRACIÓN CRÍTICA — pickup_request_docs + pickup_doc_lineas
    #  Daniel 2026-05-23: en producción la BD no tiene estas tablas porque
    #  ILUS_SKIP_MIGRATIONS=1 saltea init_pickup_tables(). Esta función SE
    #  EJECUTA AL REGISTRAR EL MÓDULO (no depende de la flag), de manera
    #  que siempre garantizamos que las tablas multidocumento existan.
    #
    #  Idempotente — todos los CREATE/ALTER tienen IF NOT EXISTS o
    #  try/except. Sin riesgo de duplicar.
    # ════════════════════════════════════════════════════════════════════
    # ⚡ PERF (Daniel 2026-05-24): flag para evitar correr ~20 ALTER TABLE
    # en CADA llamada al hot path "asociar doc al retiro" (tardaba 20-45s
    # parcialmente por esto — cada ALTER cuesta 50-200ms en Clever Cloud).
    # Al boot se corre una vez (force=True) y queda en True. Los endpoints
    # llaman ensure_multidoc_tables_runtime() sin force → retorno inmediato.
    # Si por algún motivo la tabla cayera, el INSERT del endpoint tira
    # excepción y se puede forzar via /retiros/admin/migrate-now.
    _MULTIDOC_TABLES_READY = {"v": False}

    def ensure_multidoc_tables_runtime(force=False):
        """Garantiza pickup_request_docs + pickup_doc_lineas + columnas nuevas.

        Se llama al boot del módulo (force=True, una sola vez) y desde los
        endpoints como auto-heal — pero al ser idempotente con flag global,
        el costo del hot path es ~0ms cuando ya está lista.
        """
        if _MULTIDOC_TABLES_READY["v"] and not force:
            return  # ya migrado, no perdemos tiempo en ALTER TABLE
        # 1. Tabla pickup_request_docs (multidocumento)
        try:
            mysql_execute("""
                CREATE TABLE IF NOT EXISTS pickup_request_docs (
                    id              INT AUTO_INCREMENT PRIMARY KEY,
                    request_id      INT NOT NULL,
                    document_type   VARCHAR(30) NOT NULL,
                    document_number VARCHAR(60) NOT NULL,
                    cliente_rut     VARCHAR(30),
                    cliente_nombre  VARCHAR(200),
                    observaciones_erp TEXT,
                    peso_real_kg    DECIMAL(12,3) DEFAULT 0,
                    peso_vol_kg     DECIMAL(12,3) DEFAULT 0,
                    volumen_m3      DECIMAL(12,4) DEFAULT 0,
                    n_lineas        INT DEFAULT 0,
                    erp_snapshot    LONGTEXT,
                    added_by        VARCHAR(190),
                    added_at        DATETIME DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_request_doc (request_id, document_type, document_number),
                    INDEX idx_request (request_id),
                    INDEX idx_cliente (cliente_rut),
                    FOREIGN KEY (request_id) REFERENCES `pickup_requests`(id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            print(f"[ensure_multidoc_tables] pickup_request_docs: {e}", flush=True)

        # 2. Columnas extra de pickup_request_docs (cada una try/except)
        for col_sql in [
            "ALTER TABLE pickup_request_docs ADD COLUMN con_saldo TINYINT NULL DEFAULT NULL "
            "COMMENT 'NULL=no verificado, 0=sin saldo, 1=con saldo'",
            "ALTER TABLE pickup_request_docs ADD COLUMN saldo_zz DECIMAL(12,2) NULL DEFAULT NULL "
            "COMMENT 'Saldo monetario ZZ al momento de asociar'",
            "ALTER TABLE pickup_request_docs ADD COLUMN saldo_checked_at DATETIME NULL DEFAULT NULL "
            "COMMENT 'Cuándo se verificó el saldo'",
            "ALTER TABLE pickup_request_docs ADD COLUMN email_cliente_erp VARCHAR(180) NULL "
            "COMMENT 'Email del cliente del doc capturado desde ERP al asociar'",
            "ALTER TABLE pickup_request_docs ADD COLUMN has_seleccion_lineas TINYINT(1) "
            "NOT NULL DEFAULT 0 "
            "COMMENT 'Si 1, solo se retiran las líneas marcadas en pickup_doc_lineas'",
            "ALTER TABLE pickup_request_docs ADD INDEX idx_prd_cliente_rut (cliente_rut)",
            "ALTER TABLE pickup_request_docs ADD INDEX idx_prd_doc (document_type, document_number)",
        ]:
            try: mysql_execute(col_sql)
            except Exception: pass

        # 3. Columnas extra de pickup_requests
        for col_sql in [
            f"ALTER TABLE `{REQ}` ADD COLUMN extra_emails VARCHAR(800) NULL "
            f"COMMENT 'Emails adicionales separados por coma'",
            f"ALTER TABLE `{REQ}` ADD COLUMN doc_validation_status VARCHAR(30) NULL DEFAULT 'pendiente'",
            f"ALTER TABLE `{REQ}` ADD COLUMN doc_validated_at DATETIME NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN doc_validated_by VARCHAR(190) NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN doc_erp_data MEDIUMTEXT NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN doc_validation_notes TEXT NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN peso_real_kg DECIMAL(10,2) DEFAULT NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN peso_vol_kg DECIMAL(10,2) DEFAULT NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN tiempo_estimado_min INT DEFAULT NULL",
            # Responsable de la ENTREGA del retiro (Daniel 2026-06-19): quién se
            # responsabiliza de entregar el pedido. Distinto de created_by (quién lo creó).
            f"ALTER TABLE `{REQ}` ADD COLUMN responsable_user_id INT NULL",
            f"ALTER TABLE `{REQ}` ADD COLUMN responsable_nombre VARCHAR(190) NULL",
        ]:
            try: mysql_execute(col_sql)
            except Exception: pass

        # 4. Tabla pickup_doc_lineas (selección granular)
        try:
            mysql_execute("""
                CREATE TABLE IF NOT EXISTS pickup_doc_lineas (
                    id                    INT AUTO_INCREMENT PRIMARY KEY,
                    request_id            INT NOT NULL,
                    doc_id                INT NOT NULL,
                    sku                   VARCHAR(80) NOT NULL,
                    descripcion           VARCHAR(300),
                    cantidad_doc          DECIMAL(12,3) DEFAULT 0,
                    cantidad_seleccionada DECIMAL(12,3) DEFAULT 0,
                    peso_unit_kg          DECIMAL(10,3) DEFAULT 0,
                    peso_vol_unit_kg      DECIMAL(10,3) DEFAULT 0,
                    vol_unit_m3           DECIMAL(10,5) DEFAULT 0,
                    peso_total_kg         DECIMAL(12,3) DEFAULT 0,
                    peso_vol_total_kg     DECIMAL(12,3) DEFAULT 0,
                    vol_total_m3          DECIMAL(12,5) DEFAULT 0,
                    incluida              TINYINT(1) NOT NULL DEFAULT 1,
                    nota_linea            VARCHAR(300) NULL,
                    created_at            DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at            DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_doc_sku (doc_id, sku),
                    INDEX idx_request (request_id),
                    INDEX idx_doc (doc_id),
                    INDEX idx_doc_incluida (doc_id, incluida)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            # Foreign keys separadas (algunos hosts no aceptan en CREATE)
            for fk_sql in [
                "ALTER TABLE pickup_doc_lineas ADD CONSTRAINT fk_pdl_req "
                f"FOREIGN KEY (request_id) REFERENCES `{REQ}`(id) ON DELETE CASCADE",
                "ALTER TABLE pickup_doc_lineas ADD CONSTRAINT fk_pdl_doc "
                "FOREIGN KEY (doc_id) REFERENCES pickup_request_docs(id) ON DELETE CASCADE",
            ]:
                try: mysql_execute(fk_sql)
                except Exception: pass
        except Exception as e:
            print(f"[ensure_multidoc_tables] pickup_doc_lineas: {e}", flush=True)

        # 5. Tabla pickup_blocks (bloqueos de fechas) — si tampoco existe
        try:
            mysql_execute("""
                CREATE TABLE IF NOT EXISTS pickup_blocks (
                    id           INT AUTO_INCREMENT PRIMARY KEY,
                    fecha        DATE NOT NULL,
                    hora_inicio  TIME NULL,
                    hora_fin     TIME NULL,
                    motivo       VARCHAR(200) DEFAULT '',
                    created_by   VARCHAR(190) DEFAULT NULL,
                    created_at   DATETIME DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_fecha (fecha)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception:
            pass

        # ⚡ PERF (Daniel 2026-05-24): índices críticos para acelerar consultas
        # más frecuentes del módulo retiros. Cada uno try/except (idempotente
        # — si ya existe no rompe). Estos índices reducen latencia ~30-60%
        # en pickup_detail y endpoints de polling.
        for idx_sql in [
            # pickup_logs por request_id (usado en pickup_detail LIMIT 80)
            f"ALTER TABLE `{LOG}` ADD INDEX idx_pl_req_id (request_id, id)",
            # pickup_packages por request_id (usado en pickup_detail)
            f"ALTER TABLE `{PKG}` ADD INDEX idx_pkg_req (request_id, package_number)",
            # pickup_proposals por request_id (usado en pickup_detail)
            f"ALTER TABLE `{PROP}` ADD INDEX idx_pp_req (request_id, id)",
            # pickup_attachments por request_id (usado en pickup_detail)
            f"ALTER TABLE `{ATT}` ADD INDEX idx_att_req (request_id, id)",
            # pickup_requests por document_number (usado por saldo-pendiente cruce)
            f"ALTER TABLE `{REQ}` ADD INDEX idx_req_docnum (document_number, status)",
            # pickup_doc_lineas por (request_id, incluida) — selección granular
            "ALTER TABLE pickup_doc_lineas ADD INDEX idx_pdl_req_incl (request_id, incluida)",
            # 🆕 Daniel 2026-05-24: columna marcada_sin_saldo.
            # El operador puede marcar una línea SIN SALDO en el ERP (ya entregada
            # según Random) si físicamente el cliente la viene a retirar igual.
            # Esta flag deja constancia para mostrar badge ámbar "ya rebajado en
            # ERP" en la tabla externa de productos. NO bloquea la asociación.
            "ALTER TABLE pickup_doc_lineas ADD COLUMN marcada_sin_saldo TINYINT(1) NOT NULL DEFAULT 0 "
            "COMMENT 'Operador marcó esta línea aunque ERP la reporte sin saldo (ya entregada)'",
        ]:
            try: mysql_execute(idx_sql)
            except Exception: pass

        # ── Chat en vivo cliente↔operador (Daniel 2026-06-17) ───────────
        # Mensajes del seguimiento público (burbuja estilo WhatsApp en el
        # monitor). leido_operador/leido_cliente para badges de "no leído".
        try:
            mysql_execute("""
                CREATE TABLE IF NOT EXISTS pickup_messages (
                    id              INT AUTO_INCREMENT PRIMARY KEY,
                    request_id      INT NOT NULL,
                    sender          VARCHAR(12) NOT NULL,
                    autor           VARCHAR(160),
                    cuerpo          VARCHAR(2000) NOT NULL,
                    leido_operador  TINYINT(1) NOT NULL DEFAULT 0,
                    leido_cliente   TINYINT(1) NOT NULL DEFAULT 0,
                    created_at      DATETIME DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_pm_req (request_id, id),
                    INDEX idx_pm_unread (request_id, sender, leido_operador)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as _e_pm:
            print(f"[ensure_multidoc_tables] pickup_messages: {_e_pm}", flush=True)

        # ✅ Marcamos como listas: nadie más perderá tiempo aquí en el hot path
        _MULTIDOC_TABLES_READY["v"] = True
        print("[ensure_multidoc_tables] OK (cached, los endpoints saltan en el hot path)", flush=True)

    # Ejecutar al boot del módulo (force=True → siempre corre al boot).
    # Con app_context: en import no hay `g` y sin esto el boot-run moría
    # en silencio (los endpoints lo auto-curaban después, por eso "andaba").
    with app.app_context():
        ensure_multidoc_tables_runtime(force=True)

    # ── Migración horarios v3 (Daniel 2026-05-23, reforzado 2026-05-24) ─
    # Daniel reiteró: el horario es ÚNICO Y SIN EXCEPCIÓN para todas las
    # visuales (público, interno, propuesta):
    #   Mañana: 09:00 – 12:30  (último FIN admitido 12:30)
    #   Tarde:  14:00 – 16:30  (último FIN admitido 16:30)
    #   Colación 12:30 – 14:00 BLOQUEADA TOTAL
    #   Bloques de 30 min ESTRICTOS → 7 mañana + 5 tarde = 12 slots/día
    # Idempotente: hace UPDATE absoluto a id=1 con los valores correctos.
    # Si ya están en esos valores no cambia nada (MySQL no marca affected
    # rows si el valor no cambia, pero igual es seguro re-ejecutar).
    def ensure_schedule_v3():
        try:
            mysql_execute(
                f"UPDATE `{SET}` SET "
                f"open_time='09:00:00', "
                f"close_time='16:30:00', "
                f"lunch_start='12:30:00', "
                f"lunch_end='14:00:00', "
                f"slot_minutes=30, "
                f"buffer_cierre_min=0 "
                f"WHERE id=1"
            )
        except Exception as e:
            # Si alguna columna no existe (instalación muy vieja) caemos a
            # UPDATEs uno por uno para no perder el resto.
            for sql in [
                f"UPDATE `{SET}` SET open_time='09:00:00' WHERE id=1",
                f"UPDATE `{SET}` SET close_time='16:30:00' WHERE id=1",
                f"UPDATE `{SET}` SET lunch_start='12:30:00' WHERE id=1",
                f"UPDATE `{SET}` SET lunch_end='14:00:00' WHERE id=1",
                f"UPDATE `{SET}` SET slot_minutes=30 WHERE id=1",
                f"UPDATE `{SET}` SET buffer_cierre_min=0 WHERE id=1",
            ]:
                try: mysql_execute(sql)
                except Exception: pass
            print(f"[ensure_schedule_v3] fallback individual aplicado: {e}", flush=True)

    ensure_schedule_v3()

    # Cache de settings — el tracking público y polling lo invocan a menudo.
    # TTL 30s: cualquier cambio de horario/bodega impacta en máx 30s.
    _SETTINGS_CACHE = {"row": None, "fetched_at": 0.0}
    _SETTINGS_TTL = 30.0

    def settings():
        import time as _time
        now = _time.time()
        if _SETTINGS_CACHE["row"] is not None and (now - _SETTINGS_CACHE["fetched_at"]) < _SETTINGS_TTL:
            return _SETTINGS_CACHE["row"]
        row = mysql_fetchone(f"SELECT * FROM `{SET}` WHERE id=1") or {}
        # Defensa runtime v3 (Daniel 2026-05-24): si por cualquier razón el
        # row se desincronizó (ej: alguien tocó BD manualmente), forzamos los
        # valores correctos del horario único. NO toca campos no-horario.
        _NEEDED = {
            "open_time":   "09:00:00",
            "close_time":  "16:30:00",
            "lunch_start": "12:30:00",
            "lunch_end":   "14:00:00",
            "slot_minutes": 30,
            "buffer_cierre_min": 0,
        }
        _drift = False
        if row:
            for _k, _v in _NEEDED.items():
                _cur = row.get(_k)
                if _k in ("slot_minutes", "buffer_cierre_min"):
                    if _cur is None or int(_cur or 0) != _v:
                        _drift = True; break
                else:
                    if str(_cur or "")[:5] != str(_v)[:5]:
                        _drift = True; break
        if _drift:
            try:
                mysql_execute(
                    f"UPDATE `{SET}` SET "
                    f"open_time='09:00:00', close_time='16:30:00', "
                    f"lunch_start='12:30:00', lunch_end='14:00:00', "
                    f"slot_minutes=30, buffer_cierre_min=0 WHERE id=1"
                )
                row = mysql_fetchone(f"SELECT * FROM `{SET}` WHERE id=1") or row
                _SETTINGS_CACHE["row"] = None
            except Exception as _exc:
                print(f"[pickup_settings] drift v3 no corregible: {_exc}", flush=True)
        result = row or {
            "warehouse_name": "Bodega ILUS Quilicura",
            "warehouse_addr": "Av. Presidente Eduardo Frei Montalva 9770, Bod 30, Quilicura.",
            "maps_url": "https://www.google.com/maps/search/?api=1&query=Av.%20Presidente%20Eduardo%20Frei%20Montalva%209770%20Bod%2030%20Quilicura",
            "open_time": "09:00:00",
            "close_time": "16:30:00",
            "lunch_start": "12:30:00",
            "lunch_end": "14:00:00",
            "slot_minutes": 30,
            "buffer_cierre_min": 0,
            "parallel_capacity": 2,
            "min_notice_hours": 24,
            "proposal_expiry_hours": 48,
            "work_days": "1,2,3,4,5",
            "holidays": "",
            "alert_enabled": 0,
            "alert_title": "Aviso importante",
            "alert_message": "",
            "hero_image_1": "",
            "hero_image_2": "",
            "hero_image_3": "",
            "notify_emails": "",
        }
        _SETTINGS_CACHE["row"] = result
        _SETTINGS_CACHE["fetched_at"] = now
        return result

    def _invalidate_settings_cache():
        """Llamar tras cualquier UPDATE de pickup_settings (admin guarda config)."""
        _SETTINGS_CACHE["row"] = None
        _SETTINGS_CACHE["fetched_at"] = 0.0

    def _td_to_hhmm(val):
        """Convierte TIME de MySQL (timedelta de PyMySQL) o string a 'HH:MM' cero-relleno.
        PyMySQL retorna timedelta para columnas TIME; comparar timedelta con str lanza TypeError.

        BUG 2026-06-05 (Cloud SQL): str(timedelta(hours=9)) = '9:00:00' → [:5] = '9:00:'
        (SIN cero inicial). Eso rompía la comparación lexicográfica en time_allowed
        ('09:00' < '9:00:' == True porque '0' < '9') y RECHAZABA TODOS los horarios.
        Por eso normalizamos SIEMPRE a 'HH:MM' cero-rellenado, venga timedelta o string
        ('9:00:00' / '09:00' / '9:0')."""
        try:
            if hasattr(val, "total_seconds"):  # timedelta (PyMySQL TIME)
                total = int(val.total_seconds())
                h, m = divmod(total // 60, 60)
                return f"{h:02d}:{m:02d}"
            parts = str(val).split(":")
            h = int(parts[0]); m = int(parts[1]) if len(parts) > 1 else 0
            return f"{h:02d}:{m:02d}"
        except Exception:
            return str(val)[:5] if val else "00:00"

    def _chile_holidays(year):
        """Set de feriados legales de Chile en 'YYYY-MM-DD' para `year`.

        2026-06-10: la FUENTE DE VERDAD ahora es cl_feriados.py (compartida con
        mantenciones/planificador — un solo lugar para actualizar 2027+). El
        cuerpo de abajo queda como FALLBACK por si el import fallara.

        Hardcode OFICIAL para años conocidos (fuente: feriados.cl). Para años no
        listados, fallback CALCULADO: feriados de fecha fija + Viernes/Sábado Santo
        (Computus). Actualizar la lista cada año. (Juan Daniel 2026-06-05: 'identifica
        los feriados chilenos siempre para que no los pongas como laborables'.)"""
        try:
            from cl_feriados import feriados_set
            return set(feriados_set(year))
        except Exception:
            pass
        OFICIALES = {
            2026: [
                "2026-01-01","2026-04-03","2026-04-04","2026-05-01","2026-05-21",
                "2026-06-21","2026-06-29","2026-07-16","2026-08-15","2026-09-18",
                "2026-09-19","2026-10-12","2026-10-31","2026-11-01","2026-12-08","2026-12-25",
            ],
        }
        if year in OFICIALES:
            return set(OFICIALES[year])
        # Fallback: fijos + Semana Santa (Viernes y Sábado Santo).
        fijos = [(1,1),(5,1),(5,21),(6,29),(7,16),(8,15),(9,18),(9,19),
                 (10,12),(10,31),(11,1),(12,8),(12,25)]
        out = {f"{year:04d}-{m:02d}-{d:02d}" for (m, d) in fijos}
        try:
            a = year % 19; b = year // 100; c = year % 100
            d_ = b // 4; e = b % 4; f = (b + 8) // 25
            g = (b - f + 1) // 3; h = (19*a + b - d_ - g + 15) % 30
            i = c // 4; k = c % 4; l = (32 + 2*e + 2*i - h - k) % 7
            mm = (a + 11*h + 22*l) // 451
            month = (h + l - 7*mm + 114) // 31
            day = ((h + l - 7*mm + 114) % 31) + 1
            pascua = datetime(year, month, day).date()
            out.add((pascua - timedelta(days=2)).isoformat())  # Viernes Santo
            out.add((pascua - timedelta(days=1)).isoformat())  # Sábado Santo
        except Exception:
            pass
        return out

    def date_allowed(date_str, cfg=None):
        cfg = cfg or settings()
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return False, "Fecha no valida."
        days = {int(x) for x in (cfg.get("work_days") or "1,2,3,4,5").split(",") if x.strip().isdigit()}
        holidays = {d.strip() for d in (cfg.get("holidays") or "").replace(";", ",").split(",") if d.strip()}
        holidays |= _chile_holidays(dt.year)   # feriados legales de Chile (auto)
        if dt.isoweekday() not in days:
            return False, "La bodega no recibe retiros ese día (solo días hábiles)."
        if date_str in holidays:
            return False, "La fecha seleccionada es feriado en Chile o está bloqueada."
        return True, ""

    def next_allowed_date(cfg=None):
        cfg = cfg or settings()
        base = datetime.now().date() + timedelta(days=1)
        for offset in range(0, 45):
            candidate = base + timedelta(days=offset)
            ok, _msg = date_allowed(candidate.isoformat(), cfg)
            if ok:
                return candidate.isoformat()
        return base.isoformat()

    def time_allowed(time_from, time_to, cfg=None, bypass_lunch=False):
        """Valida que el rango horario solicitado caiga en el horario de
        atención y NO cruce la colación.

        Daniel mayo 2026: el horario v3 separa Mañana / Tarde con colación
        bloqueada en medio. Un rango que cruce la colación es inválido —
        debe ser SOLO mañana o SOLO tarde.

        Daniel 2026-05-24 (operador interno): se agregó `bypass_lunch=True`
        para permitir al operador agendar manualmente cruzando la colación
        cuando la factura es grande y necesita más tiempo. El cliente
        público sigue con la restricción intacta (bypass_lunch=False).
        """
        cfg = cfg or settings()
        # _td_to_hhmm normaliza timedelta (PyMySQL TIME) o string a 'HH:MM' cero-rellenado.
        # NO usar str(...)[:5] aquí: con timedelta(hours=9) daría '9:00:' y rompería la
        # comparación lexicográfica de abajo (rechazaba TODOS los horarios). Ver _td_to_hhmm.
        open_t  = _td_to_hhmm(cfg.get("open_time")   or "09:00:00")
        close_t = _td_to_hhmm(cfg.get("close_time")  or "16:30:00")
        lunch_s = _td_to_hhmm(cfg.get("lunch_start") or "12:30:00")
        lunch_e = _td_to_hhmm(cfg.get("lunch_end")   or "14:00:00")
        if not time_from or not time_to or time_from >= time_to:
            return False, "Selecciona un rango horario valido."
        # Daniel 2026-06-10: el CIERRE es la última hora de LLEGADA admitida.
        # La grilla pública (v3, 2026-05-24) ofrece el bloque 16:30→17:00 a
        # propósito; este validador exigía que el bloque TERMINARA antes del
        # cierre (16:30) y rechazaba justo ese bloque ("El horario debe estar
        # entre 09:00 y 16:30" al elegir 16:30). Regla nueva: la LLEGADA
        # (time_from) debe caer entre apertura y cierre; el término puede
        # extenderse hasta cierre + 1 bloque (la atención del último cliente).
        try:
            _slot_min = int(cfg.get("slot_minutes") or 30)
        except (TypeError, ValueError):
            _slot_min = 30
        try:
            _cH, _cM = [int(x) for x in close_t.split(":")[:2]]
            _fin_tot = _cH * 60 + _cM + _slot_min
            _fin_max = f"{_fin_tot // 60:02d}:{_fin_tot % 60:02d}"
        except Exception:
            _fin_max = close_t
        if time_from < open_t or time_from > close_t or time_to > _fin_max:
            return False, (
                f"La hora de llegada debe estar entre {open_t} y {close_t} "
                f"(la atención del último bloque termina a las {_fin_max})."
            )
        # Cruce de colación: el rango (time_from, time_to) se solapa con
        # (lunch_s, lunch_e) cuando NO se cumple: time_to <= lunch_s OR time_from >= lunch_e
        if bypass_lunch:
            return True, ""
        try:
            if not (time_to <= lunch_s or time_from >= lunch_e):
                return False, (
                    f"El horario no puede cruzar la colacion ({lunch_s}-{lunch_e}). "
                    f"Elige un rango solo en la mañana o solo en la tarde."
                )
        except Exception:
            pass
        return True, ""

    # ══════════════════════════════════════════════════════════════════
    #  VALIDACIÓN TEMPORAL CENTRAL (Daniel 2026-05-29 — FASE 2)
    #  Fuente ÚNICA de verdad para fecha/hora de un retiro. Reemplaza el
    #  uso suelto de date_allowed() que NO validaba fecha pasada, min_notice
    #  ni timezone. Todas las rutas (solicitud pública, propuesta interna,
    #  confirmación, contrapropuesta, API disponibilidad, API calendario)
    #  deben pasar por acá.
    # ══════════════════════════════════════════════════════════════════
    def _now_chile():
        """Datetime AHORA en hora Chile, como naive (sin tzinfo).
        Railway corre en UTC; usar datetime.now() server rompe 'hoy'/'ahora'
        cerca de medianoche. Comparamos todo en naive-Chile para evitar
        además TypeError aware-vs-naive con los DATETIME de MySQL."""
        try:
            from zoneinfo import ZoneInfo
            return datetime.now(ZoneInfo("America/Santiago")).replace(tzinfo=None)
        except Exception:
            return datetime.now()

    def validate_pickup_datetime(date_str, time_from, time_to, cfg=None,
                                 mode="public", now_chile=None):
        """Valida un slot fecha+hora de retiro. Devuelve (ok: bool, error: str).

        mode:
          - 'public'   : cliente crea/confirma/contrapropone. Colación
                         bloqueada, exige min_notice_hours, prohíbe pasado.
          - 'internal' : operador propone. Permite cruzar colación
                         (bypass_lunch) y NO exige min_notice, PERO sigue
                         prohibiendo fecha/hora pasada.
          - 'calendar' : para pintar disponibilidad (igual que public).

        El VENCIMIENTO de una propuesta concreta se chequea aparte con
        proposal_is_vigente() porque depende de expires_at en la fila.
        """
        cfg = cfg or settings()
        now = now_chile or _now_chile()

        # 1) Formato fecha
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return False, "Fecha no válida."
        # 2) Formato hora
        tf = str(time_from or "")[:5]
        tt = str(time_to or "")[:5]
        try:
            h_from = datetime.strptime(tf, "%H:%M").time()
            h_to   = datetime.strptime(tt, "%H:%M").time()
        except Exception:
            return False, "Horario no válido."

        start_dt = datetime.combine(d, h_from)   # naive-Chile
        today = now.date()

        # 3) Fecha/hora pasada (TODOS los modos)
        if d < today or (d == today and start_dt <= now):
            return False, ("Esta propuesta venció o la fecha ya pasó. "
                           "Propón una nueva fecha.")

        # 4) Anticipación mínima (solo cliente: public/calendar)
        if mode in ("public", "calendar"):
            try:
                min_notice = int(cfg.get("min_notice_hours") or 24)
            except (TypeError, ValueError):
                min_notice = 24
            if (start_dt - now).total_seconds() < min_notice * 3600:
                return False, (f"Se requieren al menos {min_notice} horas de "
                               f"anticipación. Elige una fecha más adelante.")

        # 5) Día hábil + feriados (reusa date_allowed)
        ok_day, msg_day = date_allowed(date_str, cfg)
        if not ok_day:
            return False, msg_day

        # 6) Horario de bodega + colación (operador puede cruzar colación)
        ok_time, msg_time = time_allowed(tf, tt, cfg, bypass_lunch=(mode == "internal"))
        if not ok_time:
            return False, msg_time

        # 7) Duración múltiplo de slot_minutes (default 30)
        try:
            slot = int(cfg.get("slot_minutes") or 30)
        except (TypeError, ValueError):
            slot = 30
        dur_min = (datetime.combine(d, h_to) - datetime.combine(d, h_from)).total_seconds() / 60
        if dur_min <= 0:
            return False, "El horario de término debe ser posterior al de inicio."
        if slot > 0 and int(round(dur_min)) % slot != 0:
            return False, f"La duración debe ser múltiplo de {slot} minutos."

        return True, ""

    def proposal_is_vigente(proposal, now_chile=None):
        """True si la propuesta sigue 'pending' y NO venció (expires_at).
        Filas viejas sin expires_at se consideran vigentes (compat)."""
        if not proposal:
            return False
        if (proposal.get("status") or "") != "pending":
            return False
        exp = proposal.get("expires_at")
        if not exp:
            return True
        now = now_chile or _now_chile()
        # exp viene como datetime naive de MySQL (hora Chile por convención).
        if hasattr(exp, "year"):
            try:
                return now < exp
            except TypeError:
                return True
        return True

    def calc_package(length_cm, width_cm, height_cm, weight_kg):
        l, w, h = [max(float(x or 0), 0) for x in (length_cm, width_cm, height_cm)]
        kg = max(float(weight_kg or 0), 0)
        return {
            "length_cm": l,
            "width_cm": w,
            "height_cm": h,
            "weight_kg": kg,
            "volumetric_weight": round((l * w * h) / 4000, 3) if l and w and h else 0,
            "volume_m3": round((l * w * h) / 1000000, 4) if l and w and h else 0,
        }

    def quality_score(data, packages, signed=False, attachments=0):
        checks = [
            bool(data.get("document_number")),
            bool(data.get("customer_name")),
            bool(data.get("customer_rut")),
            bool(EMAIL_RE.match(data.get("contact_email") or "")),
            bool(re.match(r"^\+?[0-9\s\-]{7,20}$", data.get("contact_phone") or "")),
            bool(data.get("pickup_person_name")),
            bool(data.get("pickup_person_rut")),
            bool(data.get("requested_date")),
            bool(data.get("requested_time_from") and data.get("requested_time_to")),
            bool(packages),
            all(p["length_cm"] and p["width_cm"] and p["height_cm"] for p in packages),
            all(p["weight_kg"] for p in packages),
            attachments > 0,
            signed,
        ]
        score = int(round((sum(1 for ok in checks if ok) / len(checks)) * 100))
        risk = 0
        if not data.get("customer_rut") or not data.get("pickup_person_rut"):
            risk += 25
        if (data.get("pickup_person_relation") or "") in {"chofer", "transporte", "otro"}:
            risk += 15
        if score < 70:
            risk += 20
        return min(score, 100), min(risk, 100)

    def status_badge(status):
        return {"label": PICKUP_STATUS.get(status, status), "color": PICKUP_STATUS_COLORS.get(status, "secondary")}

    # ══════════════════════════════════════════════════════════════════
    #  CÓDIGO ALEATORIO RET-XXXXXX (Daniel 2026-05-23)
    # ══════════════════════════════════════════════════════════════════
    # Antes los códigos eran predecibles: RET-000010, RET-000011 — un
    # atacante podía iterar 1..N para listar todos los retiros.
    # Ahora: 6 chars de alfabeto sin ambigüedad (sin O/0, I/1/L) =
    # 32^6 ≈ 1.07 mil millones de combinaciones. Colisión despreciable.
    # Compatibilidad: códigos viejos RET-000010 SIGUEN sirviendo
    # (solo cambia el formato de los NUEVOS).
    _RET_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"  # 32 chars, sin O/0/I/1/L
    _RET_LEN = 6

    def _generate_pickup_code(cur=None, max_attempts=20):
        """Genera un código único RET-XXXXXX con `secrets.choice()`.

        - Reintenta hasta `max_attempts` veces si choca con UNIQUE en BD.
        - Si recibe `cur` (cursor dentro de una transacción), usa ese para
          el SELECT — así no abre conexión nueva durante un INSERT pending.
        - Si no recibe `cur`, usa `mysql_fetchone` normal.
        - Tras max_attempts intentos fallidos, levanta RuntimeError (caso
          extremo virtualmente imposible: ~32^6 colisiones).
        """
        import secrets as _secrets
        for _ in range(max_attempts):
            tail = "".join(_secrets.choice(_RET_ALPHABET) for _ in range(_RET_LEN))
            code = f"RET-{tail}"
            try:
                if cur is not None:
                    cur.execute(f"SELECT 1 FROM `{REQ}` WHERE code=%s LIMIT 1", (code,))
                    if not cur.fetchone():
                        return code
                else:
                    existing = mysql_fetchone(
                        f"SELECT 1 AS n FROM `{REQ}` WHERE code=%s LIMIT 1", (code,)
                    )
                    if not existing:
                        return code
            except Exception:
                # Si falla el SELECT (raro), igual probemos retornar el code —
                # el UNIQUE de BD lo va a rechazar y reintentaremos en INSERT.
                return code
        # Caso extremo: no se logró un código único en max_attempts.
        # Fallback: agregar timestamp suffix para diferenciarlo.
        import time as _t
        return f"RET-{tail}{int(_t.time()) % 100:02d}"


    def log_event(request_id, action, old_status=None, new_status=None, notes="", actor_type="sistema", actor_name=None):
        try:
            # ip/ua solo si hay request activo: log_event también se llama desde
            # threads daemon (_bg_post_insert) donde request.* lanza RuntimeError
            # y antes hacía perder el INSERT completo (fix 2026-06-09).
            _ip = _ua = None
            _actor_fallback = "Cliente"
            try:
                if has_request_context():
                    _ip = request.remote_addr
                    _ua = (request.user_agent.string or "")[:300]
                    _actor_fallback = g.user["nombre"] if getattr(g, "user", None) else "Cliente"
            except Exception:
                pass
            mysql_execute(
                f"""INSERT INTO `{LOG}`
                    (request_id,actor_type,actor_name,action,old_status,new_status,notes,ip,user_agent)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    request_id,
                    actor_type,
                    actor_name or _actor_fallback,
                    action,
                    old_status,
                    new_status,
                    notes,
                    _ip,
                    _ua,
                ),
            )
        except Exception as exc:
            print(f"[ILUS][PICKUP LOG] {exc}")

    # ══════════════════════════════════════════════════════════════════
    #  VALIDACIÓN DE DISPONIBILIDAD REAL DE SLOT (cupos + bloqueos + colación)
    # ══════════════════════════════════════════════════════════════════
    def _validar_disponibilidad_slot(date, time_from, time_to, exclude_request_id=None,
                                      extra_kg=0, extra_m3=0, bypass_lunch=False):
        """Valida que un slot (date + time_from..time_to) pueda usarse.

        Chequea, en este orden:
          1) Día permitido (work_days + holidays) y horario dentro de open/close
          2) Solape con colación (lunch_start/lunch_end) — saltado si bypass_lunch=True
          3) Solape con pickup_blocks (full day o franja específica)
          4) Capacidad: max_picks_per_slot, max_kg_per_slot, max_m3_per_slot
          5) Capacidad diaria: max_picks_per_day

        Args:
            date: 'YYYY-MM-DD' (str) o date
            time_from, time_to: 'HH:MM' (str)
            exclude_request_id: id de pickup_request a excluir del conteo
                                (útil al revalidar la propia confirmación)
            extra_kg, extra_m3: peso/volumen del retiro que se está agregando
                                (se suma al conteo actual del slot)
            bypass_lunch: si True, permite que el rango cruce la colación.
                          Solo el OPERADOR interno debe activar esta opción
                          (factura grande que necesita más tiempo). El cliente
                          público nunca pasa este flag → mantiene la regla.

        Returns:
            (ok: bool, motivo: str)
        """
        cfg = settings()
        try:
            date_str = date.isoformat() if hasattr(date, "isoformat") else str(date)
        except Exception:
            return False, "Fecha inválida."
        # Normalizar: PyMySQL devuelve timedelta para columnas TIME. Comparar
        # timedelta con str (open_t="09:00") lanzaría TypeError en time_allowed.
        time_from = _td_to_hhmm(time_from)
        time_to   = _td_to_hhmm(time_to)

        # 1) Día permitido (work_days + holidays) y rango horario válido
        ok_d, msg_d = date_allowed(date_str, cfg)
        if not ok_d:
            return False, msg_d
        ok_t, msg_t = time_allowed(time_from, time_to, cfg, bypass_lunch=bypass_lunch)
        if not ok_t:
            return False, msg_t

        # Parsear minutos
        try:
            sh, sm = [int(x) for x in str(time_from)[:5].split(":")]
            eh, em = [int(x) for x in str(time_to)[:5].split(":")]
            slot_s = sh * 60 + sm
            slot_e = eh * 60 + em
        except Exception:
            return False, "Horario con formato inválido."

        # Daniel 2026-05-24: FUERZA ABSOLUTA bloques de 30 min.
        # Antes leíamos `slot_minutes` de la BD; si la BD tenía legacy 60,
        # el cliente que elegía las 10:30 recibía "debe caer en bloques de
        # 60 min". HARDCODED a 30 min sin excepción.
        # Tampoco bloqueamos por divisibilidad estricta desde open_time si
        # el cliente eligió un slot válido del calendario — el frontend ya
        # genera SOLO bloques válidos. Acá solo verificamos múltiplo de 30.
        slot_min = 30
        try:
            if (slot_s % slot_min) != 0:
                return False, f"La hora de inicio debe caer en :00 o :30."
            if (slot_e % slot_min) != 0:
                return False, f"La hora de fin debe caer en :00 o :30."
            if (slot_e - slot_s) % slot_min != 0:
                return False, f"La duración debe ser múltiplo de 30 minutos."
            if (slot_e - slot_s) <= 0:
                return False, "La hora de fin debe ser posterior a la de inicio."
        except Exception:
            pass

        # 2) Solape con bloque mañana-tarde (Daniel 2026-05-25: 12:30 – 14:00)
        # 12:30-13:00 = buffer interno bodega (atender desordenados).
        # 13:00-14:00 = colación. Para el cliente público AMBOS están bloqueados.
        # bypass_lunch=True: operador puede pasarse de largo (factura grande).
        if not bypass_lunch:
            lunch_s_str = "12:30"
            lunch_e_str = "14:00"
            try:
                lh, lm = [int(x) for x in lunch_s_str.split(":")]
                leh, lem = [int(x) for x in lunch_e_str.split(":")]
                lunch_s = lh * 60 + lm
                lunch_e = leh * 60 + lem
                if lunch_s < lunch_e:
                    # Solape si NO termina antes del almuerzo y NO empieza después
                    if not (slot_e <= lunch_s or slot_s >= lunch_e):
                        return False, f"El horario solapa la colación ({lunch_s_str}-{lunch_e_str})."
            except Exception:
                pass

        # 3) Bloqueos manuales en pickup_blocks
        try:
            blk_rows = mysql_fetchall(
                "SELECT hora_inicio, hora_fin, motivo FROM pickup_blocks WHERE fecha=%s",
                (date_str,),
            ) or []
            for b in blk_rows:
                hi = b.get("hora_inicio")
                hf = b.get("hora_fin")
                motivo = (b.get("motivo") or "bloqueado").strip()
                # Día completo bloqueado
                if not hi:
                    return False, f"Día bloqueado: {motivo}"
                try:
                    bh, bm = [int(x) for x in str(hi)[:5].split(":")]
                    bs = bh * 60 + bm
                    if hf:
                        beh, bem = [int(x) for x in str(hf)[:5].split(":")]
                        be = beh * 60 + bem
                    else:
                        be = 24 * 60  # hasta cierre
                    if not (slot_e <= bs or slot_s >= be):
                        return False, f"Franja bloqueada: {motivo}"
                except Exception:
                    continue
        except Exception:
            pass  # tabla puede no existir aún

        # 4) Capacidad del slot (picks + kg + m3)
        # parallel_capacity = límite de retiros simultáneos (default 2).
        # Si no está configurado, caemos al legacy max_picks_per_slot (5)
        # para mantener compatibilidad con instalaciones viejas.
        parallel_capacity_cfg = cfg.get("parallel_capacity")
        if parallel_capacity_cfg is not None and str(parallel_capacity_cfg).strip():
            max_picks_slot = int(parallel_capacity_cfg)
        else:
            max_picks_slot = int(cfg.get("max_picks_per_slot") or 2)
        max_kg_slot = float(cfg.get("max_kg_per_slot") or 500)
        max_m3_slot = float(cfg.get("max_m3_per_slot") or 5)
        max_picks_day = int(cfg.get("max_picks_per_day") or 30)

        # Conteo en el MISMO slot (mismo time_from). Usamos confirmed_* o proposed_*
        # según el estado real. Excluimos cancelados/cerrados.
        exclude_clause = ""
        params = [date_str, str(time_from)[:5], date_str, str(time_from)[:5]]
        if exclude_request_id:
            exclude_clause = "AND id <> %s"
            params.append(int(exclude_request_id))
        slot_row = mysql_fetchone(
            f"""SELECT COUNT(*) AS n,
                       COALESCE(SUM(total_weight_kg),0) AS kg,
                       COALESCE(SUM(total_volume_m3),0) AS m3
                FROM `{REQ}`
                WHERE status NOT IN ('rechazada','cerrada','fallida')
                  AND (
                    (confirmed_date=%s AND TIME_FORMAT(confirmed_time_from,'%%H:%%i')=%s)
                    OR
                    (confirmed_date IS NULL AND proposed_date=%s
                     AND TIME_FORMAT(proposed_time_from,'%%H:%%i')=%s)
                  )
                  {exclude_clause}""",
            tuple(params),
        ) or {}
        picks_now = int(slot_row.get("n") or 0)
        kg_now = float(slot_row.get("kg") or 0)
        m3_now = float(slot_row.get("m3") or 0)

        if picks_now + 1 > max_picks_slot:
            return False, f"Slot lleno: {picks_now} retiros (máximo {max_picks_slot})."
        if kg_now + float(extra_kg or 0) > max_kg_slot:
            return False, f"Capacidad de peso excedida: {kg_now:.1f}+{float(extra_kg or 0):.1f} kg > {max_kg_slot:.0f} kg."
        if m3_now + float(extra_m3 or 0) > max_m3_slot:
            return False, f"Capacidad de volumen excedida: {m3_now:.2f}+{float(extra_m3 or 0):.2f} m³ > {max_m3_slot:.2f} m³."

        # 5) Capacidad diaria
        day_params = [date_str, date_str]
        day_exclude = ""
        if exclude_request_id:
            day_exclude = "AND id <> %s"
            day_params.append(int(exclude_request_id))
        day_row = mysql_fetchone(
            f"""SELECT COUNT(*) AS n FROM `{REQ}`
                WHERE status NOT IN ('rechazada','cerrada','fallida')
                  AND (confirmed_date=%s OR (confirmed_date IS NULL AND proposed_date=%s))
                  {day_exclude}""",
            tuple(day_params),
        ) or {}
        picks_day = int(day_row.get("n") or 0)
        if picks_day + 1 > max_picks_day:
            return False, f"Día completo: {picks_day} retiros (máximo diario {max_picks_day})."

        return True, ""

    # ── CALENDARIO (.ics + links Google/Outlook) — Daniel 2026-06-17 ─────
    #  El cliente (y el equipo) puede AGENDAR el retiro en su calendario desde
    #  el correo. Las columnas confirmed_date/time_* están en hora de PARED de
    #  Chile (NO UTC), así que para el .ics y los links convertimos a UTC con
    #  zoneinfo('America/Santiago') (DST-aware) y emitimos sufijo Z — formato
    #  inequívoco y compatible con Gmail/Outlook/Apple sin VTIMEZONE.
    def _pickup_event_dt(req, proposal=None):
        """(start, end) como datetime NAIVE en hora local Chile para el evento,
        o (None, None) si no hay fecha/hora. Prioridad confirmed > proposed >
        requested (mismo orden que el horario mostrado)."""
        from datetime import datetime as _dt, date as _date, timedelta as _tdelta
        def _as_date(v):
            if not v:
                return None
            if isinstance(v, _date) and not isinstance(v, _dt):
                return v
            if isinstance(v, _dt):
                return v.date()
            try:
                return _dt.strptime(str(v)[:10], "%Y-%m-%d").date()
            except Exception:
                return None
        def _as_hm(v):
            # v puede ser time, timedelta o str 'HH:MM[:SS]'
            s = str(v) if v is not None else ""
            if not s:
                return None
            parts = s.split(":")
            try:
                h = int(parts[0]); m = int(parts[1]) if len(parts) > 1 else 0
                return h % 24, max(0, min(59, m))
            except Exception:
                return None
        p = proposal or {}
        d = _as_date(p.get("date")) or _as_date(req.get("confirmed_date")) \
            or _as_date(req.get("proposed_date")) or _as_date(req.get("requested_date"))
        if not d:
            return None, None
        tf = p.get("time_from") or req.get("confirmed_time_from") \
            or req.get("proposed_time_from") or req.get("requested_time_from")
        tt = p.get("time_to") or req.get("confirmed_time_to") \
            or req.get("proposed_time_to") or req.get("requested_time_to")
        hm_f = _as_hm(tf)
        if not hm_f:
            return None, None
        start = _dt(d.year, d.month, d.day, hm_f[0], hm_f[1])
        hm_t = _as_hm(tt)
        if hm_t:
            end = _dt(d.year, d.month, d.day, hm_t[0], hm_t[1])
            if end <= start:
                end = start + _tdelta(hours=1)
        else:
            end = start + _tdelta(hours=1)
        return start, end

    def _pickup_event_location(req):
        cfg = settings()
        wn = cfg.get("warehouse_name") or "Bodega ILUS"
        wa = cfg.get("warehouse_addr") or ""
        return (wn + (", " + wa if wa else "")).strip()

    def _build_pickup_ics(req, proposal=None):
        """Bytes del .ics (VEVENT) del retiro, o None si no hay fecha/hora.
        UID ESTABLE por retiro (retiro-{code}@ilusfitness.com) → un reenvío
        ACTUALIZA el evento en vez de duplicarlo. Hora en UTC (sufijo Z)."""
        try:
            from datetime import datetime as _dt
            from zoneinfo import ZoneInfo as _ZI
            start, end = _pickup_event_dt(req, proposal)
            if not start or not end:
                return None
            _tz = _ZI("America/Santiago"); _utc = _ZI("UTC")
            su = start.replace(tzinfo=_tz).astimezone(_utc)
            eu = end.replace(tzinfo=_tz).astimezone(_utc)
            code = (req.get("code") or "RETIRO").strip()
            location = _pickup_event_location(req)
            persona = req.get("pickup_person_name") or req.get("contact_name") or ""
            link = _public_base_url() + "/retiros/seguimiento/" + str(req.get("public_token") or "")
            def _esc(s):
                return (str(s or "").replace("\\", "\\\\").replace(",", "\\,")
                        .replace(";", "\\;").replace("\n", "\\n").replace("\r", ""))
            desc = f"Retiro ILUS {code}."
            if persona:
                desc += f" Persona que retira: {persona}."
            if req.get("public_token"):
                desc += f" Seguimiento: {link}"
            lines = [
                "BEGIN:VCALENDAR", "VERSION:2.0",
                "PRODID:-//ILUS Sport & Health//Retiros//ES",
                "CALSCALE:GREGORIAN", "METHOD:PUBLISH",
                "BEGIN:VEVENT",
                f"UID:retiro-{code}@ilusfitness.com",
                f"DTSTAMP:{_dt.utcnow().strftime('%Y%m%dT%H%M%SZ')}",
                f"DTSTART:{su.strftime('%Y%m%dT%H%M%SZ')}",
                f"DTEND:{eu.strftime('%Y%m%dT%H%M%SZ')}",
                f"SUMMARY:{_esc('Retiro ILUS ' + code)}",
                f"LOCATION:{_esc(location)}",
                f"DESCRIPTION:{_esc(desc)}",
            ]
            if req.get("public_token"):
                lines.append(f"URL:{_esc(link)}")
            lines += [
                "STATUS:CONFIRMED",
                "BEGIN:VALARM", "TRIGGER:-PT2H", "ACTION:DISPLAY",
                f"DESCRIPTION:{_esc('Recordatorio retiro ILUS ' + code)}",
                "END:VALARM",
                "END:VEVENT", "END:VCALENDAR",
            ]
            return ("\r\n".join(lines) + "\r\n").encode("utf-8")
        except Exception as exc:
            try:
                print(f"[pickup-ics] no se pudo construir .ics: {exc}", flush=True)
            except Exception:
                pass
            return None

    def _pickup_calendar_links(req, proposal=None):
        """{'google':url, 'outlook':url} para 'Agregar a calendario' desde el
        cuerpo del correo. Vacío si no hay fecha/hora. Google usa UTC (Z)."""
        try:
            from urllib.parse import quote_plus
            from zoneinfo import ZoneInfo as _ZI
            start, end = _pickup_event_dt(req, proposal)
            if not start or not end:
                return {"google": "", "outlook": ""}
            _tz = _ZI("America/Santiago"); _utc = _ZI("UTC")
            su = start.replace(tzinfo=_tz).astimezone(_utc)
            eu = end.replace(tzinfo=_tz).astimezone(_utc)
            code = (req.get("code") or "RETIRO").strip()
            title = f"Retiro ILUS {code}"
            location = _pickup_event_location(req)
            link = _public_base_url() + "/retiros/seguimiento/" + str(req.get("public_token") or "")
            details = f"Retiro ILUS {code}." + (f" Seguimiento: {link}" if req.get("public_token") else "")
            google = (
                "https://calendar.google.com/calendar/render?action=TEMPLATE"
                "&text=" + quote_plus(title)
                + "&dates=" + su.strftime("%Y%m%dT%H%M%SZ") + "/" + eu.strftime("%Y%m%dT%H%M%SZ")
                + "&location=" + quote_plus(location)
                + "&details=" + quote_plus(details)
            )
            outlook = (
                "https://outlook.live.com/calendar/0/deeplink/compose?path=/calendar/action/compose&rru=addevent"
                "&subject=" + quote_plus(title)
                # UTC con sufijo Z (igual que Google). Sin Z, Outlook interpreta
                # la hora en la zona del DESTINATARIO → evento corrido si no está
                # en Chile. Con Z queda inequívoco para cualquier zona. (review 2026-06-17)
                + "&startdt=" + su.strftime("%Y-%m-%dT%H:%M:%SZ")
                + "&enddt=" + eu.strftime("%Y-%m-%dT%H:%M:%SZ")
                + "&location=" + quote_plus(location)
                + "&body=" + quote_plus(details)
            )
            return {"google": google, "outlook": outlook}
        except Exception:
            return {"google": "", "outlook": ""}

    def _pickup_productos_html(req):
        """Tabla HTML con los productos del retiro para el correo (Daniel
        2026-06-17: 'información completa del producto'). Cada línea muestra
        producto + SKU + cantidad + indicador de STOCK (✓ en stock / ⚠ sin
        saldo ERP, usando el flag marcada_sin_saldo ya persistido — NO consulta
        el ERP en el hilo de envío). Vacío si el retiro aún no tiene productos.
        Best-effort: '' ante cualquier error (nunca rompe el correo)."""
        try:
            import html as _h
            rid = req.get("id")
            if not rid:
                return ""
            data = _pickup_lineas_consolidadas(int(rid)) or {}
            lineas = data.get("lineas") or []
            if not lineas:
                return ""
            MAX = 12
            filas = []
            for ln in lineas[:MAX]:
                desc = (ln.get("descripcion") or ln.get("sku") or "Producto").strip()
                sku = (ln.get("sku") or "").strip()
                qty = ln.get("cantidad") or 0
                try:
                    qty_str = str(int(qty)) if float(qty) == int(float(qty)) else f"{float(qty):.2f}"
                except Exception:
                    qty_str = str(qty)
                if bool(ln.get("marcada_sin_saldo")):
                    badge = ('<span style="display:inline-block;background:#fff8e1;color:#92400e;'
                             'border:1px solid #fcd34d;border-radius:50px;font-size:10px;font-weight:800;'
                             'padding:2px 8px">⚠ sin saldo ERP</span>')
                else:
                    badge = ('<span style="display:inline-block;background:#dcfce7;color:#14532d;'
                             'border:1px solid #86efac;border-radius:50px;font-size:10px;font-weight:800;'
                             'padding:2px 8px">✓ en stock</span>')
                filas.append(
                    '<tr>'
                    f'<td style="padding:8px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#0a0a0a">{_h.escape(desc)}'
                    + (f'<br><span style="font-size:11px;color:#9ca3af">SKU {_h.escape(sku)}</span>' if sku else '')
                    + '</td>'
                    f'<td align="center" style="padding:8px 10px;border-bottom:1px solid #f0f0f0;font-size:14px;font-weight:800;color:#0a0a0a">{qty_str}</td>'
                    f'<td align="center" style="padding:8px 10px;border-bottom:1px solid #f0f0f0">{badge}</td>'
                    '</tr>'
                )
            extra = ""
            if len(lineas) > MAX:
                extra = ('<tr><td colspan="3" style="padding:8px 10px;font-size:12px;'
                         f'color:#6b7280;text-align:center">… y {len(lineas) - MAX} producto(s) más</td></tr>')
            return (
                '<table cellpadding="0" cellspacing="0" width="100%" '
                'style="background:#ffffff;border:1px solid #e5e7eb;border-radius:10px;margin:0 0 18px;overflow:hidden">'
                '<tr><td style="background:#0a0a0a;color:#fff;padding:10px 14px;font-size:12px;'
                'font-weight:800;text-transform:uppercase;letter-spacing:.06em">📦 Productos de tu retiro</td></tr>'
                '<tr><td style="padding:2px 6px 6px"><table cellpadding="0" cellspacing="0" width="100%">'
                '<tr style="background:#f8fafc">'
                '<td style="padding:8px 10px;font-size:10px;color:#6b7280;text-transform:uppercase;font-weight:700">Producto</td>'
                '<td align="center" style="padding:8px 10px;font-size:10px;color:#6b7280;text-transform:uppercase;font-weight:700">Cant.</td>'
                '<td align="center" style="padding:8px 10px;font-size:10px;color:#6b7280;text-transform:uppercase;font-weight:700">Stock</td></tr>'
                + "".join(filas) + extra +
                '</table></td></tr></table>'
            )
        except Exception as exc:
            try:
                print(f"[pickup-productos-html] {exc}", flush=True)
            except Exception:
                pass
            return ""

    def _render_pickup_vars(req, proposal=None):
        """Construye el dict de variables disponibles en plantillas de retiros."""
        cfg = settings()
        def _hm(t):
            return str(t)[:5] if t else ""
        def _fmt_horario(tf, tt):
            tf, tt = _hm(tf), _hm(tt)
            return f"{tf} - {tt}" if tf else ""
        return {
            "code":              req.get("code") or "",
            "cliente":           req.get("customer_name") or "",
            "persona_retira":    req.get("pickup_person_name") or req.get("contact_name") or "",
            "documento":         f"{(req.get('document_type') or '').upper()} {req.get('document_number') or ''}".strip(),
            "fecha_solicitada":  str(req.get("requested_date") or ""),
            "fecha_propuesta":   str((proposal or {}).get("date") or req.get("proposed_date") or ""),
            "fecha_confirmada":  str(req.get("confirmed_date") or ""),
            "horario": (
                _fmt_horario((proposal or {}).get("time_from"), (proposal or {}).get("time_to"))
                or _fmt_horario(req.get("confirmed_time_from"), req.get("confirmed_time_to"))
                or _fmt_horario(req.get("proposed_time_from"), req.get("proposed_time_to"))
                or _fmt_horario(req.get("requested_time_from"), req.get("requested_time_to"))
            ),
            "n_bultos":          str(req.get("total_packages") or 0),
            "kg":                str(req.get("total_weight_kg") or 0),
            "m3":                str(req.get("total_volume_m3") or 0),
            "warehouse_name":    cfg.get("warehouse_name") or "Bodega ILUS Quilicura",
            "warehouse_addr":    cfg.get("warehouse_addr") or "",
            # Link literal (NO url_for): esta función corre también en threads
            # daemon (notify_async / _bg_post_insert) donde url_for(_external=True)
            # lanza "Unable to build URLs outside an active request".
            "link_seguimiento":  _public_base_url() + "/retiros/seguimiento/" + str(req.get("public_token") or ""),
            # Calendario (Daniel 2026-06-17): links 'Agregar a calendario'. Vacíos
            # si el retiro aún no tiene fecha/hora definida. El editor de plantillas
            # puede incrustarlos con {{link_google_calendar}} / {{link_outlook_calendar}}.
            "link_google_calendar":  _pickup_calendar_links(req, proposal).get("google", ""),
            "link_outlook_calendar": _pickup_calendar_links(req, proposal).get("outlook", ""),
            # Tabla de productos del retiro (Daniel 2026-06-17). Vacía si el
            # retiro aún no tiene documentos/productos asociados.
            "productos_html":        _pickup_productos_html(req),
        }


    def _apply_template(text, variables):
        """Reemplaza {{var}} con su valor (string)."""
        if not text:
            return ""
        for k, v in variables.items():
            text = text.replace("{{" + k + "}}", str(v))
            text = text.replace("{{ " + k + " }}", str(v))
        # Barrido de seguridad (2026-06-14): si desde el editor se introduce un
        # {{token}} nuevo o con tipeo que no está en `variables`, NO lo dejamos
        # llegar literal al cliente. Solo limpia {{ identificador }} simple.
        # Mismo patrón que el barrido central de _render_comm_template (app.py).
        import re as _re
        text = _re.sub(r"\{\{\s*\w+\s*\}\}", "", text)
        return text


    def _get_pickup_template(estado, canal):
        """Lee plantilla de comm_templates para retiros (modulo='retiros').
        Respeta la columna 'activo' (apagar desde el editor → None → fallback),
        igual que _render_comm_template. Doble try: si la columna 'activo' no
        existe todavía (no corrió el ensure), reintenta sin ella para NO romper
        el comportamiento actual de retiros."""
        try:
            try:
                row = mysql_fetchone(
                    "SELECT asunto, cuerpo, COALESCE(activo,1) AS activo "
                    "FROM comm_templates "
                    "WHERE modulo='retiros' AND estado=%s AND canal=%s LIMIT 1",
                    (estado, canal)
                )
            except Exception:
                row = mysql_fetchone(
                    "SELECT asunto, cuerpo FROM comm_templates "
                    "WHERE modulo='retiros' AND estado=%s AND canal=%s LIMIT 1",
                    (estado, canal)
                )
            if not row:
                return None
            try:
                if "activo" in row and not int(row.get("activo", 1) or 0):
                    return None  # plantilla apagada desde el editor → fallback
            except Exception:
                pass
            return row
        except Exception:
            return None


    def _pickup_email_stepper(active_idx):
        """Stepper de 5 hitos email-safe (tabla) para el correo del cliente.
        Mismo modelo canónico PICKUP_JOURNEY que el seguimiento y las plantillas
        de BD. Verde=hecho, rojo=actual, gris=pendiente. (Daniel 2026-06-17: el
        tracking debe ir SIEMPRE en el correo, también en el fallback.)"""
        try:
            pasos = [(p["emoji"], p["label"]) for p in PICKUP_JOURNEY]
            celdas, labels = [], []
            n = len(pasos)
            for i, (ic, lb) in enumerate(pasos):
                if i < active_idx:
                    bg, fg, ex = "#16a34a", "#ffffff", ""
                elif i == active_idx:
                    bg, fg, ex = "#dc2626", "#ffffff", "box-shadow:0 0 0 4px rgba(220,38,38,.15);"
                else:
                    bg, fg, ex = "#f3f4f6", "#9ca3af", "border:1px solid #e5e7eb;"
                celdas.append(
                    f'<td align="center" width="11%" style="padding:0">'
                    f'<div style="width:34px;height:34px;line-height:34px;border-radius:17px;'
                    f'background:{bg};color:{fg};font-family:Helvetica,Arial,sans-serif;font-size:15px;'
                    f'font-weight:900;text-align:center;margin:0 auto;{ex}">{ic}</div></td>')
                if i < n - 1:
                    leg = "#16a34a" if i < active_idx else ("#dc2626" if i == active_idx else "#e5e7eb")
                    celdas.append(
                        f'<td width="11.5%" style="padding:0 2px"><div style="height:4px;'
                        f'background:{leg};border-radius:2px;font-size:0;line-height:0">&nbsp;</div></td>')
                lc = "#16a34a" if i < active_idx else ("#dc2626" if i == active_idx else "#9ca3af")
                lw = "800" if i == active_idx else "600"
                labels.append(
                    f'<td align="center" width="20%" style="font-family:Helvetica,Arial,sans-serif;'
                    f'font-size:10px;color:{lc};text-transform:uppercase;font-weight:{lw};'
                    f'letter-spacing:.04em">{lb}</td>')
            return (
                '<table cellpadding="0" cellspacing="0" width="100%" '
                'style="background:#ffffff;border:1px solid #ececef;border-radius:12px;margin:0 0 18px">'
                '<tr><td style="padding:20px 14px 16px">'
                '<table cellpadding="0" cellspacing="0" width="100%"><tr>' + "".join(celdas) +
                '</tr></table><table cellpadding="0" cellspacing="0" width="100%" style="margin-top:9px"><tr>' +
                "".join(labels) + '</tr></table></td></tr></table>')
        except Exception:
            return ""

    # Mapeo: kind del notify() → estado en comm_templates
    # Cada kind dispara una plantilla del módulo 'retiros'. Si la plantilla
    # no existe, notify() cae al template hardcoded del wrapper ILUS.
    _KIND_TO_ESTADO = {
        "created":           "solicitud_recibida",
        "proposal":          "propuesta_enviada",
        "confirmed":         "agenda_confirmada",
        "preparing":         "en_preparacion",
        "done":              "retirada",
        "rejected":          "rechazada",
        "rescheduled":       "reagendada",
        "reminder_24h":      "recordatorio_24h",
        "info_incompleta":   "informacion_incompleta",
        "failed":            "rechazada",     # reusamos plantilla de cancelación
        "closed":            "retirada",      # reusamos plantilla de retirada
        "message":           None,            # custom — sin plantilla
    }


    def notify(req, kind="created", proposal=None, custom_message=""):
        """Envía notificación al cliente usando plantillas configuradas en
        Comunicaciones → Plantillas → Retiros (DB).

        Si la plantilla no existe en BD, cae al template hardcoded original.
        """
        cfg = settings()
        # Link literal (NO url_for): notify() corre en threads daemon donde
        # url_for(_external=True) falla sin SERVER_NAME (fix 2026-06-09).
        follow_url = _public_base_url() + "/retiros/seguimiento/" + str(req.get("public_token") or "")
        variables = _render_pickup_vars(req, proposal)
        estado = _KIND_TO_ESTADO.get(kind)

        # ── CALENDARIO (.ics) — Daniel 2026-06-17 ──────────────────────
        # Adjuntamos el evento SOLO cuando el retiro tiene cita confirmada
        # (kind confirmed / reminder_24h). UID estable por retiro → un reenvío
        # ACTUALIZA el evento, no lo duplica. Best-effort: si falla, el correo
        # igual sale sin adjunto (no rompemos el envío al cliente).
        _ics_att = None
        if kind in ("confirmed", "reminder_24h"):
            try:
                _ics_bytes = _build_pickup_ics(req, proposal)
                if _ics_bytes:
                    _ics_att = [(f"retiro_{(req.get('code') or 'cita')}.ics",
                                 _ics_bytes, "text/calendar")]
            except Exception:
                _ics_att = None

        # ── EMAIL ──────────────────────────────────────────────────────
        sent_mail = False
        try:
            tpl_email = _get_pickup_template(estado, "email") if estado else None
            if tpl_email and (tpl_email.get("asunto") or tpl_email.get("cuerpo")):
                # Plantilla configurada en BD: usar con variables interpoladas
                asunto = _apply_template(tpl_email.get("asunto") or "", variables)
                cuerpo = _apply_template(tpl_email.get("cuerpo") or "", variables)
                # FIX 2026-06-19 (Daniel: "los correos me llegan SIN tracking, un
                # perfil distinto al que enviamos"). CAUSA RAÍZ: si la plantilla de
                # BD fue editada a mano (o quedó vieja) su `cuerpo` NO trae el
                # stepper, y el re-seed no la pisa por el guard updated_by. FIX
                # bulletproof: inyectamos el tracking AL VUELO si el cuerpo no lo
                # tiene ya. Marcador: el círculo del stepper usa border-radius:17px.
                # Solo hitos válidos (idx>=0; no en cancelación).
                try:
                    _aidx_db = pickup_journey_idx(estado) if estado else -99
                    if _aidx_db >= 0 and "border-radius:17px" not in cuerpo:
                        cuerpo = _pickup_email_stepper(_aidx_db) + cuerpo
                except Exception:
                    pass
                # Envolver en el wrapper HTML oficial ILUS
                # El `cuerpo` de la plantilla de retiros YA es un email completo y
                # autosuficiente (hero + stepper + datos + CTA — diseño 2026-06).
                # Por eso NO lo envolvemos con _ilus_email_html, que agregaría OTRO
                # header negro + botones + Bodega/Dirección DUPLICADOS (el "doble
                # header negro" y la repetición que reportó Daniel). Usamos el
                # wrapper de marca limpio: solo el logo ILUS grande arriba y el
                # eslogan "I LIKE U STRONG" abajo. (Juan Daniel 2026-06-05.)
                try:
                    from app import _comm_render_email_document
                    html = _comm_render_email_document(asunto, cuerpo)
                except Exception:
                    # Fallback defensivo: si el wrapper limpio no estuviera
                    # disponible, no perdemos el envío (diseño anterior).
                    html = _ilus_email_html(
                        titulo=asunto or f"Actualización retiro {req['code']}",
                        subtitulo=f"{req['code']} - {variables['documento']}",
                        saludo=variables["persona_retira"],
                        parrafos=[cuerpo],   # cuerpo ya viene como HTML
                        btn_primario_txt="Ver mi retiro en vivo",
                        btn_primario_url=follow_url,
                    )
                # Multi-email: envía al cliente declarado + extra_emails + emails del ERP
                _multi = _send_pickup_email_multi(req, f"ILUS — {asunto}", html, attachments=_ics_att)
                sent_mail = len(_multi["sent"]) > 0
                if _multi["sent"]:
                    print(f"[pickup-email] {asunto} → {_multi['total']} dest: "
                          f"{_multi['sent']} (failed: {_multi['failed']})", flush=True)
            else:
                # Fallback: plantilla hardcoded original
                titles = {
                    "created": "Solicitud de retiro recibida",
                    "proposal": "ILUS propuso una agenda de retiro",
                    "confirmed": "Agenda de retiro confirmada",
                    "preparing": "Estamos preparando tu retiro",
                    "done": "Tu retiro fue completado",
                    "reminder_24h": "Recordatorio: tu retiro es mañana",
                    "message": "Actualización de tu solicitud de retiro",
                }
                title = titles.get(kind, "Actualización de retiro ILUS")
                if kind == "created":
                    paragraphs = [
                        "Recibimos tu solicitud. Esta solicitud aún no es una reserva confirmada.",
                        "Nuestro equipo revisará documento, identidad y disponibilidad para responder con una confirmación o propuesta.",
                    ]
                elif kind == "proposal" and proposal:
                    paragraphs = [
                        "Te enviamos una propuesta de fecha y horario para tu retiro.",
                        f"Propuesta: <strong>{proposal['date']} de {str(proposal['time_from'])[:5]} a {str(proposal['time_to'])[:5]}</strong>.",
                        proposal.get("message") or "Puedes confirmar, rechazar o proponer una nueva fecha desde el enlace.",
                    ]
                elif kind == "confirmed":
                    paragraphs = [
                        "Tu agenda de retiro fue confirmada.",
                        "Recuerda presentar documento de identidad. Si retira un tercero, debe coincidir con la persona autorizada.",
                    ]
                elif kind == "preparing":
                    paragraphs = [
                        "<strong>Estamos preparando tu retiro.</strong> Nuestro equipo está alistando tus productos en bodega.",
                        f"Fecha agendada: <strong>{variables['horario']}</strong>.",
                        "Recuerda presentar documento de identidad. Si retira un tercero, debe coincidir con la persona autorizada.",
                    ]
                elif kind == "done":
                    paragraphs = [
                        "<strong>Tu retiro fue completado.</strong> ¡Gracias por preferir ILUS!",
                        "Si necesitas un comprobante o tienes alguna consulta sobre este retiro, responde a este correo o escríbenos por soporte.",
                    ]
                elif kind == "reminder_24h":
                    paragraphs = [
                        f"Te recordamos que <strong>mañana retiramos tus productos</strong>.",
                        f"Hora: <strong>{variables['horario']}</strong>.",
                        f"Dirección: {variables['warehouse_addr'] or variables['warehouse_name']}.",
                        "Por favor presentarse con documento de identidad. Si retira un tercero, debe coincidir con la persona autorizada.",
                    ]
                else:
                    paragraphs = [custom_message or "Hay una actualización disponible para tu solicitud de retiro."]
                # Tracking dentro del correo también en el fallback (Daniel
                # 2026-06-17): el modelo canónico de 5 hitos, email-safe. Solo
                # para kinds con journey válido (no en rechazada/cancelación, idx -1).
                _aidx = pickup_journey_idx(estado) if estado else 0
                if estado and _aidx >= 0:
                    paragraphs = [_pickup_email_stepper(_aidx)] + paragraphs
                html = _ilus_email_html(
                    titulo=title,
                    subtitulo=f"{req['code']} - {variables['documento']}",
                    saludo=variables["persona_retira"],
                    parrafos=paragraphs,
                    btn_primario_txt="Ver solicitud",
                    btn_primario_url=follow_url,
                    btn_secundario_txt="Cómo llegar",
                    btn_secundario_url=cfg.get("maps_url"),
                    info_lineas=[
                        ("", "Bodega", variables["warehouse_name"]),
                        ("", "Dirección", variables["warehouse_addr"]),
                        ("", "Horario solicitado", variables["horario"]),
                    ],
                )
                # Multi-email (Daniel 2026-05-23)
                _multi = _send_pickup_email_multi(req, f"ILUS - {title} {req['code']}", html, attachments=_ics_att)
                sent_mail = len(_multi["sent"]) > 0
        except Exception as exc:
            try:
                print(f"[ILUS][PICKUP EMAIL] {str(exc).encode('ascii', 'ignore').decode('ascii')}")
            except Exception:
                pass

        # ── WHATSAPP ───────────────────────────────────────────────────
        # Sólo intentamos si el canal está activo en COMM_CANALES_ACTIVOS
        # (por default Daniel sólo tiene email). El email arriba se manda
        # siempre — eso ya cubre la notificación al cliente.
        sent_wa = None
        try:
            wa_cfg = _get_wa_cfg()
            if (_canal_activo("whatsapp")
                    and wa_cfg.get("account_sid") and wa_cfg.get("auth_token") and wa_cfg.get("from_number")):
                tpl_wa = _get_pickup_template(estado, "whatsapp") if estado else None
                if tpl_wa and tpl_wa.get("cuerpo"):
                    wa_body = _apply_template(tpl_wa.get("cuerpo") or "", variables)
                else:
                    titles = {
                        "created": "Solicitud de retiro recibida",
                        "proposal": "ILUS propuso una agenda de retiro",
                        "confirmed": "Agenda de retiro confirmada",
                        "preparing": "Estamos preparando tu retiro",
                        "done": "Tu retiro fue completado",
                        "reminder_24h": "Recordatorio: tu retiro es mañana",
                        "message": "Actualización de tu retiro",
                    }
                    title = titles.get(kind, "Actualización retiro ILUS")
                    if kind == "reminder_24h":
                        wa_body = (
                            f"ILUS - {title}\n\nSolicitud: {req['code']}\n"
                            f"Mañana retiramos tus productos.\n"
                            f"Hora: {variables['horario']}\n"
                            f"Bodega: {variables['warehouse_name']}\n"
                            f"Dirección: {variables['warehouse_addr']}\n"
                            f"Seguimiento: {follow_url}"
                        )
                    else:
                        wa_body = (
                            f"ILUS - {title}\n\nSolicitud: {req['code']}\n"
                            f"Documento: {variables['documento']}\n"
                            f"Bodega: {variables['warehouse_name']}\n"
                            f"Seguimiento: {follow_url}"
                        )
                sent_wa = _send_whatsapp(
                    wa_cfg["account_sid"], wa_cfg["auth_token"], wa_cfg["from_number"],
                    req["contact_phone"], wa_body
                )
        except Exception as exc:
            print(f"[ILUS][PICKUP WA] {exc}")
        return sent_mail, sent_wa

    def notify_async(req, kind="created", proposal=None, custom_message=""):
        """Versión asíncrona de notify(): dispara el envío en un hilo daemon
        para que el operador (o el cliente) reciban su respuesta HTTP de
        inmediato — el email/WA se envía en background.

        Daniel pidió esto explícitamente: que el cambio de estado no
        bloquee el cierre del modal del calendario.

        IMPORTANTE: pasamos un snapshot dict (no la fila viva) por seguridad
        — la conexión MySQL del request termina cuando vuelve la response.
        Dentro del thread abrimos app_context() para que get_db()/g y los
        helpers de email funcionen. Los links se construyen con
        _public_base_url() (NO url_for — falla en threads sin SERVER_NAME).
        """
        try:
            import threading
            # snapshot inmutable — evita race conditions con la conexión MySQL
            req_snapshot = dict(req) if req else {}

            def _runner(_req, _kind, _proposal, _msg):
                try:
                    with app.app_context():
                        notify(_req, _kind, proposal=_proposal, custom_message=_msg)
                except Exception as exc:
                    try:
                        print(f"[ILUS][PICKUP NOTIFY_ASYNC] {exc}")
                    except Exception: pass

            t = threading.Thread(
                target=_runner,
                args=(req_snapshot, kind, proposal, custom_message),
                daemon=True,
                name=f"pickup-notify-{req_snapshot.get('code','?')}-{kind}",
            )
            t.start()
        except Exception as exc:
            print(f"[ILUS][PICKUP NOTIFY_ASYNC SPAWN] {exc}")

    # ══════════════════════════════════════════════════════════════════
    #  NOTIFICACIÓN INTERNA — cliente rechazó la propuesta
    # ══════════════════════════════════════════════════════════════════
    # Daniel 2026-05-23 — Bug #3: cuando el cliente le da "No puedo asistir",
    # el operador necesita enterarse para reagendar o cerrar la solicitud.
    # Antes solo se notificaba al CLIENTE — el operador no se enteraba salvo
    # mirando el dashboard. Ahora también se manda email al buzón soporte.
    def _notificar_operador_rechazo(req, reason):
        """Notifica al buzón soporte ILUS que un cliente rechazó la propuesta.
        Envío en background (no bloquea response). Tolerante a fallos: si el
        envío falla, solo loguea — el cliente no debe verse afectado."""
        try:
            import threading
            req_snap = dict(req) if req else {}
            reason_snap = str(reason or "")[:500]

            def _runner(_req, _reason):
                try:
                    with app.app_context():
                        brand_cfg = _get_brand_cfg() or {}
                        dest = brand_cfg.get("support_email") or "soportetec@sphs.cl"
                        if not dest:
                            return
                        code = _req.get("code") or "?"
                        cli  = _req.get("customer_name") or "?"
                        doc  = f"{_req.get('document_type','')} {_req.get('document_number','')}".strip()
                        cont = _req.get("contact_name") or "?"
                        mail = _req.get("contact_email") or "—"
                        phone= _req.get("contact_phone") or "—"
                        # Link literal (NO url_for): este runner corre en thread
                        # daemon donde url_for(_external=True) lanza "Unable to
                        # build URLs outside an active request" (fix 2026-06-09).
                        # Ficha interna /retiros/<rid>; fallback al tracking público.
                        if _req.get("id"):
                            link = _public_base_url() + "/retiros/" + str(_req["id"])
                        elif _req.get("public_token"):
                            link = _public_base_url() + "/retiros/seguimiento/" + str(_req["public_token"])
                        else:
                            link = _public_base_url()
                        subject = f"ILUS - Cliente rechazo retiro {code}"
                        html = _ilus_email_html(
                            titulo="Cliente rechazo la propuesta de retiro",
                            subtitulo=f"{code} - {doc}",
                            saludo="Equipo ILUS",
                            parrafos=[
                                f"El cliente <strong>{cli}</strong> rechazo la propuesta de fecha y horario.",
                                f"Motivo declarado: <em>{(_reason or 'No indico motivo').strip()}</em>",
                                "Sugerencia: revisa la solicitud y contacta al cliente para reagendar o cerrarla.",
                            ],
                            btn_primario_txt="Abrir solicitud",
                            btn_primario_url=link,
                            info_lineas=[
                                ("", "Contacto", cont),
                                ("", "Email",    mail),
                                ("", "Telefono", phone),
                            ],
                        )
                        _send_ilus_email(dest, subject, html, evento="pickup_reject_operador", modulo="retiros")
                except Exception as exc:
                    try:
                        print(f"[ILUS][PICKUP REJECT NOTIFY OPERADOR] {exc}")
                    except Exception: pass

            t = threading.Thread(
                target=_runner,
                args=(req_snap, reason_snap),
                daemon=True,
                name=f"pickup-reject-notify-{req_snap.get('code','?')}",
            )
            t.start()
        except Exception as exc:
            print(f"[ILUS][PICKUP REJECT NOTIFY SPAWN] {exc}")

    # ══════════════════════════════════════════════════════════════════
    #  NOTIFICACIÓN AL EQUIPO — in-app (campana) + email interno (SLA)
    # ══════════════════════════════════════════════════════════════════
    # 2026-06-09: el equipo necesita enterarse AL INSTANTE de los eventos
    # del flujo de retiros (alta pública, respuesta del cliente, retiro
    # agendado sin saldo ERP) sin mirar el dashboard.
    #  (a) In-app: mant_notificaciones para usuarios activos con rol
    #      superadmin/admin/supervisor (vía ctx['_mant_notificar']).
    #  (b) Email: pickup_settings.notify_emails (CSV) + support_email del
    #      brand. modulo='comunicacion_interna' A PROPÓSITO: las alertas
    #      internas deben fluir aunque la llave 'retiros' (cliente) esté
    #      bloqueada en el kill switch de comunicaciones.
    # Best-effort total: corre en thread con app_context y NUNCA rompe el
    # flujo principal (todo va en try/except).
    def _notificar_equipo_retiros(evento_titulo, cuerpo, rid, code,
                                  prioridad="alta", tipo="retiro_nuevo",
                                  send_email=True):
        try:
            titulo_snap = str(evento_titulo or "")[:200]
            cuerpo_snap = str(cuerpo or "")[:2000]
            rid_snap = int(rid)
            code_snap = str(code or "?")
            prio_snap = prioridad or "alta"
            tipo_snap = tipo or "retiro_nuevo"
            send_email_snap = bool(send_email)

            def _runner():
                try:
                    with app.app_context():
                        url_accion = f"/retiros/{rid_snap}"
                        # ── (a) In-app: campana del header ────────────────
                        try:
                            _mant_notificar = ctx.get("_mant_notificar")
                            _auth_table = ctx.get("AUTH_TABLE") or "app_users"
                            if _mant_notificar:
                                # Review M4 2026-06-09: los roles usan slugs
                                # compuestos (supervisor_sstt, admin_general...)
                                # que _rol_familia matchea por prefijo — un IN
                                # exacto los dejaba sin campana. LIKE prefijo.
                                rows = mysql_fetchall(
                                    f"SELECT id FROM `{_auth_table}` "
                                    f" WHERE active=1 AND (role LIKE 'superadmin%%' "
                                    f"   OR role LIKE 'admin%%' OR role LIKE 'supervisor%%')"
                                ) or []
                                creadas = 0
                                for r in rows:
                                    uid = r.get("id")
                                    if not uid:
                                        continue
                                    try:
                                        _mant_notificar(
                                            destino_user_id=int(uid),
                                            tipo=tipo_snap,
                                            titulo=titulo_snap,
                                            cuerpo=cuerpo_snap,
                                            url_accion=url_accion,
                                            prioridad=prio_snap,
                                            cliente_id=None,
                                            visita_id=None,
                                        )
                                        creadas += 1
                                    except Exception as _e_n:
                                        print(f"[ILUS][PICKUP TEAM NOTIF] in-app uid={uid}: {_e_n}", flush=True)
                                if creadas:
                                    _inval = ctx.get("_mant_notif_cache_invalidar")
                                    if _inval:
                                        try: _inval(None)
                                        except Exception: pass
                        except Exception as _e_inapp:
                            print(f"[ILUS][PICKUP TEAM NOTIF] in-app: {_e_inapp}", flush=True)

                        # ── (b) Email interno ─────────────────────────────
                        if send_email_snap:
                            try:
                                dests = []
                                try:
                                    cfg_n = settings() or {}
                                    for em in str(cfg_n.get("notify_emails") or "").replace(";", ",").split(","):
                                        em = em.strip().lower()
                                        if em and is_valid_email(em) and em not in dests:
                                            dests.append(em)
                                except Exception:
                                    pass
                                try:
                                    brand_cfg = _get_brand_cfg() or {}
                                    _sup = (brand_cfg.get("support_email") or "").strip().lower()
                                    if _sup and _sup not in dests:
                                        dests.append(_sup)
                                except Exception:
                                    pass
                                # Daniel 2026-06-17: además del CSV + soporte, avisar al
                                # CORREO PERSONAL de cada RESPONSABLE por ROL. El email es
                                # la columna `username` de app_users. Robusto a roles nuevos:
                                # incluye a quien tenga el módulo 'retiros' encendido en
                                # rol_permisos + todos los superadmin/admin/supervisor (estos
                                # NO siempre tienen fila en rol_permisos). Best-effort.
                                try:
                                    _auth_table = ctx.get("AUTH_TABLE") or "app_users"
                                    rows_rol = mysql_fetchall(
                                        f"SELECT DISTINCT u.username AS email "
                                        f"FROM `{_auth_table}` u "
                                        f"LEFT JOIN rol_permisos rp ON rp.rol_slug = u.role "
                                        f"   AND rp.modulo='retiros' AND rp.accion='ver' "
                                        f"   AND rp.permitido=1 "
                                        f"WHERE u.active=1 AND ("
                                        f"     rp.rol_slug IS NOT NULL "
                                        f"  OR u.role LIKE 'superadmin%%' "
                                        f"  OR u.role LIKE 'admin%%' "
                                        f"  OR u.role LIKE 'supervisor%%')"
                                    ) or []
                                    for r in rows_rol:
                                        em = str(r.get("email") or "").strip().lower()
                                        if em and is_valid_email(em) and em not in dests:
                                            dests.append(em)
                                except Exception as _e_rol:
                                    print(f"[ILUS][PICKUP TEAM NOTIF] dests por rol: {_e_rol}", flush=True)
                                if dests:
                                    import html as _html_esc
                                    link = _public_base_url() + f"/retiros/{rid_snap}"
                                    subject = f"ILUS - {titulo_snap}"
                                    html = _ilus_email_html(
                                        titulo=titulo_snap,
                                        subtitulo=f"Retiro {code_snap}",
                                        saludo="Equipo ILUS",
                                        parrafos=[
                                            _html_esc.escape(cuerpo_snap).replace("\n", "<br>"),
                                            "Revisa la solicitud en el panel interno de retiros.",
                                        ],
                                        btn_primario_txt="Abrir retiro",
                                        btn_primario_url=link,
                                    )
                                    for dest in dests:
                                        try:
                                            _send_ilus_email(
                                                dest, subject, html,
                                                evento="retiros_equipo",
                                                modulo="comunicacion_interna",
                                            )
                                        except Exception as _e_m:
                                            print(f"[ILUS][PICKUP TEAM NOTIF] email {_mask_email(dest)}: {_e_m}", flush=True)
                            except Exception as _e_mail:
                                print(f"[ILUS][PICKUP TEAM NOTIF] email: {_e_mail}", flush=True)
                except Exception as exc:
                    print(f"[ILUS][PICKUP TEAM NOTIF RUNNER] {exc}", flush=True)

            threading.Thread(
                target=_runner, daemon=True,
                name=f"pickup-team-notify-{code_snap}",
            ).start()
        except Exception as exc:
            print(f"[ILUS][PICKUP TEAM NOTIF SPAWN] {exc}")

    # ── D2 (2026-06-09): alerta "agendado SIN SALDO" tras confirmar ──
    # El ERP puede indicar que TODOS los docs del retiro están sin saldo
    # (todo despachado — posible guía ya emitida). NO bloqueamos la
    # confirmación (decisión Daniel 2026-05-24: indicador, no bloqueante),
    # pero el equipo debe revisarlo ANTES de preparar los bultos.
    def _alertar_si_sin_saldo(rid, code):
        try:
            rid_snap = int(rid)
            code_snap = str(code or "?")

            def _runner():
                try:
                    with app.app_context():
                        row = mysql_fetchone(
                            """SELECT COUNT(*) AS total,
                                      SUM(CASE WHEN con_saldo=0 THEN 1 ELSE 0 END) AS sin_saldo
                                 FROM pickup_request_docs
                                WHERE request_id=%s""",
                            (rid_snap,)
                        ) or {}
                        total = int(row.get("total") or 0)
                        sin_saldo = int(row.get("sin_saldo") or 0)
                        if total >= 1 and sin_saldo == total:
                            _notificar_equipo_retiros(
                                f"Retiro {code_snap} agendado SIN SALDO — revisar antes de preparar",
                                ("El ERP indica que todos los documentos asociados a este retiro "
                                 "están sin saldo (todo despachado — posible guía ya emitida). "
                                 "El retiro quedó agendado igualmente, pero requiere confirmación "
                                 "interna antes de preparar los bultos."),
                                rid_snap, code_snap,
                                prioridad="urgente", tipo="retiro_sin_saldo",
                            )
                except Exception as exc:
                    print(f"[ILUS][PICKUP SALDO ALERT] {exc}", flush=True)

            threading.Thread(
                target=_runner, daemon=True,
                name=f"pickup-saldo-alert-{code_snap}",
            ).start()
        except Exception as exc:
            print(f"[ILUS][PICKUP SALDO ALERT SPAWN] {exc}")

    # ══════════════════════════════════════════════════════════════════
    #  SANITIZACIÓN — elimina campos internos antes de enviar al cliente
    # ══════════════════════════════════════════════════════════════════
    # Daniel pidió seguridad sin brechas (mayo 2026). El tracking público
    # debe recibir un dict SIN datos que solo deba ver el operador interno.
    # Los campos `internal_notes`, `doc_validation_notes`, `doc_erp_data`,
    # `created_ip`, `created_user_agent`, `doc_validated_by` son del flujo
    # interno y no se exponen al cliente.
    #
    # IMPORTANTE: hacemos pop() sobre una copia del dict — NO mutamos la
    # fila original (mysql_fetchone devuelve un dict mutable).
    _PICKUP_INTERNAL_FIELDS = frozenset((
        "internal_notes",
        "doc_validation_notes",
        "doc_erp_data",
        "doc_validated_by",
        "doc_validated_at",
        "doc_validation_status",
        "created_ip",
        "created_user_agent",
        "risk_score",
        "information_quality_score",
        "reminder_24h_sent",
    ))

    def _strip_internal(row):
        """Devuelve una copia del dict sin campos internos. Idempotente: si
        el campo no existe, no falla. Defensa en profundidad: igual el template
        público no debería usarlos, pero el dict completo viaja por la red
        si se serializa a JSON."""
        if not row:
            return row
        try:
            out = dict(row)
        except Exception:
            return row
        for k in _PICKUP_INTERNAL_FIELDS:
            out.pop(k, None)
        return out

    def _mask_email(email):
        """Enmascara email para logs: 'daniel.aguilar@sphs.cl' → 'd***@sphs.cl'."""
        try:
            if not email or "@" not in str(email):
                return "***"
            local, dom = str(email).split("@", 1)
            return f"{local[:1]}***@{dom}" if local else f"***@{dom}"
        except Exception:
            return "***"

    def _mask_phone(phone):
        """Enmascara teléfono: '+56 9 1234 5678' → '+56 9 **** 5678'."""
        try:
            s = re.sub(r"[^\d]", "", str(phone or ""))
            if len(s) < 4:
                return "***"
            return f"***{s[-4:]}"
        except Exception:
            return "***"

    def _mask_rut(rut):
        """Enmascara RUT: '12.345.678-9' → '12.***.***-9'."""
        try:
            s = str(rut or "")
            if len(s) < 3:
                return "***"
            return f"{s[:2]}***{s[-2:]}"
        except Exception:
            return "***"

    # ══════════════════════════════════════════════════════════════════
    #  RATE LIMITING POR TOKEN — polling público
    # ══════════════════════════════════════════════════════════════════
    # El endpoint /status hace polling cada 30s desde el navegador del cliente.
    # Para evitar abuse (un cliente con un script que martille la URL), aplicamos
    # un rate limit POR TOKEN en memoria del worker (60 req/min por token).
    # NO usamos el rate_limited de app.py porque ese es por IP/user, no por
    # token público — varios clientes detrás de un NAT compartirían IP.
    _TOKEN_RL: dict = {}        # token → [timestamps]
    _TOKEN_RL_MAX = 60          # max requests / window por token
    _TOKEN_RL_WINDOW = 60.0     # ventana en segundos

    def _token_rate_ok(token):
        """Devuelve True si el token aún tiene cupo en su ventana, False si lo
        excedió. Limpia entradas viejas con probabilidad 1/100 para mantener
        el dict compacto sin agregar costo en cada request."""
        import time as _time
        now = _time.time()
        cutoff = now - _TOKEN_RL_WINDOW
        bucket = _TOKEN_RL.get(token) or []
        # Filtrar timestamps fuera de ventana
        bucket = [t for t in bucket if t >= cutoff]
        if len(bucket) >= _TOKEN_RL_MAX:
            _TOKEN_RL[token] = bucket
            return False
        bucket.append(now)
        _TOKEN_RL[token] = bucket
        # Limpieza probabilística (cheap)
        if len(_TOKEN_RL) > 1000 and (now * 100) % 100 < 1:
            for k in list(_TOKEN_RL.keys()):
                if all(t < cutoff for t in (_TOKEN_RL[k] or [])):
                    _TOKEN_RL.pop(k, None)
        return True

    # ══════════════════════════════════════════════════════════════════
    #  CACHE DE POLLING — evita martillar la BD desde el tracking público
    # ══════════════════════════════════════════════════════════════════
    # El tracking público hace polling cada 30s. Con muchos clientes mirando
    # su retiro al mismo tiempo, podría generar un peak de SELECTs. Esta
    # cache en memoria invalida-por-tiempo (10s) garantiza que aunque 100
    # navegadores polleen al mismo retiro, solo se hace 1 SELECT cada 10s.
    #
    # Diseño minimalista: dict en memoria del proceso. Si hay múltiples
    # workers (gunicorn), cada uno tiene su cache → eso está bien: igual
    # baja el peak en cada worker.
    _POLL_CACHE = {}             # token → {payload, fetched_at}
    _POLL_CACHE_TTL = 10.0       # segundos

    # Cache global del endpoint público de disponibilidad. Los datos son los
    # MISMOS para todos los visitantes (no dependen de sesión), por eso una
    # cache de 30s en memoria del worker reduce drásticamente el costo de
    # generar el grid de 30 días × ~17 slots con queries de ocupación/bloqueos.
    # TTL corto para que las nuevas reservas se reflejen rápido. Definido aquí
    # arriba (no dentro del endpoint) para que las invalidaciones de otras
    # rutas puedan referenciarlo aunque se ejecuten antes de que el endpoint
    # `pickup_disponibilidad_publica` se haya ejecutado por primera vez.
    _DISPO_CACHE: dict = {"payload": None, "ts": 0.0}
    _DISPO_TTL = 30.0  # segundos

    # ⚡ PERF (Daniel 2026-05-24): cache de saldo ERP por RUT+dias.
    # El wizard se abre/cierra varias veces — sin cache cada apertura
    # generaba una query MAEEDO+MAEDDO+MAEEN compleja (~800-1500ms).
    # TTL 60s: el operador tampoco verá retrocesos extraños.
    _SALDO_CACHE: dict = {}      # key (rut+dias+solo_con_saldo) → (payload, ts)
    _SALDO_TTL = 60.0            # segundos

    # ⚡ PERF: cache para /retiros/<rid>/docs polled muchas veces durante
    # un wizard activo. Versionado por updated_at del registro.
    _DOCS_CACHE: dict = {}       # rid → (payload, ts, version_hash)
    _DOCS_TTL = 15.0             # segundos

    def _polling_cached(token):
        import time as _time
        ent = _POLL_CACHE.get(token)
        if ent and (_time.time() - ent["fetched_at"]) < _POLL_CACHE_TTL:
            return ent["payload"]
        return None

    def _polling_store(token, payload):
        import time as _time
        _POLL_CACHE[token] = {
            "payload":    payload,
            "fetched_at": _time.time(),
        }
        # Limpia entradas viejas si la cache crece mucho
        if len(_POLL_CACHE) > 500:
            cutoff = _time.time() - _POLL_CACHE_TTL * 3
            for k in list(_POLL_CACHE.keys()):
                if _POLL_CACHE[k]["fetched_at"] < cutoff:
                    _POLL_CACHE.pop(k, None)

    def _no_store_json(payload):
        """jsonify + Cache-Control:no-store. El polling EN VIVO no debe ser
        cacheado por el navegador/CDN: cada poll debe reflejar el estado real
        (el cache de 10s de _POLL_CACHE es del lado servidor, suficiente)."""
        resp = jsonify(payload)
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp

    @app.route("/retiros/seguimiento/<token>/status", methods=["GET"])
    def pickup_public_tracking_status(token):
        """Devuelve estado actual del retiro en JSON ligero — usado por el
        polling del tracking público (cada 30s).

        Seguridad (Daniel mayo 2026):
        - Rate limit por TOKEN: max 60 req/min — evita scripts que martillen
          la URL pública (varios clientes detrás del mismo NAT comparten IP,
          por eso el rate limit va por token, no por IP).
        - Cacheado 10s en memoria: aunque 100 navegadores polleen al mismo
          retiro, solo se hace 1 SELECT cada 10s.
        - Payload mínimo: no exponemos campos internos (validation notes,
          IPs, datos personales) — solo lo que necesita el stepper visual.

        Response shape:
          { ok, status, status_label, status_color, has_pending_proposal,
            updated_at, confirmed_date, confirmed_time_from, confirmed_time_to }
        """
        # Validación del token mismo (anti-typo de URL): debe lucir como un
        # token URL-safe razonable. Esto evita queries innecesarias por
        # tokens claramente inválidos (bots probando paths).
        if not token or len(token) < 16 or len(token) > 200 \
                or not re.match(r"^[A-Za-z0-9_\-]+$", token):
            return jsonify({"ok": False, "error": "invalid_token"}), 400

        # Rate limit por token (60 req/min)
        if not _token_rate_ok(token):
            return jsonify({
                "ok": False,
                "error": "rate_limited",
                "retry_after": int(_TOKEN_RL_WINDOW),
            }), 429

        cached = _polling_cached(token)
        if cached is not None:
            return _no_store_json(cached)

        req = mysql_fetchone(
            f"SELECT id, status, confirmed_date, confirmed_time_from, "
            f"       confirmed_time_to, updated_at "
            f"FROM `{REQ}` WHERE public_token=%s LIMIT 1",
            (token,)
        )
        if not req:
            return jsonify({"ok": False, "error": "not_found"}), 404

        # ¿Hay propuesta pendiente?
        pend = mysql_fetchone(
            f"SELECT id FROM `{PROP}` WHERE request_id=%s AND status='pending' "
            f"ORDER BY id DESC LIMIT 1",
            (req["id"],)
        )

        status = req.get("status") or ""
        payload = {
            "ok":                   True,
            "status":               status,
            "status_label":         PICKUP_STATUS.get(status, status),
            "status_color":         PICKUP_STATUS_COLORS.get(status, "secondary"),
            # journey_idx: hito canónico (0-4, -1=cancelado). Lo usa el polling
            # para refrescar el stepper EN VIVO aunque el `status` exacto cambie
            # a otro estado del MISMO hito (ej: solicitud_recibida→en_revision).
            "journey_idx":          pickup_journey_idx(status),
            "has_pending_proposal": bool(pend),
            "updated_at":           str(req.get("updated_at") or "")[:19],
            "confirmed_date":       str(req.get("confirmed_date") or ""),
            "confirmed_time_from":  str(req.get("confirmed_time_from") or "")[:5],
            "confirmed_time_to":    str(req.get("confirmed_time_to") or "")[:5],
        }
        _polling_store(token, payload)
        return _no_store_json(payload)

    # ══════════════════════════════════════════════════════════════════
    #  CHAT EN VIVO cliente ↔ operador (Daniel 2026-06-17)
    #  El cliente escribe desde la pancarta del seguimiento; el operador lo
    #  ve en el monitor con burbuja "nuevo mensaje" (estilo WhatsApp) y
    #  responde. Best-effort, tabla pickup_messages.
    # ══════════════════════════════════════════════════════════════════
    def _msg_time(dt):
        """Formatea created_at (UTC en BD) a hora Chile 'dd/mm HH:MM'."""
        try:
            from zoneinfo import ZoneInfo as _ZI
            from datetime import datetime as _dt
            if dt is None:
                return ""
            if not isinstance(dt, _dt):
                dt = _dt.strptime(str(dt)[:19], "%Y-%m-%d %H:%M:%S")
            return dt.replace(tzinfo=_ZI("UTC")).astimezone(_ZI("America/Santiago")).strftime("%d/%m %H:%M")
        except Exception:
            return str(dt or "")[:16]

    def _msg_rows(rid, limit=300):
        rows = mysql_fetchall(
            "SELECT id, sender, autor, cuerpo, created_at "
            "FROM pickup_messages WHERE request_id=%s ORDER BY id ASC LIMIT %s",
            (int(rid), int(limit))
        ) or []
        return [{
            "id":     r.get("id"),
            "sender": r.get("sender") or "cliente",
            "autor":  r.get("autor") or "",
            "cuerpo": r.get("cuerpo") or "",
            "hora":   _msg_time(r.get("created_at")),
        } for r in rows]

    @app.route("/retiros/seguimiento/<token>/mensajes", methods=["GET"])
    def pickup_public_mensajes(token):
        """Lista de mensajes del chat (vista del CLIENTE). Marca como leídos por
        el cliente los mensajes del operador."""
        if not token or len(token) < 16 or len(token) > 200 \
                or not re.match(r"^[A-Za-z0-9_\-]+$", token):
            return jsonify({"ok": False, "error": "invalid_token"}), 400
        if not _token_rate_ok(token):
            return jsonify({"ok": False, "error": "rate_limited"}), 429
        req = mysql_fetchone(f"SELECT id FROM `{REQ}` WHERE public_token=%s LIMIT 1", (token,))
        if not req:
            return jsonify({"ok": False, "error": "not_found"}), 404
        try:
            mysql_execute("UPDATE pickup_messages SET leido_cliente=1 "
                          "WHERE request_id=%s AND sender='operador' AND leido_cliente=0",
                          (req["id"],))
        except Exception:
            pass
        resp = jsonify({"ok": True, "mensajes": _msg_rows(req["id"])})
        resp.headers["Cache-Control"] = "no-store"
        return resp

    @app.route("/retiros/seguimiento/<token>/mensaje", methods=["POST"])
    @_rate_limited("pickup_msg_publico", max_attempts=40, window_seconds=300, methods=("POST",))
    def pickup_public_mensaje_post(token):
        """El cliente envía un mensaje desde la pancarta del seguimiento."""
        if not token or len(token) < 16 or len(token) > 200 \
                or not re.match(r"^[A-Za-z0-9_\-]+$", token):
            return jsonify({"ok": False, "error": "invalid_token"}), 400
        body = request.get_json(silent=True) or {}
        texto = (body.get("mensaje") or request.form.get("mensaje") or "").strip()[:2000]
        if not texto:
            return jsonify({"ok": False, "error": "mensaje_vacio"}), 400
        req = mysql_fetchone(
            f"SELECT id, code, customer_name, contact_name FROM `{REQ}` "
            f"WHERE public_token=%s LIMIT 1", (token,))
        if not req:
            return jsonify({"ok": False, "error": "not_found"}), 404
        autor = (req.get("contact_name") or req.get("customer_name") or "Cliente")[:160]
        try:
            mysql_execute(
                "INSERT INTO pickup_messages (request_id, sender, autor, cuerpo, "
                "leido_operador, leido_cliente) VALUES (%s,'cliente',%s,%s,0,1)",
                (req["id"], autor, texto))
        except Exception as exc:
            print(f"[pickup-msg] insert cliente: {exc}", flush=True)
            return jsonify({"ok": False, "error": "no_guardado"}), 500
        # Notificar al equipo (campana + email) — best-effort
        try:
            _notificar_equipo_retiros(
                f"💬 Nuevo mensaje del cliente · retiro {req.get('code') or '?'}",
                f"{autor} escribió: \"{texto[:300]}\"",
                req["id"], req.get("code") or "?",
                prioridad="alta", tipo="retiro_mensaje", send_email=True)
        except Exception as _e_n:
            print(f"[pickup-msg] notif: {_e_n}", flush=True)
        return jsonify({"ok": True, "mensaje": {
            "sender": "cliente", "autor": autor, "cuerpo": texto, "hora": _msg_time(None) or ""}})

    @app.route("/retiros/<int:rid>/mensajes", methods=["GET"])
    @require_permission("retiros")
    def pickup_op_mensajes(rid):
        """Lista de mensajes del chat (vista del OPERADOR). Marca como leídos por
        el operador los mensajes del cliente."""
        try:
            mysql_execute("UPDATE pickup_messages SET leido_operador=1 "
                          "WHERE request_id=%s AND sender='cliente' AND leido_operador=0",
                          (rid,))
        except Exception:
            pass
        return jsonify({"ok": True, "mensajes": _msg_rows(rid)})

    @app.route("/retiros/<int:rid>/mensaje", methods=["POST"])
    @require_permission("retiros")
    def pickup_op_mensaje_post(rid):
        """El operador responde al cliente desde el monitor/ficha."""
        body = request.get_json(silent=True) or {}
        texto = (body.get("mensaje") or request.form.get("mensaje") or "").strip()[:2000]
        if not texto:
            return jsonify({"ok": False, "error": "mensaje_vacio"}), 400
        req = mysql_fetchone(f"SELECT id, code FROM `{REQ}` WHERE id=%s LIMIT 1", (rid,))
        if not req:
            return jsonify({"ok": False, "error": "not_found"}), 404
        autor = ((getattr(g, "user", None) or {}).get("nombre") or "Equipo ILUS")[:160]
        try:
            mysql_execute(
                "INSERT INTO pickup_messages (request_id, sender, autor, cuerpo, "
                "leido_operador, leido_cliente) VALUES (%s,'operador',%s,%s,1,0)",
                (rid, autor, texto))
        except Exception as exc:
            print(f"[pickup-msg] insert operador: {exc}", flush=True)
            return jsonify({"ok": False, "error": "no_guardado"}), 500
        return jsonify({"ok": True, "mensaje": {
            "sender": "operador", "autor": autor, "cuerpo": texto, "hora": _msg_time(None) or ""}})

    @app.route("/retiros/api/mensajes-no-leidos", methods=["GET"])
    @require_permission("retiros")
    def pickup_mensajes_no_leidos():
        """Conteo de mensajes del cliente sin leer, por retiro — alimenta la
        burbuja 'nuevo mensaje' del monitor."""
        try:
            rows = mysql_fetchall(
                "SELECT request_id, COUNT(*) AS n FROM pickup_messages "
                "WHERE sender='cliente' AND leido_operador=0 GROUP BY request_id") or []
            por = {str(r["request_id"]): int(r["n"]) for r in rows}
            resp = jsonify({"ok": True, "total": sum(por.values()), "por_retiro": por})
            resp.headers["Cache-Control"] = "no-store"
            return resp
        except Exception as exc:
            print(f"[pickup-msg] no-leidos: {exc}", flush=True)
            return jsonify({"ok": True, "total": 0, "por_retiro": {}})

    @app.route("/retiros/solicitar", methods=["GET", "POST"])
    @_rate_limited("pickup_public_request", max_attempts=40, window_seconds=3600)
    def pickup_public_request():
        # Rate limit: 40 envíos / hora por IP (ver decorador max_attempts=40).
        # Daniel pidió "sin brechas" — antes era ilimitado, lo que permitía a
        # un atacante crear cientos de retiros basura para llenar BD/disk.
        # 40 cubre clientes legítimos detrás de un NAT corporativo compartido.
        cfg = settings()
        if request.method == "POST":
            form = request.form
            data = {
                "document_type": (form.get("document_type") or "factura").strip(),
                "document_number": (form.get("document_number") or "").strip(),
                "customer_name": (form.get("customer_name") or "").strip(),
                "customer_rut": (form.get("customer_rut") or "").strip(),
                "contact_name": (form.get("contact_name") or "").strip(),
                "contact_email": (form.get("contact_email") or "").strip().lower(),
                "contact_phone": (form.get("contact_phone") or "").strip(),
                "whatsapp_phone": (form.get("whatsapp_phone") or "").strip(),
                "pickup_person_name": (form.get("pickup_person_name") or "").strip(),
                "pickup_person_rut": (form.get("pickup_person_rut") or "PENDIENTE").strip(),
                "pickup_person_phone": (form.get("pickup_person_phone") or "").strip(),
                "pickup_person_relation": (form.get("pickup_person_relation") or "autorizado").strip(),
                "requested_date": (form.get("requested_date") or "").strip(),
                "requested_time_from": (form.get("requested_time_from") or "").strip(),
                "requested_time_to": (form.get("requested_time_to") or "").strip(),
                "observations": (form.get("observations") or "").strip(),
                "invoice_total_amount": float(form.get("invoice_total_amount") or 0),
            }
            if not data["requested_date"]:
                data["requested_date"] = next_allowed_date(cfg)

            # ──────────────────────────────────────────────────────────────
            # FORMULARIO SIMPLIFICADO (a pedido):
            # La persona autorizada es OPCIONAL. Si el cliente NO marca el
            # toggle, asumimos que el cliente mismo retira → copiamos sus
            # datos a los campos pickup_person_*.
            # ──────────────────────────────────────────────────────────────
            auth_active = (form.get("auth_active") or "0") == "1"
            if not auth_active or not data.get("pickup_person_name"):
                # El cliente retira por sí mismo
                data["pickup_person_name"]     = data.get("customer_name") or ""
                data["pickup_person_rut"]      = data.get("customer_rut") or ""
                data["pickup_person_phone"]    = data.get("contact_phone") or ""
                data["pickup_person_relation"] = "dueno"
            else:
                # Cliente autorizó a un tercero — completar campos faltantes con respaldos
                if not data.get("pickup_person_phone"):
                    data["pickup_person_phone"] = data.get("contact_phone") or ""
                if not data.get("pickup_person_rut"):
                    data["pickup_person_rut"] = data.get("customer_rut") or ""

            # contact_name es el "nombre de quien retira" (puede ser el cliente o el autorizado)
            if not data.get("contact_name"):
                data["contact_name"] = data.get("pickup_person_name") or data.get("customer_name") or ""

            errors = []
            required = ["document_number", "customer_name", "customer_rut",
                        "contact_email", "contact_phone",
                        "requested_time_from", "requested_time_to"]
            if any(not data.get(k) for k in required):
                errors.append("Completa todos los campos obligatorios.")

            # ── VALIDACIÓN EMAIL (robusta: regex + doble@ + espacios) ──
            if data["contact_email"]:
                if not is_valid_email(data["contact_email"]):
                    errors.append(
                        "El email no es válido. Usa el formato nombre@dominio.cl"
                    )
                else:
                    data["contact_email"] = data["contact_email"].strip().lower()

            # ── VALIDACIÓN TELÉFONO CHILENO (+56 9 XXXX XXXX) ──
            if data["contact_phone"] and not is_valid_cl_phone(data["contact_phone"]):
                errors.append(
                    "El teléfono del contacto debe ser un móvil chileno "
                    "(+56 9 XXXX XXXX)."
                )
            elif data["contact_phone"]:
                data["contact_phone"] = format_cl_phone(data["contact_phone"])

            # ── VALIDACIÓN TELÉFONO DE QUIEN RETIRA ──
            if data["pickup_person_phone"] and not is_valid_cl_phone(data["pickup_person_phone"]):
                # Si el toggle de autorizado está activo, este error es bloqueante.
                # Si no está activo, simplemente copiamos el del contacto.
                if auth_active:
                    errors.append(
                        "El teléfono de quien retira debe ser un móvil chileno válido."
                    )
                else:
                    data["pickup_person_phone"] = data["contact_phone"]
            elif data["pickup_person_phone"]:
                data["pickup_person_phone"] = format_cl_phone(data["pickup_person_phone"])
            else:
                # Sin teléfono de pickup_person → usar el contact_phone (ya formateado)
                data["pickup_person_phone"] = data["contact_phone"]

            # ── VALIDACIÓN RUT DEL CLIENTE (módulo 11) ──
            if data["customer_rut"] and not is_valid_rut(data["customer_rut"]):
                errors.append(
                    "El RUT del cliente no es válido. "
                    "Revisa el dígito verificador."
                )
            elif data["customer_rut"]:
                data["customer_rut"] = format_rut(data["customer_rut"])

            # ── VALIDACIÓN RUT DE QUIEN RETIRA (solo si auth_active) ──
            if auth_active:
                if not data.get("pickup_person_rut") or data["pickup_person_rut"] == "PENDIENTE":
                    errors.append(
                        "Si autorizas a otra persona debes ingresar su RUT."
                    )
                elif not is_valid_rut(data["pickup_person_rut"]):
                    errors.append(
                        "El RUT de quien retira no es válido. "
                        "Revisa el dígito verificador."
                    )
                else:
                    data["pickup_person_rut"] = format_rut(data["pickup_person_rut"])
            elif data["pickup_person_rut"] and is_valid_rut(data["pickup_person_rut"]):
                data["pickup_person_rut"] = format_rut(data["pickup_person_rut"])

            # ── DECLARACIÓN DE TERCERO (RUT cliente != RUT quien retira) ──
            # Si el cliente activó el toggle y los RUTs son DIFERENTES,
            # registramos una declaración explícita en observations.
            # Esto da trazabilidad legal: el operador en bodega verá la nota.
            try:
                _rut_cli = _clean_rut(data.get("customer_rut") or "")
                _rut_ret = _clean_rut(data.get("pickup_person_rut") or "")
                _decl_acepta = (form.get("acepta_tercero") or "") in ("1", "on", "true")
                if (auth_active and _rut_cli and _rut_ret
                        and _rut_cli != _rut_ret and is_valid_rut(_rut_cli) and is_valid_rut(_rut_ret)):
                    if not _decl_acepta:
                        errors.append(
                            "Debes confirmar que el cliente autoriza a este tercero "
                            "para retirar (marca la casilla de declaración)."
                        )
                    else:
                        _decl_text = (
                            f"\n[DECLARACIÓN AUTORIZACIÓN TERCERO · "
                            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n"
                            f"El documento está a nombre de {data.get('customer_name','')} "
                            f"(RUT {data.get('customer_rut','')}). "
                            f"Quien retira es {data.get('pickup_person_name','')} "
                            f"(RUT {data.get('pickup_person_rut','')}). "
                            f"El cliente declaró explícitamente que esta persona está "
                            f"autorizada para retirar los productos en su nombre."
                        )
                        data["observations"] = (
                            (data.get("observations") or "").strip() + _decl_text
                        ).strip()
            except Exception as _decl_exc:
                # No bloqueamos por error de logueo de declaración; pero lo registramos
                print(f"[pickup_public_request] declaración tercero error: {_decl_exc}", flush=True)

            # FASE 2 (2026-05-29): validador temporal CENTRAL, modo 'public'.
            # Reemplaza el chequeo +24h manual (que usaba datetime.now() del
            # server = UTC, no Chile → fallaba cerca de medianoche), más
            # date_allowed y time_allowed sueltos. Valida TZ Chile, fecha no
            # pasada, min_notice_hours, día hábil, feriado, horario, colación
            # y duración múltiplo de slot — fuente única de verdad.
            ok_date = ok_time = True
            if data["requested_date"]:
                ok_dt, msg_dt = validate_pickup_datetime(
                    data["requested_date"], data["requested_time_from"],
                    data["requested_time_to"], cfg=cfg, mode="public")
                if not ok_dt:
                    errors.append(msg_dt)
                    ok_date = ok_time = False
            else:
                # Sin fecha elegida (raro: 1611 ya setea next_allowed_date):
                # validar al menos el rango horario base.
                ok_time, msg_time = time_allowed(data["requested_time_from"], data["requested_time_to"], cfg)
                if not ok_time:
                    errors.append(msg_time)

            # Validación dura del slot: colación, bloqueos manuales, capacidad
            # paralela. Evita que un usuario bypaseando JS mande un horario
            # inválido (colación, bloque lleno, fuera de horario, etc).
            if ok_date and ok_time and data["requested_date"]:
                try:
                    ok_slot, motivo_slot = _validar_disponibilidad_slot(
                        data["requested_date"],
                        data["requested_time_from"],
                        data["requested_time_to"],
                    )
                    if not ok_slot:
                        errors.append(motivo_slot or "El horario seleccionado no está disponible.")
                except Exception as _slot_exc:
                    # Si la validación falla por error técnico, no bloqueamos
                    # (defensa en profundidad: ya validamos lo básico arriba).
                    print(f"[pickup_public_request] validar_slot exc: {_slot_exc}", flush=True)

            # Cliente público no detalla bultos en el form: creamos 1 bulto
            # placeholder. El operador interno cubicará después con datos reales.
            packages = [calc_package(0, 0, 0, 0)]
            if not form.get("accept_terms"):
                errors.append("Debes aceptar la declaracion de responsabilidad y autorizacion.")
            if errors:
                # Detección AJAX: si el form viene por fetch(), devolvemos
                # JSON con la lista de errores. Sin esto, el cliente nunca
                # vería los errores (recibiría HTML de la página y no podría
                # parsearlos). Daniel 2026-05-24 fire-and-forget.
                _is_ajax_err = (
                    request.headers.get("X-Requested-With") == "XMLHttpRequest"
                    or "application/json" in (request.headers.get("Accept") or "")
                )
                if _is_ajax_err:
                    return jsonify({"ok": False, "errors": errors}), 400
                try:
                    _car_rows = mysql_fetchall("SELECT archivo_path, titulo, subtitulo FROM retiros_carousel WHERE activa=1 ORDER BY orden ASC, id ASC")
                    _car_imgs = [dict(r) for r in (_car_rows or [])]
                except Exception:
                    _car_imgs = []
                # D3 (2026-06-09): el re-render por error de validación NO
                # pasaba announcements → los avisos vigentes desaparecían de
                # la página tras un error. Misma query que el GET.
                try:
                    _anun_rows = mysql_fetchall(
                        "SELECT id, titulo, mensaje, tipo, icon FROM retiros_announcements "
                        "WHERE activa=1 "
                        "  AND (fecha_desde IS NULL OR fecha_desde <= NOW()) "
                        "  AND (fecha_hasta IS NULL OR fecha_hasta >= NOW()) "
                        "ORDER BY orden ASC, id DESC"
                    )
                    _anuncios_err = [dict(r) for r in (_anun_rows or [])]
                except Exception:
                    _anuncios_err = []
                return render_template("retiros/public_request.html", settings=cfg, relations=PICKUP_RELATIONS, errors=errors, fd=form, carousel_images=_car_imgs, announcements=_anuncios_err)

            files = [f for f in request.files.getlist("attachments") if f and f.filename]
            quality, risk = quality_score(data, packages, signed=True, attachments=len(files))
            token = secrets.token_urlsafe(42)
            if data["whatsapp_phone"]:
                extra = f"WhatsApp informado: {data['whatsapp_phone']}"
                data["observations"] = f"{data['observations']}\n{extra}".strip()
            total_weight = round(sum(p["weight_kg"] for p in packages), 3)
            total_vw = round(sum(p["volumetric_weight"] for p in packages), 3)
            total_m3 = round(sum(p["volume_m3"] for p in packages), 4)
            conn = get_db()
            with conn.cursor() as cur:
                cur.execute(
                    f"""INSERT INTO `{REQ}`
                        (document_type,document_number,customer_name,customer_rut,contact_name,contact_email,contact_phone,
                         pickup_person_name,pickup_person_rut,pickup_person_phone,pickup_person_relation,
                         requested_date,requested_time_from,requested_time_to,status,information_quality_score,risk_score,
                         total_packages,total_weight_kg,total_volumetric_weight,total_volume_m3,invoice_total_amount,
                         observations,public_token,signature_status,created_ip,created_user_agent)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'solicitud_recibida',%s,%s,%s,%s,%s,%s,%s,%s,%s,'firmado',%s,%s)""",
                    (
                        data["document_type"], data["document_number"], data["customer_name"], data["customer_rut"],
                        data["contact_name"], data["contact_email"], data["contact_phone"],
                        data["pickup_person_name"], data["pickup_person_rut"], data["pickup_person_phone"], data["pickup_person_relation"],
                        data["requested_date"], data["requested_time_from"], data["requested_time_to"],
                        quality, risk, len(packages), total_weight, total_vw, total_m3, data["invoice_total_amount"],
                        data["observations"], token, request.remote_addr, (request.user_agent.string or "")[:300],
                    ),
                )
                rid = cur.lastrowid
                # Código RET-XXXXXX aleatorio (Daniel mayo 2026, anti-enumeración).
                # `_generate_pickup_code` reintenta hasta no colisionar con UNIQUE.
                code = _generate_pickup_code(cur=cur)
                cur.execute(f"UPDATE `{REQ}` SET code=%s WHERE id=%s", (code, rid))
                for idx, pkg in enumerate(packages, 1):
                    cur.execute(
                        f"""INSERT INTO `{PKG}` (request_id,package_number,length_cm,width_cm,height_cm,weight_kg,volumetric_weight,volume_m3)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (rid, idx, pkg["length_cm"], pkg["width_cm"], pkg["height_cm"], pkg["weight_kg"], pkg["volumetric_weight"], pkg["volume_m3"]),
                    )
                cur.execute(
                    f"""INSERT INTO `{SIG}` (request_id,signer_name,signer_rut,accepted_terms,ip,user_agent)
                        VALUES (%s,%s,%s,1,%s,%s)""",
                    (rid, data["contact_name"], data["customer_rut"], request.remote_addr, (request.user_agent.string or "")[:300]),
                )
            conn.commit()
            # ── HARDENING UPLOAD (Daniel mayo 2026) ──────────────────────
            # 1) Whitelist estricto de extensiones
            # 2) Max 5 archivos por solicitud (defense vs abuse)
            # 3) Max 10 MB cada uno (Flask MAX_CONTENT_LENGTH cubre el total,
            #    pero quedamos blindados explícitamente acá)
            # 4) Nombre saneado con `secure_filename` y prefijo determinístico
            _allowed_ext = {"png", "jpg", "jpeg", "webp", "pdf", "doc", "docx"}
            _max_files = 5
            _max_bytes = 10 * 1024 * 1024   # 10 MB
            for f in files[:_max_files]:
                ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
                if ext not in _allowed_ext:
                    continue
                # Verificar tamaño leyendo el stream (sin cargar todo en RAM)
                try:
                    f.stream.seek(0, 2)  # ir al final
                    size_bytes = f.stream.tell()
                    f.stream.seek(0)
                except Exception:
                    size_bytes = 0
                if size_bytes > _max_bytes:
                    print(f"[pickup_public_request] adjunto rechazado por tamaño "
                          f"rid={rid} bytes={size_bytes}", flush=True)
                    continue
                safe_name = secure_filename(f.filename)
                # secure_filename puede devolver "" si el filename es 100% no-ASCII.
                # En ese caso usamos un fallback genérico.
                if not safe_name:
                    safe_name = f"upload.{ext}"
                fname = f"ret_{rid}_{int(time.time())}_{safe_name}"
                f.save(os.path.join(upload_dir, fname))
                mysql_execute(
                    f"""INSERT INTO `{ATT}` (request_id,filename,original_name,mime_type,uploaded_by)
                        VALUES (%s,%s,%s,%s,'cliente')""",
                    (rid, fname, safe_name[:240], (f.mimetype or "")[:120]),
                )
            req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
            log_event(rid, "creada", None, "solicitud_recibida", "Solicitud creada desde pagina publica", "cliente", data["contact_name"])
            # Invalidar cache del calendario público — el nuevo retiro
            # ocupa un slot, los demás clientes deben verlo al instante.
            try: _DISPO_CACHE["payload"] = None
            except Exception: pass

            # ──────────────────────────────────────────────────────────────
            # FIRE-AND-FORGET (Daniel 2026-05-24): el cliente NO debe esperar
            # SMTP. El envío de email tarda 3-8 s; antes el overlay duraba
            # 15 s porque el form era submit nativo y el browser esperaba el
            # redirect. Ahora respondemos JSON al instante (~150 ms) y todo
            # lo no-crítico (notify cliente, log declaración tercero,
            # log email_enviado/pendiente) corre en thread daemon. Si SMTP
            # falla, el cliente ya tiene su RET-XXX y un operador verá
            # "email_pendiente" en el panel interno.
            # ──────────────────────────────────────────────────────────────
            _req_snap = dict(req or {})
            _data_snap = dict(data or {})
            _auth_snap = bool(auth_active)

            def _bg_post_insert(rid_bg, req_bg, data_bg, auth_bg):
                # FIX 2026-06-09 (verificado en logs de prod): este thread
                # crasheaba con "Working outside of application context" —
                # notify()/settings()/log_event usan get_db() → g, que solo
                # existe dentro de un app_context. Lo abrimos explícitamente.
                try:
                  with app.app_context():
                    # 1. Log de declaración de tercero (auditoría legal)
                    try:
                        _cli_rut_c = _clean_rut(data_bg.get("customer_rut") or "")
                        _ret_rut_c = _clean_rut(data_bg.get("pickup_person_rut") or "")
                        if (auth_bg and _cli_rut_c and _ret_rut_c
                                and _cli_rut_c != _ret_rut_c
                                and is_valid_rut(_cli_rut_c) and is_valid_rut(_ret_rut_c)):
                            log_event(
                                rid_bg, "declaracion_tercero",
                                "solicitud_recibida", "solicitud_recibida",
                                f"Cliente {data_bg.get('customer_name','')} ({data_bg.get('customer_rut','')}) "
                                f"declaró que {data_bg.get('pickup_person_name','')} "
                                f"({data_bg.get('pickup_person_rut','')}) está autorizado para retirar.",
                                "cliente", data_bg.get("contact_name") or data_bg.get("customer_name") or "Cliente"
                            )
                    except Exception as _e_decl_log:
                        print(f"[pickup_public_request][BG] decl-log error: {_e_decl_log}", flush=True)

                    # 2. Envío de email al cliente (lo lento — SMTP)
                    sent_mail, sent_wa = notify(req_bg, "created")
                    _masked_email = _mask_email(req_bg.get('contact_email'))
                    if sent_mail:
                        log_event(rid_bg, "email_enviado", "solicitud_recibida", "solicitud_recibida",
                                  f"Correo enviado a {_masked_email}", "sistema", "Comunicaciones")
                    else:
                        log_event(rid_bg, "email_pendiente", "solicitud_recibida", "solicitud_recibida",
                                  f"No se pudo enviar correo a {_masked_email}. Revisar Comunicaciones.",
                                  "sistema", "Comunicaciones")
                    if sent_wa:
                        log_event(rid_bg, "whatsapp_enviado", "solicitud_recibida", "solicitud_recibida",
                                  "WhatsApp de solicitud enviado.", "sistema", "Comunicaciones")

                    # 3. Notificación al EQUIPO interno (in-app + email) — SLA
                    try:
                        _code_bg = req_bg.get("code") or "?"
                        _doc_bg = f"{(req_bg.get('document_type') or '').upper()} {req_bg.get('document_number') or ''}".strip()
                        # _td_to_hhmm: los TIME de MySQL llegan como timedelta
                        # ('9:00:00' → str()[:5] daría '9:00:' sin cero inicial)
                        _tf_bg = _td_to_hhmm(req_bg.get("requested_time_from")) if req_bg.get("requested_time_from") else ""
                        _tt_bg = _td_to_hhmm(req_bg.get("requested_time_to")) if req_bg.get("requested_time_to") else ""
                        _notificar_equipo_retiros(
                            f"Nueva solicitud de retiro {_code_bg}",
                            (f"Cliente: {req_bg.get('customer_name') or '?'}. "
                             f"Documento: {_doc_bg or '—'}. "
                             f"Fecha pedida: {str(req_bg.get('requested_date') or '—')} "
                             f"{_tf_bg}-{_tt_bg}."),
                            rid_bg, _code_bg,
                            prioridad="alta", tipo="retiro_nuevo",
                        )
                    except Exception as _e_team:
                        print(f"[pickup_public_request][BG] team-notify error: {_e_team}", flush=True)
                except Exception as _bg_exc:
                    print(f"[pickup_public_request][BG] crash rid={rid_bg}: {_bg_exc}", flush=True)

            try:
                threading.Thread(
                    target=_bg_post_insert,
                    args=(rid, _req_snap, _data_snap, _auth_snap),
                    daemon=True,
                ).start()
            except Exception as _t_exc:
                print(f"[pickup_public_request] no se pudo lanzar BG: {_t_exc}", flush=True)
                # Si el thread no arranca, hacemos el envío inline como fallback
                try: notify(req, "created")
                except Exception: pass

            # ── Detección AJAX: si el form usa fetch(), respondemos JSON ──
            # El JS del overlay premium pinta el código RET-XXX al instante.
            _is_ajax_req = (
                request.headers.get("X-Requested-With") == "XMLHttpRequest"
                or "application/json" in (request.headers.get("Accept") or "")
            )
            if _is_ajax_req:
                return jsonify({
                    "ok": True,
                    "code": code,
                    "token": token,
                    "tracking_url": url_for("pickup_public_tracking", token=token, created=1),
                })
            return redirect(url_for("pickup_public_tracking", token=token, created=1))
        # Anti-cache para forzar al navegador a recargar diseño actualizado
        from flask import make_response
        # Cargar imágenes activas del carrusel desde BD (retiros_carousel)
        # FIX 2026-05-16: el template espera img.src (no archivo_path).
        # Cargamos cloudinary_url y archivo_path, y construimos .src priorizando
        # Cloudinary (URL persistente). Si solo hay archivo_path, asumimos
        # filesystem en /static/uploads/retiros/ donde lo guarda el admin.
        try:
            carousel_rows = mysql_fetchall(
                "SELECT id, archivo_path, cloudinary_url, titulo, subtitulo "
                "FROM retiros_carousel "
                "WHERE activa=1 ORDER BY orden ASC, id ASC"
            )
            carousel_images = []
            for r in (carousel_rows or []):
                d = dict(r)
                # Construir URL final: Cloudinary primero (persistente)
                cld = (d.get("cloudinary_url") or "").strip()
                path = (d.get("archivo_path") or "").strip()
                if cld:
                    d["src"] = cld
                elif path:
                    # archivo_path puede venir como "retiros/abc.jpg" o "abc.jpg"
                    if path.startswith("http"):
                        d["src"] = path
                    elif path.startswith("/static/"):
                        d["src"] = path
                    elif path.startswith("static/"):
                        d["src"] = "/" + path
                    else:
                        d["src"] = "/static/uploads/retiros/" + path.lstrip("/")
                else:
                    d["src"] = ""
                if d["src"]:
                    carousel_images.append(d)
        except Exception as e_car:
            print(f"[pickup_public_request] carrusel error: {e_car}", flush=True)
            carousel_images = []
        # Cargar avisos vigentes (retiros_announcements)
        try:
            anuncios_rows = mysql_fetchall(
                "SELECT id, titulo, mensaje, tipo, icon FROM retiros_announcements "
                "WHERE activa=1 "
                "  AND (fecha_desde IS NULL OR fecha_desde <= NOW()) "
                "  AND (fecha_hasta IS NULL OR fecha_hasta >= NOW()) "
                "ORDER BY orden ASC, id DESC"
            )
            anuncios = [dict(r) for r in (anuncios_rows or [])]
        except Exception:
            anuncios = []
        resp = make_response(render_template(
            "retiros/public_request.html",
            settings=cfg, relations=PICKUP_RELATIONS, errors=[], fd={},
            carousel_images=carousel_images,
            announcements=anuncios,
        ))
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    @app.route("/retiros/buscar")
    @_rate_limited("pickup_buscar_publico", max_attempts=30, window_seconds=300, methods=("GET",))
    def pickup_buscar_publico():
        """Lookup público por código de retiro (RET-XXXXXX).
        El form público de seguimiento referencia esta ruta.
        Si encuentra el retiro, redirige a /retiros/seguimiento/<token>.

        Seguridad (Daniel mayo 2026):
        - Rate limit: 30 búsquedas / 5 min por IP. Antes era ilimitado,
          permitiendo enumerar códigos RET-XXXXXX por fuerza bruta (los
          códigos son secuenciales, por eso luego de confirmar siempre
          redirigimos al token público — no al código).
        - Validación estricta del formato del código antes de tocar BD:
          solo acepta RET-NNNNNN o NNNNNN (anti-injection y anti-noise).
        - Sanitiza el code en el flash (evita XSS si flash renderiza HTML).
        """
        code_raw = (request.args.get("code") or "").strip().upper()
        if not code_raw:
            return redirect(url_for("pickup_public_request"))
        # Validar formato: 6-12 chars alfanuméricos + guion. Si no calza,
        # rechazamos sin tocar BD (evita LIKE costoso con basura).
        # FIX 2026-06-15: regex anterior (\d+) rechazaba los códigos nuevos
        # alfanuméricos tipo RET-H29JRR (alphabet ABCDEFGHJKLMNPQRSTUVWXYZ23456789).
        if not re.match(r"^(RET[-_]?)?[A-Z0-9]{4,12}$", code_raw):
            flash("Código inválido. Usa el formato RET-XXXXXX.", "warning")
            return redirect(url_for("pickup_public_request"))
        # Buscar por código exacto primero (más rápido)
        row = mysql_fetchone(
            f"SELECT public_token FROM `{REQ}` WHERE UPPER(code)=%s LIMIT 1",
            (code_raw,)
        )
        if not row:
            # Probar quitando "RET-" si lo trae (sigue con LIKE acotado al final)
            code_alt = code_raw.replace("RET-", "").replace("RET_", "").lstrip("0") or code_raw
            row = mysql_fetchone(
                f"SELECT public_token FROM `{REQ}` "
                f"WHERE code LIKE %s ORDER BY id DESC LIMIT 1",
                (f"%{code_alt}",)
            )
        if not row or not row.get("public_token"):
            # IMPORTANTE: NO loguear el código intentado en stdout (un atacante
            # podría llenar logs con códigos basura). El flash sí lo muestra
            # al usuario que lo escribió.
            # Sanitizar antes de inyectar al flash (defensa anti-XSS)
            safe_code = re.sub(r"[^A-Z0-9\-_]", "", code_raw)[:20]
            flash(f"No encontramos un retiro con el código '{safe_code}'. Verifica e intenta nuevamente.", "warning")
            return redirect(url_for("pickup_public_request"))
        return redirect(url_for("pickup_public_tracking", token=row["public_token"]))


    @app.route("/retiros/seguimiento/<token>", methods=["GET", "POST"])
    @_rate_limited("pickup_public_tracking_post", max_attempts=20, window_seconds=600, methods=("POST",))
    def pickup_public_tracking(token):
        # Validar formato del token antes de tocar BD (evita enumeración de
        # paths inválidos por bots). Tokens son URL-safe-base64 de >= 16 chars.
        if not token or len(token) < 16 or len(token) > 200 \
                or not re.match(r"^[A-Za-z0-9_\-]+$", token):
            return "Solicitud no encontrada", 404
        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE public_token=%s LIMIT 1", (token,))
        if not req:
            return "Solicitud no encontrada", 404
        cfg = settings()
        if request.method == "POST":
            action = request.form.get("action")
            old = req["status"]
            # 2026-05-23 (Daniel) — logging quirúrgico para diagnosticar
            # "el botón Confirmar no avanza". Logueamos cada POST con
            # request_id, action y status entrante. Nada de datos personales.
            try:
                print(f"[pickup_tracking] POST req_id={req['id']} code={req.get('code')} action={action} status_before={old}", flush=True)
            except Exception:
                pass
            # ── Detección AJAX (stepper modal de confirmación) ──
            # Daniel mayo 2026: cuando el cliente confirma vía stepper,
            # el modal hace fetch() con X-Requested-With y necesita JSON
            # en vez de redirect. Si NO es AJAX, mantenemos el flujo legacy
            # con redirect+flash (form submit normal).
            _is_ajax = (
                request.headers.get("X-Requested-With") == "XMLHttpRequest"
                or (request.headers.get("Accept") or "").startswith("application/json")
            )
            def _ajax_ok(payload=None):
                p = {"ok": True}
                if payload: p.update(payload)
                try: _POLL_CACHE.pop(token, None)
                except Exception: pass
                try: _DISPO_CACHE["payload"] = None
                except Exception: pass
                return jsonify(p)
            def _ajax_err(msg, code=409, payload=None):
                p = {"ok": False, "error": str(msg)[:300]}
                if payload: p.update(payload)
                return jsonify(p), code
            # wrapper defensivo: cualquier excepción NO capturada por handlers
            # internos devuelve JSON legible (no HTML 500) al cliente.
            # Root cause histórico: timedelta vs str en time_allowed → TypeError.
            # Aunque ya fue corregido en _validar_disponibilidad_slot, este try
            # atrapa cualquier excepción futura antes de que Flask la convierta en HTML.
            try:
              if action == "confirm":
                proposal = mysql_fetchone(f"SELECT * FROM `{PROP}` WHERE request_id=%s AND status='pending' ORDER BY id DESC LIMIT 1", (req["id"],))
                if not proposal:
                    # No hay propuesta pendiente — el cliente vio el botón
                    # pero el operador ya la canceló / cambió de estado.
                    # Antes el código simplemente recargaba sin feedback.
                    try:
                        print(f"[pickup_tracking] CONFIRM sin propuesta pendiente req_id={req['id']}", flush=True)
                    except Exception: pass
                    if _is_ajax:
                        return _ajax_err(
                            "Esa propuesta ya no está vigente. Si tu retiro sigue pendiente, ILUS te enviará una nueva fecha en breve.",
                            code=410,
                            payload={"reason": "no_pending_proposal"},
                        )
                    flash(
                        "Esa propuesta ya no está vigente. Si tu retiro sigue pendiente, ILUS te enviará una nueva fecha en breve.",
                        "warning",
                    )
                    try: _POLL_CACHE.pop(token, None)
                    except Exception: pass
                    return redirect(url_for("pickup_public_tracking", token=token))
                # FASE 3 (2026-05-29): la propuesta existe pero pudo VENCER
                # (expires_at < ahora Chile). NO se puede confirmar una propuesta
                # vencida — se marca 'expired' y se pide al cliente nueva fecha.
                if proposal and not proposal_is_vigente(proposal):
                    try:
                        mysql_execute(
                            f"UPDATE `{PROP}` SET status='expired', answered_at=NOW() WHERE id=%s",
                            (proposal["id"],),
                        )
                    except Exception: pass
                    log_event(req["id"], "propuesta_vencida", old, old,
                              "Cliente intentó confirmar una propuesta vencida", "sistema", "Expiry")
                    _msg_exp = ("Esta propuesta venció o la fecha ya pasó. Propón una "
                                "nueva fecha y el equipo ILUS te responderá.")
                    try: _POLL_CACHE.pop(token, None)
                    except Exception: pass
                    if _is_ajax:
                        return _ajax_err(_msg_exp, code=410, payload={"reason": "proposal_expired"})
                    flash(_msg_exp, "warning")
                    return redirect(url_for("pickup_public_tracking", token=token))
                # ── GATE DE SEGURIDAD (C2, 2026-06-09) ──────────────────────
                # Solo se puede confirmar una propuesta hecha por ILUS
                # (proposed_by='internal'). Si la pending es la CONTRAPROPUESTA
                # del propio cliente, NO puede auto-confirmarla — la acepta el
                # equipo desde /retiros/<rid>/aceptar-contrapropuesta.
                if proposal and str(proposal.get("proposed_by") or "").lower() != "internal":
                    _msg_rev = ("Tu propuesta está en revisión por nuestro equipo. "
                                "Te confirmaremos por correo en breve.")
                    if _is_ajax:
                        return _ajax_err(_msg_rev, code=409,
                                         payload={"reason": "client_proposal_in_review"})
                    flash(_msg_rev, "info")
                    return redirect(url_for("pickup_public_tracking", token=token))
                if proposal:
                    # ═══════════════════════════════════════════════════════════
                    # FIX RACE CONDITION (2026-05-12):
                    # Entre la lectura de capacidad y el UPDATE puede entrar otro
                    # cliente que tome el último cupo. Usamos transacción explícita
                    # con SELECT ... FOR UPDATE para bloquear las filas del mismo
                    # slot mientras hacemos el conteo + UPDATE atómico.
                    #
                    # Estrategia en 2 fases:
                    #  1. Validar condiciones estáticas (día, horario, bloqueos)
                    #     sin lock — son inmutables entre clics
                    #  2. Abrir transacción → SELECT FOR UPDATE sobre el slot →
                    #     re-contar capacidad numérica → si OK, UPDATE → COMMIT.
                    #     Si OTRO cliente confirmó antes y llenó el slot, este
                    #     SELECT FOR UPDATE espera y luego ve el nuevo conteo.
                    # ═══════════════════════════════════════════════════════════

                    # FASE 1: Validar condiciones estáticas (sin lock)
                    # Daniel 2026-05-24: si la propuesta vino del operador
                    # interno (proposed_by='internal'), respetamos su decisión
                    # de cruzar colación. El cliente solo está aceptando un
                    # horario que el equipo ILUS ya validó manualmente.
                    _prop_bypass_lunch = (str(proposal.get("proposed_by") or "").lower() == "internal")
                    ok_slot, motivo = _validar_disponibilidad_slot(
                        proposal["date"], proposal["time_from"], proposal["time_to"],
                        exclude_request_id=req["id"],
                        extra_kg=float(req.get("total_weight_kg") or 0),
                        extra_m3=float(req.get("total_volume_m3") or 0),
                        bypass_lunch=_prop_bypass_lunch,
                    )
                    if not ok_slot:
                        # Marcar la propuesta como declined y notificar al cliente
                        mysql_execute(
                            f"UPDATE `{PROP}` SET status='declined', answered_at=NOW() WHERE id=%s",
                            (proposal["id"],),
                        )
                        log_event(
                            req["id"], "confirm_bloqueada", old, old,
                            f"Cliente intentó confirmar pero slot ya no disponible: {motivo}",
                            "sistema", "Validación capacidad",
                        )
                        if _is_ajax:
                            return _ajax_err(
                                "Lo sentimos, este horario ya no está disponible. "
                                "Te enviamos una nueva propuesta a la brevedad.",
                                code=409,
                                payload={"reason": "slot_unavailable", "detail": motivo},
                            )
                        flash(
                            "Lo sentimos, este horario ya no está disponible. "
                            "Te enviamos una nueva propuesta a la brevedad.",
                            "warning",
                        )
                        return redirect(url_for("pickup_public_tracking", token=token))

                    # FASE 2: Transacción con lock para garantizar atomicidad
                    confirm_ok = False
                    confirm_motivo = ""
                    conn_tx = None
                    try:
                        conn_tx = get_mysql()
                        with conn_tx.cursor() as cur_tx:
                            # 2a. SELECT FOR UPDATE sobre TODOS los retiros que comparten
                            #     el slot (mismo date+time_from). Bloquea esas filas para
                            #     que ningún otro proceso pueda confirmar/contar mientras
                            #     estamos en esta transacción.
                            extra_kg = float(req.get("total_weight_kg") or 0)
                            extra_m3 = float(req.get("total_volume_m3") or 0)
                            date_str = str(proposal["date"])[:10]
                            tf_str   = _td_to_hhmm(proposal["time_from"])

                            cur_tx.execute(
                                f"""SELECT COUNT(*) AS n,
                                            COALESCE(SUM(total_weight_kg),0) AS kg,
                                            COALESCE(SUM(total_volume_m3),0) AS m3
                                     FROM `{REQ}`
                                     WHERE status NOT IN ('rechazada','cerrada','fallida')
                                       AND id <> %s
                                       AND (
                                         (confirmed_date=%s AND TIME_FORMAT(confirmed_time_from,'%%H:%%i')=%s)
                                         OR
                                         (confirmed_date IS NULL AND proposed_date=%s
                                          AND TIME_FORMAT(proposed_time_from,'%%H:%%i')=%s)
                                       )
                                     FOR UPDATE""",
                                (req["id"], date_str, tf_str, date_str, tf_str),
                            )
                            slot_row = cur_tx.fetchone() or {}
                            picks_now = int(slot_row.get("n") or 0)
                            kg_now    = float(slot_row.get("kg") or 0)
                            m3_now    = float(slot_row.get("m3") or 0)

                            cfg_lock = settings()
                            # FASE 4 (2026-05-29): capacidad ÚNICA = parallel_capacity,
                            # la MISMA fuente que ve el calendario público y
                            # _validar_disponibilidad_slot. Antes acá se usaba
                            # max_picks_per_slot (default 5) → permitía confirmar 5
                            # retiros cuando el calendario mostraba cupo 2 = SOBREVENTA.
                            _pc_lock = cfg_lock.get("parallel_capacity")
                            if _pc_lock is not None and str(_pc_lock).strip():
                                max_picks_slot = int(_pc_lock)
                            else:
                                max_picks_slot = int(cfg_lock.get("max_picks_per_slot") or 2)
                            max_kg_slot    = float(cfg_lock.get("max_kg_per_slot") or 500)
                            max_m3_slot    = float(cfg_lock.get("max_m3_per_slot") or 5)

                            # 2b. Re-validar capacidad bajo el lock
                            if picks_now + 1 > max_picks_slot:
                                confirm_motivo = f"Slot lleno por otro retiro: {picks_now} de {max_picks_slot}."
                            elif kg_now + extra_kg > max_kg_slot:
                                confirm_motivo = f"Capacidad de peso excedida: {kg_now:.1f}+{extra_kg:.1f} > {max_kg_slot:.0f} kg."
                            elif m3_now + extra_m3 > max_m3_slot:
                                confirm_motivo = f"Capacidad de volumen excedida: {m3_now:.2f}+{extra_m3:.2f} > {max_m3_slot:.2f} m³."
                            else:
                                # 2c. UPDATE atómico dentro de la transacción
                                cur_tx.execute(
                                    f"""UPDATE `{REQ}`
                                          SET status='agenda_confirmada',
                                              confirmed_date=%s,
                                              confirmed_time_from=%s,
                                              confirmed_time_to=%s
                                        WHERE id=%s
                                          AND status <> 'agenda_confirmada'""",
                                    (proposal["date"], proposal["time_from"],
                                     proposal["time_to"], req["id"]),
                                )
                                if cur_tx.rowcount == 0:
                                    # Alguien más confirmó este mismo retiro entre clics
                                    confirm_motivo = "El retiro ya fue confirmado anteriormente."
                                else:
                                    cur_tx.execute(
                                        f"UPDATE `{PROP}` SET status='accepted', answered_at=NOW() WHERE id=%s",
                                        (proposal["id"],),
                                    )
                                    confirm_ok = True

                        if confirm_ok:
                            conn_tx.commit()
                        else:
                            conn_tx.rollback()
                    except Exception as _tx_err:
                        if conn_tx is not None:
                            try: conn_tx.rollback()
                            except Exception: pass
                        confirm_motivo = f"Error técnico al confirmar: {_tx_err}"
                        print(f"[pickup_confirm] tx error req={req['id']}: {_tx_err}", flush=True)
                    finally:
                        if conn_tx is not None:
                            try: conn_tx.close()
                            except Exception: pass

                    if not confirm_ok:
                        # Marcar propuesta como declined y avisar
                        try:
                            mysql_execute(
                                f"UPDATE `{PROP}` SET status='declined', answered_at=NOW() WHERE id=%s",
                                (proposal["id"],),
                            )
                        except Exception: pass
                        log_event(
                            req["id"], "confirm_bloqueada", old, old,
                            f"Cliente intentó confirmar pero falló (race): {confirm_motivo}",
                            "sistema", "Validación bajo lock",
                        )
                        if _is_ajax:
                            return _ajax_err(
                                "Lo sentimos, este horario ya no está disponible.",
                                code=409,
                                payload={"reason": "race_lost", "detail": confirm_motivo},
                            )
                        flash(
                            f"Lo sentimos, este horario ya no está disponible. {confirm_motivo} "
                            "Te enviamos una nueva propuesta a la brevedad.",
                            "warning",
                        )
                        return redirect(url_for("pickup_public_tracking", token=token))

                    log_event(req["id"], "cliente_confirmo", old, "agenda_confirmada", "Cliente acepto propuesta", "cliente", req["contact_name"])
                    try:
                        print(f"[pickup_tracking] CONFIRM OK req_id={req['id']} fecha={proposal['date']} {proposal['time_from']}-{proposal['time_to']}", flush=True)
                    except Exception: pass
                    # Email "agenda confirmada" al cliente (con detalles finales)
                    try:
                        req_after = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (req["id"],)) or req
                        notify_async(req_after, "confirmed")
                    except Exception as _e: print(f"[pickups][notify confirm] {_e}")
                    # Notificación al EQUIPO (in-app + email) — el cliente confirmó
                    try:
                        _notificar_equipo_retiros(
                            f"El cliente CONFIRMÓ el retiro {req.get('code') or '?'}",
                            (f"Cliente: {req.get('customer_name') or '?'}. "
                             f"Fecha confirmada: {str(proposal['date'])[:10]} "
                             f"{_td_to_hhmm(proposal['time_from'])}-{_td_to_hhmm(proposal['time_to'])}."),
                            req["id"], req.get("code") or "?",
                            prioridad="alta", tipo="retiro_respuesta",
                        )
                    except Exception as _e: print(f"[pickups][team notify confirm] {_e}")
                    # D2: si TODOS los docs ERP están sin saldo → alerta urgente
                    # al equipo (NO bloquea la confirmación).
                    try:
                        _alertar_si_sin_saldo(req["id"], req.get("code"))
                    except Exception as _e: print(f"[pickups][saldo alert confirm] {_e}")
                    if _is_ajax:
                        return _ajax_ok({
                            "message": "Retiro confirmado",
                            "fecha":    str(proposal["date"])[:10],
                            "hora_desde": str(proposal["time_from"])[:5],
                            "hora_hasta": str(proposal["time_to"])[:5],
                            "redirect_url": url_for("pickup_public_tracking", token=token),
                        })
                    flash(
                        f"¡Retiro confirmado para {proposal['date']} a las {str(proposal['time_from'])[:5]}! "
                        "Te enviamos un correo con todos los detalles.",
                        "success",
                    )
              elif action == "reject":
                # Si la solicitud ya está cerrada/rechazada, no hacer nada
                # (defensa contra doble-submit / refresh con resend POST).
                if old in ("rechazada", "cerrada"):
                    flash("Esta solicitud ya fue cancelada anteriormente.", "info")
                    return redirect(url_for("pickup_public_tracking", token=token))
                reason = (request.form.get("reason") or "").strip()[:500]
                # C3 (2026-06-09): cerrar también closed_at — gap conocido, el
                # rechazo del cliente dejaba la solicitud sin fecha de cierre.
                mysql_execute(f"UPDATE `{REQ}` SET status='rechazada', closed_at=NOW() WHERE id=%s", (req["id"],))
                # También marcar la propuesta pendiente (si hay) como declined
                try:
                    mysql_execute(
                        f"UPDATE `{PROP}` SET status='declined', answered_at=NOW() "
                        f"WHERE request_id=%s AND status='pending'",
                        (req["id"],)
                    )
                except Exception: pass
                log_event(req["id"], "cliente_rechazo", old, "rechazada", reason or "(sin motivo)", "cliente", req["contact_name"])
                try:
                    print(f"[pickup_tracking] REJECT req_id={req['id']} reason_len={len(reason)}", flush=True)
                except Exception: pass
                # Notificación al cliente (confirmación de cancelación)
                try:
                    req_after = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (req["id"],)) or req
                    notify_async(req_after, "rejected", custom_message=(
                        "Recibimos tu cancelación. Si fue un error o quieres reagendar, "
                        "puedes hacer una nueva solicitud cuando estés listo."
                    ))
                except Exception as _e: print(f"[pickups][notify reject cliente] {_e}")
                # Notificación al operador interno (avisa que el cliente rechazó)
                try:
                    _notificar_operador_rechazo(req, reason)
                except Exception as _e: print(f"[pickups][notify reject operador] {_e}")
                # In-app al equipo (campana). El email interno YA sale arriba
                # vía _notificar_operador_rechazo → send_email=False evita duplicar.
                try:
                    _notificar_equipo_retiros(
                        f"El cliente RECHAZÓ el retiro {req.get('code') or '?'}",
                        (f"Cliente: {req.get('customer_name') or '?'}. "
                         f"Motivo declarado: {reason or 'No indicó motivo'}."),
                        req["id"], req.get("code") or "?",
                        prioridad="urgente", tipo="retiro_respuesta",
                        send_email=False,
                    )
                except Exception as _e: print(f"[pickups][team notify reject] {_e}")
                flash(
                    "Tu solicitud fue cancelada. Te enviamos un correo de confirmación y avisamos al equipo ILUS.",
                    "info",
                )
              elif action == "counter":
                date, tf, tt = request.form.get("counter_date"), request.form.get("counter_time_from"), request.form.get("counter_time_to")
                # FASE 2 (2026-05-29): validador temporal CENTRAL, modo 'public'.
                # El cliente contrapropone → exige anticipación mínima (min_notice),
                # prohíbe fecha/hora pasada, valida día hábil/feriado/horario/colación
                # y duración múltiplo de slot. Reemplaza las 3 validaciones sueltas
                # anteriores (date_allowed + time_allowed + anti-pasado manual).
                ok_dt, msg_dt = validate_pickup_datetime(date, tf, tt, cfg=cfg, mode="public")
                ok_slot, motivo_slot = True, ""
                if ok_dt:
                    try:
                        ok_slot, motivo_slot = _validar_disponibilidad_slot(
                            date, tf, tt,
                            exclude_request_id=req["id"],
                            extra_kg=float(req.get("total_weight_kg") or 0),
                            extra_m3=float(req.get("total_volume_m3") or 0),
                        )
                    except Exception as _e:
                        print(f"[pickup_tracking] counter validar slot err: {_e}", flush=True)
                if ok_dt and ok_slot:
                    # FASE 3: marcar propuestas pending anteriores como 'superseded'
                    # (solo una vigente a la vez) y setear expires_at en la
                    # contrapropuesta del cliente. Queda 'en_revision' → ILUS debe
                    # aceptar o enviar nueva propuesta (lado ILUS del ping-pong).
                    try:
                        _expiry_h = int(cfg.get("proposal_expiry_hours") or 48)
                    except (TypeError, ValueError):
                        _expiry_h = 48
                    _expires_at = (_now_chile() + timedelta(hours=_expiry_h)).strftime("%Y-%m-%d %H:%M:%S")
                    mysql_execute(
                        f"UPDATE `{PROP}` SET status='superseded', answered_at=NOW() "
                        f"WHERE request_id=%s AND status='pending'", (req["id"],)
                    )
                    mysql_execute(
                        f"""INSERT INTO `{PROP}` (request_id,proposed_by,date,time_from,time_to,message,reason,status,token,expires_at)
                            VALUES (%s,'cliente',%s,%s,%s,%s,'Contrapropuesta cliente','pending',%s,%s)""",
                        (req["id"], date, tf, tt, (request.form.get("counter_message", "") or "")[:1000],
                         secrets.token_urlsafe(24), _expires_at),
                    )
                    mysql_execute(f"UPDATE `{REQ}` SET status='en_revision' WHERE id=%s", (req["id"],))
                    log_event(req["id"], "cliente_contrapropuso", old, "en_revision", f"{date} {tf}-{tt}", "cliente", req["contact_name"])
                    try:
                        print(f"[pickup_tracking] COUNTER OK req_id={req['id']} fecha={date} {tf}-{tt}", flush=True)
                    except Exception: pass
                    # Notificación al EQUIPO (in-app + email): hay contrapropuesta
                    # esperando respuesta de ILUS (aceptar o proponer otra fecha).
                    try:
                        _notificar_equipo_retiros(
                            f"El cliente contrapropuso fecha en el retiro {req.get('code') or '?'}",
                            (f"Cliente: {req.get('customer_name') or '?'}. "
                             f"El cliente contrapropuso {date} {tf}-{tt}. "
                             f"Acepta la contrapropuesta o envía una nueva fecha."),
                            req["id"], req.get("code") or "?",
                            prioridad="alta", tipo="retiro_respuesta",
                        )
                    except Exception as _e: print(f"[pickups][team notify counter] {_e}")
                    # Email AL CLIENTE confirmando recepción de su contrapropuesta
                    # (antes el cliente quedaba sin acuse por correo).
                    try:
                        req_after = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (req["id"],)) or req
                        notify_async(req_after, "message", custom_message=(
                            f"Recibimos tu propuesta de fecha {date} {tf}-{tt}. "
                            "Nuestro equipo la revisará y te confirmaremos por correo."
                        ))
                    except Exception as _e: print(f"[pickups][notify counter cliente] {_e}")
                    flash(
                        f"¡Listo! Registramos tu contrapropuesta para {date} a las {tf}. "
                        "Te confirmaremos en breve por correo.",
                        "success",
                    )
                else:
                    _err = msg_dt if not ok_dt else f"Ese horario ya no está disponible: {motivo_slot}"
                    try:
                        print(f"[pickup_tracking] COUNTER rechazada req_id={req['id']} ok_dt={ok_dt} ok_slot={ok_slot} motivo={_err}", flush=True)
                    except Exception: pass
                    if _is_ajax:
                        return _ajax_err(_err, code=409, payload={"reason": "counter_invalid"})
                    flash(_err or "No se pudo registrar la propuesta. Intenta con otro horario.", "warning")
            except Exception as _outer_err:
                try:
                    print(f"[pickup_tracking] CRASH req={req.get('id')} action={action}: {_outer_err}", flush=True)
                    import traceback; traceback.print_exc()
                except Exception:
                    pass
                if _is_ajax:
                    return jsonify({"ok": False, "error": "Error interno del servidor. Intenta nuevamente."}), 500
                flash("Error interno al procesar la acción. Intenta nuevamente.", "danger")
            # Invalidar cache de polling — el cliente verá su nuevo estado al instante
            try: _POLL_CACHE.pop(token, None)
            except Exception: pass
            # Invalidar cache global del calendario público (ocupación cambia)
            try: _DISPO_CACHE["payload"] = None
            except Exception: pass
            return redirect(url_for("pickup_public_tracking", token=token))
        packages = mysql_fetchall(f"SELECT * FROM `{PKG}` WHERE request_id=%s ORDER BY package_number", (req["id"],))
        proposals = mysql_fetchall(f"SELECT * FROM `{PROP}` WHERE request_id=%s ORDER BY id DESC", (req["id"],))
        logs = mysql_fetchall(f"SELECT * FROM `{LOG}` WHERE request_id=%s ORDER BY id DESC LIMIT 20", (req["id"],))
        attachments = mysql_fetchall(f"SELECT * FROM `{ATT}` WHERE request_id=%s ORDER BY id DESC", (req["id"],))
        # 2026-05-23 (Daniel): cálculo robusto de m³ en backend (evita Jinja
        # `namespace` que falla en algunas versiones).
        try:
            docs_asociados = mysql_fetchall(
                "SELECT id, document_type, document_number, peso_real_kg, peso_vol_kg, volumen_m3, n_lineas "
                "  FROM pickup_request_docs WHERE request_id=%s ORDER BY id ASC",
                (req["id"],)
            ) or []
        except Exception:
            docs_asociados = []
        # Fallback cascada m³: req → docs_asociados → packages
        _m3 = float(req.get("total_volume_m3") or req.get("volumen_m3") or 0)
        if _m3 == 0 and docs_asociados:
            _m3 = float(sum((d.get("volumen_m3") or 0) for d in docs_asociados))
        if _m3 == 0 and packages:
            _m3 = float(sum(
                ((p.get("length_cm") or 0) * (p.get("width_cm") or 0) * (p.get("height_cm") or 0)) / 1000000.0
                for p in packages
                if p.get("length_cm") and p.get("width_cm") and p.get("height_cm")
            ))
        # Sanitizar: eliminar campos internos antes de pasar al template público.
        req_safe = dict(_strip_internal(req))
        req_safe["m3_calculado"] = _m3
        _tracking_html = render_template("retiros/public_tracking.html",
                               req=req_safe, packages=packages, proposals=proposals,
                               logs=logs, attachments=attachments,
                               docs_asociados=docs_asociados,
                               settings=cfg, status_badge=status_badge,
                               # Daniel 2026-06-15: modelo canónico ÚNICO de tracking
                               # (mismos hitos que el correo). journey_idx = hito
                               # actual (0-4) o -1 (cancelado/fallido).
                               journey=PICKUP_JOURNEY,
                               journey_idx=pickup_journey_idx(req.get("status")),
                               created=request.args.get("created"))
        # Daniel 2026-06-15: NO cachear la página de seguimiento — el cliente
        # debe ver SIEMPRE el estado actual (si la propuesta cambia, al recargar
        # ve el hito nuevo). Sin esto el navegador servía HTML viejo (stepper
        # desactualizado). El polling de /status (cada 30s) recarga si cambia.
        from flask import make_response as _make_response
        _resp = _make_response(_tracking_html)
        _resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        _resp.headers["Pragma"] = "no-cache"
        _resp.headers["Expires"] = "0"
        return _resp

    # ══════════════════════════════════════════════════════════════════════
    #  XLSX RESUMEN PÚBLICO — descarga del cliente desde el tracking
    # ══════════════════════════════════════════════════════════════════════
    #  Solo requiere el token público del retiro. NO requiere login.
    #  Genera un XLSX con todos los datos relevantes que el cliente
    #  podría querer guardar/imprimir para su propio archivo.
    #
    #  NO incluye datos internos del operador:
    #   · No expone internal_notes / doc_validation_notes
    #   · No expone IPs ni user-agents
    #   · No expone observaciones de fraude / validaciones doc
    @app.route("/retiros/seguimiento/<token>/resumen.xlsx", methods=["GET"])
    @_rate_limited("pickup_public_xlsx", max_attempts=10, window_seconds=300, methods=("GET",))
    def pickup_public_xlsx(token):
        """Excel resumen para el cliente — accesible solo con el token público.

        Seguridad (Daniel mayo 2026):
        - Validamos formato del token antes de tocar BD.
        - Rate limit: 10 descargas / 5 min por IP (evita scraping masivo).
        - NO incluye campos internos (lo verifica `_strip_internal`).
        """
        if not token or len(token) < 16 or len(token) > 200 \
                or not re.match(r"^[A-Za-z0-9_\-]+$", token):
            return "Solicitud no encontrada", 404
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO
        except ImportError:
            return "Servicio de descarga temporalmente no disponible.", 500

        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE public_token=%s LIMIT 1", (token,))
        if not req:
            return "Solicitud no encontrada", 404
        # Sanitizar antes de pasar al builder XLSX (defensa en profundidad).
        req = _strip_internal(req)

        cfg = settings()
        packages = mysql_fetchall(
            f"SELECT * FROM `{PKG}` WHERE request_id=%s ORDER BY package_number",
            (req["id"],)
        ) or []

        # ── Helpers de formato fecha
        def _fmt_date(d):
            if not d: return ""
            try: return d.strftime("%d-%m-%Y") if hasattr(d, "strftime") else str(d)[:10]
            except Exception: return str(d)[:10]
        def _fmt_time(t):
            if not t: return ""
            try: return t.strftime("%H:%M") if hasattr(t, "strftime") else str(t)[:5]
            except Exception: return str(t)[:5]
        def _fmt_dt(d):
            if not d: return ""
            try: return d.strftime("%d-%m-%Y %H:%M") if hasattr(d, "strftime") else str(d)[:16]
            except Exception: return str(d)[:16]

        # ── Estilos
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen retiro"

        ILUS_RED = "DC2626"
        ILUS_BLACK = "0A0A0A"
        SOFT_GRAY = "F3F4F6"
        TXT_DARK = "111827"

        red_fill = PatternFill("solid", fgColor=ILUS_RED)
        black_fill = PatternFill("solid", fgColor=ILUS_BLACK)
        gray_fill = PatternFill("solid", fgColor=SOFT_GRAY)
        white_bold = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        red_bold = Font(name="Calibri", size=11, bold=True, color=ILUS_RED)
        label_bold = Font(name="Calibri", size=10, bold=True, color=TXT_DARK)
        value_font = Font(name="Calibri", size=11, color=TXT_DARK)
        title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        small_gray = Font(name="Calibri", size=9, color="6B7280", italic=True)
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        # ── Encabezado rojo grande
        ws.merge_cells("A1:B1")
        c = ws.cell(1, 1, "ILUS Sport & Health")
        c.font = title_font
        c.fill = red_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 34

        ws.merge_cells("A2:B2")
        c = ws.cell(2, 1, f"Resumen de Retiro · {req.get('code','')}")
        c.font = white_bold
        c.fill = black_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 24

        # ── Status (línea grande)
        status_lbl = PICKUP_STATUS.get(req.get("status") or "", req.get("status") or "")
        ws.merge_cells("A3:B3")
        c = ws.cell(3, 1, f"Estado actual: {status_lbl}")
        c.font = red_bold
        c.fill = gray_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 22

        # ── Sección con cabecera
        def add_section(row_start, title):
            ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=2)
            cc = ws.cell(row_start, 1, title)
            cc.font = white_bold
            cc.fill = black_fill
            cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row_start].height = 22
            return row_start + 1

        def add_row(row, label, value):
            l = ws.cell(row, 1, label)
            l.font = label_bold
            l.fill = gray_fill
            l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            l.border = thin_border
            v = ws.cell(row, 2, value if value not in (None, "") else "—")
            v.font = value_font
            v.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            v.border = thin_border
            return row + 1

        # Fallback bultos/peso/vol/m3
        n_bultos = int(req.get("total_packages") or len(packages) or 0)
        peso_kg = float(req.get("peso_real_kg") or req.get("total_weight_kg") or 0)
        peso_vol = float(req.get("peso_vol_kg") or req.get("total_volumetric_weight") or 0)
        volumen_m3 = float(req.get("total_volume_m3") or req.get("volumen_m3") or 0)

        # Bodega
        bodega_nombre = cfg.get("warehouse_name") or "ILUS Bodega"
        bodega_addr = cfg.get("warehouse_addr") or ""

        # Horarios formateados
        def _slot(date, tf, tt):
            d = _fmt_date(date)
            if tf and tt:
                return f"{d}   {_fmt_time(tf)} – {_fmt_time(tt)}"
            return d

        r = 5
        # Solicitud
        r = add_section(r, "Datos de la solicitud")
        r = add_row(r, "Código de retiro", req.get("code") or "")
        r = add_row(r, "Documento", f"{(req.get('document_type') or '').upper()} {req.get('document_number') or ''}".strip())
        r = add_row(r, "Estado", status_lbl)
        r = add_row(r, "Fecha de solicitud", _fmt_dt(req.get("created_at")))

        r += 1
        # Cliente
        r = add_section(r, "Cliente")
        r = add_row(r, "Razón social / Nombre", req.get("customer_name") or "")
        r = add_row(r, "RUT", req.get("customer_rut") or "")
        r = add_row(r, "Contacto", req.get("contact_name") or "")
        r = add_row(r, "Email contacto", req.get("contact_email") or "")
        r = add_row(r, "Teléfono contacto", req.get("contact_phone") or "")

        r += 1
        # Persona que retira
        r = add_section(r, "Persona que retira")
        r = add_row(r, "Nombre", req.get("pickup_person_name") or "")
        r = add_row(r, "RUT", req.get("pickup_person_rut") or "")
        r = add_row(r, "Teléfono", req.get("pickup_person_phone") or "")
        relation_lbl = dict(PICKUP_RELATIONS).get(req.get("pickup_person_relation") or "", req.get("pickup_person_relation") or "")
        r = add_row(r, "Relación", relation_lbl)

        r += 1
        # Agenda
        r = add_section(r, "Agenda")
        if req.get("requested_date"):
            r = add_row(r, "Fecha solicitada", _slot(req.get("requested_date"), req.get("requested_time_from"), req.get("requested_time_to")))
        if req.get("proposed_date"):
            r = add_row(r, "Propuesta de ILUS", _slot(req.get("proposed_date"), req.get("proposed_time_from"), req.get("proposed_time_to")))
        if req.get("confirmed_date"):
            r = add_row(r, "Fecha confirmada", _slot(req.get("confirmed_date"), req.get("confirmed_time_from"), req.get("confirmed_time_to")))

        r += 1
        # Bodega
        r = add_section(r, "Bodega de retiro")
        r = add_row(r, "Nombre", bodega_nombre)
        r = add_row(r, "Dirección", bodega_addr)
        if cfg.get("maps_url"):
            r = add_row(r, "Ver en mapa", cfg.get("maps_url"))

        r += 1
        # Carga
        r = add_section(r, "Carga declarada")
        r = add_row(r, "Bultos", n_bultos)
        r = add_row(r, "Peso real (kg)", f"{peso_kg:.2f}")
        r = add_row(r, "Peso volumétrico (kg)", f"{peso_vol:.2f}")
        r = add_row(r, "Volumen (m³)", f"{volumen_m3:.4f}")

        # Detalle paquetes (si hay)
        if packages:
            r += 1
            r = add_section(r, "Detalle de bultos")
            # Header tabla
            hdrs = ["#", "Largo (cm)", "Alto (cm)", "Ancho (cm)", "Kg", "P. vol"]
            # Adaptar: lo embebemos en 2 cols mergeando: para mantener layout 2col, mejor expandimos a 6
            # Mejor: descomprimimos a 6 columnas SOLO en esta sección
            for ci, h in enumerate(hdrs, 1):
                cc = ws.cell(r, ci, h)
                cc.font = white_bold
                cc.fill = red_fill
                cc.alignment = Alignment(horizontal="center", vertical="center")
                cc.border = thin_border
            r += 1
            for p in packages:
                row_vals = [
                    p.get("package_number") or "",
                    float(p.get("length_cm") or 0),
                    float(p.get("height_cm") or 0),
                    float(p.get("width_cm") or 0),
                    float(p.get("weight_kg") or 0),
                    float(p.get("volumetric_weight") or 0),
                ]
                for ci, v in enumerate(row_vals, 1):
                    cc = ws.cell(r, ci, v)
                    cc.font = value_font
                    cc.alignment = Alignment(horizontal="center", vertical="center")
                    cc.border = thin_border
                r += 1

        # Observaciones públicas del cliente (no internas)
        if req.get("observations"):
            r += 1
            r = add_section(r, "Observaciones")
            r = add_row(r, "Tus comentarios", req.get("observations") or "")

        # Footer
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        try:
            from zoneinfo import ZoneInfo as _ZI
            now_cl = datetime.now(_ZI("America/Santiago")).strftime("%d-%m-%Y %H:%M")
        except Exception:
            now_cl = datetime.now().strftime("%d-%m-%Y %H:%M")
        cc = ws.cell(r, 1, f"Documento generado el {now_cl} · ILUS Sport & Health · sistema interno de retiros")
        cc.font = small_gray
        cc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 20

        # Anchos columnas
        ws.column_dimensions[get_column_letter(1)].width = 28
        ws.column_dimensions[get_column_letter(2)].width = 48
        if packages:
            # Cuando hubo tabla de bultos extendimos a 6, ajustar
            ws.column_dimensions[get_column_letter(3)].width = 14
            ws.column_dimensions[get_column_letter(4)].width = 14
            ws.column_dimensions[get_column_letter(5)].width = 12
            ws.column_dimensions[get_column_letter(6)].width = 12

        # Guardar
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"ILUS_Retiro_{req.get('code','retiro')}.xlsx"
        from flask import send_file
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname,
        )

    @app.route("/retiros")
    @require_permission("retiros")
    def pickup_dashboard():
        filtros = {"q": request.args.get("q", "").strip()[:80],
                   "status": request.args.get("status", "").strip()[:40],
                   "date": request.args.get("date", "").strip()[:12],
                   "view": request.args.get("view", "monitor").strip()[:20]}
        where, params = ["1=1"], []
        if filtros["q"]:
            like = f"%{filtros['q']}%"; where.append("(code LIKE %s OR document_number LIKE %s OR customer_name LIKE %s OR contact_phone LIKE %s)"); params.extend([like, like, like, like])
        if filtros["status"] and filtros["status"] in PICKUP_STATUS:
            where.append("status=%s"); params.append(filtros["status"])
        if filtros["date"] and re.match(r"^\d{4}-\d{2}-\d{2}$", filtros["date"]):
            where.append("(requested_date=%s OR confirmed_date=%s OR proposed_date=%s)"); params.extend([filtros["date"], filtros["date"], filtros["date"]])
        # SELECT explícito: solo columnas que usa el dashboard (evita serializar
        # campos pesados como doc_erp_data MEDIUMTEXT a la red).
        rows = mysql_fetchall(
            f"""SELECT id, code, status, document_type, document_number,
                       customer_name, customer_rut, contact_name, contact_email,
                       contact_phone, pickup_person_name, pickup_person_rut,
                       pickup_person_phone, pickup_person_relation,
                       requested_date, requested_time_from, requested_time_to,
                       proposed_date, proposed_time_from, proposed_time_to,
                       confirmed_date, confirmed_time_from, confirmed_time_to,
                       total_packages, total_weight_kg, total_volumetric_weight,
                       total_volume_m3, peso_real_kg, peso_vol_kg,
                       tiempo_estimado_min, doc_validation_status,
                       information_quality_score, risk_score,
                       request_source, created_by_user_name, responsable_nombre,
                       created_at, updated_at
                FROM `{REQ}`
                WHERE {' AND '.join(where)}
                ORDER BY created_at DESC LIMIT 250""",
            tuple(params)
        )
        counts = mysql_fetchall(f"SELECT status, COUNT(*) AS n FROM `{REQ}` GROUP BY status")
        stats = {r["status"]: int(r["n"]) for r in counts}
        today = datetime.now().date().isoformat()
        day = mysql_fetchone(
            f"""SELECT COUNT(*) AS total, COALESCE(SUM(total_packages),0) AS bultos,
                       COALESCE(SUM(total_weight_kg),0) AS peso, COALESCE(SUM(total_volumetric_weight),0) AS pvol,
                       COALESCE(SUM(total_volume_m3),0) AS m3
                FROM `{REQ}` WHERE requested_date=%s OR confirmed_date=%s""",
            (today, today),
        ) or {}
        templates = mysql_fetchall(f"SELECT id, code, title, body, channel, active FROM `{TPL}` WHERE active=1 ORDER BY title")
        return render_template(
            "retiros/internal_dashboard.html",
            rows=rows, filtros=filtros, statuses=PICKUP_STATUS, stats=stats,
            day=day, settings=settings(), templates=templates,
            status_badge=status_badge,
            # 2026-05-26 (Daniel) — Pipeline consolidada: 12 estados → 6 columnas
            # visuales. El template usa pipeline_groups en el kanban del monitor.
            pipeline_groups=PIPELINE_GROUPS,
        )

    @app.route("/retiros/api/responsables", methods=["GET"])
    @require_permission("retiros")
    def pickup_responsables():
        """Usuarios internos que pueden ser RESPONSABLES de entregar un retiro
        (rol con acceso a retiros). Alimenta el selector del modal 'Nuevo retiro'.
        (Daniel 2026-06-19)"""
        try:
            _auth_t = ctx.get("AUTH_TABLE") or "app_users"
            rows = mysql_fetchall(
                f"SELECT DISTINCT u.id, u.nombre, u.username "
                f"FROM `{_auth_t}` u "
                f"LEFT JOIN rol_permisos rp ON rp.rol_slug=u.role "
                f"   AND rp.modulo='retiros' AND rp.accion='ver' AND rp.permitido=1 "
                f"WHERE u.active=1 AND ("
                f"     rp.rol_slug IS NOT NULL "
                f"  OR u.role LIKE 'superadmin%%' OR u.role LIKE 'admin%%' "
                f"  OR u.role LIKE 'supervisor%%') "
                f"ORDER BY u.nombre") or []
            users = [{"id": r["id"],
                      "nombre": (r.get("nombre") or r.get("username") or "Usuario")}
                     for r in rows]
            return jsonify({"ok": True, "responsables": users})
        except Exception as e:
            print(f"[pickup-responsables] {e}", flush=True)
            return jsonify({"ok": True, "responsables": []})

    # ══════════════════════════════════════════════════════════════════
    #  NUEVO RETIRO INTERNO / BACKOFFICE (Daniel 2026-05-29)
    #  Botón [+ Nuevo retiro interno] en el monitor → modal → POST aquí.
    #  Es el MISMO flujo que el público pero request_source='backoffice'.
    #  Modo CONFIRMACIÓN DIRECTA: el operador marca que el cliente ya
    #  aceptó por un canal (tel/correo/WhatsApp/presencial) → queda
    #  agenda_confirmada al instante. Reusa validate_pickup_datetime +
    #  _validar_disponibilidad_slot (no duplica lógica fecha/capacidad).
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/nuevo", methods=["POST"])
    @require_permission("retiros")
    def pickup_create_internal():
        f = request.form
        def _err(msg, code=400):
            return jsonify({"ok": False, "error": msg}), code

        customer_name      = (f.get("customer_name") or "").strip()[:200]
        document_type      = (f.get("document_type") or "").strip()[:40]
        document_number    = (f.get("document_number") or "").strip()[:60]
        pickup_person_name = (f.get("pickup_person_name") or "").strip()[:200]
        date = (f.get("date") or "").strip()
        tf   = (f.get("time_from") or "").strip()
        tt   = (f.get("time_to") or "").strip()
        canal = (f.get("canal") or "").strip().lower()[:30]

        # RESPONSABLE de la entrega (Daniel 2026-06-19): OBLIGATORIO. Quién se
        # encarga de entregar el pedido. Se resuelve el nombre desde app_users
        # (autoritativo, snapshot para que sobreviva si el usuario cambia).
        try:
            responsable_user_id = int(f.get("responsable_user_id") or 0) or None
        except (TypeError, ValueError):
            responsable_user_id = None
        responsable_nombre = ""
        if responsable_user_id:
            try:
                _auth_t = ctx.get("AUTH_TABLE") or "app_users"
                _ru = mysql_fetchone(
                    f"SELECT nombre, username FROM `{_auth_t}` WHERE id=%s AND active=1",
                    (responsable_user_id,))
                responsable_nombre = ((_ru or {}).get("nombre")
                                      or (_ru or {}).get("username") or "")[:190]
            except Exception:
                responsable_nombre = ""

        # ── FASE 8: validaciones obligatorias para confirmación directa ──
        if len(customer_name) < 2:
            return _err("Falta el nombre del cliente.")
        # Documento OPCIONAL (Daniel 2026-06-19): el retiro se crea SIN factura y
        # la factura se asocia después por cliente/rubro. Si no viene, va vacío.
        if not document_type:
            document_type = "sin_documento"
        if not responsable_user_id or not responsable_nombre:
            return _err("Falta el RESPONSABLE del retiro (quién se encarga de entregar el pedido).")
        if len(pickup_person_name) < 2:
            return _err("Falta la persona que retira.")
        if not (date and tf and tt):
            return _err("Falta la fecha y hora del retiro.")
        if canal not in ("telefono", "correo", "whatsapp", "presencial"):
            return _err("Para confirmar directo, marca el canal por el que el cliente aceptó.")

        cfg = settings()
        # Validación temporal central (modo interno: NO exige min_notice y
        # permite cruzar colación, pero NUNCA fecha/hora pasada).
        ok_dt, msg_dt = validate_pickup_datetime(date, tf, tt, cfg=cfg, mode="internal")
        if not ok_dt:
            return _err(msg_dt)

        try:    wkg = float(f.get("total_weight_kg") or 0)
        except (TypeError, ValueError): wkg = 0.0
        try:    m3 = float(f.get("total_volume_m3") or 0)
        except (TypeError, ValueError): m3 = 0.0
        try:    pv = float(f.get("total_volumetric_weight") or 0)
        except (TypeError, ValueError): pv = 0.0
        try:    bultos = int(f.get("total_packages") or 1)
        except (TypeError, ValueError): bultos = 1

        # Capacidad real del slot (misma fuente que el calendario)
        ok_slot, motivo = _validar_disponibilidad_slot(
            date, tf, tt, extra_kg=wkg, extra_m3=m3, bypass_lunch=True
        )
        if not ok_slot:
            return _err(f"Ese horario no está disponible: {motivo}", code=409)

        u = getattr(g, "user", None) or {}
        uid   = u.get("id")
        uname = u.get("nombre") or u.get("username") or "interno"
        now_cl = _now_chile().strftime("%Y-%m-%d %H:%M:%S")
        token  = secrets.token_urlsafe(42)

        conn = get_db()
        try:
            with conn.cursor() as cur:
                code = _generate_pickup_code(cur)
                cur.execute(
                    f"""INSERT INTO `{REQ}`
                        (code, request_source, created_by_user_id, created_by_user_name,
                         internal_created_at, customer_confirm_required, customer_already_agreed,
                         document_type, document_number, customer_name, customer_rut,
                         contact_name, contact_email, contact_phone,
                         pickup_person_name, pickup_person_rut, pickup_person_phone, pickup_person_relation,
                         requested_date, requested_time_from, requested_time_to,
                         proposed_date, proposed_time_from, proposed_time_to,
                         confirmed_date, confirmed_time_from, confirmed_time_to,
                         status, total_packages, total_weight_kg, total_volumetric_weight, total_volume_m3,
                         observations, public_token, signature_status, created_ip, created_user_agent,
                         responsable_user_id, responsable_nombre)
                        VALUES (%s,'backoffice',%s,%s,%s,0,1,
                                %s,%s,%s,%s,
                                %s,%s,%s,
                                %s,%s,%s,%s,
                                %s,%s,%s,
                                %s,%s,%s,
                                %s,%s,%s,
                                'agenda_confirmada',%s,%s,%s,%s,
                                %s,%s,'pendiente',%s,%s,
                                %s,%s)""",
                    (code, uid, uname, now_cl,
                     document_type, document_number, customer_name, (f.get("customer_rut") or "").strip()[:20],
                     (f.get("contact_name") or pickup_person_name)[:180],
                     (f.get("contact_email") or "").strip()[:190], (f.get("contact_phone") or "").strip()[:40],
                     pickup_person_name, (f.get("pickup_person_rut") or "").strip()[:20],
                     (f.get("pickup_person_phone") or "").strip()[:40], (f.get("pickup_person_relation") or "otro")[:30],
                     date, tf, tt,  date, tf, tt,  date, tf, tt,
                     bultos, wkg, pv, m3,
                     (f.get("observations") or "").strip()[:2000], token,
                     request.remote_addr, (request.user_agent.string or "")[:300],
                     responsable_user_id, responsable_nombre),
                )
                rid = cur.lastrowid
            conn.commit()
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            return _err(f"No se pudo crear el retiro: {str(e)[:200]}", code=500)
        finally:
            try: conn.close()
            except Exception: pass

        # FASE 9: trazabilidad — log de creación interna + confirmación directa
        log_event(rid, "retiro_interno_creado", None, "agenda_confirmada",
                  f"Retiro backoffice creado por {uname} — confirmación directa (canal: {canal})",
                  "interno", uname)
        # Notificar al cliente si dejó email (no bloqueante)
        try:
            if (f.get("contact_email") or "").strip():
                _fresh = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
                if _fresh:
                    notify_async(_fresh, "confirmed")
        except Exception as _e:
            print(f"[pickup_create_internal notify] {_e}", flush=True)
        try: _DISPO_CACHE["payload"] = None
        except Exception: pass

        return jsonify({
            "ok": True, "id": rid, "code": code,
            "message": f"Retiro {code} creado y confirmado para {date} {tf}-{tt}.",
            "redirect_url": url_for("pickup_detail", rid=rid),
        })

    # ══════════════════════════════════════════════════════════════════
    #  RECORDATORIO 24h — disparo manual o cron-style
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/admin/enviar-recordatorios-24h", methods=["POST"])
    @require_permission("admin")
    def pickup_enviar_recordatorios_24h():
        """Envía recordatorio a clientes con retiro confirmado para MAÑANA.

        Selecciona pickup_requests con status='agenda_confirmada' y
        confirmed_date = CURDATE()+1 que no tengan reminder_24h_sent
        registrado, dispara notify(req, 'reminder_24h') y marca el envío.

        Devuelve JSON: {enviados, omitidos, errores: [{code, error}]}.
        Apto para disparo manual desde el dashboard o vía cron externo.
        """
        rows = mysql_fetchall(
            f"""SELECT * FROM `{REQ}`
                WHERE status = 'agenda_confirmada'
                  AND DATE(confirmed_date) = DATE_ADD(CURDATE(), INTERVAL 1 DAY)
                  AND reminder_24h_sent IS NULL"""
        ) or []

        enviados = 0
        omitidos = 0
        errores = []
        for r in rows:
            try:
                req_dict = dict(r)
                sent_mail, sent_wa = notify(req_dict, "reminder_24h")
                if sent_mail or sent_wa:
                    mysql_execute(
                        f"UPDATE `{REQ}` SET reminder_24h_sent=NOW() WHERE id=%s",
                        (req_dict["id"],),
                    )
                    log_event(
                        req_dict["id"], "recordatorio_24h_enviado",
                        req_dict.get("status"), req_dict.get("status"),
                        "Recordatorio 24h enviado al cliente.",
                        "sistema", "Recordatorios",
                    )
                    enviados += 1
                else:
                    omitidos += 1
                    errores.append({
                        "code": req_dict.get("code"),
                        "error": "No se pudo enviar por ningún canal (email/WA).",
                    })
            except Exception as exc:
                omitidos += 1
                errores.append({
                    "code": dict(r).get("code"),
                    "error": str(exc)[:200],
                })

        return jsonify({
            "ok": True,
            "enviados": enviados,
            "omitidos": omitidos,
            "errores": errores,
            "total_candidatos": len(rows),
        })

    @app.route("/retiros/<int:rid>")
    @require_permission("view")
    def pickup_detail(rid):
        """Vista interna del operador para un retiro.

        Envuelto en try/except con logging detallado (Daniel mayo 2026):
        si algo falla, en lugar de mostrar "Internal Server Error 500"
        ciego, mostramos un mensaje amigable y dejamos el traceback en
        los logs del worker para diagnóstico inmediato.
        """
        try:
            # ⚡ PERF (Daniel 2026-05-24): SELECT explícito SIN doc_erp_data
            # (MEDIUMTEXT que NO se usa en este template — solo se setea en
            # validar_doc). Antes SELECT * traía hasta 60KB inútiles por
            # cada carga de ficha. Ahorra ~150-300ms en cold start.
            # Daniel 2026-05-24: FIX BUG 1054 — el agente PERF anterior
            # inventó columnas que NO existen en la BD (pickup_person_extra_email,
            # pickup_person_is_third_party, third_party_declaration, ip_address).
            # Volvemos a SELECT * que es ROBUSTO contra cambios de schema y el
            # template puede leer cualquier columna sin sorpresas. La penalidad
            # de leer doc_erp_data MEDIUMTEXT se mitiga descartándolo abajo
            # antes de pasar al template.
            req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
            if not req:
                flash("Solicitud de retiro no encontrada.", "danger")
                return redirect(url_for("pickup_dashboard"))

            # observations puede venir como bytes en algunas conexiones
            try:
                obs = req.get("observations")
                if obs is not None and not isinstance(obs, str):
                    req["observations"] = obs.decode("utf-8", errors="replace") if isinstance(obs, (bytes, bytearray)) else str(obs)
            except Exception:
                pass

            # ⚠️ FIX CRÍTICO (Daniel 2026-05-24 mañana):
            # El threading paralelo INTRODUCÍA un bug: get_db() de Flask retorna
            # UNA SOLA conexión por request (cacheada en `g._db`). pymysql NO es
            # thread-safe — al ejecutar 6 queries en paralelo sobre la MISMA
            # conexión se generaban packet sequence errors / lost connection,
            # haciendo que TODA la ficha del retiro fallara con 500.
            #
            # Volvemos a queries SERIALES — son ~240ms en total (medido), lo
            # cual es perfectamente aceptable y, sobre todo, FUNCIONA. La perf
            # real percibida la ganamos del SELECT explícito de arriba (sin
            # doc_erp_data MEDIUMTEXT) y del cache de saldo/sugerencias.
            packages = mysql_fetchall(
                f"SELECT * FROM `{PKG}` WHERE request_id=%s ORDER BY package_number",
                (rid,)
            ) or []
            proposals = mysql_fetchall(
                f"SELECT * FROM `{PROP}` WHERE request_id=%s ORDER BY id DESC",
                (rid,)
            ) or []
            logs = mysql_fetchall(
                f"SELECT * FROM `{LOG}` WHERE request_id=%s ORDER BY id DESC LIMIT 80",
                (rid,)
            ) or []
            attachments = mysql_fetchall(
                f"SELECT * FROM `{ATT}` WHERE request_id=%s ORDER BY id DESC",
                (rid,)
            ) or []
            tpl_rows = mysql_fetchall(
                f"SELECT * FROM `{TPL}` WHERE active=1 ORDER BY title"
            ) or []

            # Multi-documento (Daniel 2026-05-22). Try/except anidado por
            # compat con entornos sin migración aplicada.
            try:
                docs_asociados = mysql_fetchall(
                    """SELECT id, document_type, document_number, cliente_rut, cliente_nombre,
                              observaciones_erp, peso_real_kg, peso_vol_kg, volumen_m3,
                              n_lineas, added_by, added_at, con_saldo, saldo_zz, saldo_checked_at
                         FROM pickup_request_docs
                        WHERE request_id=%s
                        ORDER BY id ASC""",
                    (rid,)
                ) or []
            except Exception:
                try:
                    docs_asociados = mysql_fetchall(
                        """SELECT id, document_type, document_number, cliente_rut, cliente_nombre,
                                  observaciones_erp, peso_real_kg, peso_vol_kg, volumen_m3,
                                  n_lineas, added_by, added_at
                             FROM pickup_request_docs
                            WHERE request_id=%s
                            ORDER BY id ASC""",
                        (rid,)
                    ) or []
                except Exception as _e_docs2:
                    print(f"[pickup_detail] docs_asociados skip: {_e_docs2}", flush=True)
                    docs_asociados = []

            return render_template(
                "retiros/internal_detail.html",
                req=req, packages=packages, proposals=proposals, logs=logs,
                attachments=attachments, templates=tpl_rows,
                docs_asociados=docs_asociados,
                statuses=PICKUP_STATUS, status_badge=status_badge,
                settings=settings(),
            )
        except Exception as _e_detail:
            # Logging COMPLETO con traceback para diagnóstico inmediato
            import traceback as _tb
            print(f"[pickup_detail] EXCEPTION rid={rid}: {_e_detail}", flush=True)
            print(_tb.format_exc(), flush=True)
            flash(
                f"No se pudo cargar el detalle del retiro #{rid}. "
                f"El equipo técnico fue notificado. Detalle: {str(_e_detail)[:120]}",
                "danger"
            )
            return redirect(url_for("pickup_dashboard"))

    @app.route("/retiros/<int:rid>/status", methods=["POST"])
    @require_permission("retiros")
    def pickup_update_status(rid):
        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            return redirect(url_for("pickup_dashboard"))
        new_status, notes = request.form.get("status"), request.form.get("notes", "")
        if new_status not in PICKUP_STATUS:
            flash("Estado no valido.", "danger")
            return redirect(url_for("pickup_detail", rid=rid))
        old_status = req.get("status") or ""

        # ── GUARDA DE SECUENCIA (Daniel 2026-06-19) ────────────────────────────
        # No se puede pasar a PREPARACIÓN ni marcar RETIRADO sin una cita
        # confirmada. Evita cierres accidentales por drag-drop del monitor o el
        # <select> genérico (antes se podía saltar directo a 'retirada').
        if new_status in ("en_preparacion", "retirada") \
                and not req.get("confirmed_date") \
                and old_status not in ("agenda_confirmada", "en_preparacion"):
            flash("Antes de preparar o marcar como retirado, el retiro debe tener "
                  "una cita confirmada.", "warning")
            return redirect(url_for("pickup_detail", rid=rid))

        mysql_execute(f"UPDATE `{REQ}` SET status=%s, closed_at=IF(%s IN ('cerrada','rechazada','retirada'),NOW(),closed_at) WHERE id=%s", (new_status, new_status, rid))
        log_event(rid, "estado_actualizado", old_status, new_status, notes, "interno")

        # ── 2º CAMINO ROTO (Daniel 2026-06-19): confirmar manualmente sin fecha ──
        # Si el operador pasa a 'agenda_confirmada' desde el <select> o el monitor
        # y el retiro aún NO tiene confirmed_date, la copiamos de proposed/requested.
        # Antes quedaba status=confirmada pero confirmed_date NULL → el correo salía
        # sin fecha y el EN VIVO mostraba "Confirmada" sin día/hora.
        if new_status == "agenda_confirmada" and not req.get("confirmed_date"):
            def _eff_upd(*keys):
                for k in keys:
                    v = req.get(k)
                    if v not in (None, ""):
                        return v
                return None
            try:
                _cd = _eff_upd("proposed_date", "requested_date")
                if _cd:
                    mysql_execute(
                        f"UPDATE `{REQ}` SET confirmed_date=%s, confirmed_time_from=%s, "
                        f"confirmed_time_to=%s WHERE id=%s",
                        (_cd,
                         _eff_upd("proposed_time_from", "requested_time_from"),
                         _eff_upd("proposed_time_to", "requested_time_to"), rid))
            except Exception as _e_cd:
                print(f"[pickup-updstatus] copia confirmed_date: {_e_cd}", flush=True)

        # ─── Notificar al cliente cuando ILUS cambia estado desde el calendario ─
        # Antes este endpoint NO mandaba email/WA, dejando al cliente sin saber
        # que su retiro fue confirmado/rechazado/preparado.
        # Daniel pidió que cada estado dispare email al cliente para que vea
        # el avance del retiro. La tabla refleja TODOS los estados que disparan
        # mensaje al cliente (no solo los terminales).
        kind_map = {
            "agenda_confirmada":      "confirmed",
            "rechazada":              "rejected",
            "en_preparacion":         "preparing",
            "retirada":               "done",
            "fallida":                "failed",
            "reagendada":             "rescheduled",
            "cerrada":                "closed",
            "informacion_incompleta": "info_incompleta",
            # Nuevos: cuando un operador cambia manualmente a estos también
            # mandamos email (antes se quedaban en silencio).
            "en_revision":            "created",   # reusa plantilla de "estamos revisando"
            "esperando_cliente":      "info_incompleta",
        }
        kind = kind_map.get(new_status)
        if kind and old_status != new_status:
            try:
                # Re-leer la solicitud actualizada (estado/fechas pueden haber cambiado)
                req_after = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,)) or req
                # Async: no bloquear la respuesta del operador
                notify_async(req_after, kind)
            except Exception as _e:
                # Si falla notificación, no rompemos el cambio de estado
                print(f"[pickups][notify] error notificando {kind} a retiro #{rid}: {_e}")

        # Invalidar cache de polling para que el cliente vea el cambio en su próximo poll
        try:
            tok = req.get("public_token") if isinstance(req, dict) else None
            if tok:
                _POLL_CACHE.pop(tok, None)
        except Exception: pass
        # Invalidar cache global del calendario (cambio de status puede liberar/ocupar slot)
        try: _DISPO_CACHE["payload"] = None
        except Exception: pass

        # ── AVISO INTERNO AL EQUIPO en transiciones clave (Daniel 2026-06-19) ──
        # Antes pickup_update_status NO avisaba al equipo: el cliente se enteraba
        # por correo pero el equipo no tenía campana/email del avance ni del cierre.
        try:
            if old_status != new_status:
                _cod = req.get("code") or "?"
                _cli = req.get("customer_name") or "Cliente"
                if new_status == "en_preparacion":
                    _notificar_equipo_retiros(
                        f"📦 Retiro {_cod} en preparación",
                        f"{_cli} — bodega alistando el pedido para el retiro.",
                        rid, _cod, prioridad="media", tipo="retiro_preparacion", send_email=False)
                elif new_status == "retirada":
                    _notificar_equipo_retiros(
                        f"🎉 Retiro {_cod} COMPLETADO",
                        f"{_cli} retiró sus productos. Proceso cerrado.",
                        rid, _cod, prioridad="media", tipo="retiro_cerrado", send_email=True)
                elif new_status == "agenda_confirmada":
                    _notificar_equipo_retiros(
                        f"✅ Retiro {_cod} confirmado",
                        f"{_cli} — cita confirmada, listo para preparar.",
                        rid, _cod, prioridad="media", tipo="retiro_confirmado", send_email=False)
        except Exception as _e_team:
            print(f"[pickup-updstatus] aviso equipo: {_e_team}", flush=True)

        flash("Estado actualizado.", "success")
        return redirect(url_for("pickup_detail", rid=rid))

    @app.route("/retiros/<int:rid>/validar-doc", methods=["POST"])
    @require_permission("retiros")
    def pickup_validar_doc(rid):
        """Paso 1 del proceso interno: validar documentación del cliente.
        Marca el estado de validación, calcula peso/vol/tiempo estimado y
        deja todo listo para el paso 2 (proponer fecha).
        """
        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            return redirect(url_for("pickup_dashboard"))

        action = (request.form.get("action") or "").strip()
        notes  = (request.form.get("notes") or "").strip()[:1000]

        if action == "marcar_ok":
            # Validación exitosa
            peso_real = request.form.get("peso_real_kg") or req.get("peso_real_kg") or req.get("total_weight_kg") or 0
            peso_vol  = request.form.get("peso_vol_kg")  or req.get("peso_vol_kg")  or req.get("total_volumetric_weight") or 0
            tiempo    = request.form.get("tiempo_estimado_min") or req.get("tiempo_estimado_min") or 0
            # FIX 2026-05-22 (Daniel): m³ no se persistía aunque el frontend ya lo
            # calculaba. Ahora el operador puede sobrescribirlo y se guarda OK.
            vol_m3    = request.form.get("total_volume_m3") or req.get("total_volume_m3") or 0
            try:
                # Clamp a rangos razonables (defensa anti-input absurdo):
                # peso 0-10000 kg, vol 0-1000 m³, tiempo 0-720 min (12h)
                peso_real = max(0.0, min(float(peso_real or 0), 10000.0))
                peso_vol  = max(0.0, min(float(peso_vol or 0),  10000.0))
                tiempo    = max(0,   min(int(tiempo or 0),      720))
                vol_m3    = max(0.0, min(float(vol_m3 or 0),    1000.0))
            except Exception:
                peso_real, peso_vol, tiempo, vol_m3 = 0.0, 0.0, 0, 0.0
            mysql_execute(
                f"""UPDATE `{REQ}`
                    SET doc_validation_status='ok',
                        doc_validated_at=NOW(),
                        doc_validated_by=%s,
                        doc_validation_notes=%s,
                        peso_real_kg=%s,
                        peso_vol_kg=%s,
                        total_volume_m3=%s,
                        tiempo_estimado_min=%s,
                        status=CASE WHEN status='solicitud_recibida' THEN 'en_revision' ELSE status END
                    WHERE id=%s""",
                (g.user["nombre"] if getattr(g,"user",None) else "interno",
                 notes, peso_real, peso_vol, vol_m3, tiempo, rid)
            )
            log_event(rid, "doc_validada", req["status"], "en_revision",
                      f"Documentación validada · peso={peso_real}kg vol={peso_vol}kg m3={vol_m3} tiempo={tiempo}min",
                      "interno")
            flash("Documentación validada. Ya puedes proponer fecha al cliente.", "success")

        elif action == "marcar_incompleto":
            mysql_execute(
                f"""UPDATE `{REQ}`
                    SET doc_validation_status='incompleto',
                        doc_validation_notes=%s,
                        status='informacion_incompleta'
                    WHERE id=%s""",
                (notes, rid)
            )
            log_event(rid, "doc_incompleta", req["status"], "informacion_incompleta",
                      f"Falta información: {notes[:200]}", "interno")
            flash("Marcado como información incompleta. Notifica al cliente para completar.", "warning")

        elif action == "guardar_erp":
            # Guardar snapshot ERP + auto-rellenar peso/vol/m³
            erp_json = request.form.get("erp_data_json") or "{}"
            peso_real = request.form.get("peso_real_kg") or 0
            peso_vol  = request.form.get("peso_vol_kg") or 0
            tiempo    = request.form.get("tiempo_estimado_min") or 0
            # FIX 2026-05-22 (Daniel): el m³ que calcula el JS desde el ERP ahora
            # se persiste para que la pantalla muestre el dato correcto al recargar.
            vol_m3    = request.form.get("total_volume_m3") or 0
            mysql_execute(
                f"""UPDATE `{REQ}`
                    SET doc_erp_data=%s,
                        peso_real_kg=%s,
                        peso_vol_kg=%s,
                        total_volume_m3=%s,
                        tiempo_estimado_min=%s
                    WHERE id=%s""",
                (erp_json[:60000], peso_real or 0, peso_vol or 0, vol_m3 or 0, tiempo or 0, rid)
            )
            log_event(rid, "erp_actualizado", req["status"], req["status"],
                      f"Datos ERP cargados · peso={peso_real}kg vol={peso_vol}kg m3={vol_m3}",
                      "interno")
            flash("Datos del ERP cargados al retiro.", "success")

        # EN VIVO: invalidar cache de polling para que el seguimiento del cliente
        # refleje el cambio de estado (solicitud_recibida→en_revision, etc.) en su
        # próximo poll sin esperar el TTL de 10s. (req es SELECT *, trae public_token)
        try:
            tok = req.get("public_token") if isinstance(req, dict) else None
            if tok:
                _POLL_CACHE.pop(tok, None)
        except Exception: pass

        return redirect(url_for("pickup_detail", rid=rid))


    # ══════════════════════════════════════════════════════════════════
    #  MULTI-DOCUMENTO (Daniel 2026-05-22)
    #  Un retiro puede asociar 1..N facturas/boletas del ERP. Cada doc
    #  contribuye con su peso/vol/m³ al total del retiro.
    # ══════════════════════════════════════════════════════════════════

    def _pickup_recalc_totales(rid):
        """Suma peso/vol/m³ de todos los docs asociados y actualiza pickup_requests.
        Si no quedan docs asociados, restablece los totales a 0 (Daniel pidió que
        sea coherente con la realidad — si quitas todo, el retiro no tiene carga).

        DANIEL 2026-05-23: para docs con has_seleccion_lineas=1, sumamos
        SOLO las líneas marcadas como incluidas en pickup_doc_lineas
        (selección granular). Para docs sin selección, usamos los totales
        completos del doc (comportamiento anterior).

        Devuelve dict con los totales actualizados.
        """
        # 1) Suma de docs SIN selección granular (totales completos)
        agg = mysql_fetchone(
            """SELECT COALESCE(SUM(peso_real_kg),0)   AS peso_real,
                      COALESCE(SUM(peso_vol_kg),0)    AS peso_vol,
                      COALESCE(SUM(volumen_m3),0)     AS m3,
                      COALESCE(SUM(n_lineas),0)       AS lineas_total,
                      COUNT(*)                        AS n_docs
                 FROM pickup_request_docs
                WHERE request_id=%s
                  AND (has_seleccion_lineas IS NULL OR has_seleccion_lineas=0)""",
            (rid,)
        ) or {"peso_real": 0, "peso_vol": 0, "m3": 0, "lineas_total": 0, "n_docs": 0}

        # 2) Suma de docs CON selección granular (solo líneas incluidas)
        try:
            sel_agg = mysql_fetchone(
                """SELECT COALESCE(SUM(l.peso_total_kg),0)     AS peso_real_sel,
                          COALESCE(SUM(l.peso_vol_total_kg),0) AS peso_vol_sel,
                          COALESCE(SUM(l.vol_total_m3),0)      AS m3_sel,
                          COUNT(*)                              AS lineas_sel,
                          COUNT(DISTINCT l.doc_id)              AS n_docs_sel
                     FROM pickup_doc_lineas l
                    WHERE l.request_id=%s AND l.incluida=1""",
                (rid,)
            ) or {}
            agg["peso_real"]    = float(agg.get("peso_real") or 0) + float(sel_agg.get("peso_real_sel") or 0)
            agg["peso_vol"]     = float(agg.get("peso_vol") or 0) + float(sel_agg.get("peso_vol_sel") or 0)
            agg["m3"]           = float(agg.get("m3") or 0) + float(sel_agg.get("m3_sel") or 0)
            agg["lineas_total"] = int(agg.get("lineas_total") or 0) + int(sel_agg.get("lineas_sel") or 0)
            agg["n_docs"]       = int(agg.get("n_docs") or 0) + int(sel_agg.get("n_docs_sel") or 0)
        except Exception as _e_sel:
            # Si la tabla pickup_doc_lineas aún no existe (migration pending),
            # no rompemos — los totales completos siguen funcionando.
            print(f"[recalc-totales] sin selecciones granulares: {_e_sel}", flush=True)
        peso_real = float(agg.get("peso_real") or 0)
        peso_vol  = float(agg.get("peso_vol") or 0)
        vol_m3    = float(agg.get("m3") or 0)
        lineas_total = int(agg.get("lineas_total") or 0)
        n_docs    = int(agg.get("n_docs") or 0)
        # Si hay docs: estimar tiempo basado en líneas. Si no: 0 (sin sobreescribir
        # tiempo manual del operador — solo se autocalcula con docs presentes).
        tiempo = None
        if n_docs > 0:
            tiempo = max(15, 5 + lineas_total * 2)
            mysql_execute(
                f"""UPDATE `{REQ}`
                    SET peso_real_kg=%s, peso_vol_kg=%s, total_volume_m3=%s,
                        tiempo_estimado_min=%s
                    WHERE id=%s""",
                (peso_real, peso_vol, vol_m3, tiempo, rid)
            )
        else:
            # Sin docs: dejamos los totales en 0 pero NO tocamos tiempo (el operador
            # puede haber estimado manualmente — respetar su decisión).
            mysql_execute(
                f"""UPDATE `{REQ}`
                    SET peso_real_kg=0, peso_vol_kg=0, total_volume_m3=0
                    WHERE id=%s""",
                (rid,)
            )
        return {
            "peso_real_kg": peso_real,
            "peso_vol_kg":  peso_vol,
            "volumen_m3":   vol_m3,
            "n_docs":       n_docs,
            "lineas_total": lineas_total,
            "tiempo_estimado_min": tiempo or 0,
        }

    def _apply_lineas_seleccion_inline(rid, doc_id, lineas_in):
        """Aplica selección granular a un doc ya asociado. Usado en flujo
        DUPLICATE+lineas del endpoint /docs/agregar. Reusa el snapshot
        guardado (no llama al ERP). Idempotente.

        Daniel 2026-05-24 — extracted desde pickup_doc_lineas_guardar para
        permitir reuso en el flujo combinado agregar+lineas.
        """
        import json as _json_inl
        doc = mysql_fetchone(
            "SELECT erp_snapshot FROM pickup_request_docs WHERE id=%s AND request_id=%s",
            (doc_id, rid)
        )
        if not doc:
            raise RuntimeError("Doc no asociado")
        try:
            snap = _json_inl.loads(doc.get("erp_snapshot") or "{}")
        except Exception:
            snap = {}
        snap_lineas = {(ln.get("sku") or "").upper(): ln for ln in (snap.get("lineas") or [])}

        upsert_rows = []
        for li in lineas_in:
            sku = (li.get("sku") or "").strip()[:80]
            if not sku:
                continue
            ln_snap = snap_lineas.get(sku.upper()) or {}
            cantidad_doc = float(ln_snap.get("cantidad") or 0)
            peso_unit    = float(ln_snap.get("peso_kg_u") or 0)
            peso_vol_unit= float(ln_snap.get("peso_vol_u") or 0)
            vol_unit_cm3 = float(ln_snap.get("vol_u") or 0)
            vol_unit_m3  = vol_unit_cm3 / 1_000_000.0 if vol_unit_cm3 else 0
            try:
                qty_sel = float(li.get("cantidad_seleccionada") or 0)
            except Exception:
                qty_sel = 0
            qty_sel = max(0.0, min(qty_sel, cantidad_doc or qty_sel))
            incluida = 1 if (li.get("incluida", True) and qty_sel > 0) else 0
            descripcion = (ln_snap.get("descripcion_erp") or ln_snap.get("nombre_app") or "").strip()[:300]
            nota = (li.get("nota") or "").strip()[:300] or None
            # 🆕 Daniel 2026-05-24: ver pickup_doc_agregar — la flag persiste
            # el aviso "ya rebajado en ERP" en la tabla externa de productos.
            marcada_sin_saldo = 1 if li.get("marcada_sin_saldo") else 0
            upsert_rows.append((
                rid, doc_id, sku, descripcion,
                cantidad_doc, qty_sel,
                peso_unit, peso_vol_unit, vol_unit_m3,
                peso_unit * qty_sel,
                peso_vol_unit * qty_sel,
                vol_unit_m3 * qty_sel,
                incluida, nota,
                marcada_sin_saldo,
            ))
        if not upsert_rows:
            return 0
        conn = get_db()
        with conn.cursor() as cur:
            try:
                cur.executemany(
                    """INSERT INTO pickup_doc_lineas
                         (request_id, doc_id, sku, descripcion, cantidad_doc,
                          cantidad_seleccionada, peso_unit_kg, peso_vol_unit_kg,
                          vol_unit_m3, peso_total_kg, peso_vol_total_kg, vol_total_m3,
                          incluida, nota_linea, marcada_sin_saldo)
                       VALUES (%s,%s,%s,%s, %s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s)
                       ON DUPLICATE KEY UPDATE
                         descripcion=VALUES(descripcion),
                         cantidad_doc=VALUES(cantidad_doc),
                         cantidad_seleccionada=VALUES(cantidad_seleccionada),
                         peso_unit_kg=VALUES(peso_unit_kg),
                         peso_vol_unit_kg=VALUES(peso_vol_unit_kg),
                         vol_unit_m3=VALUES(vol_unit_m3),
                         peso_total_kg=VALUES(peso_total_kg),
                         peso_vol_total_kg=VALUES(peso_vol_total_kg),
                         vol_total_m3=VALUES(vol_total_m3),
                         incluida=VALUES(incluida),
                         nota_linea=VALUES(nota_linea),
                         marcada_sin_saldo=VALUES(marcada_sin_saldo),
                         updated_at=NOW()""",
                    upsert_rows
                )
            except Exception as _e_msm2:
                # Compat: si la columna marcada_sin_saldo aún no migró
                print(f"[_apply_lineas_seleccion_inline] sin col marcada_sin_saldo, retry: {_e_msm2}", flush=True)
                legacy_rows = [r[:-1] for r in upsert_rows]
                cur.executemany(
                    """INSERT INTO pickup_doc_lineas
                         (request_id, doc_id, sku, descripcion, cantidad_doc,
                          cantidad_seleccionada, peso_unit_kg, peso_vol_unit_kg,
                          vol_unit_m3, peso_total_kg, peso_vol_total_kg, vol_total_m3,
                          incluida, nota_linea)
                       VALUES (%s,%s,%s,%s, %s,%s, %s,%s,%s, %s,%s,%s, %s,%s)
                       ON DUPLICATE KEY UPDATE
                         descripcion=VALUES(descripcion),
                         cantidad_doc=VALUES(cantidad_doc),
                         cantidad_seleccionada=VALUES(cantidad_seleccionada),
                         peso_unit_kg=VALUES(peso_unit_kg),
                         peso_vol_unit_kg=VALUES(peso_vol_unit_kg),
                         vol_unit_m3=VALUES(vol_unit_m3),
                         peso_total_kg=VALUES(peso_total_kg),
                         peso_vol_total_kg=VALUES(peso_vol_total_kg),
                         vol_total_m3=VALUES(vol_total_m3),
                         incluida=VALUES(incluida),
                         nota_linea=VALUES(nota_linea),
                         updated_at=NOW()""",
                    legacy_rows
                )
            cur.execute(
                "UPDATE pickup_request_docs SET has_seleccion_lineas=1 WHERE id=%s",
                (doc_id,)
            )
        conn.commit()
        return len(upsert_rows)

    def _doc_already_used_by_other_request(tipo, numero, exclude_rid):
        """Devuelve {request_id, code} si OTRO retiro ya usa este doc, o None."""
        row = mysql_fetchone(
            f"""SELECT prd.request_id, pr.code
                  FROM pickup_request_docs prd
                  JOIN `{REQ}` pr ON pr.id = prd.request_id
                 WHERE prd.document_type=%s
                   AND prd.document_number=%s
                   AND prd.request_id<>%s
                 LIMIT 1""",
            (tipo, numero, exclude_rid)
        )
        return row or None

    @app.route("/retiros/<int:rid>/docs", methods=["GET"])
    @require_permission("view")
    def pickup_docs_list(rid):
        """Lista todos los documentos asociados al retiro.
        Daniel 2026-05-23 — incluye con_saldo/saldo_zz para el wizard interno:
        permite que el frontend sepa si el doc tiene saldo disponible para
        habilitar el paso "Proponer fecha". Si la columna no existe (migración
        no aplicada aún), reintentamos sin esos campos para no romper la UI.

        ⚡ PERF (Daniel 2026-05-24): cache 15s. Este endpoint se llama
        muchísimo durante un wizard activo (cada vez que se agrega/quita
        doc, refrescarDocsAsociados se llama N veces). 15s es suficiente
        para no mostrar data stale tras un cambio del operador (las
        mutaciones invalidan explícitamente el cache).
        """
        import time as _time_docs
        _hit = _DOCS_CACHE.get(rid)
        if _hit and (_time_docs.time() - _hit[1]) < _DOCS_TTL:
            return jsonify(_hit[0])
        try:
            rows = mysql_fetchall(
                """SELECT id, document_type, document_number, cliente_rut, cliente_nombre,
                          observaciones_erp, peso_real_kg, peso_vol_kg, volumen_m3,
                          n_lineas, added_by, added_at, con_saldo, saldo_zz, saldo_checked_at,
                          has_seleccion_lineas
                     FROM pickup_request_docs
                    WHERE request_id=%s
                    ORDER BY id ASC""",
                (rid,)
            ) or []
        except Exception:
            # Fallback sin las columnas nuevas (compat retroactiva)
            try:
                rows = mysql_fetchall(
                    """SELECT id, document_type, document_number, cliente_rut, cliente_nombre,
                              observaciones_erp, peso_real_kg, peso_vol_kg, volumen_m3,
                              n_lineas, added_by, added_at, con_saldo, saldo_zz, saldo_checked_at
                         FROM pickup_request_docs
                        WHERE request_id=%s
                        ORDER BY id ASC""",
                    (rid,)
                ) or []
            except Exception:
                rows = mysql_fetchall(
                    """SELECT id, document_type, document_number, cliente_rut, cliente_nombre,
                              observaciones_erp, peso_real_kg, peso_vol_kg, volumen_m3,
                              n_lineas, added_by, added_at
                         FROM pickup_request_docs
                        WHERE request_id=%s
                        ORDER BY id ASC""",
                    (rid,)
                ) or []
        out = []
        # 🔧 FIX Daniel 2026-05-24: para docs con selección granular, contar
        # cuántas líneas el operador realmente marcó (incluida=1). Antes el
        # card mostraba "10 líneas" aunque solo se hubieran asociado 2 → se
        # entendía como "está asociado todo". Ahora se ve "2 / 10".
        sel_counts = {}
        try:
            doc_ids_sel = [r.get("id") for r in rows if r.get("has_seleccion_lineas")]
            if doc_ids_sel:
                ph = ",".join(["%s"] * len(doc_ids_sel))
                rows_sc = mysql_fetchall(
                    f"SELECT doc_id, COUNT(*) AS n FROM pickup_doc_lineas "
                    f"WHERE doc_id IN ({ph}) AND incluida=1 GROUP BY doc_id",
                    tuple(doc_ids_sel)
                ) or []
                sel_counts = {r2.get("doc_id"): int(r2.get("n") or 0) for r2 in rows_sc}
        except Exception as _ce:
            print(f"[pickup-docs] sel_counts skip: {_ce}", flush=True)
        for r in rows:
            d = dict(r)
            for k in ("peso_real_kg", "peso_vol_kg", "volumen_m3", "saldo_zz"):
                if d.get(k) is not None:
                    try: d[k] = float(d[k])
                    except Exception: pass
            if d.get("added_at"):
                d["added_at"] = str(d["added_at"])[:19]
            if d.get("saldo_checked_at"):
                d["saldo_checked_at"] = str(d["saldo_checked_at"])[:19]
            d["has_seleccion_lineas"] = bool(d.get("has_seleccion_lineas"))
            # Líneas realmente incluidas (solo si hay selección granular)
            d["n_lineas_seleccionadas"] = sel_counts.get(d.get("id"))
            out.append(d)
        totales = _pickup_recalc_totales(rid)
        # Conteo de docs con saldo (para habilitar el paso 4 del wizard)
        docs_con_saldo = sum(1 for d in out if d.get("con_saldo") == 1)
        docs_sin_saldo = sum(1 for d in out if d.get("con_saldo") == 0)
        docs_no_verif  = sum(1 for d in out if d.get("con_saldo") is None)
        # Estado del retiro para refrescar el wizard sin recargar página
        req_state = mysql_fetchone(
            f"SELECT doc_validation_status, status, proposed_date, confirmed_date "
            f"FROM `{REQ}` WHERE id=%s", (rid,)
        ) or {}
        _resp_docs = {
            "ok": True,
            "docs": out,
            "totales": totales,
            "saldo_summary": {
                "con_saldo": docs_con_saldo,
                "sin_saldo": docs_sin_saldo,
                "no_verificado": docs_no_verif,
                "puede_agendar": docs_con_saldo > 0,
            },
            "request_state": {
                "doc_ok":      (req_state.get("doc_validation_status") or "") == "ok",
                "status":      req_state.get("status") or "",
                "step4_done":  bool(req_state.get("proposed_date")),
                "step5_done":  bool(req_state.get("confirmed_date")),
            },
        }
        # ⚡ PERF: cachear respuesta (las mutaciones invalidan)
        try:
            _DOCS_CACHE[rid] = (_resp_docs, _time_docs.time())
        except Exception:
            pass
        return jsonify(_resp_docs)

    def _pickup_fetch_doc_minimal_via_sql(tipo, numero):
        """Fallback: busca un documento en MAEEDO directamente, sin enrichment.
        Mismo motor que /retiros/api/buscar-erp (búsqueda avanzada).

        Devuelve un dict con la forma mínima que necesita
        _pickup_doc_agregar_impl, o None si no encontró el doc.

        Daniel 2026-05-23: "no estamos usando el mismo motor" — esto
        garantiza que SIEMPRE se pueda asociar un doc que existe en MAEEDO,
        aunque _cubicador_fetch crashee o REST/SQL Server del motor unificado
        falle. El enrichment (peso/vol/líneas) puede hacerse después.
        """
        try:
            from app import _random_sql_query, _random_sql_pool
        except ImportError:
            return None
        if _random_sql_pool() is None:
            return None

        tipo_up = (tipo or "").strip().upper()
        nudo_clean = str(numero or "").strip().lstrip("0") or "0"
        # NUDO en MAEEDO viene padded a 10 chars con prefijos posibles (VD/WEB)
        nudo_padded = nudo_clean.zfill(10)
        nudo_vd     = f"VD{nudo_clean.zfill(8)}"
        nudo_web    = f"WEB{nudo_clean.zfill(7)}"

        # TIDOs aceptados — alias VD/WEB → NVV en MAEEDO
        tido_query_list = [tipo_up]
        if tipo_up in ("VD", "WEB"):
            tido_query_list = ["NVV"]
        elif tipo_up == "NVV":
            tido_query_list = ["NVV"]

        tidos_in = "','".join(tido_query_list)

        try:
            row = None
            for nudo_try in (nudo_padded, nudo_vd, nudo_web):
                rows = _random_sql_query(f"""
                    SELECT TOP 1
                        e.IDMAEEDO,
                        LTRIM(RTRIM(e.TIDO))                AS TIDO,
                        LTRIM(RTRIM(e.NUDO))                AS NUDO,
                        LTRIM(RTRIM(COALESCE(e.ENDO,'')))   AS ENDO,
                        LTRIM(RTRIM(COALESCE(e.SUENDO,''))) AS SUENDO,
                        e.FEEMDO,
                        COALESCE(e.VANEDO, 0)               AS VANEDO,
                        COALESCE(e.VABRDO, 0)               AS VABRDO
                    FROM MAEEDO e
                    WHERE e.NUDO = %s
                      AND LTRIM(RTRIM(e.TIDO)) IN ('{tidos_in}')
                      AND (e.ESDO IS NULL OR LTRIM(RTRIM(e.ESDO)) <> 'NULO')
                """, (nudo_try,), max_rows=1)
                if rows:
                    row = rows[0]
                    break
            if not row:
                return None

            # Enrichment opcional: razón social del cliente
            endo = (row.get("ENDO") or "").strip()
            rut_base = endo.split("-")[0] if "-" in endo else endo
            cliente_nombre = ""
            try:
                cli_rows = _random_sql_query("""
                    SELECT TOP 1 LTRIM(RTRIM(COALESCE(NOKOENAMP, NOKOEN, ''))) AS razon
                      FROM MAEEN
                     WHERE RTEN = %s
                """, (endo,), max_rows=1)
                if cli_rows:
                    cliente_nombre = (cli_rows[0].get("razon") or "").strip().title()
            except Exception:
                pass

            return {
                "cliente_rut":      rut_base,
                "rut":              rut_base,
                "cliente_nombre":   cliente_nombre,
                "razon_social":     cliente_nombre,
                "observaciones":    "",
                "obs":              "",
                "_fallback_source": "maeedo_direct_sql",
                "_minimal":         True,
            }
        except Exception as e:
            print(f"[_pickup_fetch_doc_minimal_via_sql] error: {e}", flush=True)
            return None

    # ══════════════════════════════════════════════════════════════════
    #  ENDPOINT ADMIN — Forzar migraciones de retiros (Daniel 2026-05-23)
    #  Si producción tiene ILUS_SKIP_MIGRATIONS=1 y faltan tablas, este
    #  endpoint las crea sin necesidad de redeploy.
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/admin/migrate-now", methods=["POST", "GET"])
    @require_permission("retiros")
    def pickup_admin_migrate_now():
        """Ejecuta TODAS las migraciones idempotentes del módulo retiros.
        Útil cuando Railway tiene ILUS_SKIP_MIGRATIONS=1 y faltan tablas
        nuevas (multidocumento, líneas, extra_emails, etc.).
        Solo superadmin/admin/retiros con permiso.
        """
        if not (g.user and (g.permissions.get("superadmin") or g.permissions.get("admin"))):
            return jsonify({"ok": False, "error": "Solo admin/superadmin"}), 403

        resultados = {
            "multidoc_tables": "ok",
            "errores": [],
        }
        try:
            # force=True para que el endpoint admin SIEMPRE corra las migraciones
            # aunque el flag de cache las marque como listas (caso: nueva columna
            # añadida en el código y queremos aplicar sin redeploy).
            ensure_multidoc_tables_runtime(force=True)
        except Exception as e:
            resultados["multidoc_tables"] = "error"
            resultados["errores"].append(f"multidoc: {str(e)[:200]}")

        # Verificar que las tablas existan ahora
        verif = {}
        for tabla in ("pickup_requests", "pickup_request_docs", "pickup_doc_lineas",
                      "pickup_blocks", "pickup_packages", "pickup_proposals"):
            try:
                row = mysql_fetchone(f"SELECT COUNT(*) AS n FROM `{tabla}` LIMIT 1")
                verif[tabla] = {"existe": True, "filas": int(row["n"] if row else 0)}
            except Exception as e:
                verif[tabla] = {"existe": False, "error": str(e)[:100]}

        resultados["verificacion_tablas"] = verif
        resultados["timestamp"] = datetime.now().isoformat()

        return jsonify({"ok": True, "resultados": resultados})

    @app.route("/retiros/<int:rid>/docs/agregar", methods=["POST"])
    @require_permission("retiros")
    def pickup_doc_agregar(rid):
        """Agrega un documento del ERP al retiro. Valida que exista, calcula
        peso/vol/m³ desde las líneas, guarda snapshot y recalcula totales.

        Body JSON: {document_type, document_number}
        Response:  {ok:true, doc:{...}, totales:{...}}  o 4xx/5xx con error.

        🔒 GARANTIZA respuesta JSON SIEMPRE — fix Daniel 2026-05-23:
        El frontend reportaba "Unexpected token '<', '<!doctype..."
        que indica que en algún branch este endpoint devolvía HTML
        (página de error 500 default de Flask). Cualquier excepción
        no atrapada se convierte en JSON 500 con error legible.
        """
        # Wrapper anti-HTML: cualquier crash se convierte en JSON 500
        # Daniel 2026-05-23: el mensaje genérico no servía para diagnóstico.
        # Ahora el detalle real va EN el error (visible al operador), no
        # escondido en "detalle". Si es un superadmin, mostramos full
        # traceback en el campo trace para debugging rápido.
        try:
            return _pickup_doc_agregar_impl(rid)
        except Exception as _exc:
            import traceback as _tb
            tb_str = _tb.format_exc()
            err_type = type(_exc).__name__
            err_msg = str(_exc)[:300]
            print(f"[pickup_doc_agregar] CRASH rid={rid} type={err_type}: {_exc}\n{tb_str}",
                  flush=True)
            # Mensaje detallado AL OPERADOR (ya no genérico)
            user_msg = f"[{err_type}] {err_msg}"
            payload = {
                "ok": False,
                "error": user_msg,
                "error_codigo": "INTERNAL_CRASH",
                "detalle": err_msg,
                "tipo_error": err_type,
            }
            # Solo superadmin ve el traceback completo (privacidad)
            if g.user and g.permissions.get("superadmin"):
                payload["trace"] = tb_str[-1500:]
            return jsonify(payload), 500

    def _pickup_doc_agregar_impl(rid):
        """Implementación real de pickup_doc_agregar. Separada para que el
        wrapper pueda atrapar cualquier excepción y devolver JSON.

        ⚡ PERF (Daniel 2026-05-24): se redujo de 20-45s a <2s.
        Optimizaciones clave:
          1. ensure_multidoc_tables_runtime NO corre ALTER TABLE en hot path
             (flag _MULTIDOC_TABLES_READY) — antes costaba 1-4s por call
          2. snapshot reducido a 30KB (antes 50KB)
          3. log audit + auto-validate van a thread daemon
          4. eliminado SELECT extra al final (ya tenemos los datos en memoria)
          5. timing logs visibles en Railway para diagnóstico continuo
        """
        import time as _t_perf
        _t0 = _t_perf.time()
        _laps = []  # [(etapa, ms), ...]
        def _lap(etapa):
            _laps.append((etapa, int((_t_perf.time() - _t0) * 1000)))

        # AUTO-HEAL: idempotente — si las tablas YA están listas (flag),
        # devuelve en <1ms. Si no, corre las migraciones (caso del primer
        # arranque del worker).
        try:
            ensure_multidoc_tables_runtime()
        except Exception as _e_ensure:
            print(f"[pickup_doc_agregar] ensure tables falló: {_e_ensure}", flush=True)
        _lap("ensure_tables")

        req = mysql_fetchone(f"SELECT id, code FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            return jsonify({"ok": False, "error": "Retiro no existe"}), 404
        body = request.get_json(silent=True) or {}
        tipo = (body.get("document_type") or "").strip().upper()
        numero = (body.get("document_number") or "").strip()
        if not tipo or not numero:
            return jsonify({"ok": False, "error": "Falta document_type o document_number"}), 400

        # ⚡ PERF Daniel 2026-05-24 (8s → <1.5s):
        # Aceptamos OPCIONALMENTE las líneas seleccionadas en el mismo POST.
        # Antes el frontend hacía 2 POSTs secuenciales (agregar doc + guardar
        # líneas). Ahora 1 sólo: combinamos INSERT del doc + UPSERT de líneas
        # + has_seleccion_lineas=1 en la MISMA transacción. Beneficios:
        #   - 1 round-trip de red menos (~80-200ms)
        #   - 1 recalc_totales menos (~50-100ms, antes se llamaba 2 veces)
        #   - menos overhead Flask (validación + auth + middleware × 1)
        # El endpoint /docs/<id>/lineas se mantiene para EDITAR selección
        # posteriormente (no rompe nada existente).
        lineas_sel_in = body.get("lineas") or []
        if not isinstance(lineas_sel_in, list):
            lineas_sel_in = []
        _lap("validacion_input")

        # Idempotente: si ya está en este retiro, devolver 409 amistoso
        # PERO si vienen `lineas` en el body, aplicar la selección igual
        # (re-asociar con nueva selección, no error duro).
        dup = mysql_fetchone(
            "SELECT id FROM pickup_request_docs "
            "WHERE request_id=%s AND document_type=%s AND document_number=%s",
            (rid, tipo, numero)
        )
        if dup:
            # Si NO vienen líneas, comportamiento histórico (409). Si vienen,
            # actualizamos la selección y devolvemos OK con flag dup.
            if not lineas_sel_in:
                totales = _pickup_recalc_totales(rid)
                return jsonify({
                    "ok": False,
                    "error": f"El documento {tipo} {numero} ya está asociado a este retiro.",
                    "code": "DUPLICATE",
                    "doc_id": dup["id"],
                    "totales": totales,
                }), 409
            # Caso DUPLICATE con líneas → aplicar selección al doc existente
            # sin re-llamar al ERP (los snapshots ya están guardados).
            try:
                _apply_lineas_seleccion_inline(rid, dup["id"], lineas_sel_in)
                totales = _pickup_recalc_totales(rid)
                _DOCS_CACHE.pop(rid, None)
                return jsonify({
                    "ok": True,
                    "doc": {"id": dup["id"], "document_type": tipo, "document_number": numero},
                    "totales": totales,
                    "duplicate_updated": True,
                    "lineas_guardadas": len(lineas_sel_in),
                }), 200
            except Exception as _exc_dup:
                print(f"[pickup_doc_agregar] DUPLICATE+lineas falló: {_exc_dup}", flush=True)
                # cae al flujo normal — devolver 409 estándar
                totales = _pickup_recalc_totales(rid)
                return jsonify({
                    "ok": False,
                    "error": f"El documento {tipo} {numero} ya está asociado a este retiro.",
                    "code": "DUPLICATE",
                    "doc_id": dup["id"],
                    "totales": totales,
                }), 409

        # Verificar que no esté en OTRO retiro (warning, no bloqueo — Daniel decide)
        otro = _doc_already_used_by_other_request(tipo, numero, rid)
        _lap("check_dup_otro_retiro")

        # Buscar en ERP — primero motor unificado (_cubicador_fetch).
        # Si falla, fallback a query SQL directa contra MAEEDO (mismo motor
        # que la búsqueda avanzada — Daniel 2026-05-23: "no estamos usando
        # el mismo motor"). Así garantizamos que SIEMPRE se pueda asociar
        # un doc que existe en MAEEDO, aunque el cubicador no lo procese.
        hdr = None
        lineas = []
        fetch_err = None
        try:
            from app import _cubicador_fetch
        except ImportError:
            return jsonify({"ok": False, "error": "Motor ERP no disponible"}), 503

        # ⚡ PERF Daniel 2026-05-24: cache hit del modal RUT. Cuando el
        # operador EXPANDE un doc en el modal de búsqueda, ya se ejecuta
        # /api/erp/documento (cacheado 5min en _ERP_DOC_CACHE). Al hacer
        # click "Asociar" volvemos a pegar al ERP otra vez (~800-1500ms)
        # → recuperamos el cache para evitar la segunda llamada.
        # Beneficio típico: ~1200ms → ~5ms en docs ya expandidos.
        try:
            import app as _app_mod
            _ec = getattr(_app_mod, "_ERP_DOC_CACHE", None) or {}
            _ck = f"{tipo.upper()}|{numero}"
            _hit = _ec.get(_ck)
            if _hit and (time.time() - _hit[1]) < 300:
                _cached = _hit[0] or {}
                _h = _cached.get("hdr") or {}
                _l = _cached.get("lineas") or []
                if _h and _l:
                    # Reusar el snapshot tal cual — ya pasó por _cubicador_fetch
                    hdr = _h
                    lineas = _l
                    print(f"[pickup_doc_agregar] CACHE HIT {tipo}/{numero} (skip _cubicador_fetch)", flush=True)
        except Exception as _ce:
            print(f"[pickup_doc_agregar] cache lookup skip: {_ce}", flush=True)

        if not hdr:
            try:
                hdr, lineas = _cubicador_fetch(tipo, numero)
            except Exception as exc:
                fetch_err = str(exc)[:200]
                print(f"[pickup_doc_agregar] _cubicador_fetch falló para {tipo}/{numero}: {exc}",
                      flush=True)
        _lap("cubicador_fetch")

        # ⚡ FALLBACK robusto: si _cubicador_fetch no encontró nada O crasheó,
        # intentamos la MISMA query SQL directa que usa la búsqueda avanzada
        # (consulta a MAEEDO sin enrichment). Si el doc EXISTE allí, lo
        # asociamos con info mínima — luego se enriquece en background.
        if not hdr:
            try:
                hdr_min = _pickup_fetch_doc_minimal_via_sql(tipo, numero)
                if hdr_min:
                    print(f"[pickup_doc_agregar] FALLBACK SQL recuperó {tipo}/{numero} "
                          f"sin enrichment — guardando info mínima", flush=True)
                    hdr = hdr_min
                    lineas = []  # sin líneas porque no hicimos enrichment
            except Exception as sql_exc:
                print(f"[pickup_doc_agregar] FALLBACK SQL también falló: {sql_exc}",
                      flush=True)

        if not hdr:
            # Ni cubicador ni fallback SQL lo encontraron
            err_extra = f" (motor: {fetch_err})" if fetch_err else ""
            return jsonify({
                "ok": False,
                "error": f"Documento {tipo} {numero} no encontrado en ERP{err_extra}",
                "error_codigo": "DOC_NOT_FOUND",
            }), 404

        # Calcular peso/vol/m³ desde las líneas (mismo criterio que el JS frontend)
        total_kg = 0.0
        total_vol_kg = 0.0
        total_m3 = 0.0
        n_lineas_ok = 0
        for ln in (lineas or []):
            if ln.get("es_zz"):
                continue
            if ln.get("tiene_ficha") and ln.get("tiene_bultos"):
                try:
                    total_kg     += float(ln.get("peso_kg_tot") or 0)
                    total_vol_kg += float(ln.get("peso_vol_tot") or 0)
                    # vol_tot viene en cm³ — convertir a m³
                    total_m3     += float(ln.get("vol_tot") or 0) / 1_000_000.0
                    n_lineas_ok += 1
                except Exception:
                    pass

        # ⚡ PERF (Daniel 2026-05-24): snapshot reducido. Solo guardamos
        # los campos REALES que usa el wizard granular de líneas
        # (sku, descripcion, cantidad, peso/vol unitarios). Antes
        # serializábamos TODO el dict ERP incluyendo all_fields, raw_sample,
        # diagnostics — hasta 200KB por doc. Ahora ~5-15KB típico.
        try:
            import json as _json_serial
            _hdr_min = {
                "cliente_nombre":  hdr.get("cliente_nombre", ""),
                "cliente_rut":     hdr.get("cliente_rut", ""),
                "email":           hdr.get("email", ""),
                "telefono":        hdr.get("telefono", ""),
                "direccion":       hdr.get("direccion", ""),
                "comuna":          hdr.get("comuna", ""),
                "observaciones":   hdr.get("observaciones", ""),
                "fecha":           hdr.get("fecha", ""),
                "tido":            hdr.get("tido", ""),
                "nudo":            hdr.get("nudo", ""),
                "valor_neto":      hdr.get("valor_neto", 0),
                "valor_bruto":     hdr.get("valor_bruto", 0),
                "valor_iva":       hdr.get("valor_iva", 0),
                "tipo_operacion":  hdr.get("tipo_operacion", ""),
            }
            # Para cada línea solo los campos que usa el wizard granular
            _lineas_min = []
            for _ln in (lineas or []):
                if _ln.get("es_zz"):
                    # ZZ solo nos importa el saldo agregado, no detallar
                    _lineas_min.append({
                        "sku":       (_ln.get("sku") or ""),
                        "es_zz":     True,
                        "saldo_zz":  _ln.get("saldo_zz"),
                        "CAPRCO1":   _ln.get("CAPRCO1"),
                        "CAPRAD1":   _ln.get("CAPRAD1"),
                    })
                    continue
                _lineas_min.append({
                    "sku":             (_ln.get("sku") or ""),
                    "descripcion_erp": (_ln.get("descripcion_erp") or ""),
                    "nombre_app":      (_ln.get("nombre_app") or ""),
                    "cantidad":        _ln.get("cantidad", 0),
                    "peso_kg_u":       _ln.get("peso_kg_u", 0),
                    "peso_vol_u":      _ln.get("peso_vol_u", 0),
                    "vol_u":           _ln.get("vol_u", 0),
                    "peso_kg_tot":     _ln.get("peso_kg_tot", 0),
                    "peso_vol_tot":    _ln.get("peso_vol_tot", 0),
                    "vol_tot":         _ln.get("vol_tot", 0),
                    "tiene_ficha":     _ln.get("tiene_ficha", False),
                    "tiene_bultos":    _ln.get("tiene_bultos", False),
                    "es_zz":           False,
                })
            snapshot = {"hdr": _hdr_min, "lineas": _lineas_min}
            # ⚡ PERF: 30KB es más que suficiente (snapshot minimal típico
            # ronda 3-12KB). Si superara 30KB se trunca con seguridad —
            # ya filtramos los campos innecesarios arriba.
            snapshot_json = _json_serial.dumps(snapshot, default=str)[:30000]
        except Exception:
            snapshot_json = "{}"
        _lap("snapshot_json")

        cliente_rut = (hdr.get("cliente_rut") or hdr.get("rut") or "").strip()[:30]
        cliente_nombre = (hdr.get("cliente_nombre") or hdr.get("razon_social") or "").strip()[:200]
        obs = (hdr.get("observaciones") or hdr.get("obs") or "").strip()[:5000]
        # Email del cliente del doc — para sugerir como destinatario adicional.
        # Daniel 2026-05-23: "si el documento trae un correo, se puede enviar
        # tanto al correo del cliente que declaró como al del documento".
        email_doc = (hdr.get("email") or hdr.get("email_cliente") or "").strip().lower()[:180]
        if email_doc and "@" not in email_doc:
            email_doc = ""  # no es email válido
        added_by = g.user["nombre"] if getattr(g, "user", None) else "interno"

        # ── Calcular saldo ZZ del documento (Daniel 2026-05-23 wizard) ──
        # Sumamos saldo de las líneas ZZ (despacho/retiro): CAPRCO1 - CAPRAD1.
        # Si > 0: el doc tiene saldo disponible → con_saldo=1
        # Si = 0 y hay líneas ZZ: doc ya despachado → con_saldo=0 (bloqueado)
        # Si NO hay líneas ZZ: asumimos con_saldo=1 (no es bloqueante).
        # Tolerante a fallos: si no se puede calcular, queda NULL → UI advierte.
        con_saldo_val = None      # NULL = no se pudo verificar
        saldo_zz_val  = None
        try:
            saldo_total = 0.0
            tiene_zz = False
            for ln in (lineas or []):
                if not ln.get("es_zz"):
                    continue
                tiene_zz = True
                # Algunas implementaciones guardan saldo_zz; otras calculamos
                # de CAPRCO1 - CAPRAD1 si están disponibles en la línea.
                ln_saldo = ln.get("saldo_zz")
                if ln_saldo is None:
                    cc = float(ln.get("CAPRCO1") or ln.get("cantidad_total") or 0)
                    ca = float(ln.get("CAPRAD1") or ln.get("cantidad_despachada") or 0)
                    ln_saldo = max(0.0, cc - ca)
                try:
                    saldo_total += float(ln_saldo or 0)
                except Exception:
                    pass
            if tiene_zz:
                saldo_zz_val = saldo_total
                con_saldo_val = 1 if saldo_total > 0 else 0
            else:
                # Sin líneas ZZ → asumimos OK (no podemos detectar bloqueo)
                saldo_zz_val = None
                con_saldo_val = 1
        except Exception:
            con_saldo_val = None
            saldo_zz_val  = None

        # ⚡ PERF: capturar request_meta ANTES del thread daemon. log_event
        # accede a request.remote_addr + request.user_agent.string que NO
        # están disponibles fuera del contexto Flask. Si esto no se captura
        # ahora, el log se pierde silenciosamente cuando el thread corre.
        try:
            _ip_meta = request.remote_addr
            _ua_meta = (request.user_agent.string or "")[:300]
        except Exception:
            _ip_meta = None
            _ua_meta = ""

        # ⚡ PERF: INSERT + obtener lastrowid en un solo round-trip a BD.
        # Antes hacíamos INSERT (cur cerrado) + SELECT por unique key
        # = 2 round-trips. Ahora 1 round-trip ahorra ~30-80ms.
        #
        # ⚡ PERF Daniel 2026-05-24: si vienen `lineas_sel_in`, hacemos también
        # el UPSERT en pickup_doc_lineas y seteamos has_seleccion_lineas=1
        # en el MISMO INSERT (no en UPDATE separado). Todo en 1 transacción.
        from app import get_db as _get_db_doc
        new_doc_id = None
        has_sel_initial = 1 if lineas_sel_in else 0

        # Pre-armar las filas del upsert ANTES de tomar la conexión, para
        # minimizar tiempo de lock. Si el cálculo crashea, no afecta al INSERT.
        upsert_rows = []
        if lineas_sel_in:
            # snapshot ya está en memoria (la variable `lineas` del ERP fetch)
            # → no necesitamos parsear erp_snapshot otra vez (a diferencia del
            # endpoint /docs/<id>/lineas que SÍ lo parsea porque no tiene `lineas`).
            snap_by_sku = {(ln.get("sku") or "").upper(): ln for ln in (lineas or [])}
            for li in lineas_sel_in:
                sku = (li.get("sku") or "").strip()[:80]
                if not sku:
                    continue
                ln_snap = snap_by_sku.get(sku.upper()) or {}
                cantidad_doc = float(ln_snap.get("cantidad") or 0)
                peso_unit    = float(ln_snap.get("peso_kg_u") or 0)
                peso_vol_unit= float(ln_snap.get("peso_vol_u") or 0)
                vol_unit_cm3 = float(ln_snap.get("vol_u") or 0)
                vol_unit_m3  = vol_unit_cm3 / 1_000_000.0 if vol_unit_cm3 else 0
                try:
                    qty_sel = float(li.get("cantidad_seleccionada") or 0)
                except Exception:
                    qty_sel = 0
                qty_sel = max(0.0, min(qty_sel, cantidad_doc or qty_sel))
                # Si li.incluida no viene, asumimos True (semántica: viene en
                # el array de seleccionadas → el operador la quiere incluir)
                incluida = 1 if (li.get("incluida", True) and qty_sel > 0) else 0
                descripcion = (ln_snap.get("descripcion_erp") or ln_snap.get("nombre_app") or "").strip()[:300]
                nota = (li.get("nota") or "").strip()[:300] or None
                # 🆕 Daniel 2026-05-24: el frontend manda marcada_sin_saldo=true
                # cuando el operador eligió una línea que el ERP reporta como
                # ya entregada (saldo=0). Se guarda para mostrar badge en la
                # tabla externa, pero NO bloquea la asociación.
                marcada_sin_saldo = 1 if li.get("marcada_sin_saldo") else 0
                upsert_rows.append((
                    rid, sku, descripcion,
                    cantidad_doc, qty_sel,
                    peso_unit, peso_vol_unit, vol_unit_m3,
                    peso_unit * qty_sel,
                    peso_vol_unit * qty_sel,
                    vol_unit_m3 * qty_sel,
                    incluida, nota,
                    marcada_sin_saldo,
                ))
            # Si después del filtrado no quedó nada válido, NO marcamos selección
            if not upsert_rows:
                has_sel_initial = 0

        try:
            _conn_doc = _get_db_doc()
            with _conn_doc.cursor() as _cur_doc:
                _cur_doc.execute(
                    """INSERT INTO pickup_request_docs
                         (request_id, document_type, document_number,
                          cliente_rut, cliente_nombre, observaciones_erp,
                          peso_real_kg, peso_vol_kg, volumen_m3, n_lineas,
                          erp_snapshot, added_by, con_saldo, saldo_zz, saldo_checked_at,
                          email_cliente_erp, has_seleccion_lineas)
                       VALUES (%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s, %s,%s,NOW(), %s, %s)""",
                    (rid, tipo, numero,
                     cliente_rut, cliente_nombre, obs,
                     total_kg, total_vol_kg, total_m3, len(lineas or []),
                     snapshot_json, added_by, con_saldo_val, saldo_zz_val,
                     email_doc or None, has_sel_initial)
                )
                new_doc_id = _cur_doc.lastrowid
                # ⚡ UPSERT líneas en la MISMA transacción (si vinieron en body)
                if upsert_rows and new_doc_id:
                    # Inyectar doc_id en cada fila (era unknown antes del INSERT)
                    final_rows = [(r[0], new_doc_id, *r[1:]) for r in upsert_rows]
                    try:
                        _cur_doc.executemany(
                            """INSERT INTO pickup_doc_lineas
                                 (request_id, doc_id, sku, descripcion, cantidad_doc,
                                  cantidad_seleccionada, peso_unit_kg, peso_vol_unit_kg,
                                  vol_unit_m3, peso_total_kg, peso_vol_total_kg, vol_total_m3,
                                  incluida, nota_linea, marcada_sin_saldo)
                               VALUES (%s,%s,%s,%s, %s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s)""",
                            final_rows
                        )
                    except Exception as _e_msm:
                        # Compat: si la columna marcada_sin_saldo aún no migró
                        # (entorno viejo), insertamos sin esa columna.
                        print(f"[pickup_doc_agregar] sin col marcada_sin_saldo, retry: {_e_msm}", flush=True)
                        legacy_rows = [r[:-1] for r in final_rows]  # quita la última columna
                        _cur_doc.executemany(
                            """INSERT INTO pickup_doc_lineas
                                 (request_id, doc_id, sku, descripcion, cantidad_doc,
                                  cantidad_seleccionada, peso_unit_kg, peso_vol_unit_kg,
                                  vol_unit_m3, peso_total_kg, peso_vol_total_kg, vol_total_m3,
                                  incluida, nota_linea)
                               VALUES (%s,%s,%s,%s, %s,%s, %s,%s,%s, %s,%s,%s, %s,%s)""",
                            legacy_rows
                        )
            _conn_doc.commit()
        except Exception as exc:
            # Fallback si la columna con_saldo / has_seleccion_lineas aún no se
            # migró (entornos viejos). Reintentamos sin esas columnas y, si
            # había líneas, hacemos el UPSERT + UPDATE has_seleccion_lineas
            # en pasos separados (cuando la columna SÍ existe).
            try:
                _conn_doc = _get_db_doc()
                with _conn_doc.cursor() as _cur_doc:
                    _cur_doc.execute(
                        """INSERT INTO pickup_request_docs
                             (request_id, document_type, document_number,
                              cliente_rut, cliente_nombre, observaciones_erp,
                              peso_real_kg, peso_vol_kg, volumen_m3, n_lineas,
                              erp_snapshot, added_by)
                           VALUES (%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s)""",
                        (rid, tipo, numero,
                         cliente_rut, cliente_nombre, obs,
                         total_kg, total_vol_kg, total_m3, len(lineas or []),
                         snapshot_json, added_by)
                    )
                    new_doc_id = _cur_doc.lastrowid
                    # Líneas + has_seleccion_lineas en best-effort (puede fallar
                    # silencioso si la migración aún no corrió)
                    if upsert_rows and new_doc_id:
                        try:
                            final_rows = [(r[0], new_doc_id, *r[1:]) for r in upsert_rows]
                            # En fallback, asumimos que la migración nueva tampoco corrió,
                            # así que insertamos SIN marcada_sin_saldo (compat máxima).
                            legacy_rows_fb = [r[:-1] for r in final_rows]
                            _cur_doc.executemany(
                                """INSERT INTO pickup_doc_lineas
                                     (request_id, doc_id, sku, descripcion, cantidad_doc,
                                      cantidad_seleccionada, peso_unit_kg, peso_vol_unit_kg,
                                      vol_unit_m3, peso_total_kg, peso_vol_total_kg, vol_total_m3,
                                      incluida, nota_linea)
                                   VALUES (%s,%s,%s,%s, %s,%s, %s,%s,%s, %s,%s,%s, %s,%s)""",
                                legacy_rows_fb
                            )
                            _cur_doc.execute(
                                "UPDATE pickup_request_docs SET has_seleccion_lineas=1 WHERE id=%s",
                                (new_doc_id,)
                            )
                        except Exception as _e_lines_fb:
                            print(f"[pickup_doc_agregar] fallback lineas-upsert falló: {_e_lines_fb}", flush=True)
                _conn_doc.commit()
            except Exception as exc2:
                err = str(exc2)[:200]
                return jsonify({"ok": False, "error": f"Error al guardar: {err}"}), 500
        _lap("insert_doc")
        if upsert_rows:
            _lap("upsert_lineas_inline")

        # ⚡ PERF: invalidar caches AHORA (antes del thread). Si el operador
        # hace polling enseguida, debe ver el nuevo doc. Costo: <1ms (pops
        # en memoria, no toca BD).
        try:
            if cliente_rut:
                _rut_norm = re.sub(r"[^0-9kK]", "", str(cliente_rut)).upper()
                for _k in list(_SALDO_CACHE.keys()):
                    if _k.startswith(_rut_norm + "|"):
                        _SALDO_CACHE.pop(_k, None)
            _DOCS_CACHE.pop(rid, None)
        except Exception:
            pass

        # ⚡ PERF: recalcular totales SÍNCRONO (rápido: 2 SELECTs + 1 UPDATE
        # = ~50-100ms). El frontend lo usa para refrescar el panel de
        # totales del retiro en la misma respuesta.
        totales = _pickup_recalc_totales(rid)
        _lap("recalc_totales")

        # ⚡ PERF: log audit + auto-validate van a thread daemon (~200-400ms
        # en BD). El operador NO espera estos pasos — la respuesta sale
        # inmediato. Captura previa de IP/UA garantiza que funcione fuera
        # del request context Flask.
        import threading as _t_doc_agg
        _log_msg = (f"Doc {tipo} {numero} agregado · cliente={cliente_nombre} · "
                    f"peso={total_kg:.2f}kg vol={total_vol_kg:.2f}kg m3={total_m3:.3f} · "
                    f"con_saldo={con_saldo_val}")

        def _async_audit_and_autovalidate():
            # ⚠️ CRÍTICO: este thread corre FUERA del request context Flask.
            # No podemos usar mysql_execute() ni mysql_fetchone() porque ambos
            # usan get_db() → g._db que requiere contexto (RuntimeError).
            # Patrón correcto: get_mysql() directo + cerrar al final.
            # (Mismo patrón que _update_last_seen en app.py — bug histórico
            # del 2026-05-18 que costó "last_login_at congelado" para todos.)
            from app import get_mysql as _get_mysql_async
            conn = None
            try:
                conn = _get_mysql_async()
                with conn.cursor() as cur:
                    # Log de "doc_agregado"
                    try:
                        cur.execute(
                            f"""INSERT INTO `{LOG}`
                                (request_id,actor_type,actor_name,action,old_status,new_status,notes,ip,user_agent)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (rid, "interno", added_by, "doc_agregado",
                             None, None, _log_msg, _ip_meta, _ua_meta)
                        )
                    except Exception as _e_log:
                        print(f"[pickup_doc_agregar async-log] {_e_log}", flush=True)

                    # Daniel 2026-06-15: auto-validar la documentación al asociar
                    # un doc, TENGA O NO saldo. El "sin saldo" ya NO bloquea (es
                    # solo un indicador), así que no debe impedir proponer fecha.
                    # Antes solo auto-validaba CON saldo → los docs sin saldo
                    # dejaban el retiro trabado en 'pendiente' sin botón visible.
                    try:
                        _saldo_txt = ("con saldo disponible" if con_saldo_val == 1
                                      else ("sin saldo verificado" if con_saldo_val == 0
                                            else "saldo no verificado"))
                        cur.execute(
                            f"SELECT doc_validation_status FROM `{REQ}` WHERE id=%s",
                            (rid,)
                        )
                        cur_status = cur.fetchone() or {}
                        if (cur_status.get("doc_validation_status") or "") != "ok":
                            cur.execute(
                                f"""UPDATE `{REQ}`
                                    SET doc_validation_status='ok',
                                        doc_validated_at=NOW(),
                                        doc_validated_by=%s,
                                        doc_validation_notes=CONCAT(
                                          COALESCE(doc_validation_notes,''),
                                          IF(doc_validation_notes IS NULL OR doc_validation_notes='','','\\n'),
                                          'Auto-validado: doc ', %s, ' ', %s, ' (', %s, ')'
                                        )
                                    WHERE id=%s""",
                                (added_by[:120], tipo, numero, _saldo_txt, rid)
                            )
                            cur.execute(
                                f"""INSERT INTO `{LOG}`
                                    (request_id,actor_type,actor_name,action,old_status,new_status,notes,ip,user_agent)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                                (rid, "sistema", "Sistema", "doc_validacion_auto",
                                 None, "ok",
                                 f"Auto-validado al asociar {tipo} {numero} ({_saldo_txt})",
                                 _ip_meta, _ua_meta)
                            )
                    except Exception as _e_auto:
                        print(f"[pickup_doc_agregar async-autovalidate] {_e_auto}",
                              flush=True)
                conn.commit()
            except Exception as _e_outer:
                print(f"[pickup_doc_agregar async-outer] {_e_outer}", flush=True)
            finally:
                if conn is not None:
                    try: conn.close()
                    except Exception: pass

        _t_doc_agg.Thread(target=_async_audit_and_autovalidate, daemon=True).start()
        _lap("async_dispatched")

        # ⚡ PERF: construir new_row EN MEMORIA con los datos que YA tenemos.
        # Antes se hacía SELECT extra (1 round-trip = ~30-80ms). Ahora 0.
        from datetime import datetime as _dt_now
        # Conteo de líneas seleccionadas (solo si vinieron en el body)
        n_lineas_sel = sum(1 for r in upsert_rows if r[-2]) if upsert_rows else 0
        new_row = {
            "id":                new_doc_id,
            "document_type":     tipo,
            "document_number":   numero,
            "cliente_rut":       cliente_rut,
            "cliente_nombre":    cliente_nombre,
            "observaciones_erp": obs,
            "peso_real_kg":      float(total_kg),
            "peso_vol_kg":       float(total_vol_kg),
            "volumen_m3":        float(total_m3),
            "n_lineas":          len(lineas or []),
            "added_by":          added_by,
            "added_at":          _dt_now.now().strftime("%Y-%m-%d %H:%M:%S"),
            "has_seleccion_lineas":   bool(has_sel_initial),
            "n_lineas_seleccionadas": n_lineas_sel if upsert_rows else None,
            "con_saldo":              con_saldo_val,
            "saldo_zz":               saldo_zz_val,
        }

        # Logear timing total — visible en Railway logs para diagnóstico
        # continuo. Cada etapa con su latencia: si vuelve a haber un cuello
        # de botella, vemos cuál subpaso es el culpable sin adivinar.
        _total_ms = int((_t_perf.time() - _t0) * 1000)
        _laps_str = " | ".join(f"{n}={ms}ms" for n, ms in _laps)
        _combined_tag = " [COMBINED]" if upsert_rows else ""
        print(f"[pickup_doc_agregar] rid={rid} {tipo} {numero} "
              f"TOTAL={_total_ms}ms{_combined_tag} · {_laps_str}", flush=True)

        resp = {
            "ok": True,
            "doc": new_row,
            "totales": totales,
            "_perf_ms": _total_ms,  # útil para Daniel inspeccionar desde DevTools
        }
        if upsert_rows:
            resp["lineas_guardadas"] = len(upsert_rows)
            resp["lineas_incluidas"] = n_lineas_sel
        if otro:
            resp["warning_otro_retiro"] = {
                "request_id": otro["request_id"],
                "code": otro["code"],
            }
        return jsonify(resp)

    @app.route("/retiros/<int:rid>/docs/<int:doc_id>", methods=["DELETE", "POST"])
    @require_permission("retiros")
    def pickup_doc_quitar(rid, doc_id):
        """Quita un documento del retiro y recalcula totales.
        Acepta DELETE o POST (POST con _method=DELETE para clientes que no soporten DELETE).
        """
        if request.method == "POST" and (request.form.get("_method") or "").upper() != "DELETE":
            return jsonify({"ok": False, "error": "Método inválido"}), 405
        row = mysql_fetchone(
            "SELECT id, document_type, document_number "
            "FROM pickup_request_docs WHERE id=%s AND request_id=%s",
            (doc_id, rid)
        )
        if not row:
            return jsonify({"ok": False, "error": "Documento no asociado a este retiro"}), 404
        try:
            mysql_execute(
                "DELETE FROM pickup_request_docs WHERE id=%s AND request_id=%s",
                (doc_id, rid)
            )
        except Exception as exc:
            return jsonify({"ok": False, "error": f"Error al borrar: {str(exc)[:200]}"}), 500
        log_event(rid, "doc_quitado", None, None,
                  f"Doc {row['document_type']} {row['document_number']} removido del retiro",
                  "interno")
        totales = _pickup_recalc_totales(rid)
        # ⚡ PERF: invalidar caches relacionados
        try:
            _DOCS_CACHE.pop(rid, None)
        except Exception:
            pass
        return jsonify({"ok": True, "totales": totales})

    # ══════════════════════════════════════════════════════════════════
    #  SELECCIÓN GRANULAR DE LÍNEAS POR DOC (Daniel 2026-05-23)
    #
    #  "de dos documentos, que tengan diez y diez productos, podría
    #  seleccionar dos y tres productos de cada factura. Eso es el
    #  dinamismo que te estoy pidiendo".
    #
    #  Modelo:
    #  - GET  /retiros/<rid>/docs/<doc_id>/lineas
    #    → lista líneas del doc (de erp_snapshot) + estado de selección
    #      desde pickup_doc_lineas. Si nunca se guardó, devuelve TODAS
    #      las líneas marcadas con incluida=1 y qty=cantidad_doc.
    #
    #  - POST /retiros/<rid>/docs/<doc_id>/lineas
    #    Body: {lineas: [{sku, incluida, cantidad_seleccionada, nota}]}
    #    → UPSERT en pickup_doc_lineas + has_seleccion_lineas=1 en doc.
    #      Recalcula totales del retiro considerando solo líneas incluidas.
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/<int:rid>/docs/<int:doc_id>/lineas", methods=["GET"])
    @require_permission("view")
    def pickup_doc_lineas_listar(rid, doc_id):
        """Devuelve líneas del doc + estado de selección granular."""
        doc = mysql_fetchone(
            "SELECT id, document_type, document_number, erp_snapshot, "
            "       has_seleccion_lineas "
            "FROM pickup_request_docs WHERE id=%s AND request_id=%s",
            (doc_id, rid)
        )
        if not doc:
            return jsonify({"ok": False, "error": "Doc no asociado a este retiro"}), 404

        # Parsear erp_snapshot (líneas crudas guardadas al asociar)
        try:
            import json as _json_l
            snap = _json_l.loads(doc.get("erp_snapshot") or "{}")
        except Exception:
            snap = {}
        lineas_raw = snap.get("lineas") or []

        # Selecciones guardadas (si las hay)
        sel_rows = mysql_fetchall(
            "SELECT sku, cantidad_seleccionada, incluida, nota_linea "
            "FROM pickup_doc_lineas WHERE doc_id=%s",
            (doc_id,)
        ) or []
        sel_map = {(r["sku"] or "").upper(): r for r in sel_rows}

        out = []
        for ln in lineas_raw:
            if ln.get("es_zz"):
                continue  # líneas ZZ son despachos, no productos
            sku = (ln.get("sku") or "").strip()
            if not sku:
                continue
            cantidad_doc = float(ln.get("cantidad") or 0)
            peso_unit = float(ln.get("peso_kg_u") or 0)
            peso_vol_unit = float(ln.get("peso_vol_u") or 0)
            vol_unit_cm3 = float(ln.get("vol_u") or 0)
            vol_unit_m3 = vol_unit_cm3 / 1_000_000.0 if vol_unit_cm3 else 0
            sel = sel_map.get(sku.upper())
            if sel:
                incluida = bool(sel.get("incluida"))
                qty_sel = float(sel.get("cantidad_seleccionada") or 0)
                nota = sel.get("nota_linea") or ""
            elif bool(doc.get("has_seleccion_lineas")):
                # 🔧 FIX Daniel 2026-05-24: si el doc YA tiene selección
                # granular y esta línea NO está en pickup_doc_lineas, significa
                # que el operador eligió NO incluirla (al asociar desde RUT
                # solo se mandan las marcadas). Default = NO seleccionada.
                # Antes default=True hacía que al re-abrir el modal aparecieran
                # todas marcadas → confusión + posibles re-adds accidentales.
                incluida = False
                qty_sel = 0
                nota = ""
            else:
                # Sin selección previa: default = TODA la línea seleccionada
                incluida = True
                qty_sel = cantidad_doc
                nota = ""
            out.append({
                "sku":                   sku,
                "descripcion":           (ln.get("descripcion_erp") or ln.get("nombre_app") or "").strip(),
                "cantidad_doc":          cantidad_doc,
                "cantidad_seleccionada": qty_sel,
                "incluida":              incluida,
                "peso_unit_kg":          peso_unit,
                "peso_vol_unit_kg":      peso_vol_unit,
                "vol_unit_m3":           vol_unit_m3,
                "tiene_ficha":           bool(ln.get("tiene_ficha")),
                "nota":                  nota,
            })

        return jsonify({
            "ok": True,
            "doc": {
                "id": doc["id"],
                "tipo": doc["document_type"],
                "numero": doc["document_number"],
                "has_seleccion_lineas": bool(doc.get("has_seleccion_lineas")),
            },
            "lineas": out,
            "total_lineas": len(out),
        })

    # ══════════════════════════════════════════════════════════════════
    #  ENDPOINT — RESUMEN CONSOLIDADO DE PRODUCTOS A RETIRAR
    #  Daniel 2026-05-24: tabla 2 del Paso 1 — consolida todas las
    #  líneas (incluidas) de TODOS los documentos asociados al retiro,
    #  para que el operador vea de un vistazo qué se va a retirar
    #  con SKU + descripción + doc origen + totales.
    #
    #  Lógica:
    #   - Docs con has_seleccion_lineas=1 → SELECT de pickup_doc_lineas
    #     WHERE incluida=1 (selección granular guardada).
    #   - Docs SIN selección granular → leer erp_snapshot (JSON) y
    #     devolver TODAS sus líneas con cantidad_doc completa.
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/<int:rid>/lineas-resumen", methods=["GET"])
    @require_permission("view")
    def pickup_lineas_resumen(rid):
        """Resumen consolidado de productos de TODOS los docs asociados.

        Response JSON:
        {
          "ok": true,
          "lineas": [
            {"sku": "...", "descripcion": "...", "doc_tipo": "FCV",
             "doc_numero": "10599", "doc_id": 33, "cantidad": 3,
             "peso_unit_kg": 5.0, "vol_unit_m3": 0.05,
             "peso_total": 15.0, "vol_total": 0.15}
          ],
          "totales": {"n_lineas": 5, "peso_total_kg": 45.5, "vol_total_m3": 0.30}
        }
        """
        try:
            return _pickup_lineas_resumen_impl(rid)
        except Exception as e:
            print(f"[pickup_lineas_resumen] CRASH rid={rid}: {e}", flush=True)
            return jsonify({
                "ok": False,
                "error": "Error interno al consolidar productos",
                "error_codigo": "INTERNAL_CRASH",
                "detalle": str(e)[:200],
            }), 500

    def _pickup_lineas_resumen_impl(rid):
        """Wrapper HTTP: jsonify del helper puro (para la ruta /lineas-resumen)."""
        return jsonify(_pickup_lineas_consolidadas(rid))

    def _pickup_lineas_consolidadas(rid):
        """Núcleo puro (Daniel 2026-06-17): devuelve dict {ok, lineas, totales}
        SIN jsonify, para reusarlo desde el correo (producto en el email) además
        de la ruta /lineas-resumen. Cada línea trae sku/descripcion/doc_tipo/
        doc_numero/cantidad/peso/vol + marcada_sin_saldo (indicador de stock)."""
        import json as _json_res
        # Traer todos los docs asociados al retiro
        try:
            docs = mysql_fetchall(
                """SELECT id, document_type, document_number, erp_snapshot,
                          has_seleccion_lineas
                     FROM pickup_request_docs
                    WHERE request_id=%s
                    ORDER BY id ASC""",
                (rid,)
            ) or []
        except Exception:
            # Fallback si no existe la columna has_seleccion_lineas todavía
            docs = mysql_fetchall(
                """SELECT id, document_type, document_number, erp_snapshot
                     FROM pickup_request_docs
                    WHERE request_id=%s
                    ORDER BY id ASC""",
                (rid,)
            ) or []

        if not docs:
            return {
                "ok": True,
                "lineas": [],
                "totales": {"n_lineas": 0, "peso_total_kg": 0.0, "vol_total_m3": 0.0},
            }

        lineas_out = []
        peso_total_acum = 0.0
        vol_total_acum  = 0.0

        for doc in docs:
            doc_id     = doc.get("id")
            doc_tipo   = (doc.get("document_type") or "").upper()
            doc_numero = (doc.get("document_number") or "").strip()
            has_sel    = bool(doc.get("has_seleccion_lineas"))

            if has_sel:
                # Selección granular guardada — solo líneas incluidas=1
                # 🆕 Daniel 2026-05-24: traemos marcada_sin_saldo para que
                # el frontend muestre badge ámbar "ya rebajado en ERP" en
                # cada línea de la tabla externa de productos.
                try:
                    sel_rows = mysql_fetchall(
                        """SELECT sku, descripcion, cantidad_seleccionada,
                                  peso_unit_kg, vol_unit_m3,
                                  peso_total_kg, vol_total_m3,
                                  marcada_sin_saldo
                             FROM pickup_doc_lineas
                            WHERE doc_id=%s AND incluida=1
                            ORDER BY sku ASC""",
                        (doc_id,)
                    ) or []
                except Exception as _e:
                    # Fallback: columna marcada_sin_saldo aún no migró
                    print(f"[lineas_resumen] doc {doc_id} sin col marcada_sin_saldo, retry: {_e}", flush=True)
                    try:
                        sel_rows = mysql_fetchall(
                            """SELECT sku, descripcion, cantidad_seleccionada,
                                      peso_unit_kg, vol_unit_m3,
                                      peso_total_kg, vol_total_m3
                                 FROM pickup_doc_lineas
                                WHERE doc_id=%s AND incluida=1
                                ORDER BY sku ASC""",
                            (doc_id,)
                        ) or []
                    except Exception as _e2:
                        print(f"[lineas_resumen] doc {doc_id} sel error: {_e2}", flush=True)
                        sel_rows = []
                for r in sel_rows:
                    qty       = float(r.get("cantidad_seleccionada") or 0)
                    peso_unit = float(r.get("peso_unit_kg") or 0)
                    vol_unit  = float(r.get("vol_unit_m3") or 0)
                    peso_tot  = float(r.get("peso_total_kg") or (peso_unit * qty))
                    vol_tot   = float(r.get("vol_total_m3") or (vol_unit * qty))
                    lineas_out.append({
                        "sku":               (r.get("sku") or "").strip(),
                        "descripcion":       (r.get("descripcion") or "").strip(),
                        "doc_tipo":          doc_tipo,
                        "doc_numero":        doc_numero,
                        "doc_id":            doc_id,
                        "cantidad":          qty,
                        "peso_unit_kg":      peso_unit,
                        "vol_unit_m3":       vol_unit,
                        "peso_total":        peso_tot,
                        "vol_total":         vol_tot,
                        "marcada_sin_saldo": bool(r.get("marcada_sin_saldo")),
                    })
                    peso_total_acum += peso_tot
                    vol_total_acum  += vol_tot
            else:
                # Sin selección granular → leer erp_snapshot y devolver todas
                try:
                    snap = _json_res.loads(doc.get("erp_snapshot") or "{}")
                except Exception:
                    snap = {}
                for ln in (snap.get("lineas") or []):
                    if ln.get("es_zz"):
                        continue  # líneas ZZ son despachos, no productos
                    sku = (ln.get("sku") or "").strip()
                    if not sku:
                        continue
                    qty = float(ln.get("cantidad") or 0)
                    if qty <= 0:
                        continue
                    peso_unit    = float(ln.get("peso_kg_u") or 0)
                    vol_unit_cm3 = float(ln.get("vol_u") or 0)
                    vol_unit_m3  = vol_unit_cm3 / 1_000_000.0 if vol_unit_cm3 else 0.0
                    desc = (ln.get("descripcion_erp") or ln.get("nombre_app") or "").strip()
                    peso_tot = peso_unit * qty
                    vol_tot  = vol_unit_m3 * qty
                    lineas_out.append({
                        "sku":               sku,
                        "descripcion":       desc,
                        "doc_tipo":          doc_tipo,
                        "doc_numero":        doc_numero,
                        "doc_id":            doc_id,
                        "cantidad":          qty,
                        "peso_unit_kg":      peso_unit,
                        "vol_unit_m3":       vol_unit_m3,
                        "peso_total":        peso_tot,
                        "vol_total":         vol_tot,
                        "marcada_sin_saldo": False,  # snapshot completo = ERP en orden
                    })
                    peso_total_acum += peso_tot
                    vol_total_acum  += vol_tot

        return {
            "ok": True,
            "lineas": lineas_out,
            "totales": {
                "n_lineas":      len(lineas_out),
                "peso_total_kg": round(peso_total_acum, 2),
                "vol_total_m3":  round(vol_total_acum, 4),
            },
        }

    @app.route("/retiros/<int:rid>/docs/<int:doc_id>/lineas", methods=["POST"])
    @require_permission("retiros")
    def pickup_doc_lineas_guardar(rid, doc_id):
        """Guarda selección granular de líneas. UPSERT en pickup_doc_lineas.

        Body: {lineas: [{sku, incluida, cantidad_seleccionada, nota?}]}
        """
        doc = mysql_fetchone(
            "SELECT id, document_type, document_number, erp_snapshot "
            "FROM pickup_request_docs WHERE id=%s AND request_id=%s",
            (doc_id, rid)
        )
        if not doc:
            return jsonify({"ok": False, "error": "Doc no asociado"}), 404

        body = request.get_json(silent=True) or {}
        lineas_in = body.get("lineas") or []
        if not isinstance(lineas_in, list):
            return jsonify({"ok": False, "error": "lineas debe ser una lista"}), 400

        # Reusamos los unitarios del erp_snapshot para calcular totales server-side
        try:
            import json as _json_g
            snap = _json_g.loads(doc.get("erp_snapshot") or "{}")
        except Exception:
            snap = {}
        snap_lineas = {(ln.get("sku") or "").upper(): ln for ln in (snap.get("lineas") or [])}

        upsert_rows = []
        for li in lineas_in:
            sku = (li.get("sku") or "").strip()[:80]
            if not sku:
                continue
            ln_snap = snap_lineas.get(sku.upper()) or {}
            cantidad_doc = float(ln_snap.get("cantidad") or 0)
            peso_unit    = float(ln_snap.get("peso_kg_u") or 0)
            peso_vol_unit= float(ln_snap.get("peso_vol_u") or 0)
            vol_unit_cm3 = float(ln_snap.get("vol_u") or 0)
            vol_unit_m3  = vol_unit_cm3 / 1_000_000.0 if vol_unit_cm3 else 0
            try:
                qty_sel = float(li.get("cantidad_seleccionada") or 0)
            except Exception:
                qty_sel = 0
            # Clamp: no se puede pedir más que lo que hay en el doc
            qty_sel = max(0.0, min(qty_sel, cantidad_doc or qty_sel))
            incluida = 1 if li.get("incluida") and qty_sel > 0 else 0
            descripcion = (ln_snap.get("descripcion_erp") or ln_snap.get("nombre_app") or "").strip()[:300]
            nota = (li.get("nota") or "").strip()[:300] or None
            upsert_rows.append((
                rid, doc_id, sku, descripcion,
                cantidad_doc, qty_sel,
                peso_unit, peso_vol_unit, vol_unit_m3,
                peso_unit * qty_sel,
                peso_vol_unit * qty_sel,
                vol_unit_m3 * qty_sel,
                incluida, nota,
            ))

        if not upsert_rows:
            return jsonify({"ok": False, "error": "Sin líneas válidas"}), 400

        # Daniel 2026-05-24 FIX CRÍTICO: el agente anterior usó
        # `get_mysql_conn()` que NO existe en el proyecto. La función real
        # es `get_mysql()` (definida en app.py:1010). Por eso el POST
        # /docs/<id>/lineas crasheaba con "name 'get_mysql_conn' is not
        # defined" y has_seleccion_lineas quedaba en 0 → tabla externa
        # mostraba TODAS las líneas en vez de solo las marcadas.
        #
        # Además, usar get_db() (cacheado por request) es más eficiente
        # que crear nueva conexión.
        try:
            conn = get_db()
            with conn.cursor() as cur:
                cur.executemany(
                    """INSERT INTO pickup_doc_lineas
                         (request_id, doc_id, sku, descripcion, cantidad_doc,
                          cantidad_seleccionada, peso_unit_kg, peso_vol_unit_kg,
                          vol_unit_m3, peso_total_kg, peso_vol_total_kg, vol_total_m3,
                          incluida, nota_linea)
                       VALUES (%s,%s,%s,%s, %s,%s, %s,%s,%s, %s,%s,%s, %s,%s)
                       ON DUPLICATE KEY UPDATE
                         descripcion=VALUES(descripcion),
                         cantidad_doc=VALUES(cantidad_doc),
                         cantidad_seleccionada=VALUES(cantidad_seleccionada),
                         peso_unit_kg=VALUES(peso_unit_kg),
                         peso_vol_unit_kg=VALUES(peso_vol_unit_kg),
                         vol_unit_m3=VALUES(vol_unit_m3),
                         peso_total_kg=VALUES(peso_total_kg),
                         peso_vol_total_kg=VALUES(peso_vol_total_kg),
                         vol_total_m3=VALUES(vol_total_m3),
                         incluida=VALUES(incluida),
                         nota_linea=VALUES(nota_linea),
                         updated_at=NOW()""",
                    upsert_rows
                )
                cur.execute(
                    "UPDATE pickup_request_docs "
                    "SET has_seleccion_lineas=1 WHERE id=%s",
                    (doc_id,)
                )
            conn.commit()
        except Exception as exc:
            return jsonify({"ok": False, "error": f"Error al guardar: {str(exc)[:200]}"}), 500

        log_event(rid, "lineas_seleccion", None, None,
                  f"Doc {doc['document_type']} {doc['document_number']}: "
                  f"selección granular guardada ({len(upsert_rows)} líneas, "
                  f"{sum(1 for r in upsert_rows if r[-2])} incluidas)",
                  "interno")

        # Recalcular totales del retiro tomando en cuenta selecciones
        totales = _pickup_recalc_totales(rid)
        return jsonify({"ok": True, "lineas_guardadas": len(upsert_rows), "totales": totales})

    # ══════════════════════════════════════════════════════════════════
    #  BORRADO TOTAL DE SOLICITUD — SOLO SUPERADMIN
    #  Daniel 2026-05-23: necesita poder limpiar solicitudes de prueba o
    #  duplicadas sin pedir SQL al desarrollador. Se restringe a superadmin
    #  porque borra cascade (request → docs, packages, proposals, logs,
    #  attachments, signatures vía FK ON DELETE CASCADE).
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/<int:rid>", methods=["DELETE", "POST"])
    @require_permission("retiros")
    def pickup_delete(rid):
        """SOLO SUPERADMIN — borrado físico de una solicitud de retiro completa.

        Cascade automático (ON DELETE CASCADE en schema):
        - pickup_packages, pickup_proposals, pickup_logs,
          pickup_attachments, pickup_signatures, pickup_request_docs.

        Requiere:
        - Rol superadmin (admin/retiros/edit → 403).
        - Confirmación 'BORRAR' (header X-Confirm-Delete o body confirm)
          para evitar borrados accidentales (fat-finger).
        - Acepta DELETE o POST con _method=DELETE.

        Respuesta:
        - 200 {ok:True, deleted:{...}} con datos del registro borrado.
        - 403 si el usuario no es superadmin.
        - 400 si falta confirmación.
        - 404 si el retiro no existe.
        """
        # Gate superadmin (require_permission ya permite superadmin pasar,
        # pero verificamos explícito para que cualquier otro rol con
        # permiso 'retiros' NO pueda borrar — solo ver/editar).
        if not (g.user and g.permissions.get("superadmin")):
            return jsonify({
                "ok": False,
                "error": "Solo el superadministrador puede eliminar solicitudes de retiro.",
                "error_codigo": "SOLO_SUPERADMIN",
            }), 403

        # Método DELETE o POST con _method=DELETE
        if request.method == "POST":
            method_override = (request.form.get("_method") or
                               (request.get_json(silent=True) or {}).get("_method") or
                               "").upper()
            if method_override != "DELETE":
                return jsonify({"ok": False,
                                "error": "Método inválido. Usa DELETE o _method=DELETE."}), 405

        # Confirmación anti-fat-finger (header o body)
        body = request.get_json(silent=True) or {}
        confirm = (request.headers.get("X-Confirm-Delete", "") or
                   request.form.get("confirm", "") or
                   body.get("confirm", "") or "").strip().upper()
        if confirm != "BORRAR":
            return jsonify({
                "ok": False,
                "error": "Confirmación requerida. Envía 'BORRAR' como confirmación.",
                "error_codigo": "FALTA_CONFIRMACION",
            }), 400

        # Verificar que existe + obtener datos para log y respuesta
        req = mysql_fetchone(
            f"SELECT id, code, status, customer_name, customer_rut, "
            f"       contact_email, contact_phone, created_at "
            f"FROM `{REQ}` WHERE id=%s LIMIT 1",
            (rid,)
        )
        if not req:
            return jsonify({"ok": False, "error": "Solicitud no encontrada"}), 404

        # Audit log permanente — en consola (los logs DB se borran por cascade,
        # así que dejamos huella en stdout que Railway captura).
        actor_email = (g.user.get("email") or g.user.get("nombre") or "?")
        print(f"[ILUS][SUPERADMIN][DELETE] retiro id={rid} "
              f"code={req.get('code')} status={req.get('status')} "
              f"cliente='{req.get('customer_name')}' rut={req.get('customer_rut')} "
              f"por={actor_email}")

        try:
            # Cascade automático borra todas las tablas relacionadas
            mysql_execute(f"DELETE FROM `{REQ}` WHERE id=%s", (rid,))
        except Exception as exc:
            print(f"[ILUS][SUPERADMIN][DELETE-ERR] id={rid} err={str(exc)[:200]}")
            return jsonify({
                "ok": False,
                "error": f"Error al borrar: {str(exc)[:200]}",
            }), 500

        return jsonify({
            "ok": True,
            "deleted": {
                "id": rid,
                "code": req.get("code"),
                "status": req.get("status"),
                "customer_name": req.get("customer_name"),
                "customer_rut": req.get("customer_rut"),
            },
            "message": f"Solicitud {req.get('code')} eliminada permanentemente.",
        })

    # ══════════════════════════════════════════════════════════════════
    #  SALDO ERP — Documentos pendientes de despacho del cliente
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/api/cliente/<rut>/saldo-pendiente", methods=["GET"])
    @require_permission("view")
    def pickup_cliente_saldo_erp(rut):
        """Lista documentos del cliente (por RUT) que aún tienen saldo en ZZ
        (producto Despacho/Retiro) o que no han sido marcados como despachados.

        Args (query string):
            dias: int — ventana hacia atrás (default 30, max 180)
            solo_con_saldo: '1'/'0' — si '1' (default) solo devuelve docs con
                            saldo de despacho pendiente > 0. Si '0' devuelve
                            todos los emitidos en la ventana de tiempo.
                            Daniel 2026-05-23: "es mejor no asignarlo, porque
                            vamos a enfocarnos en algo que no tiene salida, ya
                            se entregó". El wizard interno usa solo_con_saldo=1
                            para enfocar al operador en lo accionable.

        Estrategia:
          - Query a MAEEDO + MAEDDO + MAEEN filtrando por RUT y fechas.
          - Detectar "no despachado" usando saldo de ZZ (CAPRCO1-CAPRAD1) > 0
            cuando hay líneas ZZ en el doc — mismo criterio que Transporte.
          - Marcar `tiene_retiro_asociado=True` si el doc ya está en pickup_requests
            o pickup_request_docs (no es definitivo "despachado" pero sí ya en flujo).

        Si la consulta SQL al ERP falla → devolver lista vacía con error legible
        (no 500). Esto permite que la UI nunca rompa por caídas del ERP.
        """
        rut_clean = re.sub(r"[^0-9kK]", "", str(rut or "")).upper()
        if not rut_clean or len(rut_clean) < 7:
            return jsonify({"ok": False, "error": "RUT inválido", "docs": []}), 400

        try:
            dias = int(request.args.get("dias", 30))
        except Exception:
            dias = 30
        dias = max(1, min(dias, 180))

        # Daniel 2026-05-23: filtrar docs SIN saldo por default. El operador se
        # enfoca solo en lo accionable (lo que el cliente puede REALMENTE retirar).
        # Los docs ya despachados completos se ocultan a menos que pase solo_con_saldo=0.
        solo_con_saldo_param = (request.args.get("solo_con_saldo") or "1").strip().lower()
        solo_con_saldo = solo_con_saldo_param in ("1", "true", "yes", "y")

        # ⚡ PERF (Daniel 2026-05-24): cache hit por RUT+dias+filter.
        # El wizard se abre/cierra varias veces — sin cache cada apertura
        # disparaba una query MAEEDO+MAEDDO+MAEEN compleja (800-1500ms).
        import time as _time_saldo
        _cache_key = f"{rut_clean}|{dias}|{int(solo_con_saldo)}"
        _hit = _SALDO_CACHE.get(_cache_key)
        if _hit and (_time_saldo.time() - _hit[1]) < _SALDO_TTL:
            # Hit fresco → response inmediato (~5ms vs ~1200ms cold)
            return jsonify(_hit[0])

        try:
            from app import _random_sql_query, _random_sql_pool
        except ImportError:
            return jsonify({"ok": True, "docs": [], "error": "Motor ERP no disponible"})

        pool = _random_sql_pool()
        if pool is None:
            return jsonify({
                "ok": True, "docs": [],
                "error": "ERP Random no está configurado en este entorno."
            })

        # RUT sin DV — formato ENDO en MAEEDO es "12345678-9", buscamos por prefijo
        rut_base = rut_clean[:-1] if len(rut_clean) >= 8 else rut_clean
        # Fecha desde
        from datetime import datetime as _dt, timedelta as _td
        fecha_desde = (_dt.now().date() - _td(days=dias)).isoformat()

        # TIDOs relevantes para "saldo pendiente de retiro".
        # Daniel 2026-06-03: EXCLUIR guías de despacho (GDV/GDP) de la búsqueda
        # por RUT — "estas figuran siempre con saldo" y ensucian el resultado.
        # El operador retira contra factura/boleta/nota de venta, no contra guía.
        tidos_in = "','".join(("FCV", "BLV", "NVI", "NVV", "VD", "WEB"))

        # ZZ SKUs típicos de despacho/retiro (ver erp_engine / transporte)
        # Para no duplicar lógica, usamos el patrón LIKE 'ZZ%' que cubre todos.
        sql = f"""
            SELECT TOP 100
                e.IDMAEEDO,
                LTRIM(RTRIM(e.TIDO)) AS TIDO,
                LTRIM(RTRIM(e.NUDO)) AS NUDO,
                LTRIM(RTRIM(e.ENDO)) AS ENDO,
                e.FEEMDO,
                e.VANEDO,
                e.VABRDO,
                LTRIM(RTRIM(COALESCE(e.ESPGDO, ''))) AS ESPGDO,
                LTRIM(RTRIM(COALESCE(e.ESDO,   ''))) AS ESDO,
                LTRIM(RTRIM(COALESCE(
                    NULLIF(LTRIM(RTRIM(en.NOKOENAMP)), ''),
                    NULLIF(LTRIM(RTRIM(en.NOKOEN)),    ''),
                    NULLIF(LTRIM(RTRIM(e.SUENDO)),     ''),
                    'Consumidor final'
                ))) AS NOMBRE,
                (SELECT COALESCE(SUM(
                           CASE WHEN d.CAPRCO1 - COALESCE(d.CAPRAD1, 0) > 0
                                THEN d.CAPRCO1 - COALESCE(d.CAPRAD1, 0)
                                ELSE 0 END), 0)
                   FROM MAEDDO d
                  WHERE d.IDMAEEDO = e.IDMAEEDO
                    AND UPPER(LTRIM(RTRIM(d.KOPRCT))) LIKE 'ZZ%%') AS saldo_zz,
                /* 🔧 FIX Daniel 2026-05-24: saldo REAL de productos (no ZZ)
                   con fórmula oficial Random. Antes solo se miraba el saldo
                   de servicios ZZ → docs aparecían "con saldo" aunque todos
                   los productos estuvieran ya despachados (incoherente). */
                (SELECT COALESCE(SUM(
                           CASE WHEN UPPER(LTRIM(RTRIM(COALESCE(d3.ESLIDO,'')))) NOT IN ('C','T','TOTAL','CERRADO','DESPACHADO')
                                 AND (d3.CAPRCO1 - COALESCE(d3.CAPRAD1,0) - COALESCE(d3.CAPREX1,0) - COALESCE(d3.CAPRNC1,0)) > 0
                                THEN d3.CAPRCO1 - COALESCE(d3.CAPRAD1,0) - COALESCE(d3.CAPREX1,0) - COALESCE(d3.CAPRNC1,0)
                                ELSE 0 END), 0)
                   FROM MAEDDO d3
                  WHERE d3.IDMAEEDO = e.IDMAEEDO
                    AND UPPER(LTRIM(RTRIM(d3.KOPRCT))) NOT LIKE 'ZZ%%') AS saldo_real_unidades,
                (SELECT COUNT(*) FROM MAEDDO d2
                  WHERE d2.IDMAEEDO = e.IDMAEEDO) AS n_lineas
            FROM MAEEDO e
            LEFT JOIN MAEEN en ON LTRIM(RTRIM(en.RTEN)) =
                  CASE
                    WHEN CHARINDEX('-', e.ENDO) > 0
                      THEN LTRIM(RTRIM(SUBSTRING(e.ENDO, 1, CHARINDEX('-', e.ENDO) - 1)))
                    ELSE LTRIM(RTRIM(COALESCE(e.ENDO, '')))
                  END
            WHERE e.ENDO LIKE %s
              AND LTRIM(RTRIM(e.TIDO)) IN ('{tidos_in}')
              AND e.FEEMDO >= %s
              AND (e.ESDO IS NULL OR LTRIM(RTRIM(e.ESDO)) <> 'NULO')
            ORDER BY e.FEEMDO DESC
        """
        try:
            rows = _random_sql_query(sql, (f"{rut_base}%", fecha_desde), max_rows=100) or []
        except Exception as exc:
            print(f"[pickup-saldo-erp] error: {exc}", flush=True)
            return jsonify({"ok": True, "docs": [], "error": "No se pudo consultar el ERP"})

        if not rows:
            return jsonify({"ok": True, "docs": [], "total": 0})

        # Cruzar con retiros ya asociados (en pickup_requests por document_number
        # o en pickup_request_docs)
        nudos_raw = [str(r.get("NUDO") or "").strip() for r in rows]
        nudos_raw = [n for n in nudos_raw if n]
        ya_asociados = set()
        if nudos_raw:
            # Buscamos también las variantes "limpias" (sin ceros a la izquierda y sin VD/WEB prefix)
            placeholders = ",".join(["%s"] * len(nudos_raw))
            try:
                rows_pr = mysql_fetchall(
                    f"SELECT document_type, document_number FROM `{REQ}` "
                    f"WHERE document_number IN ({placeholders}) "
                    f"  AND status NOT IN ('rechazada','cerrada','fallida')",
                    tuple(nudos_raw)
                ) or []
                for r in rows_pr:
                    ya_asociados.add(f"{(r.get('document_type') or '').upper()}|{r.get('document_number')}")
                rows_prd = mysql_fetchall(
                    f"SELECT document_type, document_number FROM pickup_request_docs "
                    f"WHERE document_number IN ({placeholders})",
                    tuple(nudos_raw)
                ) or []
                for r in rows_prd:
                    ya_asociados.add(f"{(r.get('document_type') or '').upper()}|{r.get('document_number')}")
            except Exception as _e2:
                pass

        # Construir lista de respuesta
        out = []
        n_con_saldo = 0
        n_sin_saldo = 0
        for r in rows:
            nudo_raw = (r.get("NUDO") or "").strip()
            tido_raw = (r.get("TIDO") or "").strip()
            # Convertir NUDO con prefijo VD/WEB a su tipo display
            tido_display = tido_raw
            if tido_raw == "NVV" and nudo_raw.startswith("VD"):
                tido_display = "VD"
                nudo_display = nudo_raw[2:].lstrip("0") or "0"
            elif tido_raw == "NVV" and nudo_raw.startswith("WEB"):
                tido_display = "WEB"
                nudo_display = nudo_raw[3:].lstrip("0") or "0"
            else:
                nudo_display = nudo_raw.lstrip("0") or "0"

            fe = r.get("FEEMDO")
            saldo_zz = float(r.get("saldo_zz") or 0)
            saldo_real_unidades = float(r.get("saldo_real_unidades") or 0)
            # 🔧 FIX Daniel 2026-05-24: "tiene_saldo" basado en productos
            # reales pendientes (líneas no-ZZ con saldo > 0), no en servicios.
            n_lineas = int(r.get("n_lineas") or 0)
            key = f"{tido_display}|{nudo_display}"
            ya_tiene_retiro = key in ya_asociados

            tiene_saldo = saldo_real_unidades > 0
            if tiene_saldo: n_con_saldo += 1
            else:           n_sin_saldo += 1

            # Filtro Daniel 2026-05-23: si el operador pidió solo_con_saldo=1,
            # saltamos los docs ya despachados (no retirables).
            if solo_con_saldo and not tiene_saldo:
                continue

            out.append({
                "tido":           tido_raw,
                "nudo":           nudo_raw,
                "tido_display":   tido_display,
                "nudo_display":   nudo_display,
                "fecha":          fe.strftime("%d/%m/%Y") if fe else "",
                "fecha_iso":      fe.strftime("%Y-%m-%d") if fe else "",
                "cliente":        (r.get("NOMBRE") or "").strip().title(),
                "rut":            rut_clean,
                "total":          float(r.get("VABRDO") or 0),
                "neto":           float(r.get("VANEDO") or 0),
                "saldo_zz":              saldo_zz,
                "saldo_real_unidades":   saldo_real_unidades,
                "tiene_saldo":           tiene_saldo,
                "estado_pago":           (r.get("ESPGDO") or "").strip(),
                "n_lineas":              n_lineas,
                "ya_tiene_retiro":       ya_tiene_retiro,
            })

        # Hint inteligente para el operador (Daniel: "tiene que enseñar")
        hint = None
        if not out and not n_con_saldo and not n_sin_saldo:
            hint = (
                "Este cliente no tiene documentos emitidos en los últimos "
                f"{dias} días. Verifica que el RUT sea correcto o amplía la ventana."
            )
        elif solo_con_saldo and not n_con_saldo and n_sin_saldo > 0:
            hint = (
                f"El cliente tiene {n_sin_saldo} documento(s) emitido(s) pero "
                "todos están ya despachados (sin saldo). Verifica con el cliente "
                "qué viene a retirar — quizás ingresó mal el documento."
            )
        elif solo_con_saldo and n_con_saldo == 1:
            hint = (
                "Solo hay 1 documento con saldo pendiente. Asócialo y avanza al paso siguiente."
            )

        _payload_saldo = {
            "ok": True,
            "docs": out,
            "total": len(out),
            "solo_con_saldo": solo_con_saldo,
            "resumen": {
                "con_saldo": n_con_saldo,
                "sin_saldo": n_sin_saldo,
                "total":     n_con_saldo + n_sin_saldo,
            },
            "hint": hint,
        }
        # ⚡ PERF: persistir en cache 60s para hits futuros
        try:
            _SALDO_CACHE[_cache_key] = (_payload_saldo, _time_saldo.time())
            # Limpieza simple: si crece demasiado, purgamos viejos
            if len(_SALDO_CACHE) > 200:
                _cutoff = _time_saldo.time() - (_SALDO_TTL * 3)
                for _k in list(_SALDO_CACHE.keys()):
                    if _SALDO_CACHE[_k][1] < _cutoff:
                        _SALDO_CACHE.pop(_k, None)
        except Exception:
            pass
        return jsonify(_payload_saldo)

    # ══════════════════════════════════════════════════════════════════
    #  BÚSQUEDA AVANZADA ERP — modal estilo mantenciones (Daniel 2026-05-23)
    #  Daniel: "me gustaría que se viera como en mantenciones un modal y
    #  dos motores de búsqueda. uno que busca por cada documento y llenando
    #  una lista de productos. Y otro con un motor de búsqueda que va a
    #  traer la totalidad de los productos. Trabaja con el mismo motor de
    #  búsqueda de mantenciones."
    #
    #  Endpoint UNIFICADO que detecta automáticamente el tipo de búsqueda:
    #    • 7-9 dígitos        → RUT (todos los docs del cliente)
    #    • 1-7 dígitos        → Número de documento
    #    • Texto              → Razón social
    #
    #  Reusa el patrón SQL Server directo (_random_sql_query) — el mismo
    #  motor que mantenciones/buscar-erp-sql, sin duplicar lógica.
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/api/buscar-erp", methods=["POST"])
    @require_permission("view")
    def pickup_buscar_erp():
        """Búsqueda inteligente en ERP Random (RUT / nombre / número).

        Detecta automáticamente:
          • 7-9 dígitos puros  → RUT
          • 1-7 dígitos        → Número doc
          • Texto              → Razón social

        Para cada documento devuelve metadata + saldo ZZ disponible (para
        que el frontend marque docs ya despachados con candado).

        Body JSON: {q: "<string>"}
        Response: {ok, modo, documentos: [...], count, query}

        FIX Daniel 2026-05-23: reescrito para usar el MISMO patrón 2-pasos
        que mant_buscar_erp_sql (que sí funciona en producción):
          1) Query mínima a MAEEDO sin JOINs (rápida, sin duplicados).
          2) Enriquecimiento del NOMBRE/saldo en una 2ª query a MAEEN/MAEDDO.
        Eliminado el LEFT JOIN MAEEN con CASE/CHARINDEX que multiplicaba filas
        cuando un RUT tenía múltiples sucursales en MAEEN.
        """
        d = request.get_json(silent=True) or {}
        q = (d.get("q") or "").strip()
        if len(q) < 3:
            return jsonify({"ok": False, "error": "Mínimo 3 caracteres", "documentos": []}), 400

        try:
            from app import _random_sql_query, _random_sql_pool
        except ImportError:
            return jsonify({"ok": True, "documentos": [],
                            "error": "Motor ERP no disponible", "sin_conexion": True})

        pool = _random_sql_pool()
        if pool is None:
            return jsonify({"ok": True, "documentos": [], "sin_conexion": True,
                            "error": "ERP Random no configurado en este entorno."})

        # Normalizar query
        q_clean   = q.replace(".", "").replace(" ", "").replace("-", "").upper()
        is_digits = q_clean.isdigit()
        # TIDOs aceptados (ventas que pueden requerir retiro)
        tidos_in  = "','".join(("FCV", "BLV", "NVI", "NVV", "GDV", "GDP", "VD", "WEB"))

        docs = []
        modo = ""
        try:
            # ── Modo RUT (7-9 dígitos) ─────────────────────────────
            #
            # ENDO en MAEEDO viene como "65206047-K" (RUT + DV separados con guión).
            # Buscamos por prefijo del RUT base (sin DV) para tolerar variaciones.
            # Si len >= 8, asumimos que el último char es el DV y lo quitamos.
            if is_digits and 7 <= len(q_clean) <= 9:
                modo = "rut"
                rut_base = q_clean[:-1] if len(q_clean) >= 8 else q_clean
                docs = _random_sql_query(f"""
                    SELECT TOP 80
                        e.IDMAEEDO,
                        LTRIM(RTRIM(e.TIDO)) AS TIDO,
                        LTRIM(RTRIM(e.NUDO)) AS NUDO,
                        LTRIM(RTRIM(e.ENDO)) AS ENDO,
                        LTRIM(RTRIM(COALESCE(e.SUENDO,''))) AS SUENDO,
                        e.FEEMDO, e.VANEDO, e.VABRDO,
                        LTRIM(RTRIM(COALESCE(e.ESPGDO,''))) AS ESPGDO
                    FROM MAEEDO e
                    WHERE (e.ENDO LIKE %s OR e.ENDO LIKE %s)
                      AND LTRIM(RTRIM(e.TIDO)) IN ('{tidos_in}')
                      AND (e.ESDO IS NULL OR LTRIM(RTRIM(e.ESDO)) <> 'NULO')
                    ORDER BY e.FEEMDO DESC
                """, (f"{rut_base}%", f"%{q_clean}%"), max_rows=80) or []

            # ── Modo Número documento (1-7 dígitos) ────────────────
            if not docs and is_digits and 1 <= len(q_clean) <= 7:
                modo = "numero"
                # NUDO en MAEEDO viene padded a 10 chars o con prefijo VD/WEB
                nudo_padded = q_clean.zfill(10)
                nudo_vd     = f"VD{q_clean.zfill(8)}"
                nudo_web    = f"WEB{q_clean.zfill(7)}"
                docs = _random_sql_query(f"""
                    SELECT TOP 30
                        e.IDMAEEDO,
                        LTRIM(RTRIM(e.TIDO)) AS TIDO,
                        LTRIM(RTRIM(e.NUDO)) AS NUDO,
                        LTRIM(RTRIM(e.ENDO)) AS ENDO,
                        LTRIM(RTRIM(COALESCE(e.SUENDO,''))) AS SUENDO,
                        e.FEEMDO, e.VANEDO, e.VABRDO,
                        LTRIM(RTRIM(COALESCE(e.ESPGDO,''))) AS ESPGDO
                    FROM MAEEDO e
                    WHERE e.NUDO IN (%s, %s, %s)
                      AND LTRIM(RTRIM(e.TIDO)) IN ('{tidos_in}')
                      AND (e.ESDO IS NULL OR LTRIM(RTRIM(e.ESDO)) <> 'NULO')
                    ORDER BY e.FEEMDO DESC
                """, (nudo_padded, nudo_vd, nudo_web), max_rows=30) or []

            # ── Modo Razón social (texto) ──────────────────────────
            # Estrategia 2-pasos:
            #   3a) RUTs candidatos en MAEEN por nombre (índice por NOKOEN)
            #   3b) Docs cuyo ENDO matchea esos RUTs (índice por ENDO)
            if not docs and not is_digits:
                modo = "nombre"
                q_like = f"%{q.upper()}%"
                ruts = _random_sql_query("""
                    SELECT TOP 20 LTRIM(RTRIM(RTEN)) AS rut,
                                  LTRIM(RTRIM(COALESCE(NOKOENAMP, NOKOEN, ''))) AS razon
                      FROM MAEEN
                     WHERE (UPPER(NOKOEN) LIKE %s OR UPPER(COALESCE(NOKOENAMP,'')) LIKE %s)
                       AND TIEN IN ('C','A')
                """, (q_like, q_like)) or []
                if ruts:
                    rut_map = {r['rut']: r['razon'] for r in ruts if r.get('rut')}
                    if rut_map:
                        like_clauses = " OR ".join(["e.ENDO LIKE %s"] * len(rut_map))
                        params = tuple(f"{rk}%" for rk in rut_map.keys())
                        docs = _random_sql_query(f"""
                            SELECT TOP 60
                                e.IDMAEEDO,
                                LTRIM(RTRIM(e.TIDO)) AS TIDO,
                                LTRIM(RTRIM(e.NUDO)) AS NUDO,
                                LTRIM(RTRIM(e.ENDO)) AS ENDO,
                                LTRIM(RTRIM(COALESCE(e.SUENDO,''))) AS SUENDO,
                                e.FEEMDO, e.VANEDO, e.VABRDO,
                                LTRIM(RTRIM(COALESCE(e.ESPGDO,''))) AS ESPGDO
                            FROM MAEEDO e
                            WHERE ({like_clauses})
                              AND LTRIM(RTRIM(e.TIDO)) IN ('{tidos_in}')
                              AND (e.ESDO IS NULL OR LTRIM(RTRIM(e.ESDO)) <> 'NULO')
                            ORDER BY e.FEEMDO DESC
                        """, params, max_rows=60) or []

            # ── Deduplicar por IDMAEEDO (defensa por si quedó duplicado) ──
            seen_ids = set()
            unique_docs = []
            for r in docs:
                idm = r.get("IDMAEEDO")
                if idm in seen_ids:
                    continue
                seen_ids.add(idm)
                unique_docs.append(r)
            docs = unique_docs

            # ── Paso 2: enriquecer cada doc con NOMBRE + saldos + n_lineas ──
            # Esto evita el LEFT JOIN problemático del SELECT principal:
            #   • MAEEN puede tener N entidades con el mismo RTEN (multiplica)
            #   • Subqueries SCALAR en SELECT son costosas y propensas a timeout
            #
            # Acá hacemos 1 query agregada para TODOS los IDMAEEDO de una vez.
            #
            # 🔧 FIX Daniel 2026-05-24 (BUG GRAVE): el listado decía "CON SALDO"
            # pero al expandir TODAS las líneas aparecían "sin saldo". Causa:
            # antes solo sumábamos saldo de líneas ZZ (servicios de despacho),
            # NO de los productos reales. Ahora calculamos el SALDO REAL DE
            # PRODUCTOS (líneas no-ZZ) usando la fórmula oficial Random:
            #     saldo = CAPRCO1 - CAPRAD1 - CAPREX1 - CAPRNC1
            # Si saldo_real > 0 → "CON SALDO" (hay productos por retirar).
            # Si saldo_real = 0 y todas las líneas no-ZZ están despachadas
            # → "SIN SALDO" (gris, doc cerrado).
            # Mantenemos `saldo_zz` por compat (lo usaba el frontend para info).
            if docs:
                idmaeedos = [r.get("IDMAEEDO") for r in docs if r.get("IDMAEEDO") is not None]
                if idmaeedos:
                    placeholders = ",".join(["%s"] * len(idmaeedos))
                    # Saldo ZZ + saldo REAL de productos + n_lineas por doc
                    try:
                        saldo_rows = _random_sql_query(f"""
                            SELECT d.IDMAEEDO,
                                   COALESCE(SUM(CASE
                                       WHEN UPPER(LTRIM(RTRIM(d.KOPRCT))) LIKE 'ZZ%%'
                                            AND (d.CAPRCO1 - COALESCE(d.CAPRAD1,0)) > 0
                                       THEN d.CAPRCO1 - COALESCE(d.CAPRAD1,0)
                                       ELSE 0
                                   END), 0) AS saldo_zz,
                                   COALESCE(SUM(CASE
                                       WHEN UPPER(LTRIM(RTRIM(d.KOPRCT))) NOT LIKE 'ZZ%%'
                                            AND UPPER(LTRIM(RTRIM(COALESCE(d.ESLIDO,'')))) NOT IN ('C','T','TOTAL','CERRADO','DESPACHADO')
                                            AND (d.CAPRCO1 - COALESCE(d.CAPRAD1,0) - COALESCE(d.CAPREX1,0) - COALESCE(d.CAPRNC1,0)) > 0
                                       THEN (d.CAPRCO1 - COALESCE(d.CAPRAD1,0) - COALESCE(d.CAPREX1,0) - COALESCE(d.CAPRNC1,0))
                                       ELSE 0
                                   END), 0) AS saldo_real_unidades,
                                   COUNT(*) AS n_lineas
                              FROM MAEDDO d
                             WHERE d.IDMAEEDO IN ({placeholders})
                             GROUP BY d.IDMAEEDO
                        """, tuple(idmaeedos), max_rows=len(idmaeedos)) or []
                        sm = {s.get("IDMAEEDO"): s for s in saldo_rows}
                    except Exception as e:
                        print(f"[pickup-buscar-erp] saldo lookup falló: {e}", flush=True)
                        sm = {}
                    for r in docs:
                        s = sm.get(r.get("IDMAEEDO")) or {}
                        r["saldo_zz"] = s.get("saldo_zz") or 0
                        r["saldo_real_unidades"] = s.get("saldo_real_unidades") or 0
                        r["n_lineas"] = s.get("n_lineas") or 0

                # Enriquecer NOMBRE por RUT desde MAEEN (1 query batch)
                # Modo nombre ya tiene rut_map en scope, pero RUT/numero NO.
                # Extraemos los RUTs únicos de los ENDO encontrados.
                ruts_needed = set()
                for r in docs:
                    endo = (r.get("ENDO") or "").strip()
                    if not endo:
                        continue
                    rut_clean = endo.split("-")[0] if "-" in endo else endo
                    if rut_clean and len(rut_clean) >= 4:
                        ruts_needed.add(rut_clean)
                ruts_needed.discard("")

                nombre_map = {}
                if ruts_needed:
                    rph = ",".join(["%s"] * len(ruts_needed))
                    try:
                        nm_rows = _random_sql_query(f"""
                            SELECT LTRIM(RTRIM(RTEN)) AS rut,
                                   LTRIM(RTRIM(COALESCE(
                                       NULLIF(LTRIM(RTRIM(NOKOENAMP)),''),
                                       NOKOEN, ''
                                   ))) AS razon
                              FROM MAEEN
                             WHERE LTRIM(RTRIM(RTEN)) IN ({rph})
                        """, tuple(ruts_needed), max_rows=len(ruts_needed) * 4) or []
                        # Si hay duplicados, conservar el último no vacío
                        for nm in nm_rows:
                            rut = (nm.get("rut") or "").strip()
                            razon = (nm.get("razon") or "").strip()
                            if rut and razon and not nombre_map.get(rut):
                                nombre_map[rut] = razon
                    except Exception as e:
                        print(f"[pickup-buscar-erp] nombre lookup falló: {e}", flush=True)

                for r in docs:
                    endo = (r.get("ENDO") or "").strip()
                    rut_clean = endo.split("-")[0] if "-" in endo else endo
                    nombre = nombre_map.get(rut_clean, "") or (r.get("SUENDO") or "").strip()
                    if not nombre:
                        nombre = "Consumidor final"
                    r["NOMBRE"] = nombre

        except PermissionError as pe:
            return jsonify({"ok": False,
                            "error": f"Bloqueado por seguridad: {pe}",
                            "documentos": []}), 403
        except Exception as exc:
            import traceback
            print(f"[pickup-buscar-erp] error inesperado: {exc}", flush=True)
            traceback.print_exc()
            return jsonify({"ok": False,
                            "error": f"Error consultando ERP: {str(exc)[:200]}",
                            "documentos": []})

        # ── Identificar docs ya asociados a algún retiro (para mostrar candado) ──
        # FIX 2026-05-23: pickup_request_docs guarda document_number en formato
        # display (sin padding de ceros del ERP), no el NUDO raw. Antes
        # comparábamos NUDO padded (`0000000123`) con document_number (`123`)
        # → NUNCA matcheaba y la marca de candado no aparecía.
        # Calculamos el display IGUAL que la sección de formateo de abajo.
        ya_asociados = set()
        try:
            nudos_display = []
            for r in docs:
                nudo_raw = (r.get("NUDO") or "").strip()
                tido_raw = (r.get("TIDO") or "").strip()
                if not nudo_raw:
                    continue
                if tido_raw == "NVV" and nudo_raw.startswith("VD"):
                    nd = nudo_raw[2:].lstrip("0") or "0"
                elif tido_raw == "NVV" and nudo_raw.startswith("WEB"):
                    nd = nudo_raw[3:].lstrip("0") or "0"
                else:
                    nd = nudo_raw.lstrip("0") or "0"
                if nd:
                    nudos_display.append(nd)
            if nudos_display:
                placeholders = ",".join(["%s"] * len(nudos_display))
                rows_prd = mysql_fetchall(
                    f"SELECT DISTINCT document_type, document_number "
                    f"FROM pickup_request_docs WHERE document_number IN ({placeholders})",
                    tuple(nudos_display)
                ) or []
                for r in rows_prd:
                    ya_asociados.add(f"{(r.get('document_type') or '').upper()}|"
                                     f"{(r.get('document_number') or '').strip()}")
        except Exception as e:
            print(f"[pickup-buscar-erp] ya_asociados fallback: {e}", flush=True)

        # ── Formatear respuesta ────────────────────────────────────
        out = []
        for r in docs:
            nudo_raw = (r.get("NUDO") or "").strip()
            tido_raw = (r.get("TIDO") or "").strip()
            # Detectar prefijo VD/WEB en NUDO
            tido_display = tido_raw
            if tido_raw == "NVV" and nudo_raw.startswith("VD"):
                tido_display = "VD"
                nudo_display = nudo_raw[2:].lstrip("0") or "0"
            elif tido_raw == "NVV" and nudo_raw.startswith("WEB"):
                tido_display = "WEB"
                nudo_display = nudo_raw[3:].lstrip("0") or "0"
            else:
                nudo_display = nudo_raw.lstrip("0") or "0"

            fe = r.get("FEEMDO")
            saldo_zz = float(r.get("saldo_zz") or 0)
            saldo_real_unidades = float(r.get("saldo_real_unidades") or 0)
            endo = (r.get("ENDO") or "").strip()
            rut_clean = endo.split("-")[0] if "-" in endo else endo
            key = f"{tido_display}|{nudo_display}"

            # 🔧 FIX Daniel 2026-05-24: "CON SALDO" cuando hay productos
            # reales pendientes (líneas no-ZZ con saldo > 0). El listado
            # ahora coincide con el detalle expandido — antes daba "CON
            # SALDO" por servicios ZZ pero todas las líneas mostraban
            # "sin saldo" → inconsistencia que confundía al operador.
            tiene_saldo_real = saldo_real_unidades > 0

            out.append({
                "idmaeedo":     r.get("IDMAEEDO"),
                "tido":         tido_raw,
                "nudo":         nudo_raw,
                "tido_display": tido_display,
                "nudo_display": nudo_display,
                "rut":          rut_clean,
                "razon_social": (r.get("NOMBRE") or "").strip().title(),
                "fecha":        fe.strftime("%d/%m/%Y") if fe else "",
                "fecha_iso":    fe.strftime("%Y-%m-%d") if fe else "",
                "valor_neto":   float(r.get("VANEDO") or 0),
                "valor_total":  float(r.get("VABRDO") or 0),
                "estado_pago":  (r.get("ESPGDO") or "").strip(),
                "saldo_zz":             saldo_zz,
                "saldo_real_unidades":  saldo_real_unidades,
                "tiene_saldo":          tiene_saldo_real,
                "n_lineas":             int(r.get("n_lineas") or 0),
                "ya_tiene_retiro":      key in ya_asociados,
            })

        return jsonify({
            "ok":         True,
            "modo":       modo,
            "documentos": out,
            "count":      len(out),
            "query":      q,
        })

    # ══════════════════════════════════════════════════════════════════
    #  ACTUALIZAR DATOS CLIENTE/CONTACTO DEL RETIRO (Daniel 2026-05-23)
    #  Daniel: "el nombre del cliente o la razón social cambie con la
    #  asignación del producto. Entonces, no te dejes llevar mucho por
    #  bloquear lo que dice el cliente"
    #  Y también: "es posible que me consulte si los datos que declaró
    #  el cliente se podrán asignar como contacto de retiro y dejar como
    #  oficial el dueño del documento"
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/<int:rid>/customer", methods=["POST"])
    @require_permission("retiros")
    def pickup_actualizar_cliente(rid):
        """Actualiza nombre/RUT del cliente y, opcionalmente, copia los
        datos actuales del cliente como persona-que-retira (oficial = dueño
        del doc, contacto/retira = quien lo declaró originalmente).

        Body JSON:
          customer_name:               nuevo nombre oficial (opcional)
          customer_rut:                nuevo RUT oficial (opcional)
          usar_cliente_como_contacto:  bool — si True copia
                                        customer_name → pickup_person_name
                                        customer_rut  → pickup_person_rut
        """
        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            return jsonify({"ok": False, "error": "Retiro no existe"}), 404

        body = request.get_json(silent=True) or {}
        new_name = (body.get("customer_name") or "").strip()[:200]
        new_rut  = (body.get("customer_rut")  or "").strip()[:30]
        usar_como_contacto = bool(body.get("usar_cliente_como_contacto"))

        sets, params = [], []
        old_name = (req.get("customer_name") or "").strip()
        old_rut  = (req.get("customer_rut")  or "").strip()

        if new_name and new_name != old_name:
            sets.append("customer_name=%s"); params.append(new_name)
        if new_rut and new_rut != old_rut:
            sets.append("customer_rut=%s"); params.append(new_rut)

        if usar_como_contacto:
            # Copiar datos cliente → persona-retira (pero solo si tenemos algo)
            target_name = new_name or old_name
            target_rut  = new_rut  or old_rut
            if target_name:
                sets.append("pickup_person_name=%s"); params.append(target_name)
            if target_rut:
                sets.append("pickup_person_rut=%s");  params.append(target_rut)

        if not sets:
            return jsonify({"ok": True, "no_changes": True})

        params.append(rid)
        try:
            mysql_execute(
                f"UPDATE `{REQ}` SET {', '.join(sets)} WHERE id=%s",
                tuple(params)
            )
        except Exception as exc:
            return jsonify({"ok": False,
                            "error": f"Error al actualizar: {str(exc)[:200]}"}), 500

        detalle_parts = []
        if new_name and new_name != old_name:
            detalle_parts.append(f"nombre: {old_name} → {new_name}")
        if new_rut and new_rut != old_rut:
            detalle_parts.append(f"RUT: {old_rut} → {new_rut}")
        if usar_como_contacto:
            detalle_parts.append("datos cliente copiados a persona-retira")
        log_event(rid, "cliente_actualizado", None, None,
                  " · ".join(detalle_parts) or "sin cambios",
                  "interno")

        # Re-leer para devolver al frontend
        updated = mysql_fetchone(
            f"SELECT customer_name, customer_rut, "
            f"       pickup_person_name, pickup_person_rut "
            f"FROM `{REQ}` WHERE id=%s",
            (rid,)
        ) or {}
        return jsonify({
            "ok": True,
            "customer_name":      updated.get("customer_name") or "",
            "customer_rut":       updated.get("customer_rut")  or "",
            "pickup_person_name": updated.get("pickup_person_name") or "",
            "pickup_person_rut":  updated.get("pickup_person_rut")  or "",
        })

    # ══════════════════════════════════════════════════════════════════
    #  INLINE EDIT — Auto-save de campos de la ficha (Daniel 2026-05-23)
    #  Objetivo: la ficha debe ser modificable "a medida que se agrega
    #  información" (palabras de Daniel). En vez de un form gigante con
    #  botón "Guardar", cada campo se autoguarda con debounce 800ms.
    #
    #  Cliente envía: PATCH /retiros/<rid>/field {field: "X", value: "Y"}
    #  Server valida whitelist + max_length, normaliza, persiste, loga.
    # ══════════════════════════════════════════════════════════════════
    def _pickup_now_chile_hms():
        """Helper: hora actual Chile como HH:MM:SS — para feedback "Guardado a las HH:MM:SS"."""
        try:
            from zoneinfo import ZoneInfo as _ZI_local
            return datetime.now(_ZI_local("America/Santiago")).strftime("%H:%M:%S")
        except Exception:
            return datetime.now().strftime("%H:%M:%S")

    # Whitelist de campos editables inline desde la ficha.
    # tuple: (max_length, label_humano para logs)
    # Solo se permiten columnas que EXISTEN en pickup_requests
    # (verificado contra el CREATE TABLE en app.py línea ~2252).
    _PICKUP_INLINE_FIELDS = {
        "customer_name":         (200, "Razón social cliente"),
        "customer_rut":          (30,  "RUT cliente"),
        "contact_name":          (160, "Nombre contacto"),
        "contact_email":         (180, "Email contacto"),
        "contact_phone":         (60,  "Teléfono contacto"),
        "pickup_person_name":    (160, "Persona que retira"),
        "pickup_person_rut":     (30,  "RUT persona que retira"),
        "pickup_person_phone":   (60,  "Teléfono persona que retira"),
        "pickup_person_relation":(40,  "Relación persona que retira"),
        "observations":          (10000, "Observaciones del cliente"),
        "internal_notes":        (10000, "Notas internas operador"),
        # Daniel 2026-05-23: emails extra separados por coma (CC del envío)
        "extra_emails":          (800, "Emails adicionales (CC)"),
    }

    # Helper compartido: junta TODOS los emails de envío para un retiro.
    # Daniel 2026-05-23: "si el documento trae un correo, se puede enviar
    # tanto al correo del cliente que declaró como al del documento.
    # Mientras más se vayan agregando correo, mejor."
    #
    # Devuelve lista de emails únicos en MAYÚSCULAS deduplicadas:
    #   1) req["contact_email"] (cliente declarado)
    #   2) parseados de req["extra_emails"] (coma-separados, agregados manualmente)
    #   3) emails de cada doc asociado (pickup_request_docs.email_cliente_erp)
    def _get_pickup_all_emails(rid_or_req):
        """Devuelve lista única de emails de envío para un retiro.

        Args: rid (int) o req (dict ya cargado).
        Returns: [str] lista de emails únicos en minúsculas.
        """
        if isinstance(rid_or_req, dict):
            req = rid_or_req
            rid = req.get("id")
        else:
            rid = int(rid_or_req)
            req = mysql_fetchone(
                f"SELECT id, contact_email, extra_emails FROM `{REQ}` WHERE id=%s",
                (rid,)
            ) or {}

        emails = set()

        def _add(e):
            if not e: return
            e = str(e).strip().lower()
            if "@" in e and len(e) >= 6 and len(e) <= 180:
                emails.add(e)

        # 1) Email principal del contacto declarado
        _add(req.get("contact_email"))

        # 2) extra_emails — separados por coma/espacio/punto-coma
        extras = req.get("extra_emails") or ""
        import re as _re_emails
        for e in _re_emails.split(r"[,;\s]+", extras):
            _add(e)

        # 3) Emails capturados desde el ERP por doc asociado
        try:
            doc_emails = mysql_fetchall(
                "SELECT DISTINCT email_cliente_erp FROM pickup_request_docs "
                "WHERE request_id=%s AND email_cliente_erp IS NOT NULL "
                "AND email_cliente_erp <> ''",
                (rid,)
            ) or []
            for d in doc_emails:
                _add(d.get("email_cliente_erp"))
        except Exception:
            pass  # columna puede no existir aún (migration pendiente)

        return sorted(emails)

    # Endpoint para que la UI sepa qué emails están activos + sugerencias
    @app.route("/retiros/<int:rid>/emails-info", methods=["GET"])
    @require_permission("view")
    def pickup_emails_info(rid):
        """Devuelve estado de emails de un retiro:
        - principal: contact_email del cliente declarado
        - extra: lista parseada de extra_emails (CC)
        - sugerencias: emails de docs ERP asociados aún NO incluidos
        - total_efectivo: lista única de TODOS los que recibirán email
        """
        req = mysql_fetchone(
            f"SELECT id, contact_email, extra_emails FROM `{REQ}` WHERE id=%s",
            (rid,)
        )
        if not req:
            return jsonify({"ok": False, "error": "Retiro no encontrado"}), 404

        principal = (req.get("contact_email") or "").strip().lower()
        extras_raw = req.get("extra_emails") or ""
        import re as _re_em
        extras_list = [e.strip().lower() for e in _re_em.split(r"[,;\s]+", extras_raw) if e.strip()]
        extras_list = sorted(set(extras_list))

        # Emails desde docs ERP
        sugerencias_set = set()
        try:
            docs_emails = mysql_fetchall(
                "SELECT DISTINCT email_cliente_erp, document_type, document_number "
                "FROM pickup_request_docs "
                "WHERE request_id=%s AND email_cliente_erp IS NOT NULL "
                "  AND email_cliente_erp <> ''",
                (rid,)
            ) or []
            for d in docs_emails:
                em = (d.get("email_cliente_erp") or "").strip().lower()
                if em and "@" in em and em != principal and em not in extras_list:
                    sugerencias_set.add(em)
        except Exception:
            pass

        total_efectivo = _get_pickup_all_emails(req)

        return jsonify({
            "ok": True,
            "principal":     principal,
            "extra":         extras_list,
            "sugerencias":   sorted(sugerencias_set),
            "total_efectivo": total_efectivo,
            "count":         len(total_efectivo),
        })

    # Wrapper de envío que multiplica el mensaje a todos los destinos
    def _send_pickup_email_multi(rid_or_req, subject, html, attachments=None):
        """Envía email a TODOS los destinatarios del retiro (multi-email).
        Devuelve dict {sent: [emails], failed: [emails]}.

        attachments (Daniel 2026-06-17): lista de adjuntos (filename, bytes,
        mimetype) — ej. el .ics del retiro. Viaja transparente vía **kwargs hasta
        _send_ilus_email_real, que ya soporta adjuntos (SMTP multipart / Resend
        base64). Si es None, el flujo es idéntico al de antes.
        """
        emails = _get_pickup_all_emails(rid_or_req)
        sent, failed = [], []
        for e in emails:
            try:
                # Juan Daniel 2026-06-05: pasar modulo="retiros" para que el correo
                # se rija por la LLAVE DE PASO de Retiros (no la de "general", que era
                # el default y bloqueaba el envío aunque Retiros estuviera abierta).
                _kw = {"evento": "pickup_created", "modulo": "retiros"}
                if attachments:
                    _kw["attachments"] = attachments
                ok = _send_ilus_email(e, subject, html, **_kw)
                if ok:
                    sent.append(e)
                else:
                    failed.append(e)
            except Exception as exc:
                print(f"[pickup-multi-email] fallo {e}: {exc}", flush=True)
                failed.append(e)
        return {"sent": sent, "failed": failed, "total": len(emails)}

    @app.route("/retiros/<int:rid>/field", methods=["PATCH", "POST"])
    @require_permission("retiros")
    def pickup_inline_field(rid):
        """Auto-save inline de un campo de la ficha.

        Body JSON: {field: "contact_email", value: "nuevo@cliente.cl"}

        Valida:
        - field está en whitelist _PICKUP_INLINE_FIELDS
        - len(value) ≤ max_length del campo
        - el retiro existe

        Devuelve: {ok:True, field, value, changed:bool, saved_at}
        """
        body = request.get_json(silent=True) or {}
        field = (body.get("field") or "").strip()
        value = body.get("value")

        # Validar field en whitelist
        if field not in _PICKUP_INLINE_FIELDS:
            return jsonify({
                "ok": False,
                "error": f"Campo '{field}' no permitido para edición inline.",
                "error_codigo": "CAMPO_NO_PERMITIDO",
            }), 400

        max_len, label = _PICKUP_INLINE_FIELDS[field]

        # Normalizar valor: None/empty → NULL, trim, max_len
        if value is None:
            value_norm = None
        else:
            value_norm = str(value).strip()
            if len(value_norm) > max_len:
                value_norm = value_norm[:max_len]
            if value_norm == "":
                value_norm = None

        # Verificar que existe + obtener old value
        req = mysql_fetchone(
            f"SELECT id, code, `{field}` AS current_value FROM `{REQ}` WHERE id=%s",
            (rid,)
        )
        if not req:
            return jsonify({"ok": False, "error": "Retiro no encontrado"}), 404

        old_value = req.get("current_value")
        # Normalizar old para comparar (None/empty equiv)
        old_norm = (str(old_value).strip() if old_value is not None else None) or None
        new_norm = value_norm or None

        if old_norm == new_norm:
            return jsonify({
                "ok": True,
                "field": field,
                "value": new_norm or "",
                "changed": False,
                "saved_at": _pickup_now_chile_hms(),
            })

        try:
            # Backtick-quote field name (whitelist garantiza que es seguro)
            mysql_execute(
                f"UPDATE `{REQ}` SET `{field}`=%s WHERE id=%s",
                (value_norm, rid)
            )
        except Exception as exc:
            return jsonify({
                "ok": False,
                "error": f"Error al guardar: {str(exc)[:200]}",
            }), 500

        # Audit log — corto, no expone PII completo
        log_event(
            rid, "ficha_inline_edit", None, None,
            f"{label}: '{(old_norm or '')[:40]}' → '{(new_norm or '')[:40]}'",
            "interno"
        )

        return jsonify({
            "ok": True,
            "field": field,
            "value": new_norm or "",
            "changed": True,
            "saved_at": _now_chile_str(),
        })

    # ══════════════════════════════════════════════════════════════════
    #  INFORME MENSUAL XLSX — Daniel pidió 22/05/2026
    # ══════════════════════════════════════════════════════════════════
    @app.route("/retiros/api/informe-mes.xlsx", methods=["GET"])
    @require_permission("view")
    def pickup_informe_mes_xlsx():
        """Excel con todos los retiros del mes solicitado (filtrable por RUT).

        Query params:
            mes: YYYY-MM (default: mes actual Chile)
            cliente_rut: opcional — filtra solo ese cliente
            status: opcional — filtra por status (e.g. 'agenda_confirmada')
            solo_pendientes: '1' → solo los que NO están en estado retirada/cerrada/rechazada
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from io import BytesIO
        except ImportError:
            return jsonify({"error": "openpyxl no instalado en el servidor."}), 500

        mes = (request.args.get("mes") or "").strip()
        cliente_rut = (request.args.get("cliente_rut") or "").strip()
        status_filter = (request.args.get("status") or "").strip()
        solo_pendientes = (request.args.get("solo_pendientes") or "").strip() in ("1", "true", "yes")

        # Resolver mes — default actual Chile
        try:
            from zoneinfo import ZoneInfo as _ZI
            _TZ = _ZI("America/Santiago")
        except Exception:
            _TZ = None
        if not mes:
            now = datetime.now(_TZ) if _TZ else datetime.now()
            mes = now.strftime("%Y-%m")
        try:
            yy, mm = mes.split("-")
            yy = int(yy); mm = int(mm)
            if not (2020 <= yy <= 2100) or not (1 <= mm <= 12):
                raise ValueError
        except Exception:
            return jsonify({"error": "Parámetro 'mes' inválido. Formato YYYY-MM."}), 400

        # Construir filtros SQL
        where = ["DATE_FORMAT(COALESCE(confirmed_date, requested_date, DATE(created_at)), '%%Y-%%m') = %s"]
        params = [mes]
        if cliente_rut:
            rut_clean = re.sub(r"[^0-9kK]", "", cliente_rut).upper()
            where.append("REPLACE(REPLACE(REPLACE(UPPER(customer_rut),'.',''),'-',''),' ','') = %s")
            params.append(rut_clean)
        if status_filter and status_filter in PICKUP_STATUS:
            where.append("status = %s")
            params.append(status_filter)
        if solo_pendientes:
            where.append("status NOT IN ('retirada','cerrada','rechazada','fallida')")

        sql = (
            f"SELECT id, code, document_type, document_number, customer_name, customer_rut, "
            f"       contact_name, contact_email, contact_phone, "
            f"       pickup_person_name, pickup_person_rut, pickup_person_phone, pickup_person_relation, "
            f"       requested_date, requested_time_from, requested_time_to, "
            f"       proposed_date, confirmed_date, confirmed_time_from, confirmed_time_to, "
            f"       status, total_weight_kg, total_volumetric_weight, total_volume_m3, "
            f"       peso_real_kg, peso_vol_kg, tiempo_estimado_min, doc_validation_status, "
            f"       created_at, closed_at "
            f"FROM `{REQ}` "
            f"WHERE {' AND '.join(where)} "
            f"ORDER BY COALESCE(confirmed_date, requested_date, DATE(created_at)) ASC, code ASC"
        )
        try:
            rows = mysql_fetchall(sql, tuple(params)) or []
        except Exception as exc:
            return jsonify({"error": f"Error consultando retiros: {str(exc)[:200]}"}), 500

        # Para cada retiro, traer docs asociados (concatenados)
        rid_list = [r["id"] for r in rows]
        docs_by_rid = {}
        if rid_list:
            ph = ",".join(["%s"] * len(rid_list))
            docs_rows = mysql_fetchall(
                f"SELECT request_id, document_type, document_number, cliente_nombre, cliente_rut "
                f"FROM pickup_request_docs WHERE request_id IN ({ph}) "
                f"ORDER BY request_id, id",
                tuple(rid_list)
            ) or []
            for d in docs_rows:
                docs_by_rid.setdefault(d["request_id"], []).append(
                    f"{d['document_type']} {d['document_number']}"
                )

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Retiros {mes}"

        RED_FILL = PatternFill("solid", fgColor="DC2626")
        WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
        HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        NUM_ALIGN = Alignment(horizontal='right')

        headers = [
            "Código", "Fecha solicitud", "Fecha confirmada",
            "Cliente", "RUT cliente",
            "Doc principal", "Docs asociados (extra)",
            "Persona retira", "RUT retira", "Teléfono retira", "Relación",
            "Status", "Validación doc",
            "Peso real (kg)", "Peso vol. (kg)", "Volumen (m³)",
            "Tiempo estimado (min)",
            "Email contacto", "Teléfono contacto",
            "Creado", "Cerrado",
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = WHITE_FONT
            c.fill = RED_FILL
            c.alignment = HDR_ALIGN

        def _fmt_date(d):
            if not d: return ""
            try: return d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
            except Exception: return str(d)[:10]

        def _fmt_dt(d):
            if not d: return ""
            try: return d.strftime("%Y-%m-%d %H:%M") if hasattr(d, "strftime") else str(d)[:16]
            except Exception: return str(d)[:16]

        for ri, r in enumerate(rows, 2):
            doc_principal = f"{(r.get('document_type') or '').upper()} {r.get('document_number') or ''}".strip()
            extras = docs_by_rid.get(r["id"], [])
            extras_str = " · ".join([e for e in extras if e and e != doc_principal])
            status_lbl = PICKUP_STATUS.get(r.get("status") or "", r.get("status") or "")
            row_vals = [
                r.get("code") or "",
                _fmt_date(r.get("requested_date")),
                _fmt_date(r.get("confirmed_date") or r.get("proposed_date")),
                r.get("customer_name") or "",
                r.get("customer_rut") or "",
                doc_principal,
                extras_str,
                r.get("pickup_person_name") or "",
                r.get("pickup_person_rut") or "",
                r.get("pickup_person_phone") or "",
                r.get("pickup_person_relation") or "",
                status_lbl,
                r.get("doc_validation_status") or "",
                float(r.get("peso_real_kg") or r.get("total_weight_kg") or 0),
                float(r.get("peso_vol_kg") or r.get("total_volumetric_weight") or 0),
                float(r.get("total_volume_m3") or 0),
                int(r.get("tiempo_estimado_min") or 0),
                r.get("contact_email") or "",
                r.get("contact_phone") or "",
                _fmt_dt(r.get("created_at")),
                _fmt_dt(r.get("closed_at")),
            ]
            for ci, v in enumerate(row_vals, 1):
                cell = ws.cell(ri, ci, v)
                if ci in (14, 15, 16, 17):
                    cell.alignment = NUM_ALIGN

        # Ancho columnas razonable
        widths = [11, 13, 14, 28, 14, 16, 28, 22, 14, 16, 14, 18, 14, 13, 13, 12, 14, 26, 16, 17, 17]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        # Hoja de filtros
        ws2 = wb.create_sheet("Filtros")
        ws2.cell(1, 1, "Filtro").font = WHITE_FONT
        ws2.cell(1, 2, "Valor").font = WHITE_FONT
        for c in range(1, 3):
            ws2.cell(1, c).fill = RED_FILL
        filt_rows = [
            ("Mes", mes),
            ("Cliente RUT", cliente_rut or "(todos)"),
            ("Status", status_filter or "(todos)"),
            ("Solo pendientes", "Sí" if solo_pendientes else "No"),
            ("Total retiros", len(rows)),
            ("Generado por", g.user["nombre"] if getattr(g, "user", None) else "sistema"),
            ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for ri, (k, v) in enumerate(filt_rows, 2):
            ws2.cell(ri, 1, k)
            ws2.cell(ri, 2, str(v))
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 36

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"ILUS_Retiros_{mes}"
        if cliente_rut:
            fname += f"_{re.sub(r'[^0-9kK]', '', cliente_rut)}"
        fname += ".xlsx"

        from flask import send_file
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname,
        )


    @app.route("/retiros/<int:rid>/proposal", methods=["POST"])
    @require_permission("retiros")
    def pickup_create_proposal(rid):
        """Crea/envía propuesta al cliente. Soporta dos modos:
          - HTML form (legacy): redirect a pickup_detail
          - AJAX (cabecera X-Requested-With='XMLHttpRequest' o Accept='application/json'):
            devuelve JSON {ok, message, redirect_url} sin recargar página.

        Optimización Daniel 2026-05-23 (wizard): el envío de email/SMS/WhatsApp
        ya estaba async, pero el redirect HTML completo es lento (full page
        reload). El frontend nuevo del wizard usa fetch + JSON para feedback
        instantáneo (<500ms).

        Saldo ERP (Daniel 2026-05-24): ya NO bloqueamos por saldo. Si ningún
        doc asociado tiene saldo (todos con_saldo=0) solo se loguea una
        advertencia para métricas y el flujo continúa — el operador asume la
        responsabilidad (puede que el cliente esté abonando o que la guía ya
        se haya emitido). Ver bloque SIN-SALDO-WARN más abajo.
        """
        # Detectar si es AJAX (sin acoplar a un único framework)
        is_ajax = (
            (request.headers.get("X-Requested-With") or "").lower() == "xmlhttprequest"
            or "application/json" in (request.headers.get("Accept") or "").lower()
            or request.is_json
        )

        def _resp_err(msg, status=400):
            if is_ajax:
                return jsonify({"ok": False, "error": msg}), status
            flash(msg, "warning")
            return redirect(url_for("pickup_detail", rid=rid))

        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            if is_ajax: return jsonify({"ok": False, "error": "Retiro no encontrado"}), 404
            return redirect(url_for("pickup_dashboard"))

        # WORKFLOW: la propuesta requiere documentación cargada.
        # Daniel 2026-06-15: si el retiro YA tiene documentos asociados NO
        # bloqueamos por la "validación formal" — los docs SIN SALDO no
        # auto-validaban y dejaban el retiro trabado sin un botón visible para
        # validar. El "sin saldo" es un indicador, no un bloqueante. Entonces:
        # con docs asociados → auto-marcamos validado y seguimos; solo
        # bloqueamos si NO hay ningún documento (ahí sí falta el Paso 2).
        if req.get("doc_validation_status") not in ("ok",):
            _ndoc = mysql_fetchone(
                "SELECT COUNT(*) AS n FROM pickup_request_docs WHERE request_id=%s",
                (rid,)
            ) or {}
            if int(_ndoc.get("n") or 0) <= 0:
                return _resp_err("Antes de proponer fecha, asocia al menos un "
                                 "documento del cliente en el Paso 2 (cargar factura/boleta).")
            try:
                mysql_execute(
                    f"UPDATE `{REQ}` SET doc_validation_status='ok', "
                    f"doc_validated_at=NOW() WHERE id=%s", (rid,)
                )
                log_event(rid, "doc_validacion_auto", req.get("status") or "",
                          req.get("status") or "",
                          "Auto-validado al proponer fecha (docs asociados; sin saldo no bloquea)",
                          "interno")
            except Exception:
                pass

        # WIZARD (Daniel 2026-05-23): bloquear si NINGÚN doc asociado tiene
        # saldo disponible. Permitimos pasar si no hay docs todavía o si
        # al menos uno tiene saldo (los sin saldo son visualizables pero
        # no cuentan). Si la columna no existe (entorno viejo), saltamos
        # esta validación — el flujo legacy sigue funcionando.
        try:
            saldo_row = mysql_fetchone(
                """SELECT
                       SUM(CASE WHEN con_saldo=1 THEN 1 ELSE 0 END) AS con_saldo,
                       SUM(CASE WHEN con_saldo=0 THEN 1 ELSE 0 END) AS sin_saldo,
                       COUNT(*) AS total
                     FROM pickup_request_docs
                    WHERE request_id=%s""",
                (rid,)
            ) or {}
            # Daniel 2026-05-24: YA NO BLOQUEAMOS por saldo.
            # Quote: "si no tiene saldo, que no pase más allá de una
            # notificación, porque a lo mejor abonen perfecto, se están
            # cuadrando, no sé, puede pasar que ya hayan hecho la guía,
            # como son desordenados, nosotros lo vamos a encargar
            # solamente de llevar las métricas. Vamos a poder filtrar
            # por documentos que no tengan saldo y se estén agendando
            # para retiro. Es un buen indicador, no un bloqueante."
            # → Solo log para métricas, NO return error.
            n_total     = int(saldo_row.get("total")     or 0)
            n_con_saldo = int(saldo_row.get("con_saldo") or 0)
            n_sin_saldo = int(saldo_row.get("sin_saldo") or 0)
            if n_total > 0 and n_con_saldo == 0 and n_sin_saldo > 0:
                print(f"[pickup-propose][SIN-SALDO-WARN] rid={rid} "
                      f"todos los {n_total} docs sin saldo verificado. "
                      f"Avanzando igualmente — operador asume responsabilidad.",
                      flush=True)
        except Exception:
            pass

        date, tf, tt = request.form.get("date"), request.form.get("time_from"), request.form.get("time_to")
        cfg = settings()
        # FASE 2 (2026-05-29): validador temporal central, modo 'internal'.
        # El OPERADOR (endpoint bajo @require_permission("retiros")) puede CRUZAR
        # la colación si la factura es grande, PERO ya NO puede proponer
        # fecha/hora PASADA (antes date_allowed lo permitía — bug). El cliente
        # público nunca pasa por acá.
        ok_dt, msg_dt = validate_pickup_datetime(date, tf, tt, cfg=cfg, mode="internal")
        if not ok_dt:
            return _resp_err(msg_dt)

        # Validar capacidad real del slot (cupos, kg, m³, bloqueos).
        # bypass_lunch=True → el operador manda por encima del cliente.
        ok_slot, motivo = _validar_disponibilidad_slot(
            date, tf, tt,
            exclude_request_id=rid,
            extra_kg=float(req.get("total_weight_kg") or 0),
            extra_m3=float(req.get("total_volume_m3") or 0),
            bypass_lunch=True,
        )
        if not ok_slot:
            log_event(rid, "propuesta_bloqueada", req["status"], req["status"],
                      f"Intento de proponer {date} {tf}-{tt}: {motivo}", "interno")
            return _resp_err(f"Slot no disponible: {motivo}", status=409)

        # ── Daniel 2026-06-15: la propuesta DEBE poder notificarse al cliente.
        #    El ping-pong con el cliente arranca con este correo, así que si el
        #    retiro no tiene NINGÚN email válido (contacto / extra / ERP), no
        #    tiene sentido avanzar a 'propuesta_enviada' — el cliente jamás se
        #    enteraría. Bloqueamos y pedimos al operador agregar el email.
        if not _get_pickup_all_emails(req):
            return _resp_err(
                "Este retiro no tiene un correo de cliente válido. Agrega el email "
                "del cliente (paso 2) antes de enviar la propuesta — el cliente lo "
                "necesita para aceptar o contraproponer.",
                status=409,
            )

        # ── AUTO-CONFIRMACIÓN POR COINCIDENCIA (Daniel 2026-06-17/19) ──────────
        # Si el operador propone EXACTAMENTE la fecha+hora que el cliente ya pidió
        # (sea en su CONTRAPROPUESTA pending, o en el FORMULARIO INICIAL con que
        # creó el retiro), NO reabrimos el ping-pong: el operador está ACEPTANDO
        # lo que el cliente pidió → confirmamos directo, sin pedirle al cliente
        # que confirme de nuevo lo que él mismo propuso. Fix del "no me quedaba
        # confirmada": el caso más común (aceptar la fecha del formulario) antes
        # NO auto-confirmaba porque la solicitud inicial no crea fila en pickup_
        # proposals. El slot ya se validó arriba (_validar_disponibilidad_slot).
        _pend_cli = mysql_fetchone(
            f"SELECT id, date, time_from, time_to FROM `{PROP}` "
            f"WHERE request_id=%s AND status='pending' AND LOWER(proposed_by)='cliente' "
            f"ORDER BY id DESC LIMIT 1", (rid,)
        )
        def _slot_coincide(p):
            try:
                return (str(p.get("date"))[:10] == str(date)[:10]
                        and _td_to_hhmm(p.get("time_from")) == str(tf)[:5]
                        and _td_to_hhmm(p.get("time_to")) == str(tt)[:5])
            except Exception:
                return False
        _match_counter = bool(_pend_cli and _slot_coincide(_pend_cli))
        # ¿La fecha/hora propuesta coincide con la que el cliente pidió en el
        # formulario inicial (requested_*)? Ese es el caso "Aceptar como propuesta".
        # OJO (review 2026-06-19): SOLO si NO hay una contrapropuesta VIVA del
        # cliente. Si el cliente ya contrapropuso otra fecha (_pend_cli existe) y
        # el operador re-propone la fecha ORIGINAL del formulario, NO auto-
        # confirmamos la vieja (pisaría la contrapropuesta viva) → ping-pong normal.
        # Si _pend_cli coincide con lo propuesto, ya lo cubre _match_counter.
        _match_requested = (
            _slot_coincide({
                "date":      req.get("requested_date"),
                "time_from": req.get("requested_time_from"),
                "time_to":   req.get("requested_time_to"),
            })
            if (req.get("requested_date") and not _pend_cli) else False
        )
        if _match_counter or _match_requested:
            try:
                mysql_execute(
                    f"UPDATE `{PROP}` SET status='accepted', answered_at=NOW() "
                    f"WHERE request_id=%s AND status='pending'", (rid,)
                )
                mysql_execute(
                    f"UPDATE `{REQ}` SET status='agenda_confirmada', "
                    f"confirmed_date=%s, confirmed_time_from=%s, confirmed_time_to=%s, "
                    f"proposed_date=%s, proposed_time_from=%s, proposed_time_to=%s "
                    f"WHERE id=%s",
                    (date, tf, tt, date, tf, tt, rid)
                )
                _origen_match = "contrapropuesta del cliente" if _match_counter else "fecha pedida por el cliente en el formulario"
                log_event(rid, "auto_confirmada_coincidencia", req.get("status") or "",
                          "agenda_confirmada",
                          f"El operador coincidió con la {_origen_match} "
                          f"({date} {tf}-{tt}) → confirmado directo, sin re-confirmación.",
                          "interno")
                # Aviso interno al equipo: se confirmó una cita
                try:
                    _notificar_equipo_retiros(
                        f"✅ Retiro {req.get('code') or '?'} CONFIRMADO",
                        f"{req.get('customer_name') or 'Cliente'} — cita {date} {tf}-{tt} "
                        f"(el operador aceptó la fecha que pidió el cliente).",
                        rid, req.get("code") or "?",
                        prioridad="media", tipo="retiro_confirmado", send_email=False)
                except Exception:
                    pass
            except Exception as _e_ac:
                print(f"[pickup-autoconfirm] error: {_e_ac}", flush=True)
                return _resp_err("No se pudo confirmar automáticamente. Reintenta.", status=500)
            # Invalidar caches para que el cliente lo vea EN VIVO
            try:
                _tok_ac = req.get("public_token") if isinstance(req, dict) else None
                if _tok_ac:
                    _POLL_CACHE.pop(_tok_ac, None)
            except Exception:
                pass
            try:
                _DISPO_CACHE["payload"] = None
            except Exception:
                pass
            # Correo de CONFIRMACIÓN al cliente (con .ics + calendario), síncrono
            _fresh_ac = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,)) or req
            _sent_ac = False
            try:
                _sent_ac, _ = notify(_fresh_ac, "confirmed")
            except Exception as _e_nc:
                print(f"[pickup-autoconfirm] notify confirmed: {_e_nc}", flush=True)
            try:
                _alertar_si_sin_saldo(rid, req.get("code") or "?")
            except Exception:
                pass
            _ac_msg = ("Coincidiste con la fecha que pidió el cliente: el retiro quedó "
                       "CONFIRMADO directo (no se le pidió confirmar de nuevo).")
            if is_ajax:
                return jsonify({"ok": True, "auto_confirmado": True, "message": _ac_msg,
                                "email_enviado": bool(_sent_ac),
                                "redirect_url": url_for("pickup_detail", rid=rid)})
            flash(_ac_msg, "success")
            return redirect(url_for("pickup_detail", rid=rid))

        # ── FASE 3 (2026-05-29): ping-pong. Antes de crear la propuesta nueva,
        #    marcar las propuestas pending anteriores como 'superseded' (solo
        #    puede haber UNA propuesta vigente a la vez). expires_at = ahora
        #    (Chile) + proposal_expiry_hours.
        try:
            _expiry_h = int(cfg.get("proposal_expiry_hours") or 48)
        except (TypeError, ValueError):
            _expiry_h = 48
        _expires_at = (_now_chile() + timedelta(hours=_expiry_h)).strftime("%Y-%m-%d %H:%M:%S")
        mysql_execute(
            f"UPDATE `{PROP}` SET status='superseded', answered_at=NOW() "
            f"WHERE request_id=%s AND status='pending'", (rid,)
        )
        # ── INSERT propuesta y UPDATE estado (lo crítico, debe completarse) ──
        mysql_execute(
            f"""INSERT INTO `{PROP}` (request_id,proposed_by,date,time_from,time_to,message,reason,status,token,expires_at)
                VALUES (%s,'internal',%s,%s,%s,%s,%s,'pending',%s,%s)""",
            (rid, date, tf, tt, request.form.get("message", ""), request.form.get("reason", ""),
             secrets.token_urlsafe(24), _expires_at),
        )
        mysql_execute(
            f"UPDATE `{REQ}` SET status='propuesta_enviada', proposed_date=%s, "
            f"proposed_time_from=%s, proposed_time_to=%s WHERE id=%s",
            (date, tf, tt, rid)
        )
        log_event(rid, "propuesta_enviada", req["status"], "propuesta_enviada",
                  f"{date} {tf}-{tt}", "interno")

        # ── Notificación al cliente — SÍNCRONA (Daniel 2026-06-15) ──
        # El correo de propuesta es el evento MÁS crítico del ping-pong.
        # Antes iba en un thread daemon con las excepciones tragadas: si el
        # envío fallaba, el operador veía "Propuesta enviada" pero al cliente
        # NO le llegaba nada (caso real reportado: "me agendé y no pasó nada").
        # Ahora lo enviamos de forma síncrona y reflejamos el resultado REAL en
        # la respuesta, para que el operador sepa con certeza si salió o no.
        _msg = request.form.get("message", "")
        fresh = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,)) or req
        try:
            n_destinatarios = len(_get_pickup_all_emails(fresh))
        except Exception:
            n_destinatarios = 0
        email_enviado = False
        try:
            _res = notify(fresh, "proposal", proposal={
                "date": date, "time_from": tf, "time_to": tt, "message": _msg,
            })
            # notify() devuelve (sent_mail, sent_wa); defensivo por si cambia.
            if isinstance(_res, (tuple, list)) and _res:
                email_enviado = bool(_res[0])
            else:
                email_enviado = bool(_res)
        except Exception as _e:
            print(f"[pickup_create_proposal notify] {_e}", flush=True)
            email_enviado = False

        # Si el correo NO salió, dejamos rastro en el tracking para que el
        # equipo lo reintente/avise (la propuesta igual queda registrada y
        # visible en el seguimiento del cliente).
        if not email_enviado:
            try:
                log_event(rid, "propuesta_email_no_enviado", "propuesta_enviada",
                          "propuesta_enviada",
                          f"Propuesta {date} {tf}-{tt} registrada pero el correo al cliente NO salió "
                          f"(destinatarios={n_destinatarios}). Revisar llave de correo de Retiros o "
                          f"el email del cliente.", "interno")
            except Exception:
                pass

        # Invalidación de caches (no crítico).
        try:
            tok_p = (fresh or {}).get("public_token") if isinstance(fresh, dict) else None
            if tok_p: _POLL_CACHE.pop(tok_p, None)
        except Exception: pass
        try: _DISPO_CACHE["payload"] = None
        except Exception: pass

        if is_ajax:
            if email_enviado:
                _ok_msg = (f"Propuesta enviada al cliente ✓ "
                           f"(correo a {n_destinatarios} destinatario"
                           f"{'s' if n_destinatarios != 1 else ''}).")
            else:
                _ok_msg = ("Propuesta registrada y visible en el seguimiento del cliente, "
                           "pero el correo NO se pudo enviar ahora. Revisa la llave de correo "
                           "de Retiros o el email del cliente.")
            return jsonify({
                "ok": True,
                "message": _ok_msg,
                "email_enviado": email_enviado,
                "destinatarios": n_destinatarios,
                "proposal": {"date": date, "time_from": tf, "time_to": tt},
                "redirect_url": url_for("pickup_detail", rid=rid),
            })
        if email_enviado:
            flash("Propuesta enviada al cliente.", "success")
        else:
            flash("Propuesta registrada, pero el correo al cliente no salió. "
                  "Revisa la llave de correo de Retiros.", "warning")
        return redirect(url_for("pickup_detail", rid=rid))

    @app.route("/retiros/<int:rid>/aceptar-contrapropuesta", methods=["POST"])
    @require_permission("retiros")
    def pickup_aceptar_contrapropuesta(rid):
        """ILUS acepta la CONTRAPROPUESTA de fecha enviada por el cliente
        (doble confirmación, C1 2026-06-09). AJAX JSON.

        Flujo: el cliente contrapropuso (proposal pending, proposed_by='cliente',
        retiro en 'en_revision') → el operador la acepta acá:
          PROP → accepted + answered_at, REQ → agenda_confirmada con
          confirmed_date/time copiados de la propuesta (en transacción).
        Luego: log_event + notify async 'confirmed' al cliente + alerta
        sin-saldo al equipo (D2). Errores siempre con mensaje amigable.
        """
        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            return jsonify({"ok": False, "error": "Retiro no encontrado."}), 404

        # Guard de estado (review M2 2026-06-09): una contrapropuesta pendiente
        # puede sobrevivir a un cierre manual vía /status (ese endpoint no las
        # marca). Aceptarla aquí "resucitaría" un retiro cerrado/rechazado.
        if str(req.get("status") or "") in ("rechazada", "cerrada", "retirada", "fallida"):
            return jsonify({
                "ok": False,
                "error": "Este retiro ya está cerrado o finalizado. Si corresponde reagendar, crea una solicitud nueva o reabre desde Cambiar estado.",
            }), 409

        proposal = mysql_fetchone(
            f"SELECT * FROM `{PROP}` WHERE request_id=%s AND status='pending' "
            f"ORDER BY id DESC LIMIT 1",
            (rid,)
        )
        if not proposal or str(proposal.get("proposed_by") or "").lower() != "cliente":
            return jsonify({
                "ok": False,
                "error": "No hay una contrapropuesta del cliente pendiente para este retiro.",
            }), 409

        # Vigencia (expires_at): una contrapropuesta vencida no se puede aceptar.
        if not proposal_is_vigente(proposal):
            try:
                mysql_execute(
                    f"UPDATE `{PROP}` SET status='expired', answered_at=NOW() WHERE id=%s",
                    (proposal["id"],),
                )
            except Exception:
                pass
            return jsonify({
                "ok": False,
                "error": "La contrapropuesta del cliente venció. Envíale una nueva propuesta de fecha.",
            }), 410

        # Disponibilidad real del slot (cupos, kg, m³, bloqueos, colación).
        ok_slot, motivo = _validar_disponibilidad_slot(
            proposal["date"], proposal["time_from"], proposal["time_to"],
            exclude_request_id=rid,
            extra_kg=float(req.get("total_weight_kg") or 0),
            extra_m3=float(req.get("total_volume_m3") or 0),
        )
        if not ok_slot:
            return jsonify({
                "ok": False,
                "error": f"Ese horario ya no está disponible: {motivo} Propón otra fecha al cliente.",
            }), 409

        # Transacción: PROP→accepted + REQ→agenda_confirmada (atómico).
        # Review M1 2026-06-09: re-validar capacidad BAJO LOCK (SELECT FOR
        # UPDATE + reconteo), mismo patrón FASE 2 del confirm público —
        # entre la validación de arriba y este commit otro cliente puede
        # confirmar el mismo slot (sobreventa real vivida el 2026-05-12).
        accept_ok = False
        accept_motivo = ""
        conn_tx = None
        try:
            conn_tx = get_mysql()
            with conn_tx.cursor() as cur_tx:
                extra_kg = float(req.get("total_weight_kg") or 0)
                extra_m3 = float(req.get("total_volume_m3") or 0)
                date_str = str(proposal["date"])[:10]
                tf_str   = _td_to_hhmm(proposal["time_from"])
                cur_tx.execute(
                    f"""SELECT COUNT(*) AS n,
                                COALESCE(SUM(total_weight_kg),0) AS kg,
                                COALESCE(SUM(total_volume_m3),0) AS m3
                         FROM `{REQ}`
                         WHERE status NOT IN ('rechazada','cerrada','fallida')
                           AND id <> %s
                           AND (
                             (confirmed_date=%s AND TIME_FORMAT(confirmed_time_from,'%%H:%%i')=%s)
                             OR
                             (confirmed_date IS NULL AND proposed_date=%s
                              AND TIME_FORMAT(proposed_time_from,'%%H:%%i')=%s)
                           )
                         FOR UPDATE""",
                    (rid, date_str, tf_str, date_str, tf_str),
                )
                slot_row  = cur_tx.fetchone() or {}
                picks_now = int(slot_row.get("n") or 0)
                kg_now    = float(slot_row.get("kg") or 0)
                m3_now    = float(slot_row.get("m3") or 0)
                cfg_lock  = settings()
                _pc_lock  = cfg_lock.get("parallel_capacity")
                if _pc_lock is not None and str(_pc_lock).strip():
                    max_picks_slot = int(_pc_lock)
                else:
                    max_picks_slot = int(cfg_lock.get("max_picks_per_slot") or 2)
                max_kg_slot = float(cfg_lock.get("max_kg_per_slot") or 500)
                max_m3_slot = float(cfg_lock.get("max_m3_per_slot") or 5)

                if picks_now + 1 > max_picks_slot:
                    accept_motivo = f"el slot se llenó ({picks_now} de {max_picks_slot})."
                elif kg_now + extra_kg > max_kg_slot:
                    accept_motivo = "se excedería la capacidad de peso del bloque."
                elif m3_now + extra_m3 > max_m3_slot:
                    accept_motivo = "se excedería la capacidad de volumen del bloque."
                else:
                    cur_tx.execute(
                        f"UPDATE `{PROP}` SET status='accepted', answered_at=NOW() "
                        f"WHERE id=%s AND status='pending'",
                        (proposal["id"],),
                    )
                    if cur_tx.rowcount:
                        cur_tx.execute(
                            f"""UPDATE `{REQ}`
                                  SET status='agenda_confirmada',
                                      confirmed_date=%s,
                                      confirmed_time_from=%s,
                                      confirmed_time_to=%s
                                WHERE id=%s""",
                            (proposal["date"], proposal["time_from"],
                             proposal["time_to"], rid),
                        )
                        accept_ok = True
                    else:
                        accept_motivo = "otro operador ya respondió esta contrapropuesta."
            if accept_ok:
                conn_tx.commit()
            else:
                conn_tx.rollback()
        except Exception as _tx_err:
            if conn_tx is not None:
                try: conn_tx.rollback()
                except Exception: pass
            accept_ok = False
            print(f"[pickup_aceptar_contrapropuesta] tx error rid={rid}: {_tx_err}", flush=True)
        finally:
            if conn_tx is not None:
                try: conn_tx.close()
                except Exception: pass

        if not accept_ok:
            _det = accept_motivo or "puede que otro operador ya la haya respondido."
            return jsonify({
                "ok": False,
                "error": f"No se pudo aceptar la contrapropuesta: {_det} Recarga la ficha e intenta de nuevo.",
            }), 409

        log_event(
            rid, "ilus_acepto_contrapropuesta", req["status"], "agenda_confirmada",
            f"{str(proposal['date'])[:10]} {_td_to_hhmm(proposal['time_from'])}-{_td_to_hhmm(proposal['time_to'])}",
            "interno",
        )

        # Invalidar caches: el cliente debe ver 'agenda_confirmada' al instante.
        try: _POLL_CACHE.pop(req.get("public_token"), None)
        except Exception: pass
        try: _DISPO_CACHE["payload"] = None
        except Exception: pass

        # Email "agenda confirmada" al cliente (async, con datos finales).
        try:
            req_after = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,)) or req
            notify_async(req_after, "confirmed")
        except Exception as _e:
            print(f"[pickup_aceptar_contrapropuesta][notify] {_e}", flush=True)

        # D2: si TODOS los docs ERP están sin saldo → alerta urgente al equipo.
        try:
            _alertar_si_sin_saldo(rid, req.get("code"))
        except Exception as _e:
            print(f"[pickup_aceptar_contrapropuesta][saldo] {_e}", flush=True)

        return jsonify({
            "ok": True,
            "message": "Contrapropuesta aceptada. El retiro quedó agendado y avisamos al cliente por correo.",
            "confirmed": {
                "date":      str(proposal["date"])[:10],
                "time_from": _td_to_hhmm(proposal["time_from"]),
                "time_to":   _td_to_hhmm(proposal["time_to"]),
            },
            "redirect_url": url_for("pickup_detail", rid=rid),
        })

    @app.route("/retiros/<int:rid>/message", methods=["POST"])
    @require_permission("retiros")
    def pickup_send_message(rid):
        req = mysql_fetchone(f"SELECT * FROM `{REQ}` WHERE id=%s", (rid,))
        if not req:
            return redirect(url_for("pickup_dashboard"))
        template_id = request.form.get("template_id", type=int)
        tpl = mysql_fetchone(f"SELECT * FROM `{TPL}` WHERE id=%s", (template_id,)) if template_id else None
        message = request.form.get("message", "").strip() or (tpl["body"] if tpl else "")
        if not message:
            flash("Selecciona o escribe un mensaje.", "warning")
            return redirect(url_for("pickup_detail", rid=rid))
        notify_async(req, "message", custom_message=message)
        log_event(rid, "mensaje_enviado", req["status"], req["status"], message, "interno")
        flash("Mensaje enviado por los canales disponibles.", "success")
        return redirect(url_for("pickup_detail", rid=rid))

    @app.route("/retiros/settings", methods=["POST"])
    @require_permission("admin")
    def pickup_settings_save():
        data = request.form
        mysql_execute(
            f"""UPDATE `{SET}`
                SET warehouse_name=%s, warehouse_addr=%s, maps_url=%s, open_time=%s, close_time=%s,
                    work_days=%s, holidays=%s, alert_enabled=%s, alert_title=%s, alert_message=%s,
                    notify_emails=%s
                WHERE id=1""",
            (
                data.get("warehouse_name", ""), data.get("warehouse_addr", ""), data.get("maps_url", ""),
                data.get("open_time", "09:00"), data.get("close_time", "16:30"),
                ",".join(data.getlist("work_days")) or "1,2,3,4,5", data.get("holidays", ""),
                1 if data.get("alert_enabled") else 0, data.get("alert_title", "Aviso importante"), data.get("alert_message", ""),
                (data.get("notify_emails") or "").strip()[:2000],
            ),
        )
        # Invalidar cache de settings (Daniel mayo 2026) — el TTL es 30s pero
        # tras un guardado del admin queremos que el público vea el cambio YA.
        try: _invalidate_settings_cache()
        except Exception: pass
        # Invalidar cache global del calendario público
        try: _DISPO_CACHE["payload"] = None
        except Exception: pass
        flash("Configuracion de retiros actualizada.", "success")
        return redirect(url_for("pickup_dashboard"))

    @app.route("/ajustes/marketing", methods=["GET", "POST"])
    @require_permission("admin")
    def marketing_settings():
        if request.method == "POST":
            current = settings()
            hero_paths = {
                "hero_image_1": current.get("hero_image_1") or "",
                "hero_image_2": current.get("hero_image_2") or "",
                "hero_image_3": current.get("hero_image_3") or "",
            }
            for key in hero_paths:
                f = request.files.get(key)
                if f and f.filename:
                    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
                    if ext in {"png", "jpg", "jpeg", "webp"}:
                        fname = f"hero_{key}_{int(time.time())}_{secure_filename(f.filename)}"
                        f.save(os.path.join(upload_dir, fname))
                        hero_paths[key] = f"uploads/retiros/{fname}"
            mysql_execute(
                f"""UPDATE `{SET}`
                    SET hero_image_1=%s, hero_image_2=%s, hero_image_3=%s
                    WHERE id=1""",
                (hero_paths["hero_image_1"], hero_paths["hero_image_2"], hero_paths["hero_image_3"]),
            )
            flash("Marketing publico actualizado.", "success")
            return redirect(url_for("marketing_settings"))
        # Cargar imágenes de login para mostrar en el tab unificado
        try:
            login_imgs = mysql_fetchall(
                "SELECT * FROM login_images ORDER BY orden ASC, id ASC"
            )
            login_imgs = [dict(r) for r in login_imgs]
        except Exception:
            login_imgs = []
        # Cargar bloqueos de retiros próximos (60 días)
        try:
            from datetime import datetime as _dt, timedelta as _td
            blocks_rows = mysql_fetchall(
                "SELECT id, fecha, hora_inicio, hora_fin, motivo, created_by, created_at "
                "FROM pickup_blocks "
                "WHERE fecha >= %s AND fecha <= %s "
                "ORDER BY fecha ASC, hora_inicio ASC",
                (_dt.now().date(), _dt.now().date() + _td(days=60))
            ) or []
            pickup_blocks_list = [dict(r) for r in blocks_rows]
        except Exception:
            pickup_blocks_list = []
        return render_template("admin/marketing.html",
                               settings=settings(),
                               login_imagenes=login_imgs,
                               pickup_blocks=pickup_blocks_list)


    @app.route("/retiros/bloqueos/nuevo", methods=["POST"])
    @require_permission("admin")
    def pickup_blocks_new():
        """Crea un bloqueo de día/franja en pickup_blocks."""
        fecha = (request.form.get("fecha") or "").strip()
        hora_ini = (request.form.get("hora_inicio") or "").strip() or None
        hora_fin = (request.form.get("hora_fin") or "").strip() or None
        motivo = (request.form.get("motivo") or "").strip()[:200]
        if not fecha:
            flash("La fecha es obligatoria.", "danger")
            return redirect(url_for("marketing_settings"))
        try:
            mysql_execute(
                "INSERT INTO pickup_blocks (fecha, hora_inicio, hora_fin, motivo, created_by) "
                "VALUES (%s, %s, %s, %s, %s)",
                (fecha, hora_ini, hora_fin, motivo,
                 g.user["nombre"] if getattr(g,"user",None) else None)
            )
            flash("Bloqueo creado correctamente.", "success")
        except Exception as exc:
            flash(f"Error al crear bloqueo: {exc}", "danger")
        return redirect(url_for("marketing_settings"))


    @app.route("/retiros/bloqueos/batch", methods=["POST"])
    @require_permission("admin")
    def pickup_blocks_new_batch():
        """Crea múltiples bloqueos en una sola operación.

        Recibe del formulario:
          fecha: YYYY-MM-DD (obligatorio)
          motivo: str (opcional)
          slots[]: lista de pares "HH:MM-HH:MM" (uno por franja seleccionada)
          full_day: si está en 1, ignora slots y bloquea día completo

        Si no se envían slots ni full_day, falla amablemente.
        """
        fecha = (request.form.get("fecha") or "").strip()
        motivo = (request.form.get("motivo") or "").strip()[:200]
        if not fecha:
            flash("La fecha es obligatoria.", "danger")
            return redirect(url_for("marketing_settings") + "#bloqueos")
        full_day = request.form.get("full_day")
        slots = request.form.getlist("slots[]") or request.form.getlist("slots")
        creado_por = g.user["nombre"] if getattr(g, "user", None) else None

        if not full_day and not slots:
            print(f"[marketing/bloqueos] POST vacío: fecha={fecha} full_day={full_day!r} slots={slots} form_keys={list(request.form.keys())}")
            flash("⚠️ No se seleccionó ninguna franja ni 'día completo'. Marca al menos una franja antes de aplicar.", "warning")
            return redirect(url_for("marketing_settings") + "#bloqueos")

        inserted = 0
        skipped = 0
        errors = []
        try:
            if full_day:
                mysql_execute(
                    "INSERT INTO pickup_blocks (fecha, hora_inicio, hora_fin, motivo, created_by) "
                    "VALUES (%s, NULL, NULL, %s, %s)",
                    (fecha, motivo or "Día completo bloqueado", creado_por)
                )
                inserted = 1
            else:
                for s in slots:
                    if "-" not in s:
                        skipped += 1
                        continue
                    hi, hf = s.split("-", 1)
                    hi = hi.strip()
                    hf = hf.strip()
                    if not hi or not hf:
                        skipped += 1
                        continue
                    try:
                        mysql_execute(
                            "INSERT INTO pickup_blocks (fecha, hora_inicio, hora_fin, motivo, created_by) "
                            "VALUES (%s, %s, %s, %s, %s)",
                            (fecha, hi, hf, motivo, creado_por)
                        )
                        inserted += 1
                    except Exception as e_ins:
                        errors.append(f"{hi}-{hf}: {e_ins}")
                        print(f"[marketing/bloqueos] INSERT falló slot={hi}-{hf} fecha={fecha}: {e_ins}")

            if inserted == 0:
                msg = "❌ No se guardó ningún bloqueo."
                if errors:
                    msg += f" Errores: {'; '.join(errors[:3])}"
                elif skipped:
                    msg += f" Se descartaron {skipped} franja(s) con formato inválido."
                flash(msg, "danger")
            else:
                flash(f"✅ {inserted} franja(s) bloqueadas para {fecha}.", "success")
        except Exception as exc:
            print(f"[marketing/bloqueos] excepción general fecha={fecha}: {exc}")
            flash(f"Error al crear bloqueos: {exc}", "danger")
        return redirect(url_for("marketing_settings") + "#bloqueos")


    @app.route("/retiros/bloqueos/<int:bid>/eliminar", methods=["POST"])
    @require_permission("admin")
    def pickup_blocks_delete(bid):
        """Elimina un bloqueo."""
        try:
            mysql_execute("DELETE FROM pickup_blocks WHERE id=%s", (bid,))
            flash("Bloqueo eliminado.", "success")
        except Exception as exc:
            flash(f"Error al eliminar: {exc}", "danger")
        return redirect(url_for("marketing_settings"))

    @app.route("/retiros/adjuntos/<int:aid>")
    @require_permission("view")
    def pickup_attachment(aid):
        row = mysql_fetchone(
            f"SELECT id, filename, original_name FROM `{ATT}` WHERE id=%s LIMIT 1",
            (aid,)
        )
        if not row:
            return "No encontrado", 404
        # Defensa anti path-traversal: si por alguna razón el filename
        # en BD viene con `..` o `/`, rechazamos el download. Los uploads
        # son saneados con `secure_filename` al guardar, así que esto
        # solo dispara en caso de manipulación directa de la BD.
        fname = row.get("filename") or ""
        if not fname or ".." in fname or "/" in fname or "\\" in fname:
            return "Archivo no válido", 400
        # Sanitizar también el download_name (lo que ve el cliente al
        # descargar): si el original_name tiene caracteres peligrosos,
        # caemos al filename normalizado.
        orig = row.get("original_name") or fname
        if any(c in orig for c in ("..", "/", "\\", "\x00")):
            orig = fname
        return send_from_directory(upload_dir, fname, as_attachment=False, download_name=orig)

    # ══════════════════════════════════════════════════════════════════
    #  CALENDARIO OPERATIVO — vista por día/franja con capacidad
    # ══════════════════════════════════════════════════════════════════

    def _slot_label(time_from):
        """Devuelve etiqueta legible de la franja (HH:MM)."""
        try: return str(time_from)[:5]
        except Exception: return "00:00"

    @app.route("/retiros/calendario")
    @require_permission("retiros")
    def pickup_calendar():
        """Vista calendario operativo de retiros."""
        return render_template("retiros/calendario.html",
            statuses=PICKUP_STATUS,
            settings=settings(),
        )

    @app.route("/retiros/api/calendario")
    @require_permission("view")
    def pickup_calendar_api():
        """Devuelve solicitudes de retiro entre 2 fechas con agregados por día/franja.

        Query params: from=YYYY-MM-DD&to=YYYY-MM-DD (defaults: hoy → +14 días)
        Response: {
          settings: {capacidad, horarios, etc.},
          dias: {
            "2026-05-08": {
              total_picks: N,
              total_kg: N,
              total_m3: N,
              ocupacion_pct: 0..100,
              slots: {
                "09:00": {picks: [...], total_kg, total_m3, capacidad_kg, capacidad_m3, ocupados, max}
              },
              picks: [{id, code, status, customer, kg, m3, time_from, time_to, doc, ...}]
            }
          }
        }
        """
        from datetime import datetime as _dt, timedelta as _td
        cfg = settings()
        try:
            d_from = _dt.strptime(request.args.get("from") or _dt.now().date().isoformat(), "%Y-%m-%d").date()
        except Exception:
            d_from = _dt.now().date()
        try:
            d_to   = _dt.strptime(request.args.get("to") or (d_from + _td(days=14)).isoformat(), "%Y-%m-%d").date()
        except Exception:
            d_to   = d_from + _td(days=14)

        rows = mysql_fetchall(
            f"""SELECT id, code, status, customer_name, customer_rut, contact_name,
                       contact_phone, contact_email, extra_emails, document_type, document_number,
                       requested_date, requested_time_from, requested_time_to,
                       proposed_date, proposed_time_from, proposed_time_to,
                       confirmed_date, confirmed_time_from, confirmed_time_to,
                       total_packages, total_weight_kg, total_volumetric_weight, total_volume_m3,
                       pickup_person_name, pickup_person_phone, pickup_person_relation,
                       information_quality_score, risk_score, observations, internal_notes,
                       public_token, created_at
                FROM `{REQ}`
                WHERE (requested_date BETWEEN %s AND %s
                       OR confirmed_date BETWEEN %s AND %s
                       OR proposed_date  BETWEEN %s AND %s)
                ORDER BY requested_date ASC, requested_time_from ASC""",
            (d_from, d_to, d_from, d_to, d_from, d_to)
        )

        # Daniel 2026-05-24 — agenda interna inteligente:
        # Cargamos emails/teléfonos detectados en docs ERP para cada retiro,
        # para que el operador vea ⚠ cuando hay diferencia con lo declarado
        # por el cliente y pueda agregarlos a CC con un click.
        erp_contacts_map = {}  # rid -> {emails:[str], phones:[str]}
        if rows:
            try:
                rids = [int(r["id"]) for r in rows]
                if rids:
                    in_clause = ",".join(["%s"] * len(rids))
                    doc_rows = mysql_fetchall(
                        f"""SELECT request_id, email_cliente_erp, telefono_cliente_erp
                            FROM pickup_request_docs
                            WHERE request_id IN ({in_clause})""",
                        tuple(rids)
                    ) or []
                    for dr in doc_rows:
                        rid_k = int(dr.get("request_id") or 0)
                        bucket = erp_contacts_map.setdefault(rid_k, {"emails": set(), "phones": set()})
                        em = (dr.get("email_cliente_erp") or "").strip().lower()
                        if em and "@" in em:
                            bucket["emails"].add(em)
                        ph = re.sub(r"[^0-9+]", "", str(dr.get("telefono_cliente_erp") or ""))
                        if ph and len(ph) >= 8:
                            bucket["phones"].add(ph)
            except Exception:
                # Si la tabla/columnas no existen aún en este entorno, seguimos sin enriquecer
                erp_contacts_map = {}

        # Capacidades
        # FASE 4 / Juan Daniel 2026-06-05: capacidad ÚNICA = parallel_capacity
        # (default 2 = "dos agendas"). El calendario mostraba X/max_picks_per_slot
        # (legacy 5), inconsistente con el cliente que ya usa 2. Ahora el operador
        # ve el DOBLE cupo (X/2) igual que el público. Mismo patrón que _validar_disponibilidad_slot.
        _pc_cal = cfg.get("parallel_capacity")
        if _pc_cal is not None and str(_pc_cal).strip():
            max_picks_slot = int(_pc_cal)
        else:
            max_picks_slot = int(cfg.get("max_picks_per_slot") or 2)
        max_kg_slot    = float(cfg.get("max_kg_per_slot") or 500)
        max_m3_slot    = float(cfg.get("max_m3_per_slot") or 5)
        max_picks_day  = int(cfg.get("max_picks_per_day") or 30)
        # v3 (Daniel 2026-05-24): horario único en bloques de 30 min.
        slot_min       = int(cfg.get("slot_minutes") or 30)

        dias = {}
        for r in rows:
            # Fecha efectiva: confirmed > proposed > requested
            fecha = r.get("confirmed_date") or r.get("proposed_date") or r.get("requested_date")
            tf    = r.get("confirmed_time_from") or r.get("proposed_time_from") or r.get("requested_time_from")
            tt    = r.get("confirmed_time_to") or r.get("proposed_time_to") or r.get("requested_time_to")
            if not fecha: continue
            fecha_str = fecha.isoformat() if hasattr(fecha,"isoformat") else str(fecha)
            slot_lbl  = _slot_label(tf)

            if fecha_str not in dias:
                dias[fecha_str] = {
                    "total_picks": 0, "total_kg": 0, "total_m3": 0,
                    "max_picks_day": max_picks_day,
                    "ocupacion_pct": 0,
                    "slots": {},
                    "picks": [],
                }
            d = dias[fecha_str]
            kg = float(r.get("total_weight_kg") or 0)
            m3 = float(r.get("total_volume_m3") or 0)
            d["total_picks"] += 1
            d["total_kg"]    += kg
            d["total_m3"]    += m3

            if slot_lbl not in d["slots"]:
                d["slots"][slot_lbl] = {
                    "ocupados": 0, "total_kg": 0, "total_m3": 0,
                    "max_picks": max_picks_slot,
                    "max_kg":    max_kg_slot,
                    "max_m3":    max_m3_slot,
                    "picks":     [],
                }
            s = d["slots"][slot_lbl]
            s["ocupados"] += 1
            s["total_kg"] += kg
            s["total_m3"] += m3

            # Daniel 2026-05-24 — qué tipo de fecha estamos pintando:
            #   "confirmed" = ambos acordaron (verde)
            #   "proposed"  = ILUS propuso, esperando cliente (ámbar)
            #   "requested" = lo que pidió el cliente (azul)
            # Esto le permite al operador distinguir VISUALMENTE en la agenda
            # sin tener que abrir cada tarjeta.
            if r.get("confirmed_date"):
                kind_date = "confirmed"
            elif r.get("proposed_date"):
                kind_date = "proposed"
            else:
                kind_date = "requested"

            # ¿El cliente RECHAZÓ una propuesta nuestra? (disputa)
            # Lo detectamos por status: si está en `reagendada` o `informacion_incompleta`
            # con propuesta pendiente, lo marcamos como en disputa.
            is_disputed = r.get("status") in ("reagendada", "informacion_incompleta")

            # Enriquecer con contactos del ERP detectados en los docs asociados
            erp_info = erp_contacts_map.get(int(r["id"]), {"emails": set(), "phones": set()})
            declared_email = (r.get("contact_email") or "").strip().lower()
            declared_phone = re.sub(r"[^0-9+]", "", str(r.get("contact_phone") or ""))
            extras_raw = (r.get("extra_emails") or "").strip()
            extras_list = [e.strip().lower() for e in re.split(r"[,;\s]+", extras_raw) if e.strip()]
            erp_emails_sorted = sorted(erp_info["emails"])
            erp_phones_sorted = sorted(erp_info["phones"])
            # Sugerencias = los que están en ERP pero NO en (declarado + extras)
            erp_email_sugeridos = [e for e in erp_emails_sorted
                                    if e != declared_email and e not in extras_list]
            erp_phone_sugeridos = [p for p in erp_phones_sorted if p != declared_phone]

            pick = {
                "id":               r["id"],
                "code":             r.get("code"),
                "status":           r.get("status"),
                "status_label":     PICKUP_STATUS.get(r.get("status"), r.get("status")),
                "status_color":     PICKUP_STATUS_COLORS.get(r.get("status"), "secondary"),
                "customer_name":    r.get("customer_name"),
                "customer_rut":     r.get("customer_rut"),
                "contact_name":     r.get("contact_name"),
                "contact_phone":    r.get("contact_phone"),
                "contact_email":    r.get("contact_email"),
                "extra_emails":     extras_list,
                "document_type":    r.get("document_type"),
                "document_number":  r.get("document_number"),
                "fecha":            fecha_str,
                "time_from":        _slot_label(tf),
                "time_to":          _slot_label(tt),
                "kind_date":        kind_date,
                "is_disputed":      is_disputed,
                "is_proposed":      bool(r.get("proposed_date") and not r.get("confirmed_date")),
                "is_confirmed":     bool(r.get("confirmed_date")),
                "total_packages":   int(r.get("total_packages") or 0),
                "total_weight_kg":  kg,
                "total_volume_m3":  m3,
                "pickup_person":    r.get("pickup_person_name"),
                "pickup_person_phone": r.get("pickup_person_phone"),
                "quality_score":    int(r.get("information_quality_score") or 0),
                "risk_score":       int(r.get("risk_score") or 0),
                "public_token":     r.get("public_token"),
                # ⚠ Indicadores de "diferencia de datos" cliente vs ERP
                "erp_emails_suggested": erp_email_sugeridos,
                "erp_phones_suggested": erp_phone_sugeridos,
                "has_data_mismatch":    bool(erp_email_sugeridos or erp_phone_sugeridos),
            }
            s["picks"].append(pick)
            d["picks"].append(pick)

        # Calcular % ocupación día
        for fecha_str, d in dias.items():
            d["ocupacion_pct"] = round(min(100, (d["total_picks"] * 100) / max(1, max_picks_day)))
            for slot_lbl, s in d["slots"].items():
                pct_picks = (s["ocupados"] * 100) / max(1, s["max_picks"])
                pct_kg    = (s["total_kg"] * 100) / max(1, s["max_kg"])
                pct_m3    = (s["total_m3"] * 100) / max(1, s["max_m3"])
                s["pct"]  = round(min(100, max(pct_picks, pct_kg, pct_m3)))
                s["pct_picks"] = round(min(100, pct_picks))
                s["pct_kg"]    = round(min(100, pct_kg))
                s["pct_m3"]    = round(min(100, pct_m3))

        return jsonify({
            "from": d_from.isoformat(),
            "to":   d_to.isoformat(),
            "settings": {
                "open_time":      str(cfg.get("open_time") or "09:00:00")[:5],
                "close_time":     str(cfg.get("close_time") or "16:30:00")[:5],
                "work_days":      cfg.get("work_days") or "1,2,3,4,5",
                "holidays":       cfg.get("holidays") or "",
                "slot_minutes":   slot_min,
                "max_picks_slot": max_picks_slot,
                "max_kg_slot":    max_kg_slot,
                "max_m3_slot":    max_m3_slot,
                "max_picks_day":  max_picks_day,
            },
            "dias": dias,
        })

    @app.route("/retiros/api/disponibilidad-publica")
    @_rate_limited("pickup_disponibilidad_publica", max_attempts=120, window_seconds=60, methods=("GET",))
    def pickup_disponibilidad_publica():
        """Devuelve disponibilidad de slots para los próximos 30 días.

        FUENTE DE VERDAD MULTI-ROL (Daniel mayo 2026):
        - Cliente externo (sin sesión) recibe shape pública: slots con estado
          + ocupación numérica, SIN nombres ni IDs de clientes.
        - Operador autenticado puede pedir `?include_owners=1` para ver quién
          ocupa cada slot (RET-XXXXXX, customer_name, detail_url). Esto es el
          mismo calendario que ve el cliente, pero enriquecido con info
          interna para que el equipo confirme "quién tiene ese horario".

        Parámetros opcionales:
          - `?include_owners=1` (solo aplica si el caller tiene `g.user` con
            permiso `retiros`, `admin` o `superadmin`. Si no, se ignora
            silenciosamente — no devolvemos 403 para preservar el contrato
            público).
          - `?date=YYYY-MM-DD` (rango de un solo día: útil para la vista
            interna del retiro o cuando un operador hace foco en una fecha).
            Por defecto: mañana + 30 días.

        Modelo de slots (v3, Daniel 2026-05-24):
        - Cada bloque visible dura 30 min (slot_minutes=30) — SIN EXCEPCIÓN.
        - El cliente puede seleccionar N bloques contiguos como rango.
        - Duración real del retiro = N × 30 min.
        - Mañana: 09:00 – 12:30 → 7 bloques agendables.
        - Tarde:  14:00 – 16:30 → 5 bloques agendables.
        - Colación 12:30 – 14:00 BLOQUEADA TOTAL (no se ven slots agendables).
        - close_time es el FIN admitido; buffer_cierre_min=0 (sin buffer).
        - `parallel_capacity` (default 2): cuántos retiros caben SIMULTÁNEAMENTE
          en el mismo bloque. Si ya hay 2 confirmados, queda "completo".
        - Si hay 1 retiro confirmado en un bloque con capacidad 2, queda como
          "ocupado" (clickeable, mostrando 1/2).

        Respuesta por slot:
            estado: "disponible" | "colacion" | "ocupado" | "completo" |
                    "fuera_horario" | "bloqueado"
            puede_iniciar / puede_finalizar: bool
            ocupacion_actual, capacidad_max, time_from, time_to
            owners: [{request_id, code, customer_name, status, status_label,
                      detail_url}]  # solo si include_owners autorizado
        """
        from datetime import datetime as _dt, timedelta as _td, date as _date_cls

        # ── Detectar contexto: ¿operador autenticado pidió owners? ─────
        # Defense in depth: validamos sesión + permiso ANTES de tocar nada.
        # Si el parámetro viene pero el usuario no califica, lo ignoramos
        # silenciosamente (no 403, para no romper el contrato público).
        wants_owners = request.args.get("include_owners") in ("1", "true", "True", "yes")
        u = getattr(g, "user", None)
        perms = getattr(g, "permissions", {}) or {}
        is_operator = bool(u) and (
            perms.get("retiros") or perms.get("admin") or perms.get("superadmin")
        )
        include_owners = bool(wants_owners and is_operator)

        # ── Rango de fechas: opcionalmente acotar a un solo día ───────
        single_date_req = (request.args.get("date") or "").strip()
        single_date_obj = None
        if single_date_req:
            try:
                single_date_obj = _date_cls.fromisoformat(single_date_req[:10])
            except Exception:
                single_date_obj = None

        # ── EXCLUIR el propio retiro del conteo de ocupación ──────────
        # Daniel 2026-05-25 (bug "No encontramos tu bloque"):
        # El stepper de confirmación re-verifica disponibilidad ANTES del
        # POST. Pero el propio retiro YA ocupa su bloque propuesto, así que
        # su propia reserva se contaba como ocupación y, con
        # parallel_capacity bajo (o si comparte slot con otro), el bloque
        # salía "completo"/"ocupado lleno" y el cliente NO podía confirmar
        # SU PROPIA hora. La solución: si el caller envía el token público
        # de su retiro (`exclude_token`), excluimos ese id del conteo —
        # exactamente igual que hace `_validar_disponibilidad_slot` y la
        # transacción FOR UPDATE del POST confirm (`id <> %s`).
        #
        # Seguridad: el token público solo lo conoce el dueño del retiro y
        # solo puede excluir SU PROPIO id. La garantía real anti-doble-cupo
        # vive en la transacción del POST confirm (FOR UPDATE), no en este
        # pre-chequeo de UX.
        exclude_token = (request.args.get("exclude_token") or "").strip()
        exclude_request_id = None
        if exclude_token and re.match(r"^[A-Za-z0-9_\-]{16,200}$", exclude_token):
            try:
                _own = mysql_fetchone(
                    f"SELECT id FROM `{REQ}` WHERE public_token=%s LIMIT 1",
                    (exclude_token,),
                )
                if _own and _own.get("id"):
                    exclude_request_id = int(_own["id"])
            except Exception:
                exclude_request_id = None

        # ── Operador autenticado: excluir el retiro EN GESTIÓN por id ──
        # FIX Daniel 2026-06-15 (bug "slot lleno por contarse a sí mismo"):
        # Al proponer hora para el retiro X desde /retiros/<X>, su PROPIA
        # reserva (requested/proposed/confirmed) estaba contando contra el
        # cupo del bloque. Con parallel_capacity=2, un bloque con SOLO el
        # propio retiro + 1 cliente más salía "completo" (2/2) y el operador
        # no podía agendar la hora que el mismo cliente pidió.
        # Solución: el operador puede pasar ?exclude_id=<id> y excluimos ese
        # retiro del conteo — mismo criterio que `_validar_disponibilidad_slot`
        # y la transacción FOR UPDATE del POST (id <> %s). Solo se honra para
        # callers autenticados como operador (is_operator); el cliente público
        # nunca puede excluir por id (sigue usando exclude_token de su propio
        # retiro). La garantía real anti-doble-cupo vive en el POST, no aquí.
        if exclude_request_id is None and is_operator:
            _exc_id_raw = (request.args.get("exclude_id") or "").strip()
            if _exc_id_raw.isdigit():
                exclude_request_id = int(_exc_id_raw)

        # Cache hit: SOLO sirve para la shape pública sin filtro de fecha
        # (el grid de 30 días es global y no cambia por cliente). Las
        # llamadas internas con owners o con date= bypasean cache para no
        # mezclar shapes — la frecuencia es mucho menor (un solo operador
        # mirando un día). Si se pidió excluir un retiro propio NUNCA usamos
        # cache (el conteo es específico de ese cliente).
        import time as _time_local
        use_cache = (not include_owners) and (single_date_obj is None) and (exclude_request_id is None)
        if use_cache and _DISPO_CACHE["payload"] is not None and \
                (_time_local.time() - _DISPO_CACHE["ts"]) < _DISPO_TTL:
            return jsonify(_DISPO_CACHE["payload"])

        cfg = settings()
        if single_date_obj is not None:
            d_from = single_date_obj
            d_to   = single_date_obj
        else:
            d_from = _dt.now().date() + _td(days=1)  # Mañana en adelante
            d_to   = d_from + _td(days=30)

        # Capacidades: parallel_capacity (default 2) sustituye a
        # max_picks_per_slot para el calendario público. max_picks_per_slot
        # legacy se conserva para el admin interno.
        parallel_capacity = int(cfg.get("parallel_capacity") or 2)
        max_kg_slot    = float(cfg.get("max_kg_per_slot") or 500)
        max_m3_slot    = float(cfg.get("max_m3_per_slot") or 5)
        max_picks_day  = int(cfg.get("max_picks_per_day") or 30)

        # Daniel 2026-05-25: HORARIO DEFINITIVO según la operación real.
        # "Cuando marqué hasta las doce y treinta, es hasta las doce y
        # treinta. No inventes, que es hasta la una, porque yo lo tengo
        # bloqueado, esa media hora la quiero tener en margen para mí,
        # para ordenar, para ver si llegó alguien desordenado."
        #
        #   Mañana: bloques cada 30 min desde 09:00 hasta 12:30.
        #           Últimos inicios admitidos: ..., 11:30, 12:00
        #           (último bloque 12:00-12:30).
        #   Buffer interno: 12:30-13:00 (la bodega usa este rato para
        #           atender desordenados que no se agendaron).
        #   Colación: 13:00 a 14:00 (no se ofrece para retiros).
        #   Tarde:  bloques cada 30 min desde 14:00 hasta 17:00.
        #           Últimos inicios admitidos: ..., 16:00, 16:30
        #           (último bloque 16:30-17:00).
        #
        # Total: 7 bloques mañana + 6 bloques tarde = 13 bloques agendables.
        slot_dur  = 30
        slot_step = 30
        buffer_cierre_min = 0

        # 12:30 marca el INICIO del bloqueado (buffer + colación).
        # Esto saca del calendario los slots 12:30, 13:00, 13:30.
        lunch_s_str = "12:30"
        lunch_e_str = "14:00"
        try:
            lH,lM = [int(x) for x in lunch_s_str.split(":")]
            leH,leM = [int(x) for x in lunch_e_str.split(":")]
            lunch_start_min = lH*60 + lM
            lunch_end_min   = leH*60 + leM
        except Exception:
            lunch_start_min = lunch_end_min = -1

        # Días hábiles
        work_days = {int(x) for x in (cfg.get("work_days") or "1,2,3,4,5").split(",") if x.strip().isdigit()}
        holidays  = {h.strip() for h in (cfg.get("holidays") or "").replace(";",",").split(",") if h.strip()}

        # Daniel 2026-05-24: HARDCODED el horario para clientes.
        # Apertura 09:00 — cierre 17:00 (último bloque 16:30 → 17:00).
        oH,oM = 9, 0
        cH,cM = 17, 0
        open_min  = oH*60 + oM
        close_min = cH*60 + cM
        # ultimo_fin_admitido = close_min - buffer_cierre (v3: buffer=0 → 16:30)
        ultimo_fin_admitido = close_min - buffer_cierre_min

        # Generar slots base del día (mismos para todas las fechas)
        slots_base = []
        m = 0
        while (open_min + m + slot_dur) <= ultimo_fin_admitido:
            t = open_min + m
            t_end = t + slot_dur
            # Detectar solapamiento con bloque de colación
            is_lunch = (lunch_start_min < lunch_end_min and
                        not (t_end <= lunch_start_min or t >= lunch_end_min))
            slots_base.append({
                "time_from":  f"{t//60:02d}:{t%60:02d}",
                "time_to":    f"{t_end//60:02d}:{t_end%60:02d}",
                "start_min":  t,
                "end_min":    t_end,
                "is_lunch":   is_lunch,
            })
            m += slot_step

        # Marcadores derivados (para el resumen del frontend)
        ultimo_inicio = slots_base[-1]["time_from"] if slots_base else ""
        ultimo_fin    = slots_base[-1]["time_to"]   if slots_base else ""
        ultimo_inicio_antes_colacion = ""
        primer_inicio_post_colacion  = ""
        if lunch_start_min < lunch_end_min:
            for s in slots_base:
                if s["end_min"] <= lunch_start_min:
                    ultimo_inicio_antes_colacion = s["time_from"]
                if not primer_inicio_post_colacion and s["start_min"] >= lunch_end_min:
                    primer_inicio_post_colacion = s["time_from"]

        # Calcular ocupación: 1 sola query a pickup_requests.
        # Para vista interna seleccionamos id+code+customer_name para que el
        # operador vea quién ocupa cada slot. Para shape pública omitimos esos
        # campos (privacidad + payload más liviano para clientes externos).
        # Cláusula opcional para excluir el propio retiro del conteo (ver
        # bloque exclude_token más arriba). Mantiene la query parametrizada.
        _excl_clause = ""
        _excl_params = []
        if exclude_request_id is not None:
            _excl_clause = " AND id <> %s"
            _excl_params = [exclude_request_id]
        if include_owners:
            rows = mysql_fetchall(
                f"""SELECT id, code, customer_name, contact_name,
                           requested_date, requested_time_from,
                           confirmed_date, confirmed_time_from,
                           proposed_date, proposed_time_from,
                           total_weight_kg, total_volume_m3, status
                    FROM `{REQ}`
                    WHERE (requested_date BETWEEN %s AND %s
                           OR confirmed_date BETWEEN %s AND %s
                           OR proposed_date  BETWEEN %s AND %s)
                      AND status NOT IN ('rechazada','cerrada','fallida'){_excl_clause}""",
                (d_from, d_to, d_from, d_to, d_from, d_to, *_excl_params)
            ) or []
        else:
            rows = mysql_fetchall(
                f"""SELECT requested_date, requested_time_from,
                           confirmed_date, confirmed_time_from,
                           proposed_date, proposed_time_from,
                           total_weight_kg, total_volume_m3, status
                    FROM `{REQ}`
                    WHERE (requested_date BETWEEN %s AND %s
                           OR confirmed_date BETWEEN %s AND %s
                           OR proposed_date  BETWEEN %s AND %s)
                      AND status NOT IN ('rechazada','cerrada','fallida'){_excl_clause}""",
                (d_from, d_to, d_from, d_to, d_from, d_to, *_excl_params)
            ) or []

        ocupacion = {}  # {fecha_str: {slot_from: {ocupados, kg, m3, owners[]}}}
        for r in rows:
            fecha = r.get("confirmed_date") or r.get("proposed_date") or r.get("requested_date")
            if not fecha: continue
            tf = r.get("confirmed_time_from") or r.get("proposed_time_from") or r.get("requested_time_from")
            if not tf: continue
            fecha_str = fecha.isoformat() if hasattr(fecha,"isoformat") else str(fecha)
            slot = str(tf)[:5]
            if fecha_str not in ocupacion: ocupacion[fecha_str] = {}
            if slot not in ocupacion[fecha_str]:
                ocupacion[fecha_str][slot] = {"ocupados":0,"kg":0,"m3":0,"owners":[]}
            ocupacion[fecha_str][slot]["ocupados"] += 1
            ocupacion[fecha_str][slot]["kg"]       += float(r.get("total_weight_kg") or 0)
            ocupacion[fecha_str][slot]["m3"]       += float(r.get("total_volume_m3") or 0)
            if include_owners and r.get("id"):
                st = r.get("status") or ""
                ocupacion[fecha_str][slot]["owners"].append({
                    "request_id":   int(r["id"]),
                    "code":         r.get("code") or f"RET-{int(r['id']):06d}",
                    "customer_name": (r.get("customer_name") or r.get("contact_name") or "Cliente")[:80],
                    "status":       st,
                    "status_label": PICKUP_STATUS.get(st, st),
                    "status_color": PICKUP_STATUS_COLORS.get(st, "secondary"),
                    "detail_url":   url_for("pickup_detail", rid=int(r["id"])),
                })

        # Bloqueos manuales (tabla pickup_blocks)
        blocks_by_date = {}
        try:
            blk_rows = mysql_fetchall(
                f"""SELECT fecha, hora_inicio, hora_fin, motivo
                    FROM pickup_blocks
                    WHERE fecha BETWEEN %s AND %s""",
                (d_from, d_to)
            ) or []
            for b in blk_rows:
                f = b["fecha"]
                fs = f.isoformat() if hasattr(f,"isoformat") else str(f)
                blocks_by_date.setdefault(fs, []).append({
                    "hora_inicio": str(b.get("hora_inicio") or "")[:5],
                    "hora_fin":    str(b.get("hora_fin") or "")[:5],
                    "motivo":      b.get("motivo") or "",
                })
        except Exception:
            pass

        # FASE 2 (2026-05-29): horizonte temporal. No ofrecer slots pasados
        # (criterio para TODOS) ni que violen min_notice (solo cliente público;
        # el operador interno puede agendar más cerca vía pickup_create_proposal).
        # Calculado en hora Chile para coincidir con el validador central.
        _now_cl = _now_chile()
        try:
            _min_notice_h = int(cfg.get("min_notice_hours") or 24)
        except (TypeError, ValueError):
            _min_notice_h = 24
        # Daniel 2026-06-16: BUFFER de visualización (60 min). El calendario NO
        # ofrece bloques que estén a punto de cruzar el límite de min_notice
        # mientras el cliente llena el formulario. CAUSA del bug: la página
        # carga la disponibilidad una vez, pero el cliente puede tardar minutos
        # en enviar; un bloque que estaba a 24h05m al cargar quedaba a 23h50m al
        # enviar → el calendario lo mostraba disponible pero el validador de 24h
        # lo rechazaba ("Se requieren al menos 24 horas"). Con el buffer, lo que
        # se OFRECE pasa cómodamente el submit (display más estricto que el POST).
        _display_buffer_min = 60
        _min_dt = _now_cl + _td(hours=_min_notice_h) + _td(minutes=_display_buffer_min)

        # Generar disponibilidad por día.
        # Si pidieron un día específico (?date=) iteramos solo ese; si no,
        # generamos los 31 días desde d_from hasta d_to inclusive.
        dias = {}
        total_dias = (d_to - d_from).days + 1
        # Feriados legales de Chile (auto) para los años del horizonte, además del
        # config. (Juan Daniel 2026-06-05: bloquear feriados chilenos en el calendario.)
        for _yr in {d_from.year, d_to.year}:
            holidays = holidays | _chile_holidays(_yr)
        for offset in range(total_dias):
            d = d_from + _td(days=offset)
            iso = d.isoformat()
            disp_dia = (d.isoweekday() in work_days) and (iso not in holidays)

            day_blocks = blocks_by_date.get(iso, [])
            full_day_block = any(not b["hora_inicio"] for b in day_blocks)
            full_day_motivo = next((b["motivo"] for b in day_blocks if not b["hora_inicio"]), "")

            day_picks = sum(s["ocupados"] for s in ocupacion.get(iso,{}).values())
            full_dia = day_picks >= max_picks_day
            dia_disponible = disp_dia and not full_dia and not full_day_block

            dias[iso] = {
                "fecha":      iso,
                "disponible": dia_disponible,
                "razon":      "" if dia_disponible else (
                                "Día bloqueado: " + full_day_motivo if full_day_block else
                                "Feriado en Chile" if iso in holidays else
                                "Fin de semana" if d.isoweekday() not in work_days else
                                "Día completo"
                              ),
                "slots":      [],
                "ocupacion_pct": min(100, round((day_picks * 100) / max(1, max_picks_day))),
            }
            if not dia_disponible: continue

            for slot_info in slots_base:
                slot_hora      = slot_info["time_from"]
                slot_to        = slot_info["time_to"]
                is_lunch       = slot_info["is_lunch"]
                slot_start_min = slot_info["start_min"]
                slot_end_min   = slot_info["end_min"]
                ocup = ocupacion.get(iso,{}).get(slot_hora, {"ocupados":0,"kg":0,"m3":0,"owners":[]})

                # Bloqueo manual por franja
                manual_block = None
                for b in day_blocks:
                    if not b["hora_inicio"]: continue
                    try:
                        bH, bM = [int(x) for x in b["hora_inicio"].split(":")]
                        bEH, bEM = [int(x) for x in (b["hora_fin"] or "23:59").split(":")]
                        bs, be = bH*60+bM, bEH*60+bEM
                        if not (slot_end_min <= bs or slot_start_min >= be):
                            manual_block = b["motivo"] or "Bloqueado"
                            break
                    except Exception:
                        continue

                ocupacion_actual = int(ocup["ocupados"])
                kg_actual = float(ocup["kg"])
                m3_actual = float(ocup["m3"])

                # FASE 2 (2026-05-29): inicio del slot en hora Chile, para
                # filtrar pasado / min_notice (coincide con el validador central).
                slot_dt = datetime.combine(d, datetime.min.time()) + _td(minutes=slot_start_min)

                # Determinar estado
                if is_lunch:
                    estado = "colacion"
                    razon  = "Horario de colación"
                elif manual_block:
                    estado = "bloqueado"
                    razon  = manual_block
                elif slot_dt <= _now_cl:
                    estado = "no_disponible"
                    razon  = "El horario ya pasó"
                elif (not include_owners) and slot_dt < _min_dt:
                    estado = "no_disponible"
                    razon  = f"Requiere {_min_notice_h}h de anticipación"
                elif ocupacion_actual >= parallel_capacity or kg_actual >= max_kg_slot or m3_actual >= max_m3_slot:
                    estado = "completo"
                    razon  = f"Cupo lleno ({ocupacion_actual}/{parallel_capacity})"
                elif ocupacion_actual > 0:
                    estado = "ocupado"
                    razon  = f"Parcial ({ocupacion_actual}/{parallel_capacity})"
                else:
                    estado = "disponible"
                    razon  = ""

                # puede_iniciar / puede_finalizar: lo mismo en este modelo
                puede = (
                    estado in ("disponible", "ocupado")
                    and (slot_end_min <= ultimo_fin_admitido)
                )
                puede_iniciar   = puede
                puede_finalizar = puede

                slot_payload = {
                    "time_from":        slot_hora,
                    "time_to":          slot_to,
                    "estado":           estado,
                    "razon":            razon,
                    "ocupacion_actual": ocupacion_actual,
                    "capacidad_max":    parallel_capacity,
                    "puede_iniciar":    puede_iniciar,
                    "puede_finalizar":  puede_finalizar,
                    # ── compat con frontend legacy ──
                    "hora":       slot_hora,
                    "disponible": (estado == "disponible" or estado == "ocupado"),
                    "ocupados":   ocupacion_actual,
                    "max":        parallel_capacity,
                    "lunch":      is_lunch,
                }
                # Solo incluimos owners en shape interna (operador autenticado
                # + permiso). En shape pública NUNCA exponemos quién está en
                # cada slot — eso sería filtración de info de clientes.
                if include_owners:
                    slot_payload["owners"] = list(ocup.get("owners") or [])
                dias[iso]["slots"].append(slot_payload)

        _payload = {
            "from": d_from.isoformat(),
            "to":   d_to.isoformat(),
            "warehouse_name": cfg.get("warehouse_name"),
            "open_time":  str(cfg.get("open_time") or "09:00")[:5],
            "close_time": str(cfg.get("close_time") or "16:30")[:5],
            "slot_minutes": slot_dur,
            "slot_step":    slot_step,
            "lunch_start":  lunch_s_str,
            "lunch_end":    lunch_e_str,
            "operacion": {
                "open_time":  str(cfg.get("open_time") or "09:00")[:5],
                "close_time": str(cfg.get("close_time") or "16:30")[:5],
                "ultimo_inicio": ultimo_inicio,
                "ultimo_fin":    ultimo_fin,
                "lunch_start":   lunch_s_str,
                "lunch_end":     lunch_e_str,
                "ultimo_inicio_antes_colacion": ultimo_inicio_antes_colacion,
                "primer_inicio_post_colacion":  primer_inicio_post_colacion,
                "slot_minutes":      slot_dur,
                "parallel_capacity": parallel_capacity,
                "buffer_cierre_min": buffer_cierre_min,
            },
            "dias": dias,
        }
        # Guardar en cache SOLO la shape pública 30-días sin owners ni date=.
        # Las llamadas internas son siempre frescas (operador puede confirmar
        # un retiro en otro tab y al refrescar el calendario debe verlo).
        if use_cache:
            _DISPO_CACHE["payload"] = _payload
            _DISPO_CACHE["ts"]      = _time_local.time()
        return jsonify(_payload)

    @app.route("/retiros/api/<int:rid>/full")
    @require_permission("view")
    def pickup_full_info(rid):
        """Ficha completa de un retiro: datos + bultos + adjuntos + logs + ERP.

        Perf (Daniel mayo 2026): SELECT explícito — antes hacía SELECT *
        que serializaba `doc_erp_data` MEDIUMTEXT (puede ser ~200KB) sobre
        la red sin razón. El endpoint enriquece con ERP por su cuenta.
        """
        r = mysql_fetchone(
            f"""SELECT id, code, status, document_type, document_number,
                       customer_name, customer_rut, contact_name, contact_email,
                       contact_phone, pickup_person_name, pickup_person_rut,
                       pickup_person_phone, pickup_person_relation,
                       requested_date, requested_time_from, requested_time_to,
                       proposed_date, proposed_time_from, proposed_time_to,
                       confirmed_date, confirmed_time_from, confirmed_time_to,
                       total_packages, total_weight_kg, total_volumetric_weight,
                       total_volume_m3, invoice_total_amount, observations,
                       internal_notes, public_token, signature_status,
                       created_at, updated_at, closed_at,
                       doc_validation_status, doc_validated_at, doc_validated_by,
                       doc_validation_notes, peso_real_kg, peso_vol_kg,
                       tiempo_estimado_min, information_quality_score, risk_score
                FROM `{REQ}` WHERE id=%s LIMIT 1""",
            (rid,)
        )
        if not r: return jsonify({"error":"No encontrado"}), 404
        d = dict(r)
        # Convertir tipos
        for k in ('requested_date','proposed_date','confirmed_date'):
            if d.get(k): d[k] = d[k].isoformat() if hasattr(d[k],'isoformat') else str(d[k])
        for k in ('requested_time_from','requested_time_to','proposed_time_from','proposed_time_to','confirmed_time_from','confirmed_time_to'):
            if d.get(k): d[k] = str(d[k])[:5]
        for k in ('created_at','updated_at','closed_at'):
            if d.get(k): d[k] = str(d[k])[:19]
        for k in ('total_weight_kg','total_volumetric_weight','total_volume_m3','invoice_total_amount'):
            if d.get(k) is not None: d[k] = float(d[k] or 0)

        # Bultos
        pkgs = mysql_fetchall(
            f"SELECT * FROM `{PKG}` WHERE request_id=%s ORDER BY package_number", (rid,)
        )
        d["packages"] = []
        for p in pkgs:
            pp = dict(p)
            for k in ('length_cm','width_cm','height_cm','weight_kg','volumetric_weight','volume_m3'):
                if pp.get(k) is not None: pp[k] = float(pp[k] or 0)
            d["packages"].append(pp)

        # Logs
        logs = mysql_fetchall(
            f"SELECT id, action, old_status, new_status, notes, actor_type, actor_name, created_at "
            f"FROM `{LOG}` WHERE request_id=%s ORDER BY created_at DESC LIMIT 30", (rid,)
        )
        d["logs"] = []
        for lg in logs:
            ll = dict(lg)
            if ll.get("created_at"): ll["created_at"] = str(ll["created_at"])[:19]
            d["logs"].append(ll)

        # Propuestas
        props = mysql_fetchall(
            f"SELECT id, proposed_by, date, time_from, time_to, message, status, created_at, answered_at "
            f"FROM `{PROP}` WHERE request_id=%s ORDER BY created_at DESC", (rid,)
        )
        d["proposals"] = []
        for p in props:
            pp = dict(p)
            if pp.get("date"): pp["date"] = pp["date"].isoformat() if hasattr(pp["date"],"isoformat") else str(pp["date"])
            for k in ("time_from","time_to"):
                if pp.get(k): pp[k] = str(pp[k])[:5]
            for k in ("created_at","answered_at"):
                if pp.get(k): pp[k] = str(pp[k])[:19]
            d["proposals"].append(pp)

        # Adjuntos
        atts = mysql_fetchall(
            f"SELECT id, original_name, mime_type, uploaded_by, created_at "
            f"FROM `{ATT}` WHERE request_id=%s ORDER BY created_at DESC", (rid,)
        )
        d["attachments"] = [dict(a) for a in atts]
        for a in d["attachments"]:
            if a.get("created_at"): a["created_at"] = str(a["created_at"])[:19]
            a["url"] = url_for("pickup_attachment", aid=a["id"])

        # Estado actual + transiciones
        d["status_label"] = PICKUP_STATUS.get(d.get("status"), d.get("status"))
        d["status_color"] = PICKUP_STATUS_COLORS.get(d.get("status"), "secondary")
        d["public_url"]   = url_for("pickup_public_tracking", token=d.get("public_token"), _external=False)

        # Intentar enriquecer con ERP (best-effort, sin romper si falla)
        d["erp"] = None
        try:
            # Mapping document_type → TIDO ERP (normalizado, sin duplicados)
            tido_map = {
                "factura":     "FCV",
                "boleta":      "BLV",
                "guia":        "GDV",
                "guia_despacho":"GDV",
                "nota_venta":  "VD",
                "venta_directa":"VD",
                "pedido":      "WEB",
                "cotizacion":  "COV",
            }
            doc_type = (d.get("document_type") or "").lower().replace(" ", "_")
            tido = tido_map.get(doc_type)
            nudo = (d.get("document_number") or "").strip()
            if tido and nudo:
                # Import directo en lugar del check frágil "in dir(__import__(...))"
                try:
                    from app import _cubicador_fetch
                except ImportError:
                    _cubicador_fetch = None
                    d["erp"] = {"error": "Motor ERP no disponible (import falló)"}

                if _cubicador_fetch:
                    try:
                        hdr, lineas = _cubicador_fetch(tido, nudo)
                        if hdr:
                            # hdr ya viene normalizado por _cubicador_fetch
                            # (cliente_nombre, cliente_rut, direccion, observaciones, etc.)
                            d["erp"] = {
                                "razon_social": (hdr.get("cliente_nombre") or hdr.get("NRAZON") or "").strip(),
                                "rut":          (hdr.get("cliente_rut") or hdr.get("NRUC") or "").strip(),
                                "direccion":    (hdr.get("direccion") or "").strip(),
                                "comuna":       (hdr.get("comuna") or "").strip(),
                                "telefono":     (hdr.get("telefono") or "").strip(),
                                "email":        (hdr.get("email") or "").strip(),
                                "observaciones": (hdr.get("observaciones") or "").strip(),
                                "tido":         tido,
                                "nudo":         nudo,
                                "lineas":       [{
                                    "sku":      (ln.get("sku") or "").strip(),
                                    "nombre":   (ln.get("descripcion_erp") or ln.get("nombre_app") or "").strip(),
                                    "cantidad": float(ln.get("cantidad") or 0),
                                    "es_zz":    bool(ln.get("es_zz")),
                                    "tiene_cubicaje": bool(ln.get("tiene_ficha") and ln.get("tiene_bultos")),
                                    "peso_kg_tot": float(ln.get("peso_kg_tot") or 0),
                                } for ln in (lineas or [])]
                            }
                        else:
                            d["erp"] = {"error": f"Documento {tido} {nudo} no encontrado en ERP"}
                    except Exception as exc:
                        d["erp"] = {"error": str(exc)[:200]}
        except Exception as exc:
            d["erp"] = {"error": str(exc)[:200]}

        return jsonify(d)
