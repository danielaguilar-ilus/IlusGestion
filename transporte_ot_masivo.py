"""
Carga masiva de OT FedEx — vinculación de N° de tracking a documentos.

Daniel sube un Excel con columnas [N° Factura, Tipo (FCV/BLV/...), N° Tracking
FedEx (OT), Observación]. El módulo:

  1) Hace match contra `transport_commitments` (nudo + tido opcional) y
     localiza el `transport_manifest_items.id` correspondiente.
  2) Detecta inteligentemente la situación de cada fila:
     · ok_nuevo    → item sin tracking previo, listo para asignar.
     · ok_re_envio → item ya tenía OTRA OT (re-envío legítimo, requiere
                     confirmación explícita del operador para sobrescribir).
     · ya_existe   → item ya tiene EXACTAMENTE este tracking — se omite
                     idempotentemente.
     · no_encontrada → la factura no existe en BD (necesita sync ERP).
     · ot_duplicada  → ese tracking ya está asignado a OTRA factura.
     · invalido      → datos malformados (sin nº de factura/tracking, etc.).
  3) En el paso "aplicar": llama a `_tr_apply_carrier_status` que
     persiste el tracking, dispara una consulta best-effort a FedEx
     (vía `_fedex_track_lookup`) y deja el estado del item al día.

Reglas ILUS respetadas:
  · NO toca app.py. Se engancha vía `register_ot_routes(app, globals())`.
  · ERP Random sigue read-only (sólo se leen commitments locales en MySQL).
  · UI con helpers ilus* (no alert/confirm/prompt nativos).
  · Paleta ILUS (rojo, negro, blanco).

Endpoints:
  GET  /transporte/ot-masivo                → render UI
  GET  /transporte/ot-masivo/plantilla      → descarga plantilla .xlsx
  POST /transporte/ot-masivo/preview        → analiza Excel, devuelve JSON
  POST /transporte/ot-masivo/aplicar        → ejecuta los UPDATE + FedEx poll
"""
import io
import re

# Tipos de documento aceptados (mismo set que `transport_commitments.tido`).
TIDO_VALIDOS = {"FCV", "FEV", "BLV", "BEV", "GDD", "NCV", "NDV", "OCV", "OEV"}

# Estados FINALES de un item: si está en uno de estos, no se llama a FedEx
# (la API no tiene info nueva que aportar).
ESTADOS_TERMINALES = {"Entregado", "Devolución"}


def _clean_factura(v):
    """N° de factura → sólo dígitos (ej: 'FCV-77' → '77', '#10644' → '10644').

    Si no hay ningún dígito devuelve ''. Quita espacios, puntos, guiones, etc.
    """
    if v is None:
        return ""
    digits = re.sub(r"\D", "", str(v))
    return digits


def _clean_tracking(v):
    """Tracking FedEx → alfanumérico mayúsculas.

    FedEx admite trackings numéricos (12-22 dígitos típicos) y algunos casos
    alfanuméricos (12 chars). Por eso conservamos letras pero limpiamos
    espacios/separadores. Vacío si queda < 10 chars.
    """
    if v is None:
        return ""
    raw = re.sub(r"[^A-Za-z0-9]", "", str(v)).upper()
    return raw


def _clean_tido(v):
    """Tipo de doc → mayúsculas 3 chars. Lo valida _match_row según corresponda."""
    if v is None:
        return ""
    s = re.sub(r"[^A-Za-z]", "", str(v)).upper()[:3]
    return s


def _clean_obs(v):
    if v is None:
        return ""
    return str(v).strip()[:240]


def register_ot_routes(app, ctx):
    """Engancha los endpoints en `app`. `ctx` = globals() de app.py."""
    mysql_fetchall = ctx["mysql_fetchall"]
    mysql_fetchone = ctx["mysql_fetchone"]
    mysql_execute = ctx["mysql_execute"]
    jsonify = ctx["jsonify"]
    request = ctx["request"]
    render_template = ctx["render_template"]
    Response = ctx["Response"]
    _tr_required = ctx["_tr_required"]
    _tr_apply_carrier_status = ctx["_tr_apply_carrier_status"]
    _fedex_track_lookup = ctx["_fedex_track_lookup"]
    _tr_log = ctx.get("_tr_log") or (lambda *a, **kw: None)

    # ───────────────────────────────────────────────────────────────────
    # Helper interno: match de una fila contra BD.
    # Devuelve un dict con `status` + datos suficientes para que el front
    # lo pinte y el endpoint /aplicar lo procese sin re-consultar BD por
    # fila (todo viaja en el payload). Usa SELECT con tido optativo.
    # ───────────────────────────────────────────────────────────────────
    def _match_row(factura, tido, tracking, obs, tracking_to_existing):
        """Match de una fila. `tracking_to_existing` = dict cargado UNA vez
        con todos los tracking ya asignados en la BD (para detectar duplicados
        en O(1) sin hacer N queries).
        """
        if not factura or not tracking:
            return {
                "factura": factura, "tido": tido, "ot": tracking, "obs": obs,
                "status": "invalido",
                "msg": "Falta n° de factura o tracking.",
            }
        if len(tracking) < 10:
            return {
                "factura": factura, "tido": tido, "ot": tracking, "obs": obs,
                "status": "invalido",
                "msg": f"Tracking inválido (largo {len(tracking)}, mínimo 10).",
            }

        # Buscar commitment con (nudo=factura) + tido si vino. Si no vino,
        # tomamos el más reciente (suele ser único — Daniel no factura el
        # mismo nudo en distintos tipos el mismo día).
        if tido and tido in TIDO_VALIDOS:
            comm = mysql_fetchone(
                "SELECT id, tido, nudo, cliente_nombre "
                "FROM transport_commitments "
                "WHERE nudo=%s AND tido=%s "
                "ORDER BY fecha_emision DESC, id DESC LIMIT 1",
                (factura, tido)
            )
        else:
            comm = mysql_fetchone(
                "SELECT id, tido, nudo, cliente_nombre "
                "FROM transport_commitments "
                "WHERE nudo=%s "
                "ORDER BY fecha_emision DESC, id DESC LIMIT 1",
                (factura,)
            )
        if not comm:
            return {
                "factura": factura, "tido": tido, "ot": tracking, "obs": obs,
                "status": "no_encontrada",
                "msg": "La factura no existe en el monitor de Transporte. "
                       "Sincronizá desde ERP primero.",
            }

        # Buscar item del manifiesto correspondiente a este commitment. Si
        # hay varios (re-manifiesto), elegimos el más reciente.
        item = mysql_fetchone(
            "SELECT id, tracking_number, estado_entrega "
            "FROM transport_manifest_items "
            "WHERE commitment_id=%s "
            "ORDER BY added_at DESC, id DESC LIMIT 1",
            (comm["id"],)
        )
        if not item:
            return {
                "factura": factura, "tido": comm["tido"], "ot": tracking, "obs": obs,
                "status": "no_encontrada",
                "commitment_id": comm["id"],
                "cliente": comm.get("cliente_nombre") or "",
                "msg": "La factura existe pero no está en ningún manifiesto. "
                       "Agregala a un manifiesto antes de asignar OT.",
            }

        # ── Conflicto: el tracking ya está en OTRO item ──
        existing_item_id = tracking_to_existing.get(tracking)
        if existing_item_id and existing_item_id != item["id"]:
            return {
                "factura": factura, "tido": comm["tido"], "ot": tracking, "obs": obs,
                "status": "ot_duplicada",
                "commitment_id": comm["id"],
                "item_id": item["id"],
                "cliente": comm.get("cliente_nombre") or "",
                "msg": (f"El tracking {tracking} ya está asignado a otra "
                        f"factura/item (item #{existing_item_id})."),
            }

        # ── ¿Tiene tracking actual? ──
        tn_actual = (item.get("tracking_number") or "").strip().upper()
        cliente = comm.get("cliente_nombre") or ""
        if not tn_actual:
            return {
                "factura": factura, "tido": comm["tido"], "ot": tracking, "obs": obs,
                "status": "ok_nuevo",
                "commitment_id": comm["id"],
                "item_id": item["id"],
                "cliente": cliente,
                "msg": "Listo para asignar.",
            }
        if tn_actual == tracking:
            return {
                "factura": factura, "tido": comm["tido"], "ot": tracking, "obs": obs,
                "status": "ya_existe",
                "commitment_id": comm["id"],
                "item_id": item["id"],
                "cliente": cliente,
                "msg": "El item ya tiene este mismo tracking. Se omite.",
            }
        # Distinto → re-envío
        return {
            "factura": factura, "tido": comm["tido"], "ot": tracking, "obs": obs,
            "status": "ok_re_envio",
            "commitment_id": comm["id"],
            "item_id": item["id"],
            "cliente": cliente,
            "ot_anterior": tn_actual,
            "msg": (f"El item ya tenía la OT {tn_actual}. "
                    f"Sobrescribir requiere confirmación del operador."),
        }

    # ───────────────────────────────────────────────────────────────────
    # GET /transporte/ot-masivo  → render UI
    # ───────────────────────────────────────────────────────────────────
    @app.route("/transporte/ot-masivo", methods=["GET"])
    @_tr_required
    def tr_ot_masivo_ui():
        return render_template("transporte/ot_masivo.html")

    # ───────────────────────────────────────────────────────────────────
    # GET /transporte/ot-masivo/plantilla → descarga .xlsx
    # ───────────────────────────────────────────────────────────────────
    @app.route("/transporte/ot-masivo/plantilla", methods=["GET"])
    @_tr_required
    def tr_ot_masivo_plantilla():
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return "openpyxl no instalado", 500
        wb = Workbook()
        ws = wb.active
        ws.title = "OT FedEx"

        headers = [
            "N° Factura",
            "Tipo (FCV/BLV/...)",
            "N° Tracking FedEx (OT)",
            "Observación opcional",
        ]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="DC2626")
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Fila de ejemplo (Daniel la borra antes de subir)
        ws.cell(row=2, column=1, value="10644")
        ws.cell(row=2, column=2, value="FCV")
        ws.cell(row=2, column=3, value="780123456789")
        ws.cell(row=2, column=4, value="Re-envío por dirección errónea")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 38

        ws2 = wb.create_sheet("Instrucciones")
        notas = [
            "CARGA MASIVA DE OT FEDEX — ILUS",
            "",
            "Cómo usar esta plantilla:",
            "",
            "1. N° Factura es OBLIGATORIO. Solo el número (sin prefijos), ej: 10644.",
            "2. Tipo (FCV/BLV/...) es OPCIONAL pero recomendado si tienes el mismo",
            "   número usado en distintos tipos de documento.",
            "3. N° Tracking FedEx es OBLIGATORIO. Pegá tal cual lo da FedEx",
            "   (mínimo 10 caracteres, sin espacios).",
            "4. Observación es opcional — texto libre (ej: 'Re-envío').",
            "",
            "Antes de subir: borra la fila de ejemplo (#10644).",
            "",
            "Al subir, el sistema te muestra una vista previa color-coded:",
            "  Verde   → ok_nuevo (factura sin OT previa, listo para asignar)",
            "  Azul    → ok_re_envio (factura ya tenía OTRA OT, es re-envío)",
            "  Gris    → ya_existe (idéntico al actual, se omite)",
            "  Ámbar   → no_encontrada (la factura no está en BD)",
            "  Rojo    → ot_duplicada / invalido (revisar antes de aplicar)",
            "",
            "Solo las filas verdes/azules confirmadas se aplican.",
        ]
        for i, txt in enumerate(notas, 1):
            c = ws2.cell(row=i, column=1, value=txt)
            if i == 1:
                c.font = Font(bold=True, size=14, color="DC2626")
        ws2.column_dimensions["A"].width = 80

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                    'attachment; filename="plantilla_ot_masivo.xlsx"'
            },
        )

    # ───────────────────────────────────────────────────────────────────
    # POST /transporte/ot-masivo/preview  → analiza Excel, NO aplica
    # ───────────────────────────────────────────────────────────────────
    @app.route("/transporte/ot-masivo/preview", methods=["POST"])
    @_tr_required
    def tr_ot_masivo_preview():
        f = request.files.get("archivo")
        if not f or not f.filename:
            return jsonify({"ok": False, "error": "No se envió archivo"}), 400
        ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
        if ext not in ("xlsx", "xls"):
            return jsonify({
                "ok": False,
                "error": "Formato no permitido. Usá .xlsx",
            }), 400
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({
                "ok": False,
                "error": "openpyxl no instalado",
            }), 500
        try:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            return jsonify({
                "ok": False,
                "error": f"No se pudo leer el Excel: {e}",
            }), 400

        # Pre-cargar mapa tracking → item_id (1 sola query, evita N+1 en preview).
        tracking_to_existing = {}
        try:
            rows_tn = mysql_fetchall(
                "SELECT id, tracking_number FROM transport_manifest_items "
                "WHERE tracking_number IS NOT NULL AND tracking_number <> ''"
            )
            for r in (rows_tn or []):
                tn = (r.get("tracking_number") or "").strip().upper()
                if tn:
                    tracking_to_existing[tn] = r["id"]
        except Exception:
            tracking_to_existing = {}

        filas = []
        # Detectar si la primera fila parece header (texto, no número).
        # Si A1 incluye "factura", "n°" o "tracking" la saltamos.
        start_row = 1
        try:
            a1 = ws.cell(row=1, column=1).value
            if a1 and isinstance(a1, str) and (
                "factur" in a1.lower() or "n°" in a1.lower()
                or "tracking" in a1.lower()
            ):
                start_row = 2
        except Exception:
            start_row = 2  # default seguro

        seen_in_file = {}  # tracking → primera fila donde apareció (intra-archivo)

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=True),
            start=start_row,
        ):
            if not row or all(c is None or str(c).strip() == "" for c in row[:4]):
                continue
            factura = _clean_factura(row[0] if len(row) > 0 else None)
            tido = _clean_tido(row[1] if len(row) > 1 else None)
            tracking = _clean_tracking(row[2] if len(row) > 2 else None)
            obs = _clean_obs(row[3] if len(row) > 3 else None)

            # Dup dentro del mismo archivo: mismo tracking en 2 filas distintas
            if tracking and tracking in seen_in_file:
                filas.append({
                    "row": row_idx,
                    "factura": factura, "tido": tido, "ot": tracking, "obs": obs,
                    "status": "ot_duplicada",
                    "msg": (f"El tracking {tracking} aparece en 2 filas del "
                            f"mismo archivo (fila {seen_in_file[tracking]} y "
                            f"esta). Revisá antes de aplicar."),
                })
                continue
            if tracking:
                seen_in_file[tracking] = row_idx

            res = _match_row(factura, tido, tracking, obs, tracking_to_existing)
            res["row"] = row_idx
            filas.append(res)

        # Resumen
        resumen = {
            "total": len(filas),
            "ok_nuevo":      sum(1 for r in filas if r["status"] == "ok_nuevo"),
            "ok_re_envio":   sum(1 for r in filas if r["status"] == "ok_re_envio"),
            "ya_existe":     sum(1 for r in filas if r["status"] == "ya_existe"),
            "no_encontrada": sum(1 for r in filas if r["status"] == "no_encontrada"),
            "ot_duplicada":  sum(1 for r in filas if r["status"] == "ot_duplicada"),
            "invalido":      sum(1 for r in filas if r["status"] == "invalido"),
        }
        resumen["aplicables_directas"] = resumen["ok_nuevo"]
        resumen["aplicables_re_envio"] = resumen["ok_re_envio"]
        resumen["requieren_revision"] = (
            resumen["no_encontrada"] + resumen["ot_duplicada"] + resumen["invalido"]
        )

        return jsonify({
            "ok": True,
            "filas": filas,
            "resumen": resumen,
        })

    # ───────────────────────────────────────────────────────────────────
    # POST /transporte/ot-masivo/aplicar  → ejecuta los UPDATE + FedEx poll
    # ───────────────────────────────────────────────────────────────────
    @app.route("/transporte/ot-masivo/aplicar", methods=["POST"])
    @_tr_required
    def tr_ot_masivo_aplicar():
        body = request.get_json(silent=True, force=True) or {}
        filas_in = body.get("filas") or []
        confirm_re_envio = bool(body.get("confirm_re_envio"))

        if not filas_in:
            return jsonify({"ok": False, "error": "No hay filas para aplicar"}), 400

        # Filtrar las que NO se deben aplicar.
        aplicar = []
        omitidas = []
        for f in filas_in:
            st = (f.get("status") or "").strip()
            if st == "ok_nuevo":
                aplicar.append(f)
            elif st == "ok_re_envio":
                if confirm_re_envio:
                    aplicar.append(f)
                else:
                    omitidas.append({**f, "omit_reason": "re_envio_sin_confirmar"})
            elif st == "ya_existe":
                omitidas.append({**f, "omit_reason": "ya_existe"})
            else:
                # no_encontrada / ot_duplicada / invalido NUNCA se aplican
                omitidas.append({**f, "omit_reason": f"status_{st}"})

        if not aplicar:
            return jsonify({
                "ok": True,
                "aplicadas": 0,
                "omitidas": len(omitidas),
                "detalle_omitidas": omitidas,
                "errores": [],
                "mensaje": "No había filas aplicables. Revisá el preview.",
            })

        # ── Re-validar contra BD justo antes de UPDATE (race-safe).
        # Cargamos el mapa tracking→item actualizado.
        tracking_to_existing = {}
        try:
            rows_tn = mysql_fetchall(
                "SELECT id, tracking_number FROM transport_manifest_items "
                "WHERE tracking_number IS NOT NULL AND tracking_number <> ''"
            )
            for r in (rows_tn or []):
                tn = (r.get("tracking_number") or "").strip().upper()
                if tn:
                    tracking_to_existing[tn] = r["id"]
        except Exception:
            tracking_to_existing = {}

        # Trackings que sí vamos a aplicar (para llamar FedEx en lote luego).
        items_to_fedex = []  # lista de (item_id, tracking, obs)
        aplicadas = 0
        errores = []

        for f in aplicar:
            item_id = f.get("item_id")
            tracking = (f.get("ot") or "").strip().upper()
            obs = (f.get("obs") or "").strip()
            factura = f.get("factura") or ""

            if not item_id or not tracking:
                errores.append({
                    "factura": factura, "ot": tracking,
                    "error": "item_id o tracking faltante",
                })
                continue

            # Re-check: si el tracking ya está en OTRO item (race), abortar fila.
            existing = tracking_to_existing.get(tracking)
            if existing and existing != item_id:
                errores.append({
                    "factura": factura, "ot": tracking,
                    "error": f"Race: tracking ya en item #{existing}",
                })
                continue

            try:
                # Persistir el tracking primero (idempotente). Si FedEx falla
                # más adelante, el tracking igual queda guardado para que el
                # cron lo procese.
                mysql_execute(
                    "UPDATE transport_manifest_items SET tracking_number=%s "
                    "WHERE id=%s",
                    (tracking, item_id),
                )
                aplicadas += 1
                tracking_to_existing[tracking] = item_id
                items_to_fedex.append((item_id, tracking, obs))
                try:
                    _tr_log(
                        "manifest_item", item_id,
                        "ot_masiva",
                        f"tracking={tracking} factura={factura} "
                        f"{'(re-envío)' if f.get('status') == 'ok_re_envio' else ''}".strip(),
                    )
                except Exception:
                    pass
            except Exception as e:
                errores.append({
                    "factura": factura, "ot": tracking,
                    "error": str(e)[:200],
                })

        # ── Llamar a FedEx en lote (hasta 30 trackings por request).
        # Best-effort: si falla la API el tracking ya quedó guardado.
        fedex_aplicado = 0
        fedex_fallo = False
        fedex_warning = ""
        if items_to_fedex:
            # Map para localizar item por tracking
            by_tn = {tn: (iid, obs) for (iid, tn, obs) in items_to_fedex}
            tns = list(by_tn.keys())
            # Llamar en chunks de 30
            for i in range(0, len(tns), 30):
                chunk = tns[i:i + 30]
                try:
                    results = _fedex_track_lookup(chunk)
                except Exception as e:
                    fedex_fallo = True
                    fedex_warning = (
                        f"FedEx Track API no respondió ({str(e)[:120]}). "
                        f"Los trackings quedaron guardados; el cron los "
                        f"actualizará automáticamente."
                    )
                    continue
                for r in (results or []):
                    tn = (r.get("tracking_number") or "").strip().upper()
                    if tn not in by_tn:
                        continue
                    iid, obs = by_tn[tn]
                    comentario = (
                        r.get("last_event") or r.get("status_label") or obs or ""
                    )
                    try:
                        _tr_apply_carrier_status(
                            iid, r["estado_ilus"], fuente="fedex",
                            tracking_number=tn,
                            payload={"fedex": {
                                "status_code":  r.get("status_code"),
                                "status_label": r.get("status_label"),
                                "eta":          r.get("eta"),
                                "scans":        (r.get("scans") or [])[:10],
                            }},
                            comentario=(comentario or None),
                        )
                        fedex_aplicado += 1
                    except Exception as e:
                        errores.append({
                            "ot": tn,
                            "error": f"_tr_apply_carrier_status: {str(e)[:200]}",
                        })

        return jsonify({
            "ok": True,
            "aplicadas": aplicadas,
            "omitidas": len(omitidas),
            "detalle_omitidas": omitidas,
            "errores": errores,
            "fedex_consultado": fedex_aplicado,
            "fedex_warning": fedex_warning if fedex_fallo else "",
            "mensaje": (
                f"✓ {aplicadas} OT(s) asignada(s). "
                f"{fedex_aplicado} estado(s) actualizado(s) vía FedEx."
                if aplicadas else
                "No se aplicó ninguna fila."
            ),
        })
