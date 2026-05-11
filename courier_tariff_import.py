"""
Importador de tarifas de couriers desde Excel.

Estructura esperada del Excel (Libro1.xlsx):
  - Una hoja por courier (FedEx, Clickex, Felca, Milling, etc.)
  - Cada hoja tiene columnas: Codigo · Sucursal · Comuna · [CP/Zona] · Días · 1..99 (kg) · rangos
  - O variante Clickex: Región · Destino · Días · días-semana · 0.5 · 1..N · rangos

Genera registros en `transport_courier_comunas` con `precios_json` que es un
dict {key: precio} donde key puede ser entero ("1", "2", ..., "99") o rango
("100-499", "500-1999", etc.) — el cual _courier_tarifa_lookup() ya entiende.

Uso:
    from courier_tariff_import import import_excel_to_db
    result = import_excel_to_db('/path/to/Libro1.xlsx', mysql_conn)
    print(result)  # → {'FedEx': 372, 'Clickex': 761, 'Felca': 303, 'Milling': 303}
"""

from __future__ import annotations

import json
import logging
import re
from typing import Optional, Iterable

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger("courier.import")


# Columnas que pueden ser "comuna" según el courier
COMUNA_COL_CANDIDATES = ("comuna", "destino", "comuna destino")
SUCURSAL_COL_CANDIDATES = ("sucursal", "destino")
ZONA_COL_CANDIDATES = ("zona", "región", "region", "codigos postales")
CODIGO_COL_CANDIDATES = ("codigo", "código", "code")
DIAS_COL_CANDIDATES = ("dias transito", "días tránsito", "días", "dias")
DIAS_ENTREGA_COL_CANDIDATES = ("lu", "ma", "mi", "ju", "vi", "sa", "do")


def _normalize_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip().lower()


def _is_peso_header(h: str) -> tuple[bool, Optional[str]]:
    """Determina si un header es de peso. Devuelve (es_peso, key_normalizada).

    Ejemplos:
      "1"           → (True, "1")
      "0,5" / "0.5" → (True, "0.5")
      "100 al 499"  → (True, "100-499")
      "500 - 1999"  → (True, "500-1999")
      "Comuna"      → (False, None)
    """
    if not h:
        return False, None
    s = str(h).strip().lower().replace(",", ".")
    # Caso simple: número entero o decimal
    if re.match(r"^\d+(\.\d+)?$", s):
        return True, s
    # Caso rango: "100 al 499", "500 - 1999", "10001 - +", "10001+"
    m = re.match(r"^(\d+)\s*(?:al|-|a|–|—)\s*(\d+|\+)$", s)
    if m:
        a, b = m.group(1), m.group(2)
        return True, f"{a}-{b}"
    # Caso rango con "+" al final
    m = re.match(r"^(\d+)\s*\+$", s)
    if m:
        return True, f"{m.group(1)}+"
    return False, None


def _normalize_comuna(name: str) -> str:
    """Normaliza nombre de comuna para matching. Quita acentos, lower, trim."""
    if not name:
        return ""
    s = str(name).strip()
    # ¥ → Ñ (codificación CP437 del ERP Random)
    s = s.replace("¥", "Ñ").replace("\xa5", "Ñ")
    return s


def _parse_sheet(ws) -> list[dict]:
    """Parsea una hoja Excel y devuelve list de dicts con:
        {comuna, sucursal, zona, codigo, dias, precios: {peso_key: precio}}
    """
    if ws.max_row < 2:
        return []

    # Leer headers
    headers_raw = [c.value for c in ws[1]]
    headers = [_normalize_header(h) for h in headers_raw]

    # Identificar columnas semánticas
    col_comuna = None
    col_sucursal = None
    col_zona = None
    col_codigo = None
    col_dias = None

    for idx, h in enumerate(headers):
        if not h:
            continue
        if col_comuna is None and any(c in h for c in COMUNA_COL_CANDIDATES):
            col_comuna = idx
        if col_sucursal is None and any(h == c for c in SUCURSAL_COL_CANDIDATES):
            col_sucursal = idx
        if col_zona is None and any(c in h for c in ZONA_COL_CANDIDATES):
            col_zona = idx
        if col_codigo is None and any(c in h for c in CODIGO_COL_CANDIDATES):
            col_codigo = idx
        if col_dias is None and any(c == h for c in DIAS_COL_CANDIDATES):
            col_dias = idx

    # Si no hay comuna pero hay destino (Clickex), usar destino como comuna
    if col_comuna is None and col_sucursal is not None:
        col_comuna = col_sucursal
        col_sucursal = None

    if col_comuna is None:
        logger.warning("Hoja %s: no se encontró columna comuna/destino", ws.title)
        return []

    # Identificar columnas de pesos
    peso_cols: list[tuple[int, str]] = []
    for idx, h in enumerate(headers):
        is_p, key = _is_peso_header(headers_raw[idx])  # usar header original
        if is_p:
            peso_cols.append((idx, key))

    if not peso_cols:
        logger.warning("Hoja %s: no se encontraron columnas de peso", ws.title)
        return []

    logger.info(
        "Hoja %s: %d columnas de peso detectadas (rango %s..%s)",
        ws.title, len(peso_cols), peso_cols[0][1], peso_cols[-1][1]
    )

    records: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col_comuna] is None:
            continue
        comuna = _normalize_comuna(row[col_comuna])
        if not comuna or comuna.lower() in ("destino", "comuna"):
            continue
        sucursal = row[col_sucursal] if col_sucursal is not None and len(row) > col_sucursal else None
        zona = row[col_zona] if col_zona is not None and len(row) > col_zona else None
        codigo = row[col_codigo] if col_codigo is not None and len(row) > col_codigo else None
        dias = row[col_dias] if col_dias is not None and len(row) > col_dias else None

        # Construir dict de precios
        precios: dict[str, float] = {}
        for col_idx, peso_key in peso_cols:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                # Algunos valores vienen como "≠" o "-" para indicar no aplica
                if isinstance(val, str):
                    val = val.strip()
                    if val in ("", "≠", "-", "—", "N/A", "n/a"):
                        continue
                    # Quitar separadores de miles si están como string
                    val = val.replace(".", "").replace(",", ".")
                    val_f = float(val)
                else:
                    val_f = float(val)
                if val_f > 0:
                    precios[peso_key] = round(val_f, 2)
            except (ValueError, TypeError):
                continue

        if not precios:
            continue

        records.append({
            "comuna":   str(comuna)[:120],
            "sucursal": str(sucursal)[:120] if sucursal else None,
            "zona":     str(zona)[:80] if zona else None,
            "codigo":   str(codigo)[:20] if codigo else None,
            "dias":     str(dias)[:20] if dias else None,
            "precios":  precios,
        })
    return records


def import_excel_to_db(file_path: str, conn, default_factor_vol: int = 6000) -> dict:
    """Importa el Excel completo. Una hoja = un courier.

    Args:
      file_path: ruta al .xlsx
      conn: conexión MySQL (DBUtils PooledDB connection)
      default_factor_vol: divisor para peso volumétrico (5000 o 6000)

    Returns:
      dict {nombre_courier: filas_importadas}
    """
    if not openpyxl:
        raise RuntimeError("openpyxl no instalado. pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    result: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        records = _parse_sheet(ws)
        if not records:
            result[sheet_name] = 0
            continue

        # Crear o actualizar el courier
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM transport_couriers WHERE LOWER(nombre)=LOWER(%s) LIMIT 1",
                (sheet_name,)
            )
            row = cur.fetchone()
            if row:
                courier_id = row["id"] if isinstance(row, dict) else row[0]
            else:
                cur.execute(
                    """INSERT INTO transport_couriers
                       (nombre, tipo, activo, factor_vol)
                       VALUES (%s, %s, 1, %s)""",
                    (sheet_name, "nacional", default_factor_vol)
                )
                courier_id = cur.lastrowid

            # Limpiar tarifas antiguas de este courier
            cur.execute(
                "DELETE FROM transport_courier_comunas WHERE courier_id=%s",
                (courier_id,)
            )

            # Insertar nuevas
            inserted = 0
            for r in records:
                try:
                    cur.execute(
                        """INSERT INTO transport_courier_comunas
                           (courier_id, codigo, sucursal, comuna, zona,
                            dias_transito, precios_json)
                           VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                        (
                            courier_id,
                            r["codigo"], r["sucursal"], r["comuna"], r["zona"],
                            r["dias"],
                            json.dumps(r["precios"], ensure_ascii=False),
                        )
                    )
                    inserted += 1
                except Exception as e:
                    logger.warning(
                        "Error insertando %s/%s: %s",
                        sheet_name, r["comuna"], e
                    )

        conn.commit()
        result[sheet_name] = inserted
        logger.info("Courier %s: %d comunas importadas", sheet_name, inserted)

    return result
