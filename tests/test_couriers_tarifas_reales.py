"""Los tarifarios de Couriers no reflejaban la realidad.

BUG REAL (2026-08-10). Daniel adjuntó el Excel que se descarga desde
Couriers → "···" → Exportar tarifas y dijo: "esto no es la realidad, no
está Shipit y hay transportes que nada que ver... necesito que solo me
entregues los transportes que están activos como lo son Transportes Felca,
Milliing, FedEx, Shipit y dejaría a Clickex por fuera nomás hasta que nos
compartan la API."

Y sobre los botones de descarga: "al presionar la tarifa solo entregue la
tarifa del courrier consultado y no todas... cuando entres al courier en la
pestaña de tarifas también se pueda bajar la tarifa por esa sección, pero
que sea individual, no unificado. El unificado es en la parte superior
derecha."

TRES BUGS REALES, confirmados leyendo el código:

  1. transporte_couriers_export ("Exportar tarifas", el consolidado) arma
     una hoja por courier a partir de transport_courier_comunas. Shipit no
     tiene filas ahí (es un agregador que cotiza en vivo, no tiene tabla
     negociada) así que quedaba afuera EN SILENCIO -- ni un aviso, ni una
     hoja vacía, directamente no aparecía. Coincide exacto con "no está
     Shipit".

  2. tr_tarifario_comparado_xlsx ("Comparar tarifas") tenía la lista de
     couriers a comparar FIJA en el código (Clickex/Felca/Milling/FedEx),
     sin mirar el interruptor Habilitado/Deshabilitado de cada courier --
     desactivar Clickex desde la tarjeta no lo sacaba de este archivo.

  3. El botón "Exportar Excel" DENTRO de la ficha de un courier (pestaña
     Tarifas) apuntaba al endpoint CONSOLIDADO (/transporte/couriers/export)
     en vez del tarifario de ESE courier -- estando parado en la ficha de
     FedEx, el botón bajaba la tarifa de TODOS. Esto es justo lo que
     describe "al presionar la tarifa solo entregue la tarifa del courrier
     consultado y no todas".

UN CUARTO BUG, real y más grave, encontrado por el propio review adversarial
de este fix (2026-08-10): el primer intento de arreglar el bug 1 introdujo
un NameError -- _sh_llenar_hoja_tarifario_shipit usaba Font(...) y
Alignment(...) sin importarlos en su propio scope (el import vivía dentro de
transporte_couriers_export, y un import local a una función NO se propaga a
una función hermana en Python). El try/except de transporte_couriers_export
tragaba la excepción en silencio y volvía a dejar a Shipit afuera del
Excel -- el MISMO síntoma ("no está Shipit"), ahora por una causa nueva.
Los tests estáticos (ast.unparse + assertIn) NO lo detectaron porque nunca
EJECUTAN la función. TestHelperShipitSeEjecutaDeVerdad de acá abajo la
ejecuta de verdad (con las dependencias externas mockeadas) para que este
tipo de bug no pueda volver a colarse en silencio.

Correr con:  py -m unittest tests.test_couriers_tarifas_reales -v
"""
import ast
import sys
import types
import unittest
from unittest import mock

APP_SRC = open("app.py", encoding="utf-8", errors="ignore").read()
FICHA_HTML = open("templates/transporte/courier_ficha.html", encoding="utf-8", errors="ignore").read()
_ARBOL = ast.parse(APP_SRC)


def _fn(nombre):
    for n in ast.walk(_ARBOL):
        if isinstance(n, ast.FunctionDef) and n.name == nombre:
            return ast.unparse(n)
    raise AssertionError(f"no se encontro {nombre} en app.py")


class TestHelperShipitSeEjecutaDeVerdad(unittest.TestCase):
    """Ejecuta _sh_llenar_hoja_tarifario_shipit de VERDAD (no solo la
    parsea) contra una hoja real de openpyxl, con las dependencias externas
    (MySQL, la API de Shipit, config del remitente) mockeadas. Esto es lo
    único que hubiera atrapado el NameError real de Font/Alignment que el
    review adversarial encontró: los asserts sobre texto fuente nunca
    ejecutan la función."""

    def _cargar_helper(self):
        import openpyxl  # dependencia real del proyecto, no se mockea

        fn_src = _fn("_sh_llenar_hoja_tarifario_shipit")

        fake_shipit_client = types.ModuleType("shipit_client")
        import shipit_client as _shc_real
        fake_shipit_client.MAX_PESO_KG = _shc_real.MAX_PESO_KG
        fake_shipit_client.EP_RATES = _shc_real.EP_RATES
        fake_shipit_client.build_rate_payload = _shc_real.build_rate_payload
        fake_shipit_client.parse_rates_response = _shc_real.parse_rates_response

        respuesta_shipit = {
            "ok": True,
            "data": {"prices": [{
                "courier": {"name": "chilexpress"},
                "name": "Standard",
                "price": 4990,
                "days": 2,
                "available_to_shipping": True,
                "destiny": {"commune_id": 1},
            }]},
        }

        ns = {
            "mysql_fetchall": mock.Mock(return_value=[
                {"comuna": "Santiago", "n": 10},
                {"comuna": "Providencia", "n": 5},
            ]),
            "_shipit_commune_id": mock.Mock(return_value=(1, "ok")),
            "_tr_sender_cfg": mock.Mock(return_value={"city": "Quilicura"}),
            "_shipit_request": mock.Mock(return_value=respuesta_shipit),
            "_xlsx_cell": lambda v: v,
            "chile_fmt_filter": lambda dt, fmt: "10/08/2026 12:00",
            "datetime": __import__("datetime").datetime,
        }
        # shipit_client se importa DENTRO de la función real con
        # "import shipit_client as _shc" -- para no depender de la red ni de
        # un token real, se inyecta el modulo fake en sys.modules antes de
        # ejecutar, y se restaura despues (evita ensuciar otros tests).
        sys.modules["shipit_client"] = fake_shipit_client
        try:
            exec(compile(ast.parse(fn_src), "<_sh_llenar_hoja_tarifario_shipit>", "exec"), ns)
        finally:
            sys.modules["shipit_client"] = _shc_real
        return ns["_sh_llenar_hoja_tarifario_shipit"], openpyxl

    def test_no_lanza_nameerror_y_devuelve_true_con_datos(self):
        """Este es el test que hubiera atrapado el bug real: Font/Alignment
        sin importar en el scope de la función lanzaban NameError acá."""
        helper, openpyxl = self._cargar_helper()
        wb = openpyxl.Workbook()
        ws = wb.active
        from openpyxl.styles import PatternFill, Font as _RealFont
        red = PatternFill("solid", fgColor="CC0000")
        blanco = _RealFont(color="FFFFFF", bold=True)

        resultado = helper(ws, "Shipit", red, blanco)

        self.assertTrue(resultado, "debería haber escrito al menos una fila con datos")
        self.assertIn("Shipit", str(ws.cell(1, 1).value))
        # La cabecera (fila 5) debe tener el formato aplicado -- confirma
        # que Font/Alignment realmente se usaron, no solo que no reventó.
        self.assertTrue(ws.cell(5, 1).font.bold)
        self.assertEqual(ws.cell(5, 1).alignment.horizontal, "center")

    def test_respeta_pesos_y_limite_explicitos(self):
        """tr_tarifario_de_un_courier_xlsx le pasa pesos/limite desde la
        query string -- confirma que el helper realmente los usa en vez de
        los valores por defecto."""
        helper, openpyxl = self._cargar_helper()
        wb = openpyxl.Workbook()
        ws = wb.active
        from openpyxl.styles import PatternFill, Font as _RealFont
        red = PatternFill("solid", fgColor="CC0000")
        blanco = _RealFont(color="FFFFFF", bold=True)

        helper(ws, "Shipit", red, blanco, pesos=[2, 7], limite=1)

        encabezados = [ws.cell(5, ci).value for ci in range(1, 8) if ws.cell(5, ci).value]
        self.assertIn("2 kg", encabezados)
        self.assertIn("7 kg", encabezados)
        self.assertNotIn("1 kg", encabezados)


class TestExportConsolidadoIncluyeShipit(unittest.TestCase):
    """transporte_couriers_export: el bug real "no está Shipit"."""

    def test_llama_al_helper_de_shipit_en_vivo(self):
        f = _fn("transporte_couriers_export")
        self.assertIn("_sh_llenar_hoja_tarifario_shipit(", f)
        self.assertIn("'shipit' in", f)

    def test_sigue_filtrando_solo_couriers_activos(self):
        """No se toca lo que ya estaba bien: el consolidado ya filtraba
        activo=1, eso no era el bug."""
        f = _fn("transporte_couriers_export")
        self.assertIn("WHERE activo=1", f)

    def test_no_revienta_si_shipit_no_logra_cotizar(self):
        """Si la API de Shipit falla, el resto del Excel debe seguir
        armandose -- un try/except alrededor del helper, no un crash de
        toda la descarga."""
        f = _fn("transporte_couriers_export")
        self.assertIn("except Exception", f)


class TestHelperShipitCompartido(unittest.TestCase):
    """_sh_llenar_hoja_tarifario_shipit: cotiza en vivo, no toca tablas."""

    def test_existe_y_devuelve_booleano_de_exito(self):
        f = _fn("_sh_llenar_hoja_tarifario_shipit")
        self.assertIn("return huvo_datos", f)

    def test_usa_la_api_real_de_shipit_no_una_tabla(self):
        f = _fn("_sh_llenar_hoja_tarifario_shipit")
        self.assertIn("_shipit_request(", f)
        self.assertIn("EP_RATES", f)
        self.assertNotIn("FROM transport_courier_comunas", f)


class TestComparadorRespetaElInterruptorActivo(unittest.TestCase):
    """tr_tarifario_comparado_xlsx: ya no hardcodea Clickex/Felca/Milling/FedEx."""

    def test_ya_no_hay_lista_fija_de_couriers(self):
        f = _fn("tr_tarifario_comparado_xlsx")
        self.assertNotIn('[("clickex", "Clickex")', f)

    def test_arma_la_lista_desde_los_couriers_activos(self):
        f = _fn("tr_tarifario_comparado_xlsx")
        self.assertIn("WHERE activo=1", f)
        self.assertIn("slug_para_courier(", f)

    def test_shipit_sigue_aparte_no_por_la_tabla_fija(self):
        """Shipit se cotiza en vivo en su propia columna -- slug_para_courier
        no debe intentar meterlo en SLUGS (no tiene tabla que cargar)."""
        f = _fn("tr_tarifario_comparado_xlsx")
        i = f.find("_couriers_activos = mysql_fetchall")
        self.assertGreater(i, 0)
        bloque = f[i:i + 700]
        self.assertNotIn('"shipit"', bloque)


class TestTarifarioIndividualUsaElHelperCompartido(unittest.TestCase):
    """El review adversarial encontró que el docstring del helper decía
    'compartida por tr_tarifario_de_un_courier_xlsx' pero esa función tenía
    su PROPIA copia independiente de la cotización de Shipit -- riesgo real
    de que un fix futuro (timeout, pesos, bug de la API) se aplique a una
    copia y no a la otra. Se refactorizó para que de verdad comparta."""

    def test_llama_al_helper_en_vez_de_duplicar_la_cotizacion(self):
        f = _fn("tr_tarifario_de_un_courier_xlsx")
        self.assertIn("_sh_llenar_hoja_tarifario_shipit(ws, nombre, RED, BLANCO", f)
        self.assertNotIn("ThreadPoolExecutor", f,
            "quedó una copia propia de la cotización de Shipit en vez de usar el helper")

    def test_conserva_la_validacion_http_de_query_params(self):
        """La validación de 'pesos'/'limite' con sus 400 propios debe seguir
        viviendo en la ruta (donde tiene sentido devolver un error HTTP),
        no en el helper compartido (que también usa el consolidado, sin
        query string)."""
        f = _fn("tr_tarifario_de_un_courier_xlsx")
        self.assertIn("Parámetro 'pesos' inválido", f)
        self.assertIn("request.args.get('pesos')", f)


class TestBotonDeLaFichaEsIndividual(unittest.TestCase):
    """El botón de descarga DENTRO de la ficha (pestaña Tarifas) debe bajar
    SOLO el tarifario de ese courier -- el consolidado vive en Couriers →
    "···" → Exportar tarifas, no acá."""

    def test_ya_no_apunta_al_export_consolidado(self):
        i = FICHA_HTML.find('id="tab-tarifas"')
        j = FICHA_HTML.find('id="tab-logos"')
        self.assertGreater(i, 0)
        self.assertGreater(j, i)
        bloque = FICHA_HTML[i:j]
        self.assertNotIn("'/transporte/couriers/export'", bloque)

    def test_apunta_al_tarifario_de_este_courier(self):
        i = FICHA_HTML.find('id="tab-tarifas"')
        j = FICHA_HTML.find('id="tab-logos"')
        bloque = FICHA_HTML[i:j]
        self.assertIn("/transporte/couriers/{{ courier.id }}/tarifario.xlsx", bloque)


class TestBotonDeLaTarjetaYaEraIndividual(unittest.TestCase):
    """Guarda de no-regresión: el botón "Excel" de la TARJETA (couriers.html)
    ya llamaba al tarifario individual desde el 2026-08-05 -- confirmar que
    sigue así y que no se duplicó sin querer."""

    def test_tarjeta_sigue_usando_el_endpoint_individual(self):
        with open("templates/transporte/couriers.html", encoding="utf-8", errors="ignore") as fh:
            html = fh.read()
        self.assertIn("descargarTarifarioCourier(", html)
        i = html.find("function descargarTarifarioCourier")
        self.assertGreater(i, 0)
        bloque = html[i:i + 400]
        self.assertIn("/tarifario.xlsx", bloque)


if __name__ == "__main__":
    unittest.main(verbosity=2)
